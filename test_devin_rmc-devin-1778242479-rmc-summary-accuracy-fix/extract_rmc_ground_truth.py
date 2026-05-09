"""
Extract RMC Summary ground truth from the manually filled workbook.
Reads ALL columns (A-CJ, 88 cols) from 'RMC summary' sheet and dumps
them into a JSON + prints a detailed analysis.
"""
import sys, os, json
sys.path.insert(0, r'c:\Users\mohamad.al\OneDrive - AL KHAYYAT INVESTMENTS\Desktop\IPP')

import openpyxl
from openpyxl.utils import get_column_letter

base = r'c:\Users\mohamad.al\OneDrive - AL KHAYYAT INVESTMENTS\Desktop\IPP'
filled_path = os.path.join(base, 'Files_need_to_study', 'Filled_Output', '1 Base RMC _ 2026 February.xlsx')
unfilled_path = os.path.join(base, 'Files_need_to_study', 'Unfilled', '1 Base RMC _ 2026 February.xlsx')

print(f"Loading FILLED workbook: {os.path.getsize(filled_path):,} bytes...", flush=True)
wb_f = openpyxl.load_workbook(filled_path, data_only=True, read_only=True)
ws_f = wb_f['RMC summary']

# ═══════════════════════════════════════════════════════════════
# STEP 1: Extract ALL headers (rows 2-4)
# ═══════════════════════════════════════════════════════════════
print("\n=== HEADER ROWS (rows 2-4) ===", flush=True)
max_col = 88  # Based on analysis: 88 columns
headers = {}
for row in [2, 3, 4]:
    for c in range(1, max_col + 1):
        v = ws_f.cell(row=row, column=c).value
        if v is not None:
            cl = get_column_letter(c)
            headers[(row, c)] = str(v).strip()
            print(f"  [{row},{cl}({c})]: {v}", flush=True)

# ═══════════════════════════════════════════════════════════════
# STEP 2: Detect what each column header is at row 4
# ═══════════════════════════════════════════════════════════════
print("\n=== COLUMN MAP (row 4 headers) ===", flush=True)
col_headers = {}
for c in range(1, max_col + 1):
    v = ws_f.cell(row=4, column=c).value
    if v is not None:
        col_headers[c] = str(v).strip()
        cl = get_column_letter(c)
        print(f"  Col {cl}({c}): {v}", flush=True)
    else:
        # Try row 3
        v3 = ws_f.cell(row=3, column=c).value
        if v3 is not None:
            col_headers[c] = str(v3).strip()
            cl = get_column_letter(c)
            print(f"  Col {cl}({c}): [row3] {v3}", flush=True)

# ═══════════════════════════════════════════════════════════════
# STEP 3: Extract ALL data rows
# ═══════════════════════════════════════════════════════════════
print("\n=== DATA ROWS (row 5+) ===", flush=True)
data_rows = []
row_count = 0
empty_count = 0
for r in range(5, ws_f.max_row + 1):
    order = ws_f.cell(row=r, column=2).value  # B = Order No
    if order is None:
        empty_count += 1
        if empty_count > 10:
            break
        continue
    empty_count = 0
    row_count += 1
    
    row_data = {'_row': r}
    for c in range(1, max_col + 1):
        v = ws_f.cell(row=r, column=c).value
        if v is not None:
            cl = get_column_letter(c)
            if isinstance(v, (int, float)):
                row_data[f'{cl}({c})'] = round(v, 6) if isinstance(v, float) else v
            else:
                row_data[f'{cl}({c})'] = str(v).strip()
    data_rows.append(row_data)

print(f"  Total data rows: {row_count}", flush=True)

# Print first 5 rows in detail
for i, row in enumerate(data_rows[:5]):
    print(f"\n  Row {row['_row']} (order={row.get('B(2)', '?')}):", flush=True)
    for k, v in sorted(row.items()):
        if k == '_row': continue
        print(f"    {k}: {v}", flush=True)

# ═══════════════════════════════════════════════════════════════
# STEP 4: Column-level statistics
# ═══════════════════════════════════════════════════════════════
print("\n=== COLUMN FILL STATISTICS ===", flush=True)
col_stats = {}
for c in range(1, max_col + 1):
    cl = get_column_letter(c)
    key = f'{cl}({c})'
    filled = sum(1 for row in data_rows if key in row)
    col_stats[key] = filled
    header = col_headers.get(c, '?')
    if filled > 0:
        print(f"  {key}: {filled}/{row_count} filled ({100*filled//row_count}%) — {header}", flush=True)

# ═══════════════════════════════════════════════════════════════
# STEP 5: Save detailed JSON
# ═══════════════════════════════════════════════════════════════
output = {
    'headers': {f'{get_column_letter(c)}({c})': h for (r, c), h in headers.items() if r == 4},
    'col_stats': col_stats,
    'data_rows': data_rows,
    'total_rows': row_count,
}

out_path = os.path.join(base, 'rmc_summary_ground_truth.json')
with open(out_path, 'w') as f:
    json.dump(output, f, indent=2, default=str)
print(f"\nSaved to: {out_path}", flush=True)

# ═══════════════════════════════════════════════════════════════
# STEP 6: Also extract headers from UNFILLED template
# ═══════════════════════════════════════════════════════════════
print("\n=== UNFILLED TEMPLATE RMC SUMMARY ===", flush=True)
print(f"Loading UNFILLED workbook...", flush=True)

# Use strip_template approach for faster loading
from engine.base_rmc.loaders import _strip_template_xml
from engine.base_rmc.context import RMCContext
import tempfile

tmp_ctx = RMCContext()
stripped = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False).name
_strip_template_xml(unfilled_path, stripped, tmp_ctx)

wb_u = openpyxl.load_workbook(stripped, data_only=True, read_only=True)
ws_u = wb_u['RMC summary']

unfilled_orders = []
for r in range(5, ws_u.max_row + 1):
    order = ws_u.cell(row=r, column=2).value
    if order is None:
        continue
    remarks = ws_u.cell(row=r, column=7).value
    combined = ws_u.cell(row=r, column=1).value
    design = ws_u.cell(row=r, column=3).value
    customer = ws_u.cell(row=r, column=4).value
    material = ws_u.cell(row=r, column=5).value
    structure = ws_u.cell(row=r, column=6).value
    unfilled_orders.append({
        'row': r,
        'combined_key': str(combined).strip() if combined else '',
        'order': str(order).strip(),
        'design': str(design).strip() if design else '',
        'customer': str(customer).strip() if customer else '',
        'material': str(material).strip() if material else '',
        'structure': str(structure).strip() if structure else '',
        'remarks': str(remarks).strip() if remarks else '',
    })

print(f"  Unfilled template has {len(unfilled_orders)} orders in RMC Summary", flush=True)
for o in unfilled_orders[:5]:
    print(f"    Row {o['row']}: {o['order']} — {o['remarks'][:60]}", flush=True)

# Check what's pre-filled vs what needs filling
print("\n=== PRE-FILLED vs NEEDS FILLING ===", flush=True)
for c in range(8, max_col + 1):
    cl = get_column_letter(c)
    pre_filled = 0
    for r in range(5, 5 + len(unfilled_orders)):
        v = ws_u.cell(row=r, column=c).value
        if v is not None:
            pre_filled += 1
    if pre_filled > 0:
        header = col_headers.get(c, '?')
        print(f"  {cl}({c}): {pre_filled} pre-filled — {header}", flush=True)

wb_f.close()
wb_u.close()
os.unlink(stripped)
print("\nDone!", flush=True)
