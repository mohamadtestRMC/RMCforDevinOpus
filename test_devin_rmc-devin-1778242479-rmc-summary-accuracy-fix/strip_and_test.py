"""Strip huge empty-row sheets using proper XML parsing."""
import zipfile, re, os, time
from xml.etree import ElementTree as ET

src = r'c:\Users\mohamad.al\OneDrive - AL KHAYYAT INVESTMENTS\Desktop\IPP\Files_need_to_study\Unfilled\1 Base RMC _ 2026 February.xlsx'
dst = r'c:\Users\mohamad.al\OneDrive - AL KHAYYAT INVESTMENTS\Desktop\IPP\output\template_fast.xlsx'
os.makedirs(os.path.dirname(dst), exist_ok=True)

STRIP = {
    'xl/worksheets/sheet26.xml': 300,
    'xl/worksheets/sheet27.xml': 300,
}
REMOVE = {'xl/calcChain.xml'}

NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'

t0 = time.time()
print('Stripping template with XML parser...', flush=True)

with zipfile.ZipFile(src, 'r') as zin, zipfile.ZipFile(dst, 'w', zipfile.ZIP_DEFLATED) as zout:
    for item in zin.infolist():
        if item.filename in REMOVE:
            print(f'  Removed: {item.filename}', flush=True)
            continue

        data = zin.read(item.filename)

        if item.filename in STRIP:
            max_row = STRIP[item.filename]
            before_len = len(data)

            # Parse XML properly
            root = ET.fromstring(data)

            # Fix dimension element
            dim = root.find(f'{{{NS}}}dimension')
            if dim is not None:
                ref = dim.get('ref', '')
                # Change "A1:K1048576" to "A1:K300"
                parts = ref.split(':')
                if len(parts) == 2:
                    col_part = ''.join(c for c in parts[1] if c.isalpha())
                    dim.set('ref', f'{parts[0]}:{col_part}{max_row}')

            # Find sheetData and remove rows > max_row
            sheet_data = root.find(f'{{{NS}}}sheetData')
            if sheet_data is not None:
                rows_to_remove = []
                for row_elem in sheet_data.findall(f'{{{NS}}}row'):
                    r_num = int(row_elem.get('r', '0'))
                    if r_num > max_row:
                        rows_to_remove.append(row_elem)
                for row_elem in rows_to_remove:
                    sheet_data.remove(row_elem)
                print(f'  Removed {len(rows_to_remove)} rows from {item.filename}', flush=True)

            # Serialize back with namespace declarations
            ET.register_namespace('', NS)
            ET.register_namespace('r', 'http://schemas.openxmlformats.org/officeDocument/2006/relationships')
            ET.register_namespace('mc', 'http://schemas.openxmlformats.org/markup-compatibility/2006')
            ET.register_namespace('x14ac', 'http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac')
            data = ET.tostring(root, xml_declaration=True, encoding='UTF-8')
            print(f'  Stripped {item.filename}: {before_len:,} -> {len(data):,} bytes', flush=True)

        zout.writestr(item, data)

elapsed = time.time() - t0
print(f'Strip done in {elapsed:.1f}s', flush=True)
print(f'  Original: {os.path.getsize(src):,} bytes', flush=True)
print(f'  Stripped: {os.path.getsize(dst):,} bytes', flush=True)

# Load test
import openpyxl
print('\nLoading stripped template...', flush=True)
t1 = time.time()
wb = openpyxl.load_workbook(dst, keep_links=False)
elapsed2 = time.time() - t1
print(f'Loaded in {elapsed2:.1f}s!', flush=True)
for sn in ['OPN_WIP', 'CLS_WIP', 'Jobtrack', 'Print', 'Lam', 'BFL', 'Slit', 'RMC summary']:
    if sn in wb.sheetnames:
        ws = wb[sn]
        print(f'  {sn}: max_row={ws.max_row}, max_col={ws.max_column}', flush=True)
wb.close()
print('SUCCESS!', flush=True)
