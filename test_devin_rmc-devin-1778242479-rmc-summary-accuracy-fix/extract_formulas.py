"""Extract FORMULAS (not values) from every sheet of the filled Base RMC.
This is the key to reverse-engineering every cell."""
import openpyxl
import os
from openpyxl.utils import get_column_letter

base = r'c:\Users\mohamad.al\OneDrive - AL KHAYYAT INVESTMENTS\Desktop\IPP\Files_need_to_study'
filled = os.path.join(base, 'Filled_Output', '1 Base RMC _ 2026 February.xlsx')

# Open WITHOUT data_only to see formulas
wb = openpyxl.load_workbook(filled, data_only=False)
# Also open WITH data_only to see computed values
wb_val = openpyxl.load_workbook(filled, data_only=True)

output_lines = []

def log(s):
    output_lines.append(s)
    print(s)

# Analyze each sheet
for sn in wb.sheetnames:
    ws = wb[sn]
    ws_val = wb_val[sn]
    log(f"\n{'='*80}")
    log(f"SHEET: {sn} (rows={ws.max_row}, cols={ws.max_column})")
    log(f"{'='*80}")
    
    # For very large sheets, limit analysis
    max_r = min(ws.max_row or 0, 15)  # first 15 rows for structure
    max_c = min(ws.max_column or 0, 90)
    
    # Special handling for key sheets - show more rows
    if sn in ('RMC summary', 'FG', 'Spout&Valve', 'Embossing', 'HCI Rew',
              'Overall Wastage - Process Wise', 'Pivot_Lam Rates'):
        max_r = min(ws.max_row or 0, 25)
    
    for r in range(1, max_r + 1):
        row_cells = []
        for c in range(1, max_c + 1):
            cell = ws.cell(row=r, column=c)
            val = cell.value
            val_computed = ws_val.cell(row=r, column=c).value
            
            if val is not None:
                col_letter = get_column_letter(c)
                if isinstance(val, str) and val.startswith('='):
                    # It's a formula - show both formula and value
                    row_cells.append(f'{col_letter}={val[:80]} → {repr(val_computed)[:40]}')
                else:
                    row_cells.append(f'{col_letter}={repr(val)[:50]}')
        
        if row_cells:
            log(f"  Row {r}: {' | '.join(row_cells[:20])}")
            if len(row_cells) > 20:
                log(f"         + {' | '.join(row_cells[20:40])}")
                if len(row_cells) > 40:
                    log(f"         + {' | '.join(row_cells[40:60])}")
                    if len(row_cells) > 60:
                        log(f"         + {' | '.join(row_cells[60:])}")

wb.close()
wb_val.close()

# Save output
out_path = os.path.join(os.path.dirname(base), 'formula_extraction.txt')
with open(out_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(output_lines))
print(f"\nSaved to {out_path}")
