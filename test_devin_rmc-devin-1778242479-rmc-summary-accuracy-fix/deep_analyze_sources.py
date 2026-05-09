"""Analyze Purchase Register, Stores, Granules, Components, Ink Consumption."""
import openpyxl
import os

base = r'c:\Users\mohamad.al\OneDrive - AL KHAYYAT INVESTMENTS\Desktop\IPP\Files_need_to_study'

def show_sheet(ws, name, max_rows=10, max_cols=20):
    print(f'\n  --- Sheet: "{name}" (rows={ws.max_row}, cols={ws.max_column}) ---')
    for r in range(1, min((ws.max_row or 0)+1, max_rows+1)):
        row_data = []
        for c in range(1, min((ws.max_column or 0)+1, max_cols+1)):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                row_data.append(f'C{c}={repr(val)[:50]}')
        if row_data:
            print(f'    Row {r}: {"; ".join(row_data[:15])}')

# ===== PURCHASE REGISTER =====
print('='*80)
print('PURCHASE REGISTER - UNFILLED')
print('='*80)
fp = os.path.join(base, 'Unfilled', '2 Purchase Register - 2021 - 2026 _Feb 26.xlsx')
wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
print(f'  SHEETS: {wb.sheetnames}')
for sn in wb.sheetnames[:3]:
    show_sheet(wb[sn], sn, max_rows=6, max_cols=20)
wb.close()

# ===== STORES / RM FILM STOCK =====
print('\n' + '='*80)
print('RM FILM STOCK MAIN FILE - UNFILLED')
print('='*80)
fp = os.path.join(base, 'Unfilled', '3 RM FILM STOCK MAIN FILE - WORKING - 2026.xlsx')
wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
print(f'  SHEETS: {wb.sheetnames}')
for sn in wb.sheetnames[:3]:
    show_sheet(wb[sn], sn, max_rows=6, max_cols=20)
wb.close()

# ===== GRANULES RECIPE =====
print('\n' + '='*80)
print('GRANULES RECIPE - UNFILLED (first sheet)')
print('='*80)
fp = os.path.join(base, 'Unfilled', '4 Granules Recipe - February 2026.xlsx')
wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
print(f'  SHEETS: {wb.sheetnames[:5]}...')
sn = wb.sheetnames[0]
show_sheet(wb[sn], sn, max_rows=10, max_cols=15)
wb.close()

# ===== INK CONSUMPTION =====
print('\n' + '='*80)
print('INK CONSUMPTION - UNFILLED')
print('='*80)
fp = os.path.join(base, 'Unfilled', '5 Ink Consumption February 2026.xlsx')
wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
print(f'  SHEETS: {wb.sheetnames}')
for sn in wb.sheetnames:
    show_sheet(wb[sn], sn, max_rows=6, max_cols=15)
wb.close()

# ===== COMPONENTS CONSUMPTION =====
print('\n' + '='*80)
print('COMPONENTS CONSUMPTIONS DISPENSED DETAILS - UNFILLED')
print('='*80)
fp = os.path.join(base, 'Unfilled', '12 Components Consumptions Dispensed Details.xlsx')
wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
print(f'  SHEETS: {wb.sheetnames}')
for sn in wb.sheetnames[:3]:
    show_sheet(wb[sn], sn, max_rows=8, max_cols=15)
wb.close()

# ===== OPENING WIP =====
print('\n' + '='*80)
print('OPENING WIP STOCK - UNFILLED')
print('='*80)
fp = os.path.join(base, 'Unfilled', '9 Opening WIP Stock.xlsx')
wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
print(f'  SHEETS: {wb.sheetnames}')
for sn in wb.sheetnames:
    show_sheet(wb[sn], sn, max_rows=8, max_cols=15)
wb.close()

# ===== CLOSING WIP =====
print('\n' + '='*80)
print('CLOSING WIP STOCK - UNFILLED')
print('='*80)
fp = os.path.join(base, 'Unfilled', '10 Closing WIP Stock.xlsx')
wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
print(f'  SHEETS: {wb.sheetnames}')
for sn in wb.sheetnames:
    show_sheet(wb[sn], sn, max_rows=8, max_cols=15)
wb.close()
