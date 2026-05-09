"""Analyze all Excel files in Files_need_to_study for reverse engineering."""
import openpyxl
import os

base = r'c:\Users\mohamad.al\OneDrive - AL KHAYYAT INVESTMENTS\Desktop\IPP\Files_need_to_study'

for folder in ['Unfilled', 'Filled_Output', 'How_to_link']:
    fpath = os.path.join(base, folder)
    if not os.path.exists(fpath):
        continue
    print(f'\n{"="*60}')
    print(f' {folder}')
    print(f'{"="*60}')
    for f in sorted(os.listdir(fpath)):
        if f.endswith('.xlsx'):
            fp = os.path.join(fpath, f)
            try:
                wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
                sheets = wb.sheetnames
                print(f'\n  FILE: {f}')
                print(f'  SIZE: {os.path.getsize(fp):,} bytes')
                print(f'  SHEETS ({len(sheets)}): {sheets}')
                for sn in sheets:
                    ws = wb[sn]
                    print(f'    Sheet "{sn}": rows={ws.max_row}, cols={ws.max_column}')
                wb.close()
            except Exception as e:
                print(f'\n  FILE: {f} -- ERROR: {e}')
