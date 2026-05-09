import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell.read_only import EmptyCell
import sys

file_path = r"c:\Users\mohamad.al\OneDrive - AL KHAYYAT INVESTMENTS\Desktop\IPP\Files_need_to_study\Unfilled\1 Base RMC _ 2026 February.xlsx"

print(f"Opening: {file_path}")
sys.stdout.flush()

wb = openpyxl.load_workbook(file_path, data_only=False, read_only=True)

print(f"File opened successfully!")
print(f"Total sheets: {len(wb.sheetnames)}")
print(f"Sheet names: {wb.sheetnames}")
print("=" * 100)
sys.stdout.flush()

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'#' * 100}")
    print(f"## SHEET: '{sheet_name}'")
    print(f"{'#' * 100}")
    sys.stdout.flush()

    max_rows_to_show = 35
    formulas_found = []
    row_count = 0
    col_fill_counts = {}
    total_filled = 0
    total_empty = 0
    max_col_seen = 0
    
    print(f"\n  --- CELL DATA (first {max_rows_to_show} rows) ---")
    sys.stdout.flush()
    
    for row in ws.iter_rows():
        row_count += 1
        row_data = []
        col_idx = 0
        for cell in row:
            col_idx += 1
            if isinstance(cell, EmptyCell):
                col_letter = get_column_letter(col_idx)
                if col_letter not in col_fill_counts:
                    col_fill_counts[col_letter] = {'filled': 0, 'empty': 0, 'formulas': 0}
                col_fill_counts[col_letter]['empty'] += 1
                total_empty += 1
                continue
            
            try:
                actual_col = cell.column
            except:
                actual_col = col_idx
            
            col_letter = get_column_letter(actual_col)
            if actual_col > max_col_seen:
                max_col_seen = actual_col
            
            if col_letter not in col_fill_counts:
                col_fill_counts[col_letter] = {'filled': 0, 'empty': 0, 'formulas': 0}
            
            val = cell.value
            if val is not None:
                col_fill_counts[col_letter]['filled'] += 1
                total_filled += 1
                cell_ref = f"{col_letter}{row_count}"
                if isinstance(val, str) and val.startswith('='):
                    col_fill_counts[col_letter]['formulas'] += 1
                    if len(formulas_found) < 100:
                        formulas_found.append((cell_ref, val))
                
                if row_count <= max_rows_to_show:
                    display_val = repr(val)
                    if len(display_val) > 70:
                        display_val = display_val[:70] + "..."
                    row_data.append(f"{cell_ref}={display_val}")
            else:
                col_fill_counts[col_letter]['empty'] += 1
                total_empty += 1
        
        if row_count <= max_rows_to_show:
            if row_data:
                print(f"    Row {row_count}: {' | '.join(row_data)}")
            else:
                print(f"    Row {row_count}: [EMPTY ROW]")
    
    print(f"\n  --- SHEET STATS ---")
    print(f"    Total rows: {row_count}, Max columns: {max_col_seen}")
    print(f"    Total filled cells: {total_filled}, Total empty cells: {total_empty}")
    
    print(f"\n  --- COLUMN FILL ANALYSIS ---")
    for col_letter in sorted(col_fill_counts.keys(), key=lambda x: column_index_from_string(x)):
        info = col_fill_counts[col_letter]
        status = "EMPTY_COL" if info['filled'] == 0 else ("ALL_FILLED" if info['empty'] == 0 else "PARTIAL")
        print(f"    Col {col_letter}: filled={info['filled']}, empty={info['empty']}, formulas={info['formulas']} [{status}]")
    
    if formulas_found:
        print(f"\n  --- FORMULAS FOUND ({len(formulas_found)} shown) ---")
        for cell_ref, formula in formulas_found[:60]:
            print(f"    {cell_ref}: {formula[:150]}")
        if len(formulas_found) > 60:
            print(f"    ... and more ...")
    else:
        print(f"\n  --- NO FORMULAS FOUND ---")
    
    sys.stdout.flush()

wb.close()
print("\n\n" + "=" * 100)
print("ANALYSIS COMPLETE")
print("=" * 100)
sys.stdout.flush()
