"""Check suppliers for each failing row."""
import openpyxl

BASE = "Template_Files"

# Build MRR->Supplier map from Stores
wb_st = openpyxl.load_workbook(f"{BASE}/Stores Recordings.xlsx", data_only=True)
ws_st = wb_st.active
mrr_sup = {}
for r in range(3, ws_st.max_row + 1):
    mrr = ws_st.cell(row=r, column=16).value
    sup = str(ws_st.cell(row=r, column=6).value or "").strip().upper()
    if mrr and sup:
        try:
            mrr_sup[int(float(mrr))] = sup
        except:
            pass
wb_st.close()

# Check the specific rows that FAIL in our test
fail_rows = [9, 42, 43, 45, 46, 54]

wb = openpyxl.load_workbook(f"{BASE}/Jobtrack Feb With MRR.xlsx", data_only=True)
ws = wb.active

for r in fail_rows:
    uid = ws.cell(row=r, column=1).value
    order = ws.cell(row=r, column=11).value
    process = str(ws.cell(row=r, column=6).value or "")

    for label, mr_col, rate_col, name_col in [
        ("Film", 54, 55, 47),
        ("Fresh1", 78, 79, 71),
        ("Fresh2", 88, 89, 81),
    ]:
        mr = ws.cell(row=r, column=mr_col).value
        rate = ws.cell(row=r, column=rate_col).value
        name = ws.cell(row=r, column=name_col).value
        if not mr or not rate:
            continue

        suppliers = []
        if str(mr) == "INH":
            suppliers.append("INH (in-house)")
        else:
            for ms in str(mr).split("/"):
                try:
                    mn = int(float(ms.strip()))
                    s = mrr_sup.get(mn, "UNKNOWN")
                    suppliers.append(f"{mn}={s}")
                except:
                    pass

        sup_str = ", ".join(suppliers)
        print(f"Row {r}: {label}, UID={uid}, Order={order}, Material={name}, "
              f"MR={mr}, Rate={rate}, Suppliers=[{sup_str}]")

wb.close()
