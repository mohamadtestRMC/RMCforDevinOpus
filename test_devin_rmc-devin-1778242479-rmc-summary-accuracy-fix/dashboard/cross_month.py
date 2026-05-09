"""
Cross-Month Validation — Compare Jan vs Feb RMC outputs for consistency.
"""
from __future__ import annotations
import logging
import pandas as pd
import openpyxl
import io
from typing import Dict, List, Any, Optional

logger = logging.getLogger(__name__)


def _sf(v):
    try: return float(v) if v is not None else 0.0
    except: return 0.0


def load_rmc_summary(file_bytes, label: str = "Month") -> pd.DataFrame:
    """Load RMC Summary from a filled Base RMC workbook."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = None
    for n in ['RMC summary', 'RMC Summary']:
        if n in wb.sheetnames:
            ws = wb[n]; break
    if not ws:
        wb.close()
        return pd.DataFrame()

    rows = []
    for r in range(7, ws.max_row + 1):
        order = ws.cell(row=r, column=2).value
        if not order: continue
        rows.append({
            'Order': str(order).strip(),
            'Design': str(ws.cell(row=r, column=3).value or ''),
            'OPN_WIP_Val': _sf(ws.cell(row=r, column=17).value),
            'Print_Film_Val': _sf(ws.cell(row=r, column=18).value),
            'Lam_Fresh_Val': _sf(ws.cell(row=r, column=19).value),
            'Slit_Val': _sf(ws.cell(row=r, column=20).value),
            'Ink_Val': _sf(ws.cell(row=r, column=21).value),
            'Lam_Chem_Val': _sf(ws.cell(row=r, column=22).value),
            'BP_SV_Val': _sf(ws.cell(row=r, column=23).value),
            'CLS_WIP_Val': _sf(ws.cell(row=r, column=24).value),
            'FG_Output': _sf(ws.cell(row=r, column=25).value),
            'Total_Cost': _sf(ws.cell(row=r, column=26).value),
            'RMC_Kg': _sf(ws.cell(row=r, column=27).value),
        })
    wb.close()
    df = pd.DataFrame(rows)
    df['_month'] = label
    return df


def compare_months(df_a: pd.DataFrame, df_b: pd.DataFrame,
                   label_a: str = "Jan", label_b: str = "Feb") -> Dict[str, Any]:
    """Compare two months' RMC Summary data."""
    result = {
        'label_a': label_a, 'label_b': label_b,
        'summary_a': {}, 'summary_b': {},
        'common_orders': [], 'new_orders': [], 'dropped_orders': [],
        'deltas': [], 'rate_drift': [],
        'consistency_score': 0.0,
    }

    if df_a.empty or df_b.empty:
        return result

    orders_a = set(df_a['Order'].str.upper())
    orders_b = set(df_b['Order'].str.upper())
    common = orders_a & orders_b
    new = orders_b - orders_a
    dropped = orders_a - orders_b

    result['common_orders'] = sorted(common)
    result['new_orders'] = sorted(new)
    result['dropped_orders'] = sorted(dropped)

    # Summary stats per month
    for label, df in [(label_a, df_a), (label_b, df_b)]:
        result[f'summary_{label.lower()[:1]}'] = {
            'orders': len(df),
            'total_cost': df['Total_Cost'].sum(),
            'total_output': df['FG_Output'].sum(),
            'avg_rmc': df['Total_Cost'].sum() / max(df['FG_Output'].sum(), 1),
            'orders_with_rmc': (df['RMC_Kg'] > 0).sum(),
        }

    # Per-order deltas for common orders
    val_cols = ['Total_Cost', 'FG_Output', 'RMC_Kg', 'Print_Film_Val',
                'Lam_Fresh_Val', 'Ink_Val', 'Lam_Chem_Val']
    deltas = []
    consistent = 0

    for order in sorted(common):
        row_a = df_a[df_a['Order'].str.upper() == order].iloc[0]
        row_b = df_b[df_b['Order'].str.upper() == order].iloc[0]
        delta = {'Order': order}
        has_big_diff = False

        for col in val_cols:
            va = float(row_a.get(col, 0))
            vb = float(row_b.get(col, 0))
            diff = vb - va
            pct = (diff / va * 100) if va != 0 else (100 if vb != 0 else 0)
            delta[f'{col}_{label_a}'] = round(va, 2)
            delta[f'{col}_{label_b}'] = round(vb, 2)
            delta[f'{col}_delta'] = round(diff, 2)
            delta[f'{col}_pct'] = round(pct, 1)
            if abs(pct) > 10 and abs(diff) > 1:
                has_big_diff = True

        if not has_big_diff:
            consistent += 1
        deltas.append(delta)

    result['deltas'] = deltas
    result['consistency_score'] = (consistent / max(len(common), 1)) * 100

    # Rate drift: avg material rates per month
    rate_cols = ['Print_Film_Val', 'Lam_Fresh_Val', 'Ink_Val', 'Lam_Chem_Val']
    for col in rate_cols:
        avg_a = df_a[df_a[col] > 0][col].mean() if (df_a[col] > 0).any() else 0
        avg_b = df_b[df_b[col] > 0][col].mean() if (df_b[col] > 0).any() else 0
        if avg_a > 0:
            drift_pct = (avg_b - avg_a) / avg_a * 100
            result['rate_drift'].append({
                'category': col.replace('_', ' '),
                f'avg_{label_a}': round(avg_a, 2),
                f'avg_{label_b}': round(avg_b, 2),
                'drift_pct': round(drift_pct, 1),
            })

    return result
