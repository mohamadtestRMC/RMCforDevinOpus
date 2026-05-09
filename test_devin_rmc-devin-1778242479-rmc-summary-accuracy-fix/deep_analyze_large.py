"""Deep analysis of the large critical files - headers & sample data."""
import openpyxl
import os

base = r'c:\Users\mohamad.al\OneDrive - AL KHAYYAT INVESTMENTS\Desktop\IPP\Files_need_to_study'

def analyze_sheet_detail(ws, sheet_name, max_rows=10, max_cols=30):
    print(f'\n  --- Sheet: "{sheet_name}" (rows={ws.max_row}, cols={ws.max_column}) ---')
    actual_cols = min(ws.max_column or 0, max_cols)
    actual_rows = min(ws.max_row or 0, max_rows)
    for r in range(1, actual_rows + 1):
        row_data = []
        for c in range(1, actual_cols + 1):
            cell = ws.cell(row=r, column=c)
            val = cell.value
            if val is not None:
                row_data.append(f'C{c}={repr(val)[:50]}')
        if row_data:
            print(f'    Row {r}: {"; ".join(row_data[:20])}')

# ========== 1. BASE RMC UNFILLED ==========
print('='*80)
print('1. BASE RMC - UNFILLED (Key sheets only)')
print('='*80)
fp = os.path.join(base, 'Unfilled', '1 Base RMC _ 2026 February.xlsx')
wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
print(f'  ALL SHEETS: {wb.sheetnames}')
# Analyze key sheets
for sn in ['BFL', 'Slit', 'Lam', 'Print', 'RMC summary', 'Printing Work',
           'Bag&Pouch', 'Embossing', 'Spout&Valve', 'PTR Rew', 'HCI Rew',
           'Jobtrack', 'FG', 'Overall Wastage - Process Wise', 'OPN_WIP', 'CLS_WIP']:
    if sn in wb.sheetnames:
        analyze_sheet_detail(wb[sn], sn, max_rows=8, max_cols=25)
wb.close()

# ========== 2. BASE RMC FILLED ==========
print('\n' + '='*80)
print('2. BASE RMC - FILLED (Key sheets - compare structure)')
print('='*80)
fp = os.path.join(base, 'Filled_Output', '1 Base RMC _ 2026 February.xlsx')
wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
print(f'  ALL SHEETS: {wb.sheetnames}')
for sn in ['BFL', 'Slit', 'Lam', 'Print', 'RMC summary', 'Printing Work',
           'Bag&Pouch', 'Embossing', 'Spout&Valve', 'PTR Rew', 'HCI Rew',
           'Jobtrack', 'FG', 'Overall Wastage - Process Wise']:
    if sn in wb.sheetnames:
        analyze_sheet_detail(wb[sn], sn, max_rows=8, max_cols=25)
wb.close()
