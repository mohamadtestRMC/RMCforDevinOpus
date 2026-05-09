import openpyxl
import os

base = r'c:\Users\mohamad.al\OneDrive - AL KHAYYAT INVESTMENTS\Desktop\IPP\Files_need_to_study'

path = os.path.join(base, 'Filled_Output', '1 Base RMC _ 2026 February.xlsx')
wb = openpyxl.load_workbook(path, data_only=False, read_only=True)

print("ALL SHEET NAMES:")
for i, sn in enumerate(wb.sheetnames):
    print(f"  {i}: {sn}")

print("\n\n" + "="*80)
print("RMC SUMMARY SHEET - FULL DUMP")
print("="*80)

ws = wb['RMC summary']
for ri, row in enumerate(ws.iter_rows(max_col=30, values_only=False), 1):
    vals = []
    for cell in row:
        v = cell.value
        if v is not None:
            vals.append(f"C{cell.column}={repr(v)[:120]}")
    if vals:
        print(f"Row {ri}: {vals}")
    if ri > 200:
        print("... truncated at row 200")
        break

wb.close()

# Now do the same for unfilled
print("\n\n" + "="*80)
print("UNFILLED RMC - Sheet names")
print("="*80)
path2 = os.path.join(base, 'Unfilled', '1 Base RMC _ 2026 February.xlsx')
wb2 = openpyxl.load_workbook(path2, data_only=False, read_only=True)
print("ALL SHEET NAMES:")
for i, sn in enumerate(wb2.sheetnames):
    print(f"  {i}: {sn}")

if 'RMC summary' in wb2.sheetnames:
    print("\nRMC Summary in UNFILLED:")
    ws2 = wb2['RMC summary']
    for ri, row in enumerate(ws2.iter_rows(max_col=30, values_only=False), 1):
        vals = []
        for cell in row:
            v = cell.value
            if v is not None:
                vals.append(f"C{cell.column}={repr(v)[:120]}")
        if vals:
            print(f"Row {ri}: {vals}")
        if ri > 200:
            break
wb2.close()

# Also check reference doc
print("\n\n" + "="*80)
print("REFERENCE DOC - How_to_link")
print("="*80)
path3 = os.path.join(base, 'How_to_link', 'Base RMC Documents (2).xlsx')
wb3 = openpyxl.load_workbook(path3, data_only=False, read_only=True)
print("ALL SHEET NAMES:")
for i, sn in enumerate(wb3.sheetnames):
    print(f"  {i}: {sn}")

for sn in wb3.sheetnames:
    ws3 = wb3[sn]
    print(f"\n--- Sheet: '{sn}' ---")
    for ri, row in enumerate(ws3.iter_rows(max_col=30, values_only=False), 1):
        vals = []
        for cell in row:
            v = cell.value
            if v is not None:
                vals.append(f"C{cell.column}={repr(v)[:150]}")
        if vals:
            print(f"Row {ri}: {vals}")
        if ri > 60:
            print("... truncated")
            break
wb3.close()
