"""Deep investigation: Check why process caches are empty/wrong."""
import sys, os, json
sys.path.insert(0, r'c:\Users\mohamad.al\OneDrive - AL KHAYYAT INVESTMENTS\Desktop\IPP')
import openpyxl

base = r'c:\Users\mohamad.al\OneDrive - AL KHAYYAT INVESTMENTS\Desktop\IPP'
engine_path = os.path.join(base, 'output', 'Base_RMC_Feb2026_FILLED.xlsx')

print("Loading engine output...", flush=True)
wb = openpyxl.load_workbook(engine_path, data_only=True, read_only=True)
print(f"Sheets: {wb.sheetnames}", flush=True)

# Check each process sheet to see if data was written
for sheet_name in ['Print', 'Lam', 'BFL', 'Slit', 'Bag&Pouch', 'FG',
                   'Spout&Valve', 'HCI Rew', 'PTR Rew', 'OPN_WIP', 'CLS_WIP']:
    if sheet_name not in wb.sheetnames:
        print(f"\n{sheet_name}: NOT FOUND", flush=True)
        continue
    ws = wb[sheet_name]
    print(f"\n{'='*50}", flush=True)
    print(f"{sheet_name}: max_row={ws.max_row}, max_col={ws.max_column}", flush=True)

    # Count rows with data in col B (order)
    data_rows = 0
    unique_orders = set()
    sample_rows = []
    for r in range(5, min(ws.max_row + 1, 2000)):
        order = ws.cell(row=r, column=2).value
        if order and str(order).strip():
            data_rows += 1
            unique_orders.add(str(order).strip().upper())
            if len(sample_rows) < 2:
                row_data = {}
                for c in range(1, min(ws.max_column + 1, 25)):
                    v = ws.cell(row=r, column=c).value
                    if v is not None:
                        row_data[c] = v
                sample_rows.append({'row': r, 'data': row_data})
    
    print(f"  Data rows: {data_rows}, Unique orders: {len(unique_orders)}", flush=True)
    for s in sample_rows:
        print(f"  Sample row {s['row']}: {s['data']}", flush=True)

    # For Print specifically, check film_value columns
    if sheet_name == 'Print':
        print(f"\n  --- Print column analysis ---", flush=True)
        # Check what's in key columns for first 5 data rows
        for r in range(5, min(5 + data_rows, 15)):
            order = ws.cell(row=r, column=2).value
            if not order: continue
            cols_of_interest = {
                'B(2)': ws.cell(row=r, column=2).value,
                'G(7)': ws.cell(row=r, column=7).value,  # input_size or input_kgs?
                'H(8)': ws.cell(row=r, column=8).value,
                'I(9)': ws.cell(row=r, column=9).value,
                'J(10)': ws.cell(row=r, column=10).value,
                'K(11)': ws.cell(row=r, column=11).value,
                'L(12)': ws.cell(row=r, column=12).value,
                'Q(17)': ws.cell(row=r, column=17).value,
                'R(18)': ws.cell(row=r, column=18).value,
            }
            print(f"    Row {r}: {cols_of_interest}", flush=True)

    # For FG, check what's in key columns
    if sheet_name == 'FG':
        print(f"\n  --- FG column analysis ---", flush=True)
        for r in range(3, min(3 + 10, ws.max_row + 1)):
            order = ws.cell(row=r, column=1).value
            if not order: continue
            cols = {}
            for c in range(1, min(ws.max_column + 1, 10)):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    cols[c] = v
            print(f"    Row {r}: {cols}", flush=True)

# Also check RMC Summary to confirm data was written
ws_rmc = wb['RMC summary']
print(f"\n{'='*50}", flush=True)
print(f"RMC Summary: checking rows 7-12", flush=True)
for r in range(7, 12):
    order = ws_rmc.cell(row=r, column=2).value
    if not order: continue
    vals = {
        'B': order,
        'R(18)': ws_rmc.cell(row=r, column=18).value,
        'S(19)': ws_rmc.cell(row=r, column=19).value,
        'U(21)': ws_rmc.cell(row=r, column=21).value,
        'Y(25)': ws_rmc.cell(row=r, column=25).value,
        'Z(26)': ws_rmc.cell(row=r, column=26).value,
    }
    print(f"  Row {r}: {vals}", flush=True)

wb.close()
print("\nDone!", flush=True)
