"""FAST: Extract formulas from RMC summary + key process sheets only."""
import openpyxl, os
from openpyxl.utils import get_column_letter

base = r'c:\Users\mohamad.al\OneDrive - AL KHAYYAT INVESTMENTS\Desktop\IPP\Files_need_to_study'
filled = os.path.join(base, 'Filled_Output', '1 Base RMC _ 2026 February.xlsx')

# Only open formula view (skip data_only to save time)
print("Loading formulas workbook...")
wb = openpyxl.load_workbook(filled, data_only=False, read_only=True)

TARGET_SHEETS = {
    'RMC summary': (1, 12, 1, 88),   # rows 1-12, all 88 cols
    'BFL': (1, 10, 1, 22),
    'Print': (1, 10, 1, 29),
    'Lam': (1, 10, 1, 85),
    'Slit': (1, 10, 1, 29),
    'Bag&Pouch': (1, 10, 1, 33),
    'Spout&Valve': (1, 11, 1, 44),
    'HCI Rew': (1, 10, 1, 20),
    'PTR Rew': (1, 10, 1, 38),
    'Embossing': (1, 10, 1, 26),
    'FG': (1, 10, 1, 15),
    'OPN_WIP': (1, 10, 1, 11),
    'CLS_WIP': (1, 10, 1, 12),
    'Printing Work': (1, 10, 1, 34),
    'Overall Wastage - Process Wise': (1, 28, 1, 24),
    'Pivot_Lam Rates': (1, 10, 1, 12),
}

for sn, (r1, r2, c1, c2) in TARGET_SHEETS.items():
    if sn not in wb.sheetnames:
        continue
    ws = wb[sn]
    print(f"\n{'='*80}")
    print(f"SHEET: {sn}")
    print(f"{'='*80}")
    
    for r in range(r1, r2 + 1):
        cells = []
        for c in range(c1, c2 + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                cl = get_column_letter(c)
                s = repr(val)[:80] if not (isinstance(val, str) and val.startswith('=')) else val[:90]
                cells.append(f"{cl}{r}={s}")
        if cells:
            # Print in chunks
            for i in range(0, len(cells), 8):
                prefix = f"  Row {r}: " if i == 0 else "         "
                print(f"{prefix}{' | '.join(cells[i:i+8])}")

wb.close()
print("\nDone!")
