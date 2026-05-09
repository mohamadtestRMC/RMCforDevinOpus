"""Quick extra diagnostics for LAM rows and balance tracing."""
import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

BASE = r"c:\Users\mohamad.al\OneDrive - AL KHAYYAT INVESTMENTS\Desktop\IPP\Template3"

jt_path = BASE + r"\Job Track Feb 26.xlsx"
filled_path = BASE + r"\Jobtrack_Filled_MRR_20260429_1929.xlsx"

wb = openpyxl.load_workbook(jt_path, data_only=True)
ws = wb.active
wb_f = openpyxl.load_workbook(jt_path)
ws_f = wb_f.active
wb_filled = openpyxl.load_workbook(filled_path, data_only=True)
ws_filled = wb_filled.active

# Row 111 is LAM - check Fresh1 columns
print("=" * 60)
print("ROW 111 - LAM FRESH COLUMNS")
print("=" * 60)
row = 111
print(f"  Fresh1 Name (BS/71): {ws.cell(row=row, column=71).value}")
print(f"  Fresh1 Size (BT/72): {ws.cell(row=row, column=72).value}")
print(f"  Fresh1 Mic (BU/73): {ws.cell(row=row, column=73).value}")
print(f"  Fresh1 Qty formula (BW/75): {ws_f.cell(row=row, column=75).value}")
print(f"  Fresh1 Qty value (BW/75): {ws.cell(row=row, column=75).value}")
print(f"  Fresh1 Balance (BX/76): {ws.cell(row=row, column=76).value}")
print(f"  Total Fresh1 (BY/77): {ws.cell(row=row, column=77).value}")
print(f"  Adhesive (CM/91): {ws.cell(row=row, column=91).value}")
print(f"  FILLED Fresh1 MR# (BZ/78): {ws_filled.cell(row=row, column=78).value}")
print(f"  FILLED Fresh1 Rate (CA/79): {ws_filled.cell(row=row, column=79).value}")
print(f"  FILLED Fresh1 Value (CB/80): {ws_filled.cell(row=row, column=80).value}")

# Row 625 - full data
print("\n" + "=" * 60)
print("ROW 625 - PRINTING - B01077 details")
print("=" * 60)
row = 625
print(f"  UID (A): {ws.cell(row=row, column=1).value}")
print(f"  Process (F): {ws.cell(row=row, column=6).value}")
print(f"  Order No (K): {ws.cell(row=row, column=11).value}")
print(f"  Input Name (AU/47): {ws.cell(row=row, column=47).value}")
print(f"  Input Size (AV/48): {ws.cell(row=row, column=48).value}")
print(f"  Input Mic (AW/49): {ws.cell(row=row, column=49).value}")
print(f"  AY formula: {ws_f.cell(row=row, column=51).value}")
print(f"  AY value: {ws.cell(row=row, column=51).value}")
print(f"  AZ formula: {ws_f.cell(row=row, column=52).value}")
print(f"  AZ value: {ws.cell(row=row, column=52).value}")
print(f"  BA value: {ws.cell(row=row, column=53).value}")
print(f"  FILLED MR# (BB): {ws_filled.cell(row=row, column=54).value}")
print(f"  FILLED Rate (BC): {ws_filled.cell(row=row, column=55).value}")

# Row 747 analysis - 365 comes from 500.1 balance
# The formula =365 means the whole qty is 365
# AZ=-135 means balance=-135
# Total = 365 + (-135) = 230
# The 365 itself - where does it come from?
# User says: "365 balance from =107+500.1 -> from 500.1"
# This means row 625 has AY=107+500.1 and AZ=-365
# So 365 is the leftover from row 625 which consumed 107 (from MRR 84526) + 500.1 (from MRR 85460)
# Since 365 < 500.1, the balance comes from MRR 85460
# Therefore row 747 should only have MRR 85460 (which it does!)
print("\n" + "=" * 60)
print("ROW 747 ANALYSIS - Balance tracing")
print("=" * 60)
print("Row 625: AY=107+500.1, AZ=-365")
print("  -> 107 matches Store MRR 84526 (Issue Qty=107)")
print("  -> 500.1 matches Store MRR 85460 (Issue Qty=500.1)")
print("  -> Balance of -365 means 365 was passed forward")
print("  -> 365 < 500.1, so it comes from MRR 85460")
print("")
print("Row 747: AY=365, AZ=-135")
print(f"  -> FILLED MR#: {ws_filled.cell(row=747, column=54).value}")
print("  -> 365 is the balance from row 625's 500.1 (MRR 85460)")
print("  -> So MRR should be 85460 ONLY (CORRECT!)")
print("")
print("Row 625 should have: 84526/85460 (both MRRs contribute)")
print(f"  -> FILLED MR#: {ws_filled.cell(row=625, column=54).value}")
print("  -> Total = 107+500.1-365 = 242.1")
print("  -> Rate should be weighted: (107*4.404 + 500.1*4.587) / (107+500.1)")

# Compute weighted rate
total_qty = 107 + 500.1
weighted = (107 * 4.404 + 500.1 * 4.587) / total_qty
print(f"  -> Weighted rate = {weighted:.6f}")

# Rows 907, 1025 - LAM rows, check fresh columns
print("\n" + "=" * 60)
print("ROWS 907, 1025 - LAM FRESH COLUMNS")
print("=" * 60)
for row in [907, 1025]:
    print(f"\nRow {row}:")
    print(f"  Fresh1 Name (BS/71): {ws.cell(row=row, column=71).value}")
    print(f"  Fresh1 Qty formula (BW/75): {ws_f.cell(row=row, column=75).value}")
    print(f"  Fresh1 Qty value (BW/75): {ws.cell(row=row, column=75).value}")
    print(f"  Fresh1 Balance (BX/76): {ws.cell(row=row, column=76).value}")
    print(f"  Fresh2 Name (CC/81): {ws.cell(row=row, column=81).value}")
    print(f"  FILLED Fresh1 MR# (BZ/78): {ws_filled.cell(row=row, column=78).value}")
    print(f"  FILLED Fresh1 Rate (CA/79): {ws_filled.cell(row=row, column=79).value}")
    print(f"  Adhesive (CM/91): {ws.cell(row=row, column=91).value}")
    print(f"  FILLED Adh Rate (CO/93): {ws_filled.cell(row=row, column=93).value}")

# Line 2350 - check why 2 MRRs when QTY=100
print("\n" + "=" * 60)
print("ROW 2350 - MRR SELECTION ANALYSIS")
print("=" * 60)
row = 2350
print(f"  Fresh1 Name: MET PET")
print(f"  Fresh1 Qty (BW): {ws_f.cell(row=row, column=75).value} = {ws.cell(row=row, column=75).value}")
print(f"  Fresh1 Balance (BX): {ws_f.cell(row=row, column=76).value}")
print(f"  Fresh1 Size (BT): {ws.cell(row=row, column=72).value}")
print(f"  Fresh1 Mic (BU): {ws.cell(row=row, column=73).value}")
print(f"  FILLED MR# (BZ): {ws_filled.cell(row=row, column=78).value}")
print(f"  FILLED Rate (CA): {ws_filled.cell(row=row, column=79).value}")
print()
print("Stores for WO=N01067 with MET PET + LAMINATION:")
print("  MRR 85226: Issue Qty=317.5 (MET PET UPF)")
print("  MRR 81732: Issue Qty=100.0 (MET PET UPF)")
print()
print("BW=100 (no formula, plain number) -> should match MRR 81732 exactly (100.0)")
print("  -> Only MRR 81732 should be selected, NOT 85226!")
print("  BUG: _pick_mrrs_by_total_qty gets both MRRs from lookup_mrr_with_qty")
print("  since it doesn't use formula matching (BW=100 is not a formula)")
print("  -> The fallback lookup returns ALL MRRs for the WO+material+process")
print("  -> _pick_mrrs_by_total_qty should pick only 81732 since its qty=100 exactly matches")

# Check what _pick_mrrs_by_total_qty would do
mrr_qty = {81732: 100.0, 85226: 317.5}
target = 100.0
print(f"\n  _pick_mrrs_by_total_qty({mrr_qty}, target={target}):")
# Single exact match check
for mrr, qty in mrr_qty.items():
    diff = abs(float(qty) - target)
    print(f"    MRR {mrr}: qty={qty}, diff from target={diff}")
    if diff <= 1.0:
        print(f"    -> EXACT MATCH! Should return [{mrr}]")

wb.close()
wb_f.close()
wb_filled.close()
print("\nDONE!")
