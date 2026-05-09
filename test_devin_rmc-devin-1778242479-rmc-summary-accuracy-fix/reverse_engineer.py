"""
REVERSE ENGINEER: For each mismatched row, figure out EXACTLY how the friend 
got his number. Try every possible combination of MRRs, quantities, and methods.
Output: Excel with step-by-step calculation trace.
"""
import io, os, shutil, openpyxl, pandas as pd, itertools
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from engine.fill_jobtrack import COLS, _safe_str
from engine.mrr_lookup import load_stores_recordings, lookup_mrr_with_qty
from engine.rate_lookup import (
    load_purchase_register, lookup_film_rate_weighted, 
    filter_mrr_by_pr, _find_col
)
from engine.supplier_rates import build_mrr_supplier_map

BASE = "Template_Files"

# Load everything
stores_df = load_stores_recordings(f"{BASE}/Stores Recordings.xlsx")
pr_df = load_purchase_register(f"{BASE}/Purchase Register - 2021 - 2026 _Feb 26.xlsx")
mrr_sup = build_mrr_supplier_map(stores_df)

# Get Issue Date info from Stores
issue_col = None
for c in stores_df.columns:
    if 'issue' in str(c).lower():
        issue_col = c
        break

# Get sub/material col from Stores
sub_col = None
for c in stores_df.columns:
    if 'sub' in str(c).lower():
        sub_col = c
        break

wo_col = None
for c in stores_df.columns:
    cl = str(c).lower()
    if 'w/o' in cl or 'work' in cl or 'w.o' in cl:
        wo_col = c
        break

mrr_col = None
for c in stores_df.columns:
    if 'm.r.r' in str(c).lower() or 'mrr' in str(c).lower():
        mrr_col = c
        break

qty_col = None
for c in stores_df.columns:
    cl = str(c).lower()
    if 'issue' in cl and ('qty' in cl or 'q' in cl) and 'date' not in cl:
        qty_col = c
        break
if not qty_col:
    for c in stores_df.columns:
        if 'qty' in str(c).lower():
            qty_col = c
            break

mic_col_st = None
for c in stores_df.columns:
    if 'mic' in str(c).lower():
        mic_col_st = c
        break

process_col = None
for c in stores_df.columns:
    if 'process' in str(c).lower():
        process_col = c
        break

print(f"Stores columns used: sub={sub_col}, wo={wo_col}, mrr={mrr_col}, qty={qty_col}, issue_date={issue_col}, mic={mic_col_st}, process={process_col}")

# PR columns
tracking_col = _find_col(pr_df, 'tracking')
rate_pr_col = [c for c in pr_df.columns if str(c).strip().lower() == 'rate'][0]
material_col = _find_col(pr_df, 'material')
month_col = _find_col(pr_df, 'month')

# Load GT
tmp_gt = f"{BASE}/_gt_rev.xlsx"
shutil.copy2(f"{BASE}/Jobtrack Feb With MRR.xlsx", tmp_gt)
wb_gt = openpyxl.load_workbook(tmp_gt, data_only=True)
ws_gt = wb_gt.active

# The 4 problem rows
rows_to_check = [
    (42, 'Fresh1', COLS['Fresh1_Name'], COLS['Fresh1_Size'], COLS['Fresh1_Mic'], 
     COLS['Fresh1_MR'], COLS['Fresh1_Rate'], 'LAMINATION'),
    (43, 'Fresh1', COLS['Fresh1_Name'], COLS['Fresh1_Size'], COLS['Fresh1_Mic'],
     COLS['Fresh1_MR'], COLS['Fresh1_Rate'], 'LAMINATION'),
    (45, 'Film', COLS['Input_Name'], COLS['Input_Size'], COLS['Input_Mic'],
     COLS['Film_MR'], COLS['Film_Rate'], 'PRINTING'),
    (46, 'Film', COLS['Input_Name'], COLS['Input_Size'], COLS['Input_Mic'],
     COLS['Film_MR'], COLS['Film_Rate'], 'PRINTING'),
]

# ══════════════════════════════════════════════════════════════
# DEEP INVESTIGATION
# ══════════════════════════════════════════════════════════════
all_row_data = []

for row, slot, name_col, size_col, mic_col, mr_col, rate_col, proc in rows_to_check:
    uid = ws_gt.cell(row=row, column=1).value
    order = str(ws_gt.cell(row=row, column=COLS['Order_No']).value or '').strip()
    mat_name = _safe_str(ws_gt.cell(row=row, column=name_col).value)
    mat_size = ws_gt.cell(row=row, column=size_col).value
    mat_mic = ws_gt.cell(row=row, column=mic_col).value
    gt_mr = str(ws_gt.cell(row=row, column=mr_col).value or '')
    gt_rate = float(ws_gt.cell(row=row, column=rate_col).value or 0)
    
    print(f"\n{'='*80}")
    print(f"ROW {row} | {uid} | {order} | {mat_name} Size={mat_size} Mic={mat_mic}")
    print(f"GT: MR={gt_mr}, Rate={gt_rate}")
    print(f"{'='*80}")
    
    # Get ALL Stores entries for this order + material (raw data)
    mrr_qty = lookup_mrr_with_qty(stores_df, mat_name, mat_mic, mat_size, order, proc)
    if not mrr_qty:
        mrr_qty = lookup_mrr_with_qty(stores_df, mat_name, mat_mic, None, order, proc)
    if not mrr_qty:
        mrr_qty = lookup_mrr_with_qty(stores_df, mat_name, mat_mic, None, order)
    
    # Get individual MRR info
    mrr_info = {}
    for m, q in mrr_qty.items():
        rate = lookup_film_rate_weighted(pr_df, {m: q}, mat_name, mat_size, mat_mic)
        if rate == 0:
            rate = lookup_film_rate_weighted(pr_df, {m: q}, mat_name, None, mat_mic)
        sup = mrr_sup.get(int(m), '?')
        
        # Get PR month
        mask = pd.to_numeric(pr_df[tracking_col], errors='coerce') == m
        pr_rows = pr_df[mask]
        pr_month = str(pr_rows.iloc[0][month_col]).strip() if not pr_rows.empty and month_col else '?'
        
        # Get Stores issue dates for this MRR
        st_mask = pd.to_numeric(stores_df[mrr_col], errors='coerce') == m
        st_rows = stores_df[st_mask]
        issue_dates = []
        if issue_col and not st_rows.empty:
            for _, sr in st_rows.iterrows():
                d = sr[issue_col]
                if pd.notna(d):
                    issue_dates.append(str(d)[:10])
        
        mrr_info[m] = {
            'qty': q, 'rate': rate, 'supplier': sup, 
            'pr_month': pr_month, 'issue_dates': issue_dates
        }
        print(f"  MRR {m}: Qty={q}, Rate={rate}, Supplier={sup}, PR_Month={pr_month}, IssueDates={issue_dates[:3]}")
    
    # TRY EVERY POSSIBLE SUBSET to find which one matches GT
    print(f"\n  Trying all subsets to match GT rate {gt_rate}...")
    mrr_keys = list(mrr_info.keys())
    best_match = None
    best_diff = 999
    
    for size in range(1, len(mrr_keys) + 1):
        for combo in itertools.combinations(mrr_keys, size):
            # Method 1: Qty-weighted average
            total_q = sum(mrr_info[m]['qty'] for m in combo)
            if total_q > 0:
                w_avg = sum(mrr_info[m]['rate'] * mrr_info[m]['qty'] for m in combo) / total_q
                diff = abs(w_avg - gt_rate)
                if diff < best_diff:
                    best_diff = diff
                    best_match = {
                        'mrrs': combo, 'method': 'weighted_avg', 'rate': w_avg,
                        'diff': diff, 'total_qty': total_q
                    }
            
            # Method 2: Simple average
            rates = [mrr_info[m]['rate'] for m in combo if mrr_info[m]['rate'] > 0]
            if rates:
                s_avg = sum(rates) / len(rates)
                diff = abs(s_avg - gt_rate)
                if diff < best_diff:
                    best_diff = diff
                    best_match = {
                        'mrrs': combo, 'method': 'simple_avg', 'rate': s_avg,
                        'diff': diff, 'total_qty': total_q
                    }
    
    if best_match:
        match_str = "EXACT" if best_match['diff'] < 0.001 else f"Close (diff={best_match['diff']:.6f})"
        print(f"\n  BEST MATCH: {match_str}")
        print(f"    MRRs: {best_match['mrrs']}")
        print(f"    Method: {best_match['method']}")
        print(f"    Computed Rate: {best_match['rate']:.6f}")
        print(f"    GT Rate:       {gt_rate:.6f}")
        
        # Show the calculation
        print(f"\n  CALCULATION:")
        for m in best_match['mrrs']:
            info = mrr_info[m]
            print(f"    MRR {m}: {info['qty']} kg × {info['rate']:.4f} AED = {info['qty'] * info['rate']:.4f}")
        if best_match['method'] == 'weighted_avg':
            num = sum(mrr_info[m]['rate'] * mrr_info[m]['qty'] for m in best_match['mrrs'])
            den = sum(mrr_info[m]['qty'] for m in best_match['mrrs'])
            print(f"    Total: {num:.4f} / {den:.1f} = {num/den:.6f}")
        
        # What rule would get this subset?
        subset_suppliers = set(mrr_info[m]['supplier'] for m in best_match['mrrs'])
        subset_months = set(mrr_info[m]['pr_month'] for m in best_match['mrrs'])
        excluded = [m for m in mrr_keys if m not in best_match['mrrs']]
        excluded_sups = set(mrr_info[m]['supplier'] for m in excluded) if excluded else set()
        excluded_months = set(mrr_info[m]['pr_month'] for m in excluded) if excluded else set()
        
        print(f"\n  PATTERN ANALYSIS:")
        print(f"    Included suppliers: {subset_suppliers}")
        print(f"    Excluded suppliers: {excluded_sups}")
        print(f"    Included PR months: {subset_months}")
        print(f"    Excluded PR months: {excluded_months}")
        
        # What's different about excluded MRRs?
        if excluded:
            print(f"    Excluded MRRs detail:")
            for m in excluded:
                info = mrr_info[m]
                print(f"      MRR {m}: Qty={info['qty']}, Rate={info['rate']}, "
                      f"Supplier={info['supplier']}, PR_Month={info['pr_month']}")
    
    # Also show what our engine computes
    filtered = filter_mrr_by_pr(pr_df, mrr_qty.copy(), mat_name, mat_size, mat_mic)
    eng_rate = lookup_film_rate_weighted(pr_df, filtered, mat_name, mat_size, mat_mic)
    if eng_rate == 0:
        eng_rate = lookup_film_rate_weighted(pr_df, mrr_qty.copy(), mat_name, None, mat_mic)
    print(f"\n  ENGINE RESULT: {eng_rate:.6f} (using {list(filtered.keys())})")
    
    all_row_data.append({
        'row': row, 'uid': uid, 'order': order, 'mat_name': mat_name,
        'mat_size': mat_size, 'mat_mic': mat_mic, 'gt_mr': gt_mr,
        'gt_rate': gt_rate, 'eng_rate': eng_rate,
        'mrr_info': mrr_info, 'best_match': best_match,
        'slot': slot
    })

wb_gt.close()
os.remove(tmp_gt)

# ══════════════════════════════════════════════════════════════
# CREATE EXCEL
# ══════════════════════════════════════════════════════════════
print("\n\nCreating Excel report...")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Gap Analysis"

# Styles
tf = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
hf = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
hfill = PatternFill('solid', fgColor='4F46E5')
gfill = PatternFill('solid', fgColor='D1FAE5')
rfill = PatternFill('solid', fgColor='FEE2E2')
yfill = PatternFill('solid', fgColor='FEF3C7')
bfill = PatternFill('solid', fgColor='DBEAFE')
df = Font(name='Calibri', size=11)
bf = Font(name='Calibri', size=11, bold=True)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_a = Alignment(horizontal='left', vertical='center', wrap_text=True)
border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1'),
)

r = 1
for rd in all_row_data:
    # Title
    ws.merge_cells(f'A{r}:J{r}')
    c = ws.cell(row=r, column=1, 
                value=f"ROW {rd['row']} — {rd['uid']} — Order {rd['order']} — "
                      f"{rd['slot']}={rd['mat_name']} Size={rd['mat_size']} Mic={rd['mat_mic']}")
    c.font = tf
    c.fill = PatternFill('solid', fgColor='1E293B')
    c.alignment = left_a
    ws.row_dimensions[r].height = 28
    r += 1
    
    # Summary row
    ws.cell(row=r, column=1, value="Ground Truth Rate:").font = bf
    ws.cell(row=r, column=2, value=rd['gt_rate']).font = Font(name='Calibri', size=13, bold=True, color='059669')
    ws.cell(row=r, column=2).fill = gfill
    ws.cell(row=r, column=3, value="Engine Rate:").font = bf
    ws.cell(row=r, column=4, value=rd['eng_rate']).font = Font(name='Calibri', size=13, bold=True, color='DC2626')
    ws.cell(row=r, column=4).fill = rfill
    ws.cell(row=r, column=5, value="GT MR#:").font = bf
    ws.cell(row=r, column=6, value=rd['gt_mr']).font = bf
    r += 1
    
    # All MRRs header
    r += 1
    headers = ['MRR #', 'Qty (kg)', 'Rate (AED)', 'Supplier', 'PR Month', 
               'Issue Dates', 'In GT?', 'Qty × Rate', 'Used by Engine?', 'Used by Friend?']
    for i, h in enumerate(headers):
        c = ws.cell(row=r, column=i+1, value=h)
        c.font = hf
        c.fill = hfill
        c.alignment = center
        c.border = border
    r += 1
    
    # GT MRR list
    gt_mrrs = set()
    if rd['gt_mr']:
        for x in str(rd['gt_mr']).split('/'):
            try:
                gt_mrrs.add(int(float(x.strip())))
            except:
                pass
    
    best_mrrs = set(rd['best_match']['mrrs']) if rd['best_match'] else set()
    
    for m in sorted(rd['mrr_info'].keys()):
        info = rd['mrr_info'][m]
        in_gt = '✓ YES' if m in gt_mrrs else '✗ NO'
        in_best = '✓ YES' if m in best_mrrs else '✗ NO'
        
        vals = [
            m, info['qty'], info['rate'], info['supplier'], info['pr_month'],
            ', '.join(info['issue_dates'][:2]) if info['issue_dates'] else '—',
            in_gt, info['qty'] * info['rate'], 
            '✓' if m in set(rd.get('filtered_mrrs', rd['mrr_info'].keys())) else '?',
            in_best
        ]
        for i, v in enumerate(vals):
            c = ws.cell(row=r, column=i+1, value=v)
            c.font = df
            c.alignment = center
            c.border = border
            if in_gt == '✓ YES':
                if i in (0, 6):
                    c.fill = gfill
            elif i == 6:
                c.fill = rfill
            if in_best == '✓ YES' and i == 9:
                c.fill = bfill
        r += 1
    
    # Calculation trace
    r += 1
    bm = rd['best_match']
    if bm:
        ws.cell(row=r, column=1, value="HOW FRIEND CALCULATED:").font = Font(name='Calibri', size=12, bold=True, color='4F46E5')
        ws.cell(row=r, column=1).fill = bfill
        r += 1
        
        ws.cell(row=r, column=1, value="Method:").font = bf
        ws.cell(row=r, column=2, value=bm['method'].replace('_', ' ').title()).font = bf
        ws.cell(row=r, column=2).fill = yfill
        
        match_label = "EXACT MATCH" if bm['diff'] < 0.001 else f"Close (diff={bm['diff']:.6f})"
        ws.cell(row=r, column=3, value="Accuracy:").font = bf
        ws.cell(row=r, column=4, value=match_label).font = bf
        ws.cell(row=r, column=4).fill = gfill if bm['diff'] < 0.001 else yfill
        r += 1
        
        # Step by step
        ws.cell(row=r, column=1, value="Step").font = hf
        ws.cell(row=r, column=1).fill = hfill
        ws.cell(row=r, column=2, value="MRR").font = hf
        ws.cell(row=r, column=2).fill = hfill
        ws.cell(row=r, column=3, value="Qty").font = hf
        ws.cell(row=r, column=3).fill = hfill
        ws.cell(row=r, column=4, value="× Rate").font = hf
        ws.cell(row=r, column=4).fill = hfill
        ws.cell(row=r, column=5, value="= Amount").font = hf
        ws.cell(row=r, column=5).fill = hfill
        r += 1
        
        total_num = 0
        total_den = 0
        for idx, m in enumerate(bm['mrrs']):
            info = rd['mrr_info'][m]
            amount = info['qty'] * info['rate']
            total_num += amount
            total_den += info['qty']
            ws.cell(row=r, column=1, value=f"Step {idx+1}").font = df
            ws.cell(row=r, column=2, value=m).font = df
            ws.cell(row=r, column=3, value=info['qty']).font = df
            ws.cell(row=r, column=4, value=f"× {info['rate']:.4f}").font = df
            ws.cell(row=r, column=5, value=round(amount, 4)).font = df
            for i in range(1, 6):
                ws.cell(row=r, column=i).border = border
                ws.cell(row=r, column=i).alignment = center
            r += 1
        
        # Total
        ws.cell(row=r, column=1, value="TOTAL").font = bf
        ws.cell(row=r, column=3, value=total_den).font = bf
        ws.cell(row=r, column=5, value=round(total_num, 4)).font = bf
        for i in range(1, 6):
            ws.cell(row=r, column=i).fill = yfill
            ws.cell(row=r, column=i).border = border
        r += 1
        
        ws.cell(row=r, column=1, value="RESULT").font = Font(name='Calibri', size=13, bold=True)
        ws.cell(row=r, column=2, value=f"{total_num:.4f} ÷ {total_den:.1f}").font = bf
        ws.cell(row=r, column=3, value="=").font = bf
        ws.cell(row=r, column=4, value=round(bm['rate'], 6)).font = Font(name='Calibri', size=13, bold=True, color='059669')
        ws.cell(row=r, column=4).fill = gfill
        ws.cell(row=r, column=5, value=f"GT={rd['gt_rate']:.6f}").font = bf
        r += 1
        
        # WHY DIFFERENT
        r += 1
        ws.cell(row=r, column=1, value="WHY DIFFERENT FROM ENGINE:").font = Font(name='Calibri', size=12, bold=True, color='DC2626')
        ws.cell(row=r, column=1).fill = rfill
        r += 1
        
        excluded = [m for m in rd['mrr_info'] if m not in set(bm['mrrs'])]
        if excluded:
            ws.cell(row=r, column=1, value="Friend EXCLUDED these MRRs:").font = bf
            r += 1
            for m in excluded:
                info = rd['mrr_info'][m]
                ws.cell(row=r, column=1, value=f"  MRR {m}").font = df
                ws.cell(row=r, column=2, value=f"Supplier: {info['supplier']}").font = df
                ws.cell(row=r, column=3, value=f"PR Month: {info['pr_month']}").font = df
                ws.cell(row=r, column=4, value=f"Rate: {info['rate']}").font = df
                for i in range(1, 5):
                    ws.cell(row=r, column=i).fill = rfill
                r += 1
        
        included_not_in_gt = [m for m in rd['mrr_info'] if m not in gt_mrrs and m in best_mrrs]
        if included_not_in_gt:
            ws.cell(row=r, column=1, value="Engine INCLUDED but friend didn't list in MR# column:").font = bf
            r += 1
    
    r += 2

# Column widths
widths = [18, 14, 14, 14, 14, 22, 12, 16, 16, 16]
for i, w in enumerate(widths):
    ws.column_dimensions[get_column_letter(i+1)].width = w

path = f"{BASE}/Gap_Analysis_Detail.xlsx"
wb.save(path)
print(f"\nSaved to: {path}")
