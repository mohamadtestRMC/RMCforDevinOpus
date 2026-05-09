import openpyxl
from openpyxl.utils import get_column_letter
import sys

FILE_PATH = r"Files_need_to_study\Filled_Output\1 Base RMC _ 2026 February.xlsx"

print("=" * 100)
print("DEEP ANALYSIS OF FILLED BASE RMC REPORT")
print("=" * 100)

wb = openpyxl.load_workbook(FILE_PATH, data_only=False)

print(f"\nFile: {FILE_PATH}")
print(f"Total sheets: {len(wb.sheetnames)}")
print(f"Sheet names: {wb.sheetnames}")
print()

for idx, sheet_name in enumerate(wb.sheetnames):
    ws = wb[sheet_name]
    print("\n" + "#" * 100)
    print(f"SHEET {idx+1}: '{sheet_name}'")
    print("#" * 100)
    print(f"  Dimensions: {ws.dimensions}")
    print(f"  Min row: {ws.min_row}, Max row: {ws.max_row}")
    print(f"  Min col: {ws.min_column}, Max col: {ws.max_column}")
    print(f"  Sheet state: {'hidden' if ws.sheet_state == 'hidden' else 'visible'}")

    merged = list(ws.merged_cells.ranges)
    if merged:
        print(f"\n  MERGED CELLS ({len(merged)} ranges):")
        for mc in merged:
            print(f"    {mc}")

    formulas = []
    for row in ws.iter_rows(min_row=ws.min_row, max_row=ws.max_row,
                            min_col=ws.min_column, max_col=ws.max_column):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formulas.append((cell.coordinate, cell.value))

    if sheet_name == "RMC Summary":
        print(f"\n  *** FULL DUMP OF 'RMC Summary' SHEET ***")
        print(f"  Total formulas found: {len(formulas)}")
        print()

        print("  --- ALL CELL DATA (row by row) ---")
        for row_num in range(ws.min_row, ws.max_row + 1):
            row_data = []
            for col_num in range(ws.min_column, ws.max_column + 1):
                cell = ws.cell(row=row_num, column=col_num)
                val = cell.value
                col_letter = get_column_letter(col_num)
                coord = f"{col_letter}{row_num}"
                nf = cell.number_format if cell.number_format != 'General' else None
                font_bold = cell.font.bold if cell.font else None
                fill_color = None
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb and cell.fill.fgColor.rgb != '00000000':
                    fill_color = cell.fill.fgColor.rgb

                extras = []
                if nf:
                    extras.append(f"fmt={nf}")
                if font_bold:
                    extras.append("BOLD")
                if fill_color:
                    extras.append(f"fill={fill_color}")

                extra_str = f" [{', '.join(extras)}]" if extras else ""
                if val is not None:
                    row_data.append(f"    {coord}: {repr(val)}{extra_str}")

            if row_data:
                print(f"  ROW {row_num}:")
                for rd in row_data:
                    print(rd)

        if formulas:
            print(f"\n  --- ALL FORMULAS IN 'RMC Summary' ---")
            for coord, formula in formulas:
                print(f"    {coord}: {formula}")

            unique_patterns = set()
            for _, f in formulas:
                import re
                pattern = re.sub(r'[A-Z]+\d+', 'REF', f)
                pattern = re.sub(r'\d+\.?\d*', 'NUM', pattern)
                unique_patterns.add(pattern)
            print(f"\n  --- UNIQUE FORMULA PATTERNS ({len(unique_patterns)}) ---")
            for p in sorted(unique_patterns):
                print(f"    {p}")

    else:
        print(f"\n  --- HEADERS (first 3 rows) ---")
        for row_num in range(ws.min_row, min(ws.min_row + 3, ws.max_row + 1)):
            row_vals = {}
            for col_num in range(ws.min_column, ws.max_column + 1):
                cell = ws.cell(row=row_num, column=col_num)
                if cell.value is not None:
                    col_letter = get_column_letter(col_num)
                    row_vals[f"{col_letter}{row_num}"] = repr(cell.value)
            if row_vals:
                print(f"  Row {row_num}: {row_vals}")

        data_start = ws.min_row + 1
        data_end = min(data_start + 9, ws.max_row + 1)
        print(f"\n  --- DATA ROWS {data_start} to {data_end - 1} ---")
        for row_num in range(data_start, data_end):
            row_vals = {}
            for col_num in range(ws.min_column, ws.max_column + 1):
                cell = ws.cell(row=row_num, column=col_num)
                if cell.value is not None:
                    col_letter = get_column_letter(col_num)
                    row_vals[f"{col_letter}{row_num}"] = repr(cell.value)
            if row_vals:
                print(f"  Row {row_num}: {row_vals}")

        if formulas:
            print(f"\n  --- FORMULAS ({len(formulas)} total) ---")
            shown = 0
            for coord, formula in formulas:
                if shown < 30:
                    print(f"    {coord}: {formula}")
                    shown += 1
            if len(formulas) > 30:
                print(f"    ... and {len(formulas) - 30} more formulas")

            unique_patterns = set()
            for _, f in formulas:
                import re
                pattern = re.sub(r'[A-Z]+\d+', 'REF', f)
                pattern = re.sub(r'\d+\.?\d*', 'NUM', pattern)
                unique_patterns.add(pattern)
            print(f"\n  --- UNIQUE FORMULA PATTERNS ({len(unique_patterns)}) ---")
            for p in sorted(unique_patterns):
                print(f"    {p}")

        last_rows_start = max(ws.max_row - 4, data_end)
        if last_rows_start > data_end:
            print(f"\n  --- LAST 5 ROWS ({last_rows_start} to {ws.max_row}) ---")
            for row_num in range(last_rows_start, ws.max_row + 1):
                row_vals = {}
                for col_num in range(ws.min_column, ws.max_column + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    if cell.value is not None:
                        col_letter = get_column_letter(col_num)
                        row_vals[f"{col_letter}{row_num}"] = repr(cell.value)
                if row_vals:
                    print(f"  Row {row_num}: {row_vals}")

wb.close()
print("\n" + "=" * 100)
print("ANALYSIS COMPLETE")
print("=" * 100)
