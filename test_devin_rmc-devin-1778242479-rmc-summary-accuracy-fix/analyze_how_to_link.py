import openpyxl
from openpyxl.utils import get_column_letter
import re

FILE_PATH = r"c:\Users\mohamad.al\OneDrive - AL KHAYYAT INVESTMENTS\Desktop\IPP\Files_need_to_study\How_to_link\Base RMC Documents (2).xlsx"

print("=" * 100)
print("DEEP ANALYSIS: Base RMC Documents (2).xlsx")
print("=" * 100)

# Load with data_only=False to see formulas
wb_formulas = openpyxl.load_workbook(FILE_PATH, data_only=False)
# Load with data_only=True to see computed values
wb_values = openpyxl.load_workbook(FILE_PATH, data_only=True)

print(f"\n{'='*100}")
print("SHEET NAMES:")
print(f"{'='*100}")
for i, name in enumerate(wb_formulas.sheetnames, 1):
    print(f"  {i}. '{name}'")

print(f"\nTotal sheets: {len(wb_formulas.sheetnames)}")

# Check for defined names (named ranges)
print(f"\n{'='*100}")
print("DEFINED NAMES / NAMED RANGES:")
print(f"{'='*100}")
if wb_formulas.defined_names:
    for dn in wb_formulas.defined_names.definedName:
        print(f"  Name: '{dn.name}' -> {dn.attr_text}")
else:
    print("  No defined names found.")

# Analyze each sheet
for sheet_name in wb_formulas.sheetnames:
    ws_f = wb_formulas[sheet_name]
    ws_v = wb_values[sheet_name]
    
    print(f"\n{'#'*100}")
    print(f"# SHEET: '{sheet_name}'")
    print(f"{'#'*100}")
    
    print(f"\n  Dimensions: {ws_f.dimensions}")
    print(f"  Max Row: {ws_f.max_row}")
    print(f"  Max Column: {ws_f.max_column} ({get_column_letter(ws_f.max_column) if ws_f.max_column else 'N/A'})")
    print(f"  Min Row: {ws_f.min_row}")
    print(f"  Min Column: {ws_f.min_column}")
    
    # Merged cells
    print(f"\n  --- MERGED CELLS ---")
    if ws_f.merged_cells.ranges:
        for mc in ws_f.merged_cells.ranges:
            print(f"    {mc}")
    else:
        print(f"    None")
    
    # Column headers (first row)
    print(f"\n  --- COLUMN HEADERS (Row 1) ---")
    if ws_f.max_column and ws_f.max_row:
        for col in range(1, ws_f.max_column + 1):
            cell = ws_f.cell(row=1, column=col)
            if cell.value is not None:
                print(f"    {get_column_letter(col)}1: {repr(cell.value)}")
    
    # ALL DATA - Every cell with content
    print(f"\n  --- ALL CELL DATA (Formula view & Value view) ---")
    formulas_found = []
    cross_sheet_refs = []
    external_refs = []
    
    if ws_f.max_row and ws_f.max_column:
        for row in range(1, ws_f.max_row + 1):
            row_has_data = False
            row_data = []
            for col in range(1, ws_f.max_column + 1):
                cell_f = ws_f.cell(row=row, column=col)
                cell_v = ws_v.cell(row=row, column=col)
                cell_ref = f"{get_column_letter(col)}{row}"
                
                if cell_f.value is not None:
                    row_has_data = True
                    formula_val = cell_f.value
                    cached_val = cell_v.value
                    
                    is_formula = isinstance(formula_val, str) and formula_val.startswith('=')
                    
                    if is_formula:
                        row_data.append(f"      {cell_ref}: FORMULA: {formula_val}  |  CACHED_VALUE: {repr(cached_val)}")
                        formulas_found.append((cell_ref, formula_val))
                        
                        # Check for cross-sheet references
                        if "!" in formula_val:
                            cross_sheet_refs.append((cell_ref, formula_val))
                        # Check for external file references
                        if "[" in formula_val and "]" in formula_val:
                            external_refs.append((cell_ref, formula_val))
                    else:
                        if formula_val != cached_val and cached_val is not None:
                            row_data.append(f"      {cell_ref}: {repr(formula_val)}  (cached: {repr(cached_val)})")
                        else:
                            row_data.append(f"      {cell_ref}: {repr(formula_val)}")
            
            if row_has_data:
                print(f"    ROW {row}:")
                for rd in row_data:
                    print(rd)
    
    # Summary of formulas
    print(f"\n  --- FORMULAS SUMMARY ---")
    print(f"  Total formulas found: {len(formulas_found)}")
    if formulas_found:
        for ref, formula in formulas_found:
            print(f"    {ref}: {formula}")
    
    # Cross-sheet references
    print(f"\n  --- CROSS-SHEET REFERENCES ---")
    if cross_sheet_refs:
        for ref, formula in cross_sheet_refs:
            print(f"    {ref}: {formula}")
    else:
        print(f"    None")
    
    # External file references
    print(f"\n  --- EXTERNAL FILE REFERENCES ---")
    if external_refs:
        for ref, formula in external_refs:
            print(f"    {ref}: {formula}")
    else:
        print(f"    None")
    
    # Special formula patterns
    print(f"\n  --- SPECIAL FORMULA PATTERNS ---")
    vlookup_formulas = [(r, f) for r, f in formulas_found if 'VLOOKUP' in f.upper()]
    index_match = [(r, f) for r, f in formulas_found if 'INDEX' in f.upper() or 'MATCH' in f.upper()]
    sumif_formulas = [(r, f) for r, f in formulas_found if 'SUMIF' in f.upper()]
    sumproduct_formulas = [(r, f) for r, f in formulas_found if 'SUMPRODUCT' in f.upper()]
    if_formulas = [(r, f) for r, f in formulas_found if 'IF(' in f.upper()]
    concatenate_formulas = [(r, f) for r, f in formulas_found if 'CONCATENATE' in f.upper() or '&' in f]
    
    if vlookup_formulas:
        print(f"    VLOOKUP formulas ({len(vlookup_formulas)}):")
        for r, f in vlookup_formulas:
            print(f"      {r}: {f}")
    if index_match:
        print(f"    INDEX/MATCH formulas ({len(index_match)}):")
        for r, f in index_match:
            print(f"      {r}: {f}")
    if sumif_formulas:
        print(f"    SUMIF formulas ({len(sumif_formulas)}):")
        for r, f in sumif_formulas:
            print(f"      {r}: {f}")
    if sumproduct_formulas:
        print(f"    SUMPRODUCT formulas ({len(sumproduct_formulas)}):")
        for r, f in sumproduct_formulas:
            print(f"      {r}: {f}")
    if if_formulas:
        print(f"    IF formulas ({len(if_formulas)}):")
        for r, f in if_formulas:
            print(f"      {r}: {f}")
    if concatenate_formulas:
        print(f"    CONCATENATE/& formulas ({len(concatenate_formulas)}):")
        for r, f in concatenate_formulas:
            print(f"      {r}: {f}")
    
    if not any([vlookup_formulas, index_match, sumif_formulas, sumproduct_formulas, if_formulas, concatenate_formulas]):
        print(f"    None found")

print(f"\n{'='*100}")
print("ANALYSIS COMPLETE")
print(f"{'='*100}")

wb_formulas.close()
wb_values.close()
