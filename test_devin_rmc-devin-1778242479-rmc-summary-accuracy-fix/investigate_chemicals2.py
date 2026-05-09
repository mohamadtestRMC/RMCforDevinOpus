"""
Investigation part 2: 
- Stores Recordings with correct headers
- Cross-reference DA (hardener name) with PR to understand the lookup rule
- Check the Jobtrack With MRR ground truth for adhesive/hardener rate values
"""
import pandas as pd
import openpyxl

T2 = "Template2"
T1 = "Template_Files"

# ── 1. Stores Recordings - find correct header row ──
print("=" * 80)
print("1. STORES RECORDINGS - Finding correct header row")
print("=" * 80)

# Try different header rows
for hr in [0, 1, 2, 3, 4]:
    stores = pd.read_excel(f"{T2}/Stores Recordings.xlsx", sheet_name=0, header=hr, nrows=3)
    cols_str = [str(c).lower() for c in stores.columns]
    has_material = any('material' in c for c in cols_str)
    has_category = any('categ' in c for c in cols_str)
    print(f"  header={hr}: has_material={has_material}, has_category={has_category}")
    if has_material or has_category:
        print(f"    Columns: {list(stores.columns)[:15]}...")

# Use the correct header row from Template_Files stores (which works)
stores1 = pd.read_excel(f"{T1}/Stores Recordings.xlsx", sheet_name=0, header=2)
print(f"\nTemplate_Files stores header=2 columns: {list(stores1.columns)[:15]}...")

# Check raw cell values in Template2 stores
wb = openpyxl.load_workbook(f"{T2}/Stores Recordings.xlsx", data_only=True)
ws = wb.active
print(f"\nTemplate2 Stores - first 5 rows, first 10 cols:")
for r in range(1, 6):
    vals = [ws.cell(row=r, column=c).value for c in range(1, 11)]
    print(f"  Row {r}: {vals}")

# Find the header row
for r in range(1, 10):
    v = ws.cell(row=r, column=1).value
    if v and 'material' in str(v).lower():
        print(f"\n  Header row found at row {r}")
        break

wb.close()

# ── 2. Load stores properly ──
print("\n" + "=" * 80)
print("2. STORES - Loading from engine to match what the code uses")
print("=" * 80)

import sys
sys.path.insert(0, '.')
from engine.mrr_lookup import load_stores_recordings

stores_df = load_stores_recordings(f"{T2}/Stores Recordings.xlsx")
print(f"Stores loaded: {len(stores_df)} rows")
print(f"Columns: {list(stores_df.columns)[:15]}...")

# Find category column
cat_col = None
mat_col = None
for c in stores_df.columns:
    cl = str(c).strip().lower()
    if 'categ' in cl:
        cat_col = c
    if c == 'Material' or (cl == 'material' and 'sub' not in cl):
        mat_col = c

print(f"Category col: {cat_col}")
print(f"Material col: {mat_col}")

if cat_col:
    cats = stores_df[cat_col].dropna().astype(str).str.strip().str.upper().unique()
    print(f"\nUnique categories: {sorted(cats)}")

    for kw in ['ADHESIVE', 'HARDENER', 'SOLVENT']:
        rows = stores_df[stores_df[cat_col].astype(str).str.strip().str.upper().str.contains(kw)]
        if len(rows) > 0 and mat_col:
            mats = rows[mat_col].dropna().astype(str).str.strip().unique()
            print(f"\n{kw} materials in Stores: {sorted(mats)}")
        else:
            print(f"\n{kw}: {len(rows)} rows")

# ── 3. Cross-reference: DA column hardener names vs PR ADHESIVE category ──
print("\n" + "=" * 80)
print("3. CROSS-REFERENCE: DA hardener names vs Purchase Register")
print("=" * 80)

pr = pd.read_excel(f"{T2}/Purchase Register - 2021 - 2025 _Nov.xlsx", sheet_name=0, header=2)
pr_cat = None
pr_mat = None
for c in pr.columns:
    cl = str(c).strip().lower()
    if 'categ' in cl:
        pr_cat = c
    if cl == 'material':
        pr_mat = c

# The DA column contains: CT85, CR84, CR 88-300, S110
da_hardeners = ['CT85', 'CR84', 'CR 88-300', 'S110']

print("\nLooking up each DA hardener name in PR:")
for h in da_hardeners:
    h_upper = h.strip().upper()
    # Check in ADHESIVE category
    adh_mask = pr[pr_cat].astype(str).str.strip().str.upper().str.contains('ADHESIVE')
    mat_match = pr[pr_mat].astype(str).str.strip().str.upper() == h_upper
    found = pr[adh_mask & mat_match]
    print(f"\n  '{h}' in ADHESIVE category: {len(found)} rows")
    if len(found) > 0:
        # Show month, rate, amount
        for _, row in found.head(3).iterrows():
            print(f"    Month={row.get('month', '?')}, Rate={row.get('Rate', '?')}, "
                  f"Amount={row.get('Amount', '?')}, Qty={row.get('Actual Quantity', '?')}")

# ── 4. Verify the actual RULES by checking Jobtrack WITH MRR ground truth ──
print("\n" + "=" * 80)
print("4. GROUND TRUTH: Jobtrack WITH MRR - Chemical rates vs PR values")
print("=" * 80)

jt_wb = openpyxl.load_workbook(f"{T2}/Jobtrack With MRR.xlsx", data_only=True)
jt_ws = jt_wb.active

from engine.rate_lookup import load_purchase_register, lookup_adhesive_rate, lookup_hardener_rate, lookup_solvent_rate

pr_df = load_purchase_register(f"{T2}/Purchase Register - 2021 - 2025 _Nov.xlsx")

# Get report month
report_month = None
for r in range(5, jt_ws.max_row + 1):
    dv = jt_ws.cell(row=r, column=4).value
    if dv and hasattr(dv, 'month'):
        report_month = f"{dv.month}-{dv.year}"
        break
print(f"Report month: {report_month}")

# Compare each LAM row's adhesive/hardener/solvent rates
print("\nRow-by-row comparison: Ground Truth (GT) vs Our Engine:")
print(f"{'Row':>5} | {'Adh Name':>10} | {'GT Adh Rate':>12} | {'Engine Adh':>12} | {'DA(Hardener)':>12} | {'GT Hard Rate':>12} | {'Engine Hard':>12} | {'GT Sol Rate':>12} | {'Engine Sol':>12}")
print("-" * 120)

mismatches = []
for row in range(5, jt_ws.max_row + 1):
    process = jt_ws.cell(row=row, column=6).value
    if not process or str(process).strip().upper() != 'LAM':
        continue
    
    adh_name = jt_ws.cell(row=row, column=91).value  # CM
    gt_adh_rate = jt_ws.cell(row=row, column=93).value  # CO
    gt_hard_rate = jt_ws.cell(row=row, column=97).value  # CS
    gt_sol_rate = jt_ws.cell(row=row, column=101).value  # CW
    da_val = jt_ws.cell(row=row, column=105).value  # DA = hardener name
    
    if not adh_name:
        continue
    
    adh_name_str = str(adh_name).strip()
    
    # Engine lookups
    eng_adh = lookup_adhesive_rate(pr_df, adh_name_str, report_month=report_month)
    eng_hard = lookup_hardener_rate(pr_df, adh_name_str, report_month=report_month)
    eng_sol = lookup_solvent_rate(pr_df, report_month=report_month)
    
    gt_adh_f = float(gt_adh_rate) if gt_adh_rate else 0
    gt_hard_f = float(gt_hard_rate) if gt_hard_rate else 0
    gt_sol_f = float(gt_sol_rate) if gt_sol_rate else 0
    
    adh_ok = abs(gt_adh_f - eng_adh) < 0.01 if gt_adh_f > 0 else True
    hard_ok = abs(gt_hard_f - eng_hard) < 0.01 if gt_hard_f > 0 else True
    sol_ok = abs(gt_sol_f - eng_sol) < 0.01 if gt_sol_f > 0 else True
    
    flag = ""
    if not adh_ok:
        flag += " ADH-MISMATCH"
    if not hard_ok:
        flag += " HARD-MISMATCH"
    if not sol_ok:
        flag += " SOL-MISMATCH"
    
    print(f"{row:>5} | {adh_name_str:>10} | {gt_adh_f:>12.4f} | {eng_adh:>12.4f} | {str(da_val):>12} | {gt_hard_f:>12.4f} | {eng_hard:>12.4f} | {gt_sol_f:>12.4f} | {eng_sol:>12.4f}{flag}")
    
    if flag:
        mismatches.append({
            'row': row, 'adh': adh_name_str, 'da': da_val,
            'gt_adh': gt_adh_f, 'eng_adh': eng_adh,
            'gt_hard': gt_hard_f, 'eng_hard': eng_hard,
            'gt_sol': gt_sol_f, 'eng_sol': eng_sol,
            'issues': flag
        })

print(f"\nTotal mismatches: {len(mismatches)}")
for m in mismatches:
    print(f"  Row {m['row']}: {m['issues']} | Adh={m['adh']}, DA={m['da']}")

# ── 5. KEY INSIGHT: How does DA column relate to the hardener rate lookup? ──
print("\n" + "=" * 80)
print("5. KEY INSIGHT: DA column = HARDENER NAME for rate lookup")
print("=" * 80)

print("\nThe DA column (col 105, header 'HARDNER') contains the actual hardener material name.")
print("This is the name we should use to look up the hardener rate from PR.")
print("\nCurrent ADH_HARDENER_PAIRS mapping:")
from engine.rate_lookup import ADH_HARDENER_PAIRS
for k, v in ADH_HARDENER_PAIRS.items():
    print(f"  {k} -> {v}")

print("\nActual pairings found in data (Adh Name -> DA value):")
pairings = {}
for row in range(5, jt_ws.max_row + 1):
    process = jt_ws.cell(row=row, column=6).value
    if not process or str(process).strip().upper() != 'LAM':
        continue
    adh = jt_ws.cell(row=row, column=91).value
    da = jt_ws.cell(row=row, column=105).value
    if adh and da:
        key = str(adh).strip().upper()
        val = str(da).strip().upper()
        if key not in pairings:
            pairings[key] = set()
        pairings[key].add(val)

for k, vs in sorted(pairings.items()):
    print(f"  {k} -> {sorted(vs)}")
    # Check if our current mapping matches
    if k in ADH_HARDENER_PAIRS:
        mapped = ADH_HARDENER_PAIRS[k].upper()
        if mapped not in vs:
            print(f"    WARNING: Our mapping says '{mapped}' but data shows {sorted(vs)}")

jt_wb.close()

print("\n" + "=" * 80)
print("INVESTIGATION COMPLETE")
print("=" * 80)
