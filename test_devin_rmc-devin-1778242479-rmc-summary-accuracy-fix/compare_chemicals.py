"""
COMPREHENSIVE CHEMICAL COMPARISON: Fill vs Ground Truth
- Template2 (Nov 2025): Fill "Without MRR" and compare to "With MRR"
- Template_Files (Feb 2026): Fill "Without MRR" and compare to "With MRR"
- Goal: 100% accuracy on Adhesive, Hardener, Solvent rates/values
"""
import pandas as pd
import openpyxl
import sys, os
sys.path.insert(0, '.')

from engine.rate_lookup import (
    load_purchase_register, lookup_adhesive_rate, 
    lookup_hardener_rate, lookup_solvent_rate,
    _find_col, _filter_by_month, _qty_weighted_rate,
    ADH_HARDENER_PAIRS
)

COLS = {
    'UID': 1, 'Date': 4, 'Process': 6, 'Order_No': 11,
    'Adh_Name': 91, 'Adh_Kgs': 92, 'Adh_Rate': 93, 'Adh_Value': 94,
    'Adh_Solids': 95,
    'Hard_Kgs': 96, 'Hard_Rate': 97, 'Hard_Value': 98,
    'Hard_Solids': 99,
    'Sol_Qty': 100, 'Sol_Rate': 101, 'Sol_Value': 102,
    'DA_Hardener': 105,  # KEY: Hardener material name
}

def safe_float(val):
    if val is None or val == '' or (isinstance(val, float) and pd.isna(val)):
        return 0.0
    if isinstance(val, str) and val.startswith('='):
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0

def compare_chemical_rates(dataset_name, jt_with_mrr_path, jt_without_mrr_path, pr_path, report_month):
    """Compare chemical rates between ground truth and engine output."""
    print(f"\n{'#' * 100}")
    print(f"# DATASET: {dataset_name}  |  Report Month: {report_month}")
    print(f"{'#' * 100}")
    
    # Load ground truth (With MRR)
    gt_wb = openpyxl.load_workbook(jt_with_mrr_path, data_only=True)
    gt_ws = gt_wb.active
    
    # Load template (Without MRR) to see what's pre-filled
    tmpl_wb = openpyxl.load_workbook(jt_without_mrr_path, data_only=True)
    tmpl_ws = tmpl_wb.active
    
    # Load PR
    pr_df = load_purchase_register(pr_path)
    
    # ── Collect ALL LAM rows with chemical data ──
    rows_data = []
    for row in range(5, gt_ws.max_row + 1):
        process = gt_ws.cell(row=row, column=COLS['Process']).value
        if not process or str(process).strip().upper() != 'LAM':
            continue
        
        uid = gt_ws.cell(row=row, column=COLS['UID']).value
        adh_name = gt_ws.cell(row=row, column=COLS['Adh_Name']).value
        
        if not adh_name:
            continue
        
        # Ground truth values
        gt = {
            'row': row,
            'uid': str(uid).strip() if uid else '',
            'adh_name': str(adh_name).strip(),
            'adh_kgs': safe_float(gt_ws.cell(row=row, column=COLS['Adh_Kgs']).value),
            'gt_adh_rate': safe_float(gt_ws.cell(row=row, column=COLS['Adh_Rate']).value),
            'gt_adh_value': safe_float(gt_ws.cell(row=row, column=COLS['Adh_Value']).value),
            'hard_kgs': safe_float(gt_ws.cell(row=row, column=COLS['Hard_Kgs']).value),
            'gt_hard_rate': safe_float(gt_ws.cell(row=row, column=COLS['Hard_Rate']).value),
            'gt_hard_value': safe_float(gt_ws.cell(row=row, column=COLS['Hard_Value']).value),
            'sol_qty': safe_float(gt_ws.cell(row=row, column=COLS['Sol_Qty']).value),
            'gt_sol_rate': safe_float(gt_ws.cell(row=row, column=COLS['Sol_Rate']).value),
            'gt_sol_value': safe_float(gt_ws.cell(row=row, column=COLS['Sol_Value']).value),
            'da_hardener_gt': str(gt_ws.cell(row=row, column=COLS['DA_Hardener']).value or '').strip(),
        }
        
        # Template values (what's available in Without MRR)
        gt['tmpl_adh_name'] = str(tmpl_ws.cell(row=row, column=COLS['Adh_Name']).value or '').strip()
        gt['tmpl_da_hardener'] = str(tmpl_ws.cell(row=row, column=COLS['DA_Hardener']).value or '').strip()
        gt['tmpl_adh_kgs'] = safe_float(tmpl_ws.cell(row=row, column=COLS['Adh_Kgs']).value)
        gt['tmpl_hard_kgs'] = safe_float(tmpl_ws.cell(row=row, column=COLS['Hard_Kgs']).value)
        gt['tmpl_sol_qty'] = safe_float(tmpl_ws.cell(row=row, column=COLS['Sol_Qty']).value)
        
        rows_data.append(gt)
    
    gt_wb.close()
    tmpl_wb.close()
    
    print(f"\nTotal LAM rows with adhesive: {len(rows_data)}")
    
    # ── Template analysis: What's pre-filled? ──
    print("\n--- Template ('Without MRR') Pre-filled Check ---")
    has_adh_name = sum(1 for r in rows_data if r['tmpl_adh_name'])
    has_da = sum(1 for r in rows_data if r['tmpl_da_hardener'])
    has_adh_kgs = sum(1 for r in rows_data if r['tmpl_adh_kgs'] > 0)
    has_hard_kgs = sum(1 for r in rows_data if r['tmpl_hard_kgs'] > 0)
    has_sol_qty = sum(1 for r in rows_data if r['tmpl_sol_qty'] > 0)
    print(f"  Adh Name (CM): {has_adh_name}/{len(rows_data)}")
    print(f"  DA Hardener Name: {has_da}/{len(rows_data)}")
    print(f"  Adh Kgs (CN): {has_adh_kgs}/{len(rows_data)}")
    print(f"  Hard Kgs (CR): {has_hard_kgs}/{len(rows_data)}")
    print(f"  Sol Qty (CV): {has_sol_qty}/{len(rows_data)}")
    
    # ── Now compute engine rates and compare ──
    print("\n--- Engine vs Ground Truth Comparison ---")
    
    # Pre-compute solvent rate (one rate for the whole month)
    sol_rate = lookup_solvent_rate(pr_df, report_month=report_month)
    print(f"  Engine Solvent Rate: {sol_rate:.4f}")
    
    total = 0
    adh_match = 0
    adh_mismatch = 0
    hard_match = 0
    hard_mismatch = 0
    sol_match = 0
    sol_mismatch = 0
    
    mismatches = []
    
    # Collect unique adhesive names and their DA pairings
    adh_da_pairs = {}
    
    for r in rows_data:
        total += 1
        adh_name = r['adh_name']
        da_name = r['da_hardener_gt']
        
        if adh_name and da_name:
            key = adh_name.upper()
            if key not in adh_da_pairs:
                adh_da_pairs[key] = set()
            adh_da_pairs[key].add(da_name.upper())
        
        # ── Adhesive Rate ──
        eng_adh_rate = lookup_adhesive_rate(pr_df, adh_name, report_month=report_month)
        gt_adh_rate = r['gt_adh_rate']
        
        adh_ok = abs(gt_adh_rate - eng_adh_rate) < 0.01 if gt_adh_rate > 0 else eng_adh_rate == 0
        if adh_ok:
            adh_match += 1
        else:
            adh_mismatch += 1
        
        # ── Hardener Rate (current engine: uses ADH_HARDENER_PAIRS) ──
        eng_hard_rate = lookup_hardener_rate(pr_df, adh_name, report_month=report_month)
        gt_hard_rate = r['gt_hard_rate']
        
        hard_ok = abs(gt_hard_rate - eng_hard_rate) < 0.01 if gt_hard_rate > 0 else eng_hard_rate == 0
        if hard_ok:
            hard_match += 1
        else:
            hard_mismatch += 1
        
        # ── Solvent Rate ──
        gt_sol_rate = r['gt_sol_rate']
        eng_sol_rate = sol_rate if r['sol_qty'] > 0 else 0
        
        sol_ok = abs(gt_sol_rate - eng_sol_rate) < 0.01 if gt_sol_rate > 0 else True
        if sol_ok:
            sol_match += 1
        else:
            sol_mismatch += 1
        
        if not adh_ok or not hard_ok or not sol_ok:
            mismatches.append({
                **r,
                'eng_adh_rate': eng_adh_rate,
                'eng_hard_rate': eng_hard_rate,
                'eng_sol_rate': eng_sol_rate,
                'adh_ok': adh_ok,
                'hard_ok': hard_ok,
                'sol_ok': sol_ok,
            })
    
    # ── Results ──
    print(f"\n{'='*80}")
    print(f"  RESULTS: {dataset_name}")
    print(f"{'='*80}")
    print(f"  Adhesive Rate:  {adh_match}/{total} match ({adh_mismatch} mismatches)")
    print(f"  Hardener Rate:  {hard_match}/{total} match ({hard_mismatch} mismatches)")
    print(f"  Solvent Rate:   {sol_match}/{total} match ({sol_mismatch} mismatches)")
    
    if mismatches:
        print(f"\n--- MISMATCHES ({len(mismatches)}) ---")
        for m in mismatches:
            flags = []
            if not m['adh_ok']:
                flags.append(f"ADH: GT={m['gt_adh_rate']:.4f} vs ENG={m['eng_adh_rate']:.4f}")
            if not m['hard_ok']:
                flags.append(f"HARD: GT={m['gt_hard_rate']:.4f} vs ENG={m['eng_hard_rate']:.4f}")
            if not m['sol_ok']:
                flags.append(f"SOL: GT={m['gt_sol_rate']:.4f} vs ENG={m['eng_sol_rate']:.4f}")
            print(f"  Row {m['row']}: Adh={m['adh_name']}, DA={m['da_hardener_gt']} | {' | '.join(flags)}")
    
    # ── Adhesive/Hardener Pairings Found ──
    print(f"\n--- Adhesive -> DA Hardener Pairings ---")
    for adh, das in sorted(adh_da_pairs.items()):
        current_map = ADH_HARDENER_PAIRS.get(adh, 'NOT MAPPED')
        print(f"  {adh} -> {sorted(das)} | Current engine map: {current_map}")
    
    # ── Now test: what if we use DA column directly for hardener lookup? ──
    print(f"\n{'='*80}")
    print(f"  TESTING: Direct DA Column Lookup (proposed fix)")
    print(f"{'='*80}")
    
    hard_match2 = 0
    hard_mismatch2 = 0
    mismatches2 = []
    
    for r in rows_data:
        da_name = r['da_hardener_gt']
        gt_hard_rate = r['gt_hard_rate']
        
        if not da_name or gt_hard_rate == 0:
            hard_match2 += 1
            continue
        
        # Direct lookup: use DA name in ADHESIVE category
        eng_hard_rate2 = _direct_hardener_lookup(pr_df, da_name, report_month)
        
        ok = abs(gt_hard_rate - eng_hard_rate2) < 0.01
        if ok:
            hard_match2 += 1
        else:
            hard_mismatch2 += 1
            mismatches2.append({
                'row': r['row'], 'adh': r['adh_name'], 'da': da_name,
                'gt': gt_hard_rate, 'eng': eng_hard_rate2
            })
    
    print(f"  Hardener (DA direct): {hard_match2}/{total} match ({hard_mismatch2} mismatches)")
    for m in mismatches2:
        print(f"    Row {m['row']}: DA={m['da']}, GT={m['gt']:.4f}, ENG={m['eng']:.4f}")
    
    return rows_data, mismatches


def _direct_hardener_lookup(pr_df, hardener_name, report_month):
    """Direct hardener rate lookup by material name in ADHESIVE category."""
    if not hardener_name:
        return 0.0
    
    category_col = _find_col(pr_df, 'categery', 'category')
    material_col = _find_col(pr_df, 'material')
    rate_col = [c for c in pr_df.columns if str(c).strip().lower() == 'rate']
    rate_col = rate_col[0] if rate_col else 'Rate'
    
    h_upper = str(hardener_name).strip().upper()
    
    # Name normalization for PR matching
    name_map = {
        'CR84': 'CR 84',     # PR has "CR 84" with space
        'CR 84': 'CR 84',
    }
    lookup_name = name_map.get(h_upper, h_upper)
    
    mask = pd.Series([False] * len(pr_df))
    if category_col:
        mask = pr_df[category_col].astype(str).str.lower().str.contains('adhesive', na=False)
    
    if material_col:
        mat_mask = pr_df[material_col].astype(str).str.upper().str.strip().apply(
            lambda x: x == lookup_name or lookup_name in x or x in lookup_name
        )
        mask = mask & mat_mask
    
    filtered = pr_df[mask]
    if filtered.empty:
        return 0.0
    
    filtered = _filter_by_month(pr_df, filtered, report_month)
    return _qty_weighted_rate(filtered, rate_col)


# ══════════════════════════════════════════════════════════════════
# RUN COMPARISONS
# ══════════════════════════════════════════════════════════════════

# Template2 (November 2025)
compare_chemical_rates(
    "Template2 (November 2025)",
    "Template2/Jobtrack With MRR.xlsx",
    "Template2/Jobtrack Without MRR.xlsx",
    "Template2/Purchase Register - 2021 - 2025 _Nov.xlsx",
    "11-2025"
)

# Template_Files (February 2026)
compare_chemical_rates(
    "Template_Files (February 2026)",
    "Template_Files/Jobtrack Feb With MRR.xlsx",
    "Template_Files/Jobtrack Feb Without MRR.xlsx",
    "Template_Files/Purchase Register - 2021 - 2026 _Feb 26.xlsx",
    "2-2026"
)


# ══════════════════════════════════════════════════════════════════
# DEEP DIVE: Check ALL possible adhesive/hardener combinations
# ══════════════════════════════════════════════════════════════════
print(f"\n{'#' * 100}")
print(f"# DEEP DIVE: ALL Adhesive/Hardener combinations across BOTH datasets")
print(f"{'#' * 100}")

for ds_name, jt_path, pr_path, month in [
    ("Nov 2025", "Template2/Jobtrack With MRR.xlsx", "Template2/Purchase Register - 2021 - 2025 _Nov.xlsx", "11-2025"),
    ("Feb 2026", "Template_Files/Jobtrack Feb With MRR.xlsx", "Template_Files/Purchase Register - 2021 - 2026 _Feb 26.xlsx", "2-2026"),
]:
    print(f"\n--- {ds_name} ---")
    wb = openpyxl.load_workbook(jt_path, data_only=True)
    ws = wb.active
    pr_df = load_purchase_register(pr_path)
    
    combos = {}
    for row in range(5, ws.max_row + 1):
        process = ws.cell(row=row, column=6).value
        if not process or str(process).strip().upper() != 'LAM':
            continue
        adh = ws.cell(row=row, column=91).value
        da = ws.cell(row=row, column=105).value
        gt_adh_r = safe_float(ws.cell(row=row, column=93).value)
        gt_hard_r = safe_float(ws.cell(row=row, column=97).value)
        gt_sol_r = safe_float(ws.cell(row=row, column=101).value)
        sol_qty = safe_float(ws.cell(row=row, column=100).value)
        
        if not adh:
            continue
        
        key = (str(adh).strip().upper(), str(da).strip().upper() if da else 'NONE')
        if key not in combos:
            combos[key] = {
                'count': 0,
                'gt_adh_rates': set(),
                'gt_hard_rates': set(),
                'gt_sol_rates': set(),
                'has_solvent': False,
            }
        combos[key]['count'] += 1
        if gt_adh_r > 0:
            combos[key]['gt_adh_rates'].add(round(gt_adh_r, 4))
        if gt_hard_r > 0:
            combos[key]['gt_hard_rates'].add(round(gt_hard_r, 4))
        if gt_sol_r > 0:
            combos[key]['gt_sol_rates'].add(round(gt_sol_r, 4))
            combos[key]['has_solvent'] = True
    
    for (adh, da), info in sorted(combos.items()):
        eng_adh = lookup_adhesive_rate(pr_df, adh, report_month=month)
        eng_hard = _direct_hardener_lookup(pr_df, da if da != 'NONE' else None, month)
        eng_sol = lookup_solvent_rate(pr_df, report_month=month)
        
        gt_adh_str = '/'.join(str(x) for x in info['gt_adh_rates']) or '-'
        gt_hard_str = '/'.join(str(x) for x in info['gt_hard_rates']) or '-'
        gt_sol_str = '/'.join(str(x) for x in info['gt_sol_rates']) or '-'
        
        adh_ok = all(abs(x - eng_adh) < 0.01 for x in info['gt_adh_rates']) if info['gt_adh_rates'] else True
        hard_ok = all(abs(x - eng_hard) < 0.01 for x in info['gt_hard_rates']) if info['gt_hard_rates'] else True
        sol_ok = all(abs(x - eng_sol) < 0.01 for x in info['gt_sol_rates']) if info['gt_sol_rates'] else True
        
        status = "OK" if (adh_ok and hard_ok and sol_ok) else "MISMATCH"
        details = []
        if not adh_ok:
            details.append(f"ADH: GT={gt_adh_str} vs {eng_adh:.4f}")
        if not hard_ok:
            details.append(f"HARD: GT={gt_hard_str} vs {eng_hard:.4f}")
        if not sol_ok:
            details.append(f"SOL: GT={gt_sol_str} vs {eng_sol:.4f}")
        
        print(f"  ADH={adh:>12} DA={da:>12} x{info['count']:>3} | "
              f"Adh: GT={gt_adh_str:>10} Eng={eng_adh:>10.4f} | "
              f"Hard: GT={gt_hard_str:>10} Eng={eng_hard:>10.4f} | "
              f"Sol: GT={gt_sol_str:>10} Eng={eng_sol:>10.4f} | "
              f"{status} {' '.join(details)}")
    
    wb.close()

print(f"\n{'#' * 100}")
print("COMPARISON COMPLETE")
print(f"{'#' * 100}")
