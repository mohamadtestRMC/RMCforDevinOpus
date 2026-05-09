"""Debug: Check what's in the RMC Summary sheet of the output file."""
import sys, os
sys.path.insert(0, r'c:\Users\mohamad.al\OneDrive - AL KHAYYAT INVESTMENTS\Desktop\IPP')
import openpyxl

base = r'c:\Users\mohamad.al\OneDrive - AL KHAYYAT INVESTMENTS\Desktop\IPP'

# Check BOTH: output AND unfilled template
for label, path in [
    ("OUTPUT", os.path.join(base, 'output', 'Base_RMC_Feb2026_FILLED.xlsx')),
    ("TEMPLATE", os.path.join(base, 'output', 'template_fast.xlsx')),
]:
    if not os.path.exists(path):
        print(f"{label}: NOT FOUND at {path}")
        continue
    print(f"\n{'='*60}")
    print(f"{label}: {os.path.basename(path)}")
    print(f"{'='*60}")
    
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    
    if 'RMC summary' not in wb.sheetnames:
        print("  RMC summary sheet NOT FOUND!")
        print(f"  Available sheets: {wb.sheetnames}")
        wb.close()
        continue
    
    ws = wb['RMC summary']
    print(f"  max_row={ws.max_row}, max_column={ws.max_column}")
    
    # Check rows 1-10 for content
    print("\n  --- First 10 rows ---")
    for r in range(1, 11):
        row_vals = {}
        for c in range(1, 30):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                row_vals[c] = repr(v)[:40]
        if row_vals:
            print(f"  Row {r}: {row_vals}")
    
    # Check for orders in column B at various positions
    print("\n  --- Scanning column B for orders ---")
    found = 0
    first_order_row = None
    for r in range(1, min(ws.max_row + 1, 700)):
        v = ws.cell(row=r, column=2).value
        if v is not None and str(v).strip():
            vs = str(v).strip()
            if any(c.isdigit() for c in vs) and len(vs) >= 5:
                if found < 5:
                    print(f"  Row {r}: B={v}, A={ws.cell(row=r, column=1).value}, G={ws.cell(row=r, column=7).value}")
                if first_order_row is None:
                    first_order_row = r
                found += 1
    print(f"  Total orders found: {found} (first at row {first_order_row})")
    
    # Also check column A for combined keys
    print("\n  --- Scanning column A ---")
    a_found = 0
    for r in range(1, min(ws.max_row + 1, 700)):
        v = ws.cell(row=r, column=1).value
        if v is not None and str(v).strip():
            if a_found < 3:
                print(f"  Row {r}: A={v}")
            a_found += 1
    print(f"  Total col A values: {a_found}")
    
    wb.close()

print("\nDone!", flush=True)
