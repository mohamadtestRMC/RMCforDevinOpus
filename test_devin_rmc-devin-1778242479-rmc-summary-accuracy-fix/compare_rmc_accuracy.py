"""
Compare engine-generated RMC Summary vs manually filled ground truth.
Run AFTER test_base_rmc.py completes to check accuracy.
"""
import sys, os, json
sys.path.insert(0, r'c:\Users\mohamad.al\OneDrive - AL KHAYYAT INVESTMENTS\Desktop\IPP')
import pandas as pd
import openpyxl

base = r'c:\Users\mohamad.al\OneDrive - AL KHAYYAT INVESTMENTS\Desktop\IPP'
engine_path = os.path.join(base, 'output', 'Base_RMC_Feb2026_FILLED.xlsx')
filled_path = os.path.join(base, 'Files_need_to_study', 'Filled_Output', '1 Base RMC _ 2026 February.xlsx')

if not os.path.exists(engine_path):
    print("ERROR: Engine output not found. Run test_base_rmc.py first!", flush=True)
    sys.exit(1)

# ═══════════════════════════════════════════════════════════════
# Load engine output RMC Summary
# ═══════════════════════════════════════════════════════════════
print("Loading ENGINE output...", flush=True)
wb_eng = openpyxl.load_workbook(engine_path, data_only=True)
ws_eng = wb_eng['RMC summary']

engine_data = {}
for r in range(7, ws_eng.max_row + 1):
    order = ws_eng.cell(row=r, column=2).value
    if not order:
        continue
    order = str(order).strip().upper()
    engine_data[order] = {
        'row': r,
        'opn_wip_val': ws_eng.cell(row=r, column=17).value or 0,
        'print_film_val': ws_eng.cell(row=r, column=18).value or 0,
        'lam_fresh_val': ws_eng.cell(row=r, column=19).value or 0,
        'slit_input_val': ws_eng.cell(row=r, column=20).value or 0,
        'print_ink_val': ws_eng.cell(row=r, column=21).value or 0,
        'lam_chem_val': ws_eng.cell(row=r, column=22).value or 0,
        'bp_sv_val': ws_eng.cell(row=r, column=23).value or 0,
        'cls_wip_val': ws_eng.cell(row=r, column=24).value or 0,
        'fg_output': ws_eng.cell(row=r, column=25).value or 0,
        'total_cost': ws_eng.cell(row=r, column=26).value or 0,
        'rmc_per_kg': ws_eng.cell(row=r, column=27).value or 0,
    }
wb_eng.close()
print(f"  Engine: {len(engine_data)} orders", flush=True)

# ═══════════════════════════════════════════════════════════════
# Load ground truth (filled reference)
# ═══════════════════════════════════════════════════════════════
print("Loading GROUND TRUTH...", flush=True)
gt_path = os.path.join(base, 'rmc_summary_gt.json')
if os.path.exists(gt_path):
    with open(gt_path) as f:
        gt = json.load(f)
    gt_rows = gt['data_rows']
else:
    print("  Ground truth JSON not found, loading from Excel...", flush=True)
    df_ref = pd.read_excel(filled_path, sheet_name='RMC summary', header=None, engine='openpyxl')
    gt_rows = []
    for i in range(len(df_ref)):
        v = df_ref.iloc[i, 1]
        if pd.notna(v) and isinstance(v, str) and any(c.isdigit() for c in str(v)):
            for j in range(i, len(df_ref)):
                order = df_ref.iloc[j, 1]
                if pd.isna(order): continue
                row = {}
                for c in range(88):
                    cv = df_ref.iloc[j, c]
                    if pd.notna(cv):
                        from openpyxl.utils import get_column_letter
                        row[get_column_letter(c+1)] = cv
                gt_rows.append(row)
            break

ref_data = {}
for row in gt_rows:
    order = str(row.get('B', '')).strip().upper()
    if not order:
        continue
    def gf(col):
        v = row.get(col, 0)
        try: return float(v)
        except: return 0
    ref_data[order] = {
        'opn_wip_val': gf('Q'),
        'print_film_val': gf('R'),
        'lam_fresh_val': gf('S'),
        'slit_input_val': gf('T'),
        'print_ink_val': gf('U'),
        'lam_chem_val': gf('V'),
        'bp_sv_val': gf('W'),
        'cls_wip_val': gf('X'),
        'fg_output': gf('Y'),
        'total_cost': gf('Z'),
        'rmc_per_kg': gf('AA'),
    }
print(f"  Reference: {len(ref_data)} orders", flush=True)

# ═══════════════════════════════════════════════════════════════
# Compare
# ═══════════════════════════════════════════════════════════════
print(f"\n{'='*80}", flush=True)
print("RMC SUMMARY ACCURACY REPORT", flush=True)
print(f"{'='*80}", flush=True)

# Key columns to compare
compare_cols = [
    ('total_cost', 'Total Cost (Z)', 1.0),
    ('rmc_per_kg', 'RMC/Kg (AA)', 0.5),
    ('fg_output', 'FG Output (Y)', 0.5),
    ('opn_wip_val', 'OPN WIP Val (Q)', 0.5),
    ('print_film_val', 'Print Film Val (R)', 0.5),
    ('lam_fresh_val', 'Lam Fresh Val (S)', 0.5),
    ('slit_input_val', 'Slit Input Val (T)', 0.5),
    ('print_ink_val', 'Print Ink Val (U)', 0.5),
    ('lam_chem_val', 'Lam Chem Val (V)', 0.5),
    ('cls_wip_val', 'CLS WIP Val (X)', 0.5),
]

all_orders = set(engine_data.keys()) | set(ref_data.keys())
common = set(engine_data.keys()) & set(ref_data.keys())
print(f"\nOrders: Engine={len(engine_data)}, Reference={len(ref_data)}, Common={len(common)}", flush=True)

for col_key, col_name, tolerance in compare_cols:
    matches = 0
    mismatches = []
    zero_both = 0
    engine_only = 0
    ref_only = 0

    for order in common:
        ev = float(engine_data[order].get(col_key, 0) or 0)
        rv = float(ref_data[order].get(col_key, 0) or 0)

        if ev == 0 and rv == 0:
            zero_both += 1
            matches += 1
        elif abs(ev - rv) <= tolerance:
            matches += 1
        elif abs(ev - rv) / max(abs(rv), 0.01) < 0.01:  # Within 1%
            matches += 1
        else:
            if ev != 0 and rv == 0:
                engine_only += 1
            elif ev == 0 and rv != 0:
                ref_only += 1
            mismatches.append((order, ev, rv, ev - rv))

    total = len(common)
    accuracy = matches / total * 100 if total > 0 else 0
    print(f"\n  {col_name}:", flush=True)
    print(f"    ✅ Match: {matches}/{total} ({accuracy:.1f}%)", flush=True)
    print(f"    ⚪ Both zero: {zero_both}", flush=True)
    if mismatches:
        print(f"    ❌ Mismatch: {len(mismatches)} (engine_only={engine_only}, ref_only={ref_only})", flush=True)
        # Show top 5 mismatches
        mismatches.sort(key=lambda x: abs(x[3]), reverse=True)
        for order, ev, rv, diff in mismatches[:5]:
            print(f"       {order}: engine={ev:.2f}, ref={rv:.2f}, diff={diff:.2f}", flush=True)

# Overall RMC/Kg accuracy
print(f"\n{'='*80}", flush=True)
rmc_matches = 0
rmc_total = 0
for order in common:
    ev = float(engine_data[order].get('rmc_per_kg', 0) or 0)
    rv = float(ref_data[order].get('rmc_per_kg', 0) or 0)
    if rv > 0:
        rmc_total += 1
        if abs(ev - rv) < 0.5 or abs(ev - rv) / rv < 0.01:
            rmc_matches += 1

if rmc_total > 0:
    print(f"OVERALL RMC/Kg ACCURACY: {rmc_matches}/{rmc_total} = {rmc_matches/rmc_total*100:.1f}%", flush=True)
else:
    print("No reference RMC/Kg values to compare", flush=True)
print(f"{'='*80}", flush=True)
