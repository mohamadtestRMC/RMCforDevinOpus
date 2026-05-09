"""Deep analysis: Extract headers, sample data, and formulas from key files."""
import openpyxl
import os

base = r'c:\Users\mohamad.al\OneDrive - AL KHAYYAT INVESTMENTS\Desktop\IPP\Files_need_to_study'

def analyze_sheet(ws, sheet_name, max_rows=15, max_cols=30):
    """Extract headers and sample data from a sheet."""
    print(f'\n  --- Sheet: "{sheet_name}" (rows={ws.max_row}, cols={ws.max_column}) ---')
    actual_cols = min(ws.max_column or 0, max_cols)
    actual_rows = min(ws.max_row or 0, max_rows)
    
    for r in range(1, actual_rows + 1):
        row_data = []
        for c in range(1, actual_cols + 1):
            cell = ws.cell(row=r, column=c)
            val = cell.value
            if val is not None:
                row_data.append(f'[{r},{c}]={repr(val)[:60]}')
        if row_data:
            print(f'    Row {r}: {"; ".join(row_data[:15])}')


# 1. Base RMC Documents (How_to_link)
print('\n' + '='*80)
print('BASE RMC DOCUMENTS (How_to_link)')
print('='*80)
fp = os.path.join(base, 'How_to_link', 'Base RMC Documents (2).xlsx')
wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
for sn in wb.sheetnames:
    analyze_sheet(wb[sn], sn, max_rows=25, max_cols=10)
wb.close()

# 2. Material Names (Filled_Output)
print('\n' + '='*80)
print('MATERIAL NAMES (Filled_Output)')
print('='*80)
fp = os.path.join(base, 'Filled_Output', 'Material Names.xlsx')
wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
for sn in wb.sheetnames:
    analyze_sheet(wb[sn], sn, max_rows=40, max_cols=5)
wb.close()

# 3. Base RMC Documents (Filled_Output)
print('\n' + '='*80)
print('BASE RMC DOCUMENTS (Filled_Output)')
print('='*80)
fp = os.path.join(base, 'Filled_Output', 'Base RMC Documents.xlsx')
wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
for sn in wb.sheetnames:
    analyze_sheet(wb[sn], sn, max_rows=25, max_cols=10)
wb.close()

# 4. MEGAPACK Rate
print('\n' + '='*80)
print('MEGAPACK RATE (Unfilled)')
print('='*80)
fp = os.path.join(base, 'Unfilled', '6 MEGAPACK Rate.xlsx')
wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
for sn in wb.sheetnames:
    analyze_sheet(wb[sn], sn, max_rows=35, max_cols=10)
wb.close()

# 5. Dispense Ink Stock Opening
print('\n' + '='*80)
print('DISPENSE INK STOCK OPENING (Unfilled)')
print('='*80)
fp = os.path.join(base, 'Unfilled', '7 Dispense Ink Stock Opening.xlsx')
wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
for sn in wb.sheetnames:
    analyze_sheet(wb[sn], sn, max_rows=25, max_cols=20)
wb.close()

# 6. Dispensed Stock Movement
print('\n' + '='*80)
print('DISPENSED STOCK MOVEMENT (Unfilled)')
print('='*80)
fp = os.path.join(base, 'Unfilled', '8 Dispensed Stock Movement.xlsx')
wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
for sn in wb.sheetnames:
    analyze_sheet(wb[sn], sn, max_rows=20, max_cols=10)
wb.close()

# 7. Price of Tin Tie, Valve & Spout
print('\n' + '='*80)
print('PRICE OF TIN TIE, VALVE & SPOUT (Unfilled)')
print('='*80)
fp = os.path.join(base, 'Unfilled', '11 Price of Tin Tie, Valve & Spout 2026 updated.xlsx')
wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
for sn in wb.sheetnames:
    analyze_sheet(wb[sn], sn, max_rows=25, max_cols=10)
wb.close()
