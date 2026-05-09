"""
Fast RMC Pipeline — reads with iter_rows (1.5s), computes in memory, writes via xlsxwriter.
Total runtime target: under 30 seconds.
"""
from __future__ import annotations

import json
import re
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
from openpyxl import load_workbook

from .config import RMCPaths

# ---------------------------------------------------------------------------
# Column-index map for RMC summary (1-based Excel columns)
# ---------------------------------------------------------------------------
RMC_REF_COLS = {
    "order": 2, "design": 3, "customer": 4, "sales_code": 5,
    "material": 6, "remarks": 7, "structure": 8,
    "opn_wip_kg": 9, "print_film_kg": 10, "lam_fresh_kg": 11,
    "other_film_kg": 12, "dry_ink_kg": 13, "adh_hard_kg": 14,
    "zip_pe_valve_kg": 15, "cls_wip_kg": 16,
    "opn_wip_val": 17, "print_film_val": 18, "lam_fresh_val": 19,
    "other_film_val": 20, "ink_sol_val": 21, "adh_hard_val": 22,
    "zip_pe_valve_val": 23, "cls_wip_val": 24,
    "output_kg": 25, "total_cost": 26, "rmc_per_kg": 27,
}


def _sf(v) -> float:
    """Safe float conversion."""
    if v is None:
        return 0.0
    try:
        f = float(v)
        return 0.0 if np.isnan(f) or np.isinf(f) else f
    except (ValueError, TypeError):
        return 0.0


def _ss(v) -> str:
    """Safe string."""
    return str(v).strip() if v is not None else ""


# ---------------------------------------------------------------------------
class OrderIndex:
    """Fast order-based SUMIF lookup built from rows + column headers."""

    def __init__(self, headers: List[str], rows: List[tuple], order_col_idx: int):
        self.headers = headers
        self.all_rows = rows  # preserve original order for writing
        self._col_map = {h: i for i, h in enumerate(headers)}
        self._by_order: Dict[str, List[tuple]] = {}
        for row in rows:
            if order_col_idx < len(row):
                o = _ss(row[order_col_idx])
                if o:
                    self._by_order.setdefault(o, []).append(row)

    def _ci(self, col_name: str) -> int:
        return self._col_map.get(col_name, -1)

    def sumif(self, order: str, col_name: str) -> float:
        ci = self._ci(col_name)
        if ci < 0:
            return 0.0
        total = 0.0
        for row in self._by_order.get(order, []):
            if ci < len(row):
                total += _sf(row[ci])
        return total

    def vlookup(self, order: str, col_name: str) -> float:
        ci = self._ci(col_name)
        if ci < 0:
            return 0.0
        entries = self._by_order.get(order, [])
        if entries and ci < len(entries[0]):
            return _sf(entries[0][ci])
        return 0.0


# ---------------------------------------------------------------------------
class FastRMCPipeline:

    def __init__(self, paths: RMCPaths, output_file: Path) -> None:
        self.paths = paths
        self.output_file = output_file
        self.output_file.parent.mkdir(parents=True, exist_ok=True)
        self.log: List[str] = []
        self.metrics: Dict[str, Any] = {}

    def _log(self, msg: str) -> None:
        self.log.append(msg)
        print(msg, flush=True)

    # ------------------------------------------------------------------ read
    def _read_sheet(
        self, wb, sheet_name: str, header_row: int, max_row: int
    ) -> Tuple[List[str], List[tuple]]:
        """Read sheet using fast iter_rows. Returns (headers, data_rows)."""
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(
            min_row=header_row, max_row=max_row, values_only=True
        ))
        if not all_rows:
            return [], []
        raw_headers = all_rows[0]
        headers: List[str] = []
        seen: Dict[str, int] = {}
        for h in raw_headers:
            name = str(h).strip().replace("\n", " ") if h else "_blank"
            if name in seen:
                seen[name] += 1
                name = f"{name}_{seen[name]}"
            else:
                seen[name] = 0
            headers.append(name)
        data_rows = [r for r in all_rows[1:] if any(c is not None for c in r)]
        return headers, data_rows

    def _build_index(
        self, wb, sheet_name: str, header_row: int, max_row: int, order_col: str
    ) -> Optional[OrderIndex]:
        if sheet_name not in wb.sheetnames:
            self._log(f"  WARN: '{sheet_name}' not found")
            return None
        headers, rows = self._read_sheet(wb, sheet_name, header_row, max_row)
        if not headers:
            return None
        oci = headers.index(order_col) if order_col in headers else 0
        idx = OrderIndex(headers, rows, oci)
        self._log(f"  {sheet_name}: {len(rows)} rows, {len(headers)} cols")
        return idx

    # ------------------------------------------------------------------ main
    def run(self, progress_cb=None) -> Dict[str, Any]:
        t0 = time.time()

        def step(n, total, msg):
            self._log(f"[{n}/{total}] {msg}")
            if progress_cb:
                progress_cb(n, total, msg)

        step(1, 7, "Opening filled reference (read_only + iter_rows)...")
        wb = load_workbook(
            str(self.paths.filled_base_rmc), read_only=True, data_only=True
        )

        step(2, 7, "Building sheet indexes...")
        idx = self._build_all_indexes(wb)

        step(3, 7, "Reading RMC summary reference for validation...")
        rmc_ref_rows = self._read_rmc_ref(wb)
        wb.close()
        self._log(f"  Read phase done in {time.time()-t0:.1f}s")

        step(4, 7, "Extracting formula offsets & special cases...")
        offsets, transfer_orders, other_film_orders, combined_orders = (
            self._extract_offsets_from_formulas()
        )

        step(5, 7, "Computing RMC summary...")
        rmc_rows = self._compute_rmc_summary(
            rmc_ref_rows, idx, offsets, transfer_orders,
            other_film_orders, combined_orders,
        )

        step(6, 7, "Writing output xlsx via xlsxwriter...")
        self._write_output(idx, rmc_rows)

        step(7, 7, "Validating...")
        val = self._validate(rmc_ref_rows, rmc_rows)

        elapsed = time.time() - t0
        self.metrics["elapsed_seconds"] = round(elapsed, 1)
        self.metrics.update(val)

        report_path = self.output_file.parent / "validation_report.json"
        report_path.write_text(
            json.dumps({"metrics": self.metrics, "log": self.log}, indent=2, default=str),
            encoding="utf-8",
        )
        self._log(f"Total: {elapsed:.1f}s  Accuracy: {val.get('accuracy_pct', 0):.1f}%")
        return {
            "output_file": str(self.output_file),
            "report_file": str(report_path),
            "metrics": self.metrics,
            "log": self.log,
        }

    def _extract_offsets_from_formulas(self):
        """Read formulas (not data_only) to extract offset constants and special cases.

        Formulas may contain chained additions/subtractions like:
          =(((SUMIF(...))+845.2)+92)+801.57
          =(...)-361.50)-767.45)-785.98
        We extract ALL numeric offsets and sum them.
        """
        wb = load_workbook(
            str(self.paths.filled_base_rmc), read_only=True, data_only=False
        )
        ws = wb["RMC summary"]
        all_rows = list(ws.iter_rows(min_row=7, max_row=650, values_only=True))
        wb.close()

        offset_cols = {
            8: "opn_wip_kg", 9: "print_film_kg", 10: "lam_fresh_kg",
            12: "dry_ink_kg", 13: "adh_hard_kg", 14: "zip_pe_valve_kg", 15: "cls_wip_kg",
            16: "opn_wip_val", 17: "print_film_val", 18: "lam_fresh_val",
            20: "ink_sol_val", 21: "adh_hard_val", 22: "zip_pe_valve_val", 23: "cls_wip_val",
        }

        # Match all numeric offsets after closing parens: )+123.45 or )-123.45
        offset_token = re.compile(r'\)\s*([+-])\s*([\d.]+)')

        offsets: Dict[str, Dict[str, float]] = {}
        transfer_orders: set = set()
        other_film_orders: set = set()
        combined_orders: set = set()

        for row in all_rows:
            if not row or row[1] is None:
                continue
            order = _ss(row[1])
            remarks = _ss(row[6])

            if "transfer" in remarks.lower():
                transfer_orders.add(order)
                continue

            row_offsets: Dict[str, float] = {}
            for ci, name in offset_cols.items():
                val = row[ci] if ci < len(row) else None
                if val is None or not str(val).startswith("="):
                    continue
                fstr = str(val)
                # Find all offset tokens: )+num or )-num
                tokens = offset_token.findall(fstr)
                total_offset = 0.0
                for sign, num_str in tokens:
                    try:
                        v = float(num_str)
                        total_offset += v if sign == "+" else -v
                    except ValueError:
                        pass
                if abs(total_offset) > 0.001:
                    row_offsets[name] = total_offset

            other_f = str(row[11] or "")
            if "SUMIFS" in other_f:
                other_film_orders.add(order)

            output_f = str(row[24] or "")
            if "VLOOKUP(A" in output_f and "VLOOKUP(B" in output_f:
                combined_orders.add(order)

            if row_offsets:
                offsets[order] = row_offsets

        self._log(
            f"  Offsets: {len(offsets)} orders, Transfers: {len(transfer_orders)}, "
            f"Other Film SUMIFS: {len(other_film_orders)}, Combined: {len(combined_orders)}"
        )
        return offsets, transfer_orders, other_film_orders, combined_orders

    def _build_all_indexes(self, wb) -> Dict[str, OrderIndex]:
        idx: Dict[str, OrderIndex] = {}
        config = [
            ("BFL",         6, 200, "Order No"),
            ("Print",       6, 400, "Order  No"),
            ("Lam",         6, 700, "Order No"),
            ("Slit",        6, 550, "Order No"),
            ("Bag&Pouch",   6, 100, "Order No"),
            ("Spout&Valve", 6, 20,  "Order No"),
            ("PTR Rew",     6, 200, "Order No"),
            ("HCI Rew",     6, 150, "Order No"),
            ("OPN_WIP",     5, 300, "W/O"),
            ("CLS_WIP",     5, 300, "W/O"),
        ]
        for sn, hr, mr, oc in config:
            result = self._build_index(wb, sn, hr, mr, oc)
            if result:
                idx[sn] = result

        # FG: special layout — col 0 = "Row Labels", col 6 = "Final FG"
        if "FG" in wb.sheetnames:
            ws = wb["FG"]
            fg_rows_raw = list(ws.iter_rows(min_row=4, max_row=650, values_only=True))
            fg_headers = ["Row Labels", "Raw Output", "_c3", "_c4", "_c5", "HCI Wastage", "Final FG"]
            fg_data = []
            for row in fg_rows_raw:
                if row and row[0] is not None:
                    padded = tuple(row[i] if i < len(row) else None for i in range(7))
                    fg_data.append(padded)
            idx["FG"] = OrderIndex(fg_headers, fg_data, 0)
            self._log(f"  FG: {len(fg_data)} rows")

        return idx

    def _read_rmc_ref(self, wb) -> List[Dict[str, Any]]:
        """Read RMC summary reference values for validation."""
        ws = wb["RMC summary"]
        all_rows = list(ws.iter_rows(min_row=7, max_row=650, values_only=True))
        result = []
        for row in all_rows:
            if not row or row[1] is None:  # col B (index 1) = Order No
                continue
            d: Dict[str, Any] = {}
            for key, col_1based in RMC_REF_COLS.items():
                ci = col_1based - 1  # 0-based
                d[key] = row[ci] if ci < len(row) else None
            # Column A (index 0) holds combined-order reference like "L00328/L00334"
            d["_combined_ref"] = _ss(row[0]) if row[0] else ""
            result.append(d)
        return result

    # -------------------------------------------------------------- compute
    def _compute_rmc_summary(
        self,
        rmc_ref_rows: List[Dict],
        idx: Dict[str, OrderIndex],
        offsets: Dict[str, Dict[str, float]],
        transfer_orders: set,
        other_film_orders: set,
        combined_orders: set,
    ) -> List[Dict[str, Any]]:

        prn = idx.get("Print")
        lam = idx.get("Lam")
        slit = idx.get("Slit")
        bp = idx.get("Bag&Pouch")
        sv = idx.get("Spout&Valve")
        bfl = idx.get("BFL")
        hci = idx.get("HCI Rew")
        ptr = idx.get("PTR Rew")
        opn = idx.get("OPN_WIP")
        cls = idx.get("CLS_WIP")
        fg = idx.get("FG")

        result = []
        for ref in rmc_ref_rows:
            order = _ss(ref.get("order"))
            if not order:
                continue

            r: Dict[str, Any] = {}
            r["Order No"] = order
            r["Design Name"] = _ss(ref.get("design"))
            r["Customer Name"] = _ss(ref.get("customer"))
            r["Sales Code"] = _ss(ref.get("sales_code"))
            r["Material"] = _ss(ref.get("material"))
            r["Remarks"] = _ss(ref.get("remarks"))
            r["Structure"] = _ss(ref.get("structure"))

            # Transfer orders get zeroed out
            if order in transfer_orders:
                for col in [
                    "Opening WIP (Kg)", "Printing Film Input (Kgs)", "Lam Fresh Mat (Kgs)",
                    "Other Film Input (Kgs)", "Dry Ink (Kgs)", "Adh+ Hard Solids (Kgs)",
                    "Zipper + PE strip+ Valve  (Kgs)", "Closing WIP (Kg)",
                    "Opening WIP Value (AED)", "Printing Film Value (AED)",
                    "Lam Fresh Mat Value (AED)", "Other Film Value (AED)",
                    "Ink & Sol Value (AED)", "Adh+ Hard +Sol Value (AED)",
                    "Zipper + PE strip +Valve Value (AED)", "Closing WIP Value (AED)",
                    "Prod / Output (Kg)", "Total Cost", "Prod RMC / Kg",
                    "Input Output check",
                    "BFL Wastage Qty", "Print Wastage Qty", "Lam Wastage Qty",
                    "Slit Wastage Qty", "B&P Wastage Qty", "S&V Wastage Qty",
                    "HCI Wastage Qty", "PTR Wastage Qty",
                    "BFL Wastage Val", "Print Wastage Val", "Lam Wastage Val",
                    "Slit Wastage Val", "B&P Wastage Val", "S&V Wastage Val",
                    "HCI Wastage Val", "PTR Wastage Val",
                ]:
                    r[col] = 0.0
                result.append(r)
                continue

            # Get offsets for this order
            ofs = offsets.get(order, {})

            # ---- Quantities (Kg) — SUMIF + offset ----
            r["Opening WIP (Kg)"] = (
                (opn.sumif(order, "Qty") if opn else 0.0) + ofs.get("opn_wip_kg", 0.0)
            )
            r["Printing Film Input (Kgs)"] = (
                (prn.sumif(order, "Film Input (Kgs)") if prn else 0.0) + ofs.get("print_film_kg", 0.0)
            )
            r["Lam Fresh Mat (Kgs)"] = (
                (lam.sumif(order, "Fresh Mat Qty") if lam else 0.0) + ofs.get("lam_fresh_kg", 0.0)
            )
            r["Dry Ink (Kgs)"] = (
                (prn.sumif(order, "Dry Ink (Kgs)") if prn else 0.0) + ofs.get("dry_ink_kg", 0.0)
            )
            r["Adh+ Hard Solids (Kgs)"] = (
                (lam.sumif(order, "Adh+Hard Solids Qty") if lam else 0.0) + ofs.get("adh_hard_kg", 0.0)
            )

            bp_qty = bp.sumif(order, "PE STRIP + ZIPPER Qty") if bp else 0.0
            sv_qty = sv.sumif(order, "TIN TIE+Valve+Spout Qty") if sv else 0.0
            r["Zipper + PE strip+ Valve  (Kgs)"] = (
                bp_qty + sv_qty + ofs.get("zip_pe_valve_kg", 0.0)
            )

            r["Closing WIP (Kg)"] = (
                (cls.sumif(order, "Qty") if cls else 0.0) + ofs.get("cls_wip_kg", 0.0)
            )

            # Other Film: only for orders with SUMIFS formula
            if order in other_film_orders and slit:
                r["Other Film Input (Kgs)"] = slit.sumif(order, "Input (Kgs)")
            else:
                r["Other Film Input (Kgs)"] = 0.0

            # ---- Values (AED) — SUMIF + offset ----
            r["Opening WIP Value (AED)"] = (
                (opn.sumif(order, "Value") if opn else 0.0) + ofs.get("opn_wip_val", 0.0)
            )
            r["Printing Film Value (AED)"] = (
                (prn.sumif(order, "Film Value") if prn else 0.0) + ofs.get("print_film_val", 0.0)
            )
            r["Lam Fresh Mat Value (AED)"] = (
                (lam.sumif(order, "Fresh Mat Value") if lam else 0.0) + ofs.get("lam_fresh_val", 0.0)
            )
            r["Ink & Sol Value (AED)"] = (
                (prn.sumif(order, "Ink Value") if prn else 0.0) + ofs.get("ink_sol_val", 0.0)
            )
            r["Adh+ Hard +Sol Value (AED)"] = (
                (lam.sumif(order, "Adh+Hard +Solv Val") if lam else 0.0) + ofs.get("adh_hard_val", 0.0)
            )

            bp_val = bp.sumif(order, "PE STRIP + ZIPPER Value") if bp else 0.0
            sv_val = sv.sumif(order, "TIN TIE+Valve+Spout Value") if sv else 0.0
            r["Zipper + PE strip +Valve Value (AED)"] = (
                bp_val + sv_val + ofs.get("zip_pe_valve_val", 0.0)
            )

            r["Closing WIP Value (AED)"] = (
                (cls.sumif(order, "Value") if cls else 0.0) + ofs.get("cls_wip_val", 0.0)
            )

            if order in other_film_orders and slit:
                r["Other Film Value (AED)"] = slit.sumif(order, "Slitting Input Val (AED)")
            else:
                r["Other Film Value (AED)"] = 0.0

            # ---- Output from FG ----
            if order in combined_orders and fg:
                # Combined orders: look up both the combined ref (col A) and order (col B)
                # The combined ref is stored as col A in the reference
                combined_ref = _ss(ref.get("_combined_ref"))
                r["Prod / Output (Kg)"] = (
                    fg.vlookup(order, "Final FG")
                    + (fg.vlookup(combined_ref, "Final FG") if combined_ref else 0.0)
                )
            else:
                r["Prod / Output (Kg)"] = fg.vlookup(order, "Final FG") if fg else 0.0

            # ---- Total Cost & RMC ----
            total = (
                r["Opening WIP Value (AED)"]
                + r["Printing Film Value (AED)"]
                + r["Lam Fresh Mat Value (AED)"]
                + r["Other Film Value (AED)"]
                + r["Ink & Sol Value (AED)"]
                + r["Adh+ Hard +Sol Value (AED)"]
                + r["Zipper + PE strip +Valve Value (AED)"]
                - r["Closing WIP Value (AED)"]
            )
            r["Total Cost"] = total
            output_kg = r["Prod / Output (Kg)"]
            r["Prod RMC / Kg"] = total / output_kg if output_kg > 0 else 0.0

            # ---- Wastage ----
            r["BFL Wastage Qty"] = bfl.sumif(order, "Wastage  (Kgs)") if bfl else 0.0
            r["Print Wastage Qty"] = prn.sumif(order, "Wastage Qty (Calc)") if prn else 0.0
            r["Lam Wastage Qty"] = lam.sumif(order, "Wastage (Calc)") if lam else 0.0
            r["Slit Wastage Qty"] = slit.sumif(order, "Wastage (Kgs)") if slit else 0.0
            r["B&P Wastage Qty"] = bp.sumif(order, "Wastage (Calc)") if bp else 0.0
            r["S&V Wastage Qty"] = sv.sumif(order, "Wastage (Calc)") if sv else 0.0
            r["HCI Wastage Qty"] = hci.sumif(order, "Sum of Wastage (Calc)") if hci else 0.0
            r["PTR Wastage Qty"] = ptr.sumif(order, "Total Wastage Qty") if ptr else 0.0

            r["BFL Wastage Val"] = bfl.sumif(order, "Wastage  value (AED)") if bfl else 0.0
            r["Print Wastage Val"] = prn.sumif(order, "Wastage Value (AED)") if prn else 0.0
            r["Lam Wastage Val"] = lam.sumif(order, "Wastage (AED)") if lam else 0.0
            r["Slit Wastage Val"] = slit.sumif(order, "Wastage Val (AED)") if slit else 0.0
            r["B&P Wastage Val"] = bp.sumif(order, "Wastage (AED)") if bp else 0.0
            r["S&V Wastage Val"] = sv.sumif(order, "Wastage (AED)") if sv else 0.0
            r["HCI Wastage Val"] = hci.sumif(order, "Wastage Value (AED)") if hci else 0.0
            r["PTR Wastage Val"] = ptr.sumif(order, "Total Wastage Value") if ptr else 0.0

            # I/O check
            total_input = (
                r["Opening WIP (Kg)"]
                + r["Printing Film Input (Kgs)"]
                + r["Lam Fresh Mat (Kgs)"]
                + r["Other Film Input (Kgs)"]
            )
            total_waste = sum(r.get(k, 0.0) for k in [
                "BFL Wastage Qty", "Print Wastage Qty", "Lam Wastage Qty",
                "Slit Wastage Qty", "B&P Wastage Qty", "S&V Wastage Qty",
                "HCI Wastage Qty", "PTR Wastage Qty",
            ])
            r["Input Output check"] = total_input - output_kg - r["Closing WIP (Kg)"] - total_waste

            result.append(r)

        self.metrics["rmc_summary_orders"] = len(result)
        self._log(f"  Computed {len(result)} RMC summary rows")
        return result

    # ---------------------------------------------------------------- write
    def _write_output(self, idx: Dict[str, OrderIndex], rmc_rows: List[Dict]) -> None:
        import xlsxwriter
        from datetime import datetime

        out_path = str(self.output_file)
        xwb = xlsxwriter.Workbook(out_path)

        hdr_fmt = xwb.add_format({
            "bold": True, "bg_color": "#4472C4", "font_color": "white",
            "border": 1, "text_wrap": True, "valign": "vcenter",
        })
        num_fmt = xwb.add_format({"num_format": "#,##0.00"})
        date_fmt = xwb.add_format({"num_format": "dd-mmm-yyyy"})

        rmc_col_order = [
            "Order No", "Design Name", "Customer Name", "Sales Code",
            "Material", "Remarks", "Structure",
            "Opening WIP (Kg)", "Printing Film Input (Kgs)", "Lam Fresh Mat (Kgs)",
            "Other Film Input (Kgs)", "Dry Ink (Kgs)", "Adh+ Hard Solids (Kgs)",
            "Zipper + PE strip+ Valve  (Kgs)", "Closing WIP (Kg)",
            "Opening WIP Value (AED)", "Printing Film Value (AED)", "Lam Fresh Mat Value (AED)",
            "Other Film Value (AED)", "Ink & Sol Value (AED)", "Adh+ Hard +Sol Value (AED)",
            "Zipper + PE strip +Valve Value (AED)", "Closing WIP Value (AED)",
            "Prod / Output (Kg)", "Total Cost", "Prod RMC / Kg",
            "Input Output check",
            "BFL Wastage Qty", "Print Wastage Qty", "Lam Wastage Qty",
            "Slit Wastage Qty", "B&P Wastage Qty", "S&V Wastage Qty",
            "HCI Wastage Qty", "PTR Wastage Qty",
            "BFL Wastage Val", "Print Wastage Val", "Lam Wastage Val",
            "Slit Wastage Val", "B&P Wastage Val", "S&V Wastage Val",
            "HCI Wastage Val", "PTR Wastage Val",
        ]
        text_cols = set(rmc_col_order[:7])

        # -- RMC summary sheet --
        ws = xwb.add_worksheet("RMC summary")
        ws.set_column(0, 0, 12)
        ws.set_column(1, 1, 45)
        ws.set_column(2, 2, 25)
        ws.set_column(3, 6, 15)
        ws.set_column(7, len(rmc_col_order) - 1, 14)
        ws.freeze_panes(6, 1)
        ws.write(0, 0, "INTEGRATED PLASTICS PACKAGING")
        ws.write(1, 0, "Raw Material Consumption Summary")
        for ci, cn in enumerate(rmc_col_order):
            ws.write(5, ci, cn, hdr_fmt)
        for ri, row in enumerate(rmc_rows):
            for ci, cn in enumerate(rmc_col_order):
                val = row.get(cn, "")
                if cn in text_cols:
                    ws.write(6 + ri, ci, _ss(val))
                else:
                    ws.write_number(6 + ri, ci, _sf(val), num_fmt)

        # -- Process & WIP sheets (preserve original row order) --
        for sn, oi in idx.items():
            if sn == "FG":
                continue
            wsx = xwb.add_worksheet(sn[:31])
            sr = 5 if sn not in ("OPN_WIP", "CLS_WIP") else 4
            for ci, h in enumerate(oi.headers):
                wsx.write(sr, ci, h, hdr_fmt)
            for ri, row_tuple in enumerate(oi.all_rows):
                for ci, val in enumerate(row_tuple):
                    if val is None:
                        continue
                    if isinstance(val, datetime):
                        wsx.write_datetime(sr + 1 + ri, ci, val, date_fmt)
                    elif isinstance(val, (int, float)):
                        wsx.write_number(sr + 1 + ri, ci, val, num_fmt)
                    else:
                        wsx.write(sr + 1 + ri, ci, str(val))

        # -- FG sheet --
        fg_oi = idx.get("FG")
        if fg_oi:
            ws_fg = xwb.add_worksheet("FG")
            for ci, h in enumerate(fg_oi.headers):
                ws_fg.write(2, ci, h, hdr_fmt)
            for ri, row_tuple in enumerate(fg_oi.all_rows):
                for ci, val in enumerate(row_tuple):
                    if val is None:
                        continue
                    if isinstance(val, (int, float)):
                        ws_fg.write_number(3 + ri, ci, val, num_fmt)
                    else:
                        ws_fg.write(3 + ri, ci, str(val))

        xwb.close()
        self._log(f"  Output: {out_path}")

    # -------------------------------------------------------------- validate
    def _validate(
        self, rmc_ref_rows: List[Dict], rmc_rows: List[Dict]
    ) -> Dict[str, Any]:

        ref_by_order = {_ss(r.get("order")): r for r in rmc_ref_rows if _ss(r.get("order"))}

        check_pairs = [
            ("Opening WIP (Kg)", "opn_wip_kg"),
            ("Printing Film Input (Kgs)", "print_film_kg"),
            ("Lam Fresh Mat (Kgs)", "lam_fresh_kg"),
            ("Other Film Input (Kgs)", "other_film_kg"),
            ("Dry Ink (Kgs)", "dry_ink_kg"),
            ("Adh+ Hard Solids (Kgs)", "adh_hard_kg"),
            ("Zipper + PE strip+ Valve  (Kgs)", "zip_pe_valve_kg"),
            ("Closing WIP (Kg)", "cls_wip_kg"),
            ("Opening WIP Value (AED)", "opn_wip_val"),
            ("Printing Film Value (AED)", "print_film_val"),
            ("Lam Fresh Mat Value (AED)", "lam_fresh_val"),
            ("Other Film Value (AED)", "other_film_val"),
            ("Ink & Sol Value (AED)", "ink_sol_val"),
            ("Adh+ Hard +Sol Value (AED)", "adh_hard_val"),
            ("Zipper + PE strip +Valve Value (AED)", "zip_pe_valve_val"),
            ("Closing WIP Value (AED)", "cls_wip_val"),
            ("Prod / Output (Kg)", "output_kg"),
            ("Total Cost", "total_cost"),
            ("Prod RMC / Kg", "rmc_per_kg"),
        ]

        total = 0
        exact = 0
        close = 0
        mismatches = []

        for row in rmc_rows:
            order = row.get("Order No", "")
            ref = ref_by_order.get(order)
            if not ref:
                continue
            for comp_col, ref_key in check_pairs:
                cv = _sf(row.get(comp_col))
                rv = _sf(ref.get(ref_key))
                total += 1
                d = abs(cv - rv)
                if d <= 0.01:
                    exact += 1
                elif d <= 1.0:
                    close += 1
                else:
                    mismatches.append({
                        "order": order,
                        "col": comp_col,
                        "computed": round(cv, 4),
                        "reference": round(rv, 4),
                        "diff": round(cv - rv, 4),
                    })

        acc = exact / total * 100 if total > 0 else 0
        close_acc = (exact + close) / total * 100 if total > 0 else 0
        mismatches.sort(key=lambda x: abs(x["diff"]), reverse=True)

        self._log(
            f"  {exact}/{total} exact ({acc:.1f}%), "
            f"{close} close (<1), {len(mismatches)} mismatches (>1)"
        )
        return {
            "total_checks": total,
            "exact_matches": exact,
            "close_lt1": close,
            "mismatches_gt1": len(mismatches),
            "accuracy_pct": round(acc, 2),
            "close_accuracy_pct": round(close_acc, 2),
            "top_mismatches": mismatches[:100],
        }
