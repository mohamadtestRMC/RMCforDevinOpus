from __future__ import annotations

from pathlib import Path
from typing import Iterable, Optional, Tuple

import openpyxl


def load_wb(path: Path, data_only: bool = False) -> openpyxl.Workbook:
    return openpyxl.load_workbook(path, data_only=data_only)


def first_sheet(workbook: openpyxl.Workbook) -> openpyxl.worksheet.worksheet.Worksheet:
    return workbook[workbook.sheetnames[0]]


def find_header_row(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    must_include: Iterable[str],
    max_scan_rows: int = 40,
) -> Optional[int]:
    wanted = [x.strip().lower() for x in must_include]
    for r in range(1, min(max_scan_rows, ws.max_row) + 1):
        vals = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            vals.append(str(v).strip().lower())
        if all(any(w in cell for cell in vals) for w in wanted):
            return r
    return None


def clear_data_rows(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    start_row: int,
    from_col: int = 1,
    to_col: Optional[int] = None,
    end_row: Optional[int] = None,
) -> None:
    max_col = to_col or ws.max_column
    max_row = end_row or ws.max_row
    for r in range(start_row, max_row + 1):
        for c in range(from_col, max_col + 1):
            ws.cell(r, c).value = None


def copy_values_block(
    src_ws: openpyxl.worksheet.worksheet.Worksheet,
    dst_ws: openpyxl.worksheet.worksheet.Worksheet,
    src_start_row: int,
    dst_start_row: int,
    max_cols: Optional[int] = None,
    max_scan_rows: int = 10000,
    blank_streak_stop: int = 100,
) -> Tuple[int, int]:
    cols = max_cols or src_ws.max_column
    row_count = 0
    blank_streak = 0
    for sr in range(src_start_row, src_start_row + max_scan_rows):
        row_has_any = False
        for c in range(1, cols + 1):
            v = src_ws.cell(sr, c).value
            if v is not None:
                row_has_any = True
            dst_ws.cell(dst_start_row + row_count, c).value = v
        if row_has_any:
            blank_streak = 0
            row_count += 1
        else:
            blank_streak += 1
            if blank_streak >= blank_streak_stop:
                break
    return row_count, cols


def clone_formula_grid(
    src_ws: openpyxl.worksheet.worksheet.Worksheet,
    dst_ws: openpyxl.worksheet.worksheet.Worksheet,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
) -> int:
    copied = 0
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            v = src_ws.cell(r, c).value
            if isinstance(v, str) and v.startswith("="):
                dst_ws.cell(r, c).value = v
                copied += 1
    return copied

