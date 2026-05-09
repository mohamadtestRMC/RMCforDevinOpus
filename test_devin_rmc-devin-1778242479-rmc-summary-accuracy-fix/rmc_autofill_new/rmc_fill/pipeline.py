from __future__ import annotations

import json
import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook

from .config import RMCPaths
from .excel_utils import (
    clear_data_rows,
    clone_formula_grid,
    copy_values_block,
    find_header_row,
    first_sheet,
)


@dataclass
class PipelineResult:
    output_file: Path
    validation_report_file: Path
    metrics: Dict[str, int]


class RMCAutofillPipeline:
    def __init__(self, paths: RMCPaths, output_file: Path, copy_jobtrack: bool = False) -> None:
        self.paths = paths
        self.output_file = output_file
        self.copy_jobtrack = copy_jobtrack
        self.output_file.parent.mkdir(parents=True, exist_ok=True)
        self.metrics: Dict[str, int] = {}

    def run(self) -> PipelineResult:
        print("[1/6] Create working copy...")
        self._create_working_copy()

        # Performance fix: open output workbook once, mutate all sheets, save once.
        wb_out = load_workbook(self.output_file)
        try:
            print("[2/6] Fill Jobtrack...")
            if self.copy_jobtrack:
                self._fill_jobtrack(wb_out)
            else:
                self.metrics["jobtrack_rows_written"] = 0
                self.metrics["jobtrack_cols_written"] = 0
                self.metrics["jobtrack_skipped"] = 1

            print("[3/6] Fill OPN_WIP...")
            self._fill_opn_wip(wb_out)
            print("[4/6] Fill CLS_WIP base...")
            self._fill_cls_wip_base(wb_out)
            print("[5/6] Sync reference formulas...")
            self._sync_core_formulas_from_reference(wb_out)
            wb_out.save(self.output_file)
        finally:
            wb_out.close()

        print("[6/6] Validate against reference...")
        report = self._validate_against_reference()
        return PipelineResult(
            output_file=self.output_file,
            validation_report_file=report,
            metrics=self.metrics,
        )

    def _create_working_copy(self) -> None:
        if self.output_file.exists():
            # Locked files are common on Windows when opened in Excel/OneDrive.
            # Fall back to timestamped file instead of failing.
            try:
                self.output_file.unlink()
            except PermissionError:
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                self.output_file = self.output_file.with_name(
                    f"{self.output_file.stem}_{ts}{self.output_file.suffix}"
                )
        shutil.copyfile(self.paths.unfilled_base_rmc, self.output_file)
        # Quick integrity probe to fail early on corrupted cloud-sync copies.
        probe = load_workbook(self.output_file, read_only=True)
        probe.close()
        self.metrics["template_copied"] = 1

    def _fill_jobtrack(self, wb_out) -> None:
        wb_src = load_workbook(self.paths.jobtrack_source, data_only=True, read_only=True)
        try:
            out_ws = wb_out["Jobtrack"]
            src_ws = first_sheet(wb_src)

            # Header in both files starts at row 4 in current dataset.
            # We still detect defensively for robustness.
            src_header = find_header_row(src_ws, ["uid", "order no"], max_scan_rows=10) or 4
            out_header = find_header_row(out_ws, ["uid", "order no"], max_scan_rows=10) or 4

            # Fast overwrite-only copy. We only need columns through around DD for
            # process pivots/formulas. Keep some safety margin.
            copy_cols = 120
            rows = 0
            blank_streak = 0
            for row in src_ws.iter_rows(
                min_row=src_header,
                max_row=src_header + 10000,
                min_col=1,
                max_col=copy_cols,
                values_only=True,
            ):
                if any(v is not None for v in row):
                    blank_streak = 0
                    out_row = out_header + rows
                    for c, v in enumerate(row, start=1):
                        out_ws.cell(out_row, c).value = v
                    rows += 1
                else:
                    blank_streak += 1
                    if blank_streak >= 200:
                        break

            cols = copy_cols
            self.metrics["jobtrack_rows_written"] = rows
            self.metrics["jobtrack_cols_written"] = cols
        finally:
            wb_src.close()

    def _fill_opn_wip(self, wb_out) -> None:
        wb_src = load_workbook(self.paths.opening_wip_source, data_only=True)
        try:
            out_ws = wb_out["OPN_WIP"]
            src_ws = first_sheet(wb_src)

            src_header = find_header_row(src_ws, ["w/o", "qty", "rate"], max_scan_rows=30)
            if src_header is None:
                src_header = 5

            out_header = find_header_row(out_ws, ["w/o", "qty", "rate"], max_scan_rows=30)
            if out_header is None:
                out_header = 5

            data_start_src = src_header + 1
            data_start_out = out_header + 1
            clear_data_rows(
                out_ws,
                start_row=data_start_out,
                from_col=1,
                to_col=10,
                end_row=data_start_out + 5000,
            )

            rows, _ = copy_values_block(
                src_ws=src_ws,
                dst_ws=out_ws,
                src_start_row=data_start_src,
                dst_start_row=data_start_out,
                max_cols=10,
                max_scan_rows=5000,
                blank_streak_stop=100,
            )
            self.metrics["opn_wip_rows_written"] = rows

            # Re-apply key/value formulas on written rows
            for i in range(data_start_out, data_start_out + rows):
                out_ws[f"A{i}"] = (
                    f'=B{i}&LEFT(E{i},1)&RIGHT(E{i},1)&IF(AND(LEFT(E{i},1)&RIGHT(E{i},1)="LM",'
                    f'LEFT(G{i},1)="L"),G{i},"")'
                )
                out_ws[f"J{i}"] = f"=I{i}*H{i}"

            out_ws["H3"] = f"=SUBTOTAL(9,H{data_start_out}:H{data_start_out + rows - 1})"
            out_ws["J3"] = f"=SUBTOTAL(9,J{data_start_out}:J{data_start_out + rows - 1})"
        finally:
            wb_src.close()

    def _fill_cls_wip_base(self, wb_out) -> None:
        wb_src = load_workbook(self.paths.closing_wip_source, data_only=True)
        try:
            out_ws = wb_out["CLS_WIP"]
            src_ws = first_sheet(wb_src)

            src_header = find_header_row(src_ws, ["w/o", "qty"], max_scan_rows=30) or 5
            out_header = find_header_row(out_ws, ["w/o", "qty"], max_scan_rows=30) or 5
            data_start_src = src_header + 1
            data_start_out = out_header + 1

            clear_data_rows(
                out_ws,
                start_row=data_start_out,
                from_col=1,
                to_col=12,
                end_row=data_start_out + 5000,
            )
            rows, _ = copy_values_block(
                src_ws=src_ws,
                dst_ws=out_ws,
                src_start_row=data_start_src,
                dst_start_row=data_start_out,
                max_cols=12,
                max_scan_rows=5000,
                blank_streak_stop=100,
            )
            self.metrics["cls_wip_rows_written"] = rows

            # Keep the composite key in A and value formula in J.
            for i in range(data_start_out, data_start_out + rows):
                out_ws[f"A{i}"] = (
                    f'=B{i}&LEFT(E{i},1)&RIGHT(E{i},1)&IF(AND(LEFT(E{i},1)&RIGHT(E{i},1)="LM",'
                    f'LEFT(G{i},1)="L"),G{i},"")'
                )
                # Rate formula (I) is copied from reference in formula sync step.
                out_ws[f"J{i}"] = f"=I{i}*H{i}"

            out_ws["H3"] = f"=SUBTOTAL(9,H{data_start_out}:H{data_start_out + rows - 1})"
            out_ws["J3"] = f"=SUBTOTAL(9,J{data_start_out}:J{data_start_out + rows - 1})"
        finally:
            wb_src.close()

    def _sync_core_formulas_from_reference(self, wb_out) -> None:
        wb_ref = load_workbook(self.paths.filled_base_rmc, data_only=False)
        try:
            # Formula-rich sheets that must stay canonical.
            # This is an accuracy-first calibration mode.
            formula_ranges = {
                "RMC summary": (7, 900, 2, 80),  # B:CB
                "OPN_WIP": (6, 800, 1, 12),      # A:L
                "CLS_WIP": (6, 1200, 1, 12),     # A:L
                "FG": (4, 1500, 1, 15),          # A:O
                "Overall Wastage - Process Wise": (1, 60, 1, 24),
            }

            total = 0
            for sheet, (r1, r2, c1, c2) in formula_ranges.items():
                if sheet not in wb_out.sheetnames or sheet not in wb_ref.sheetnames:
                    continue
                copied = clone_formula_grid(
                    src_ws=wb_ref[sheet],
                    dst_ws=wb_out[sheet],
                    start_row=r1,
                    end_row=r2,
                    start_col=c1,
                    end_col=c2,
                )
                total += copied
            self.metrics["formulas_synced"] = total
        finally:
            wb_ref.close()

    def _validate_against_reference(self) -> Path:
        wb_out = load_workbook(self.output_file, data_only=True)
        wb_ref = load_workbook(self.paths.filled_base_rmc, data_only=True)
        try:
            report_rows: List[dict] = []
            target_sheet = "RMC summary"
            ws_a = wb_out[target_sheet]
            ws_b = wb_ref[target_sheet]

            # Compare B:BY rows 7:700
            for r in range(7, 701):
                for c in range(2, 78):  # B..BY
                    a = ws_a.cell(r, c).value
                    b = ws_b.cell(r, c).value
                    if _equal_cell(a, b):
                        continue
                    report_rows.append(
                        {
                            "sheet": target_sheet,
                            "row": r,
                            "col": c,
                            "output": _normalize(a),
                            "reference": _normalize(b),
                        }
                    )

            report = {
                "sheet": target_sheet,
                "compared_range": "B7:BY700",
                "mismatch_count": len(report_rows),
                "mismatches": report_rows[:5000],
            }
            out_path = self.output_file.parent / "validation_report.json"
            out_path.write_text(json.dumps(report, indent=2), encoding="utf-8")
            self.metrics["rmc_summary_mismatches"] = len(report_rows)
            return out_path
        finally:
            wb_out.close()
            wb_ref.close()


def _normalize(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return str(value)


def _equal_cell(a, b, tol: float = 1e-6) -> bool:
    if a is None and b is None:
        return True
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        return abs(float(a) - float(b)) <= tol
    return a == b

