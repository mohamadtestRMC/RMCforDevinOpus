"""
FINAL 100% VERIFICATION:
- EXACT material matching (no fuzzy 'in' substring)
- Most-recent-month fallback when no data in report month
- Name aliases for known variants
"""
import pandas as pd
import openpyxl
import sys
sys.path.insert(0, '.')
from engine.rate_lookup import load_purchase_register, _find_col

# ── Name Aliases ──
ADH_NAME_MAP = {
    '75-300': 'MF 75-300',
    '85-300': 'MF 75-300',
}

HARDENER_NAME_MAP = {
    'CR84': 'CR 84',
    'CR 88-300': 'CR 800-300',
}

def lookup_chemical_rate_v2(pr_df, material_name, category_kw, report_month, name_aliases=None):
    """
    FINAL VERSION: Chemical rate lookup with:
    1. EXACT material matching (no fuzzy substring matching)
    2. Most-recent-month fallback
    3. Name alias support
    """
    cat_col = _find_col(pr_df, 'categery', 'category')
    mat_col = _find_col(pr_df, 'material')
    month_col = None
    amt_col = None
    qty_col = None
    for c in pr_df.columns:
        cl = str(c).strip().lower()
        if cl == 'month': month_col = c
        if cl == 'amount': amt_col = c
        if cl == 'actual quantity': qty_col = c
    
    if not cat_col or not mat_col or not month_col:
        return 0.0
    
    lookup_name = str(material_name).strip().upper()
    if name_aliases and lookup_name in name_aliases:
        lookup_name = name_aliases[lookup_name]
    
    # Category filter
    cat_mask = pr_df[cat_col].astype(str).str.lower().str.contains(category_kw, na=False)
    
    # EXACT material match (no fuzzy substring)
    mat_mask = pr_df[mat_col].astype(str).str.upper().str.strip() == lookup_name
    
    rows = pr_df[cat_mask & mat_mask]
    if rows.empty:
        return 0.0
    
    # Try report month first
    month_rows = rows[rows[month_col].astype(str).str.strip() == report_month]
    
    if month_rows.empty:
        # Fallback: most recent month <= report month
        parts = report_month.split('-')
        report_key = int(parts[1]) * 100 + int(parts[0])
        
        best_month = None
        best_key = 0
        for m in rows[month_col].astype(str).str.strip().unique():
            mp = m.split('-')
            if len(mp) == 2:
                try:
                    mk = int(mp[1]) * 100 + int(mp[0])
                    if mk <= report_key and mk > best_key:
                        best_key = mk
                        best_month = m
                except:
                    pass
        
        if best_month:
            month_rows = rows[rows[month_col].astype(str).str.strip() == best_month]
        else:
            return 0.0
    
    # Calculate: Total Amount / Total Actual Quantity
    if amt_col and qty_col:
        total_amt = pd.to_numeric(month_rows[amt_col], errors='coerce').fillna(0).sum()
        total_qty = pd.to_numeric(month_rows[qty_col], errors='coerce').fillna(0).sum()
        if total_qty > 0:
            return total_amt / total_qty
    
    return 0.0


# ══════════════════════════════════════════════════════════════
# FULL VERIFICATION ON BOTH DATASETS
# ══════════════════════════════════════════════════════════════

all_ok = True

for ds_name, jt_path, pr_path, month in [
    ("Template2 (Nov 2025)", "Template2/Jobtrack With MRR.xlsx", 
     "Template2/Purchase Register - 2021 - 2025 _Nov.xlsx", "11-2025"),
    ("Template_Files (Feb 2026)", "Template_Files/Jobtrack Feb With MRR.xlsx", 
     "Template_Files/Purchase Register - 2021 - 2026 _Feb 26.xlsx", "2-2026"),
]:
    print(f"\n{'=' * 80}")
    print(f"  {ds_name}")
    print(f"{'=' * 80}")
    
    wb = openpyxl.load_workbook(jt_path, data_only=True)
    ws = wb.active
    pr_df = load_purchase_register(pr_path)
    
    total = 0
    adh_ok = 0
    hard_ok = 0
    sol_ok = 0
    
    for row in range(5, ws.max_row + 1):
        process = ws.cell(row=row, column=6).value
        if not process or str(process).strip().upper() != 'LAM':
            continue
        
        adh_name = ws.cell(row=row, column=91).value
        if not adh_name:
            continue
        
        total += 1
        adh_str = str(adh_name).strip()
        da_str = str(ws.cell(row=row, column=105).value or '').strip()
        
        gt_adh_rate = float(ws.cell(row=row, column=93).value or 0)
        gt_hard_rate = float(ws.cell(row=row, column=97).value or 0)
        gt_sol_rate = float(ws.cell(row=row, column=101).value or 0)
        sol_qty = float(ws.cell(row=row, column=100).value or 0)
        
        # Compute
        eng_adh = lookup_chemical_rate_v2(pr_df, adh_str, 'adhesive', month, ADH_NAME_MAP)
        eng_hard = lookup_chemical_rate_v2(pr_df, da_str, 'adhesive', month, HARDENER_NAME_MAP) if da_str else 0
        eng_sol = lookup_chemical_rate_v2(pr_df, 'ETHYL ACETATE', 'solvent', month) if sol_qty > 0 else 0
        
        a_ok = abs(gt_adh_rate - eng_adh) < 0.01 if gt_adh_rate > 0 else True
        h_ok = abs(gt_hard_rate - eng_hard) < 0.01 if gt_hard_rate > 0 else True
        s_ok = abs(gt_sol_rate - eng_sol) < 0.01 if gt_sol_rate > 0 else True
        
        if a_ok: adh_ok += 1
        if h_ok: hard_ok += 1
        if s_ok: sol_ok += 1
        
        status = "OK" if (a_ok and h_ok and s_ok) else "FAIL"
        
        detail = ""
        if not a_ok:
            detail += f" ADH:GT={gt_adh_rate:.4f}/ENG={eng_adh:.4f}"
        if not h_ok:
            detail += f" HARD:GT={gt_hard_rate:.4f}/ENG={eng_hard:.4f}"
        if not s_ok:
            detail += f" SOL:GT={gt_sol_rate:.4f}/ENG={eng_sol:.4f}"
        
        print(f"  Row {row:>3} [{status}] Adh={adh_str:>10} DA={da_str:>12} | "
              f"Adh={eng_adh:>8.4f} Hard={eng_hard:>8.4f} Sol={eng_sol:>8.4f}{detail}")
        
        if status == "FAIL":
            all_ok = False
    
    wb.close()
    
    print(f"\n  TOTALS: Adh={adh_ok}/{total}, Hard={hard_ok}/{total}, Sol={sol_ok}/{total}")
    overall = (adh_ok + hard_ok + sol_ok)
    overall_total = total * 3
    pct = overall * 100 / overall_total if overall_total > 0 else 0
    print(f"  OVERALL: {overall}/{overall_total} ({pct:.1f}%)")

print(f"\n{'#' * 80}")
if all_ok:
    print("  >>> ALL RATES MATCH: 100% ACCURACY <<<")
else:
    print("  >>> SOME MISMATCHES REMAIN <<<")
print(f"{'#' * 80}")
