"""
Investigation part 3: 
- Why report_month is None for Nov Jobtrack
- How to filter PR by month for Nov 2025
- Verify correct month-filtered rates match ground truth
- Check DA column also exists in Without MRR template
"""
import pandas as pd
import openpyxl

T2 = "Template2"

# ── 1. Check dates in the Jobtrack ──
print("=" * 80)
print("1. JOBTRACK DATE ANALYSIS - Why report_month is None")
print("=" * 80)

jt_wb = openpyxl.load_workbook(f"{T2}/Jobtrack With MRR.xlsx", data_only=True)
jt_ws = jt_wb.active

print("First 20 rows, Date column (D=4) and UID column (A=1):")
for r in range(5, 25):
    date_val = jt_ws.cell(row=r, column=4).value
    uid_val = jt_ws.cell(row=r, column=1).value
    process = jt_ws.cell(row=r, column=6).value
    print(f"  Row {r}: UID={uid_val}, Date={date_val} (type={type(date_val).__name__}), Process={process}")

jt_wb.close()

# ── 2. Check PR months available ──
print("\n" + "=" * 80)
print("2. PURCHASE REGISTER - Available months for Nov 2025")
print("=" * 80)

pr = pd.read_excel(f"{T2}/Purchase Register - 2021 - 2025 _Nov.xlsx", sheet_name=0, header=2)

# Find month column
month_col = None
for c in pr.columns:
    if str(c).strip().lower() == 'month':
        month_col = c
        break

if month_col:
    months = pr[month_col].dropna().astype(str).str.strip().unique()
    print(f"Available months: {sorted(months)}")
    
    # Check Nov 2025 specifically
    nov_months = [m for m in months if '11' in m or 'nov' in m.lower()]
    print(f"\nNov-related months: {nov_months}")
    
    # What format? Check all
    for m in sorted(months):
        count = len(pr[pr[month_col].astype(str).str.strip() == m])
        print(f"  {m}: {count} rows")

# ── 3. Now compute correct rates with month filter ──
print("\n" + "=" * 80)
print("3. COMPUTING CORRECT RATES WITH MONTH FILTER")
print("=" * 80)

import sys
sys.path.insert(0, '.')
from engine.rate_lookup import (
    load_purchase_register, lookup_adhesive_rate, 
    lookup_hardener_rate, lookup_solvent_rate,
    _find_col, _filter_by_month, _qty_weighted_rate
)

pr_df = load_purchase_register(f"{T2}/Purchase Register - 2021 - 2025 _Nov.xlsx")

# Test with explicit month "11-2025"
report_month = "11-2025"
print(f"Testing with report_month = {report_month}")

# Test each adhesive
for adh_name in ['MB655', '75-300', 'S621']:
    rate = lookup_adhesive_rate(pr_df, adh_name, report_month=report_month)
    print(f"\n  Adhesive '{adh_name}': rate = {rate:.4f}")

# Test each hardener via the current engine
for adh_name in ['MB655', '75-300', 'S621']:
    rate = lookup_hardener_rate(pr_df, adh_name, report_month=report_month)
    print(f"  Hardener for '{adh_name}': rate = {rate:.4f}")

# Test solvent
sol_rate = lookup_solvent_rate(pr_df, report_month=report_month)
print(f"\n  Solvent (Ethyl Acetate): rate = {sol_rate:.4f}")

# ── 4. Manual rate calculation for verification ──
print("\n" + "=" * 80)
print("4. MANUAL RATE VERIFICATION for November 2025")
print("=" * 80)

cat_col = _find_col(pr_df, 'categery', 'category')
mat_col = _find_col(pr_df, 'material')
rate_col = [c for c in pr_df.columns if str(c).strip().lower() == 'rate']
rate_col = rate_col[0] if rate_col else 'Rate'

# Filter for November 2025
month_col2 = None
for c in pr_df.columns:
    if str(c).strip().lower() == 'month':
        month_col2 = c
        break

nov_mask = pr_df[month_col2].astype(str).str.strip() == '11-2025'
nov_pr = pr_df[nov_mask]
print(f"Total PR rows for Nov 2025: {len(nov_pr)}")

# ADHESIVE - MF 75-300
for mat_name in ['MB655', 'MF 75-300', 'CT85', 'CR 84', 'CR84', 'CR 88-300', 
                 'S110', 'S621', 'ETHYL ACETATE']:
    cat_mask = nov_pr[cat_col].astype(str).str.lower().str.contains('adhesive|solvent', regex=True, na=False)
    mat_mask = nov_pr[mat_col].astype(str).str.upper().str.strip() == mat_name.upper()
    rows = nov_pr[cat_mask & mat_mask]
    
    if len(rows) > 0:
        amt_col = None
        qty_col = None
        for c in rows.columns:
            cl = str(c).strip().lower()
            if cl == 'amount':
                amt_col = c
            if cl == 'actual quantity':
                qty_col = c
        
        if amt_col and qty_col:
            total_amt = pd.to_numeric(rows[amt_col], errors='coerce').fillna(0).sum()
            total_qty = pd.to_numeric(rows[qty_col], errors='coerce').fillna(0).sum()
            calc_rate = total_amt / total_qty if total_qty > 0 else 0
            cat_vals = rows[cat_col].unique()
            print(f"\n  {mat_name}: {len(rows)} rows, Cat={cat_vals}")
            print(f"    Total Amount={total_amt:.2f}, Total Qty={total_qty:.2f}")
            print(f"    Calculated Rate (Amount/Qty) = {calc_rate:.4f}")
            # Show individual rows
            for _, row in rows.iterrows():
                print(f"    Tracking={row.get('Tracking N o.', '?')}, "
                      f"Rate={row.get(rate_col, '?')}, "
                      f"Qty={row.get('Actual Quantity', '?')}, "
                      f"Amt={row.get('Amount', '?')}")
    else:
        print(f"\n  {mat_name}: NO rows in Nov 2025")

# ── 5. Check DA column in Without MRR template ──
print("\n" + "=" * 80)
print("5. DA COLUMN IN 'WITHOUT MRR' TEMPLATE (what we need to fill)")
print("=" * 80)

jt_wb2 = openpyxl.load_workbook(f"{T2}/Jobtrack Without MRR.xlsx", data_only=True)
jt_ws2 = jt_wb2.active

# Check DA column header and data
da_col = 105
print(f"DA column ({da_col}) header: {jt_ws2.cell(row=4, column=da_col).value}")

da_count = 0
for r in range(5, jt_ws2.max_row + 1):
    process = jt_ws2.cell(row=r, column=6).value
    if process and str(process).strip().upper() == 'LAM':
        da = jt_ws2.cell(row=r, column=da_col).value
        adh = jt_ws2.cell(row=r, column=91).value
        if da:
            da_count += 1
            if da_count <= 15:
                print(f"  Row {r}: DA={da}, Adhesive={adh}")

print(f"\nTotal LAM rows with DA value: {da_count}")

# Unique DA values in Without MRR
da_vals = set()
for r in range(5, jt_ws2.max_row + 1):
    process = jt_ws2.cell(row=r, column=6).value
    if process and str(process).strip().upper() == 'LAM':
        da = jt_ws2.cell(row=r, column=da_col).value
        if da:
            da_vals.add(str(da).strip())
print(f"Unique DA values: {sorted(da_vals)}")

jt_wb2.close()

# ── 6. Compare Feb data (Template_Files) for same patterns ──
print("\n" + "=" * 80)
print("6. COMPARE WITH FEB DATA (Template_Files)")
print("=" * 80)

jt_feb = openpyxl.load_workbook(f"{T1}/Jobtrack Feb With MRR.xlsx", data_only=True)
jt_feb_ws = jt_feb.active

# Check DA column
da_col_idx = 105
print(f"DA column header in Feb: {jt_feb_ws.cell(row=4, column=da_col_idx).value}")

feb_pairings = {}
for r in range(5, jt_feb_ws.max_row + 1):
    process = jt_feb_ws.cell(row=r, column=6).value
    if process and str(process).strip().upper() == 'LAM':
        da = jt_feb_ws.cell(row=r, column=da_col_idx).value
        adh = jt_feb_ws.cell(row=r, column=91).value
        if da and adh:
            key = str(adh).strip().upper()
            val = str(da).strip().upper()
            if key not in feb_pairings:
                feb_pairings[key] = set()
            feb_pairings[key].add(val)

print("\nFeb Adhesive -> DA (Hardener) pairings:")
for k, vs in sorted(feb_pairings.items()):
    print(f"  {k} -> {sorted(vs)}")

jt_feb.close()

print("\n" + "=" * 80)
print("INVESTIGATION COMPLETE")
print("=" * 80)
