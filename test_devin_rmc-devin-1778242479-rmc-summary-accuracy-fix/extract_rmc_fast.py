"""Fast RMC Summary extraction using pandas + openpyxl formula read."""
import sys, os, json, time
sys.path.insert(0, r'c:\Users\mohamad.al\OneDrive - AL KHAYYAT INVESTMENTS\Desktop\IPP')
import pandas as pd

base = r'c:\Users\mohamad.al\OneDrive - AL KHAYYAT INVESTMENTS\Desktop\IPP'
filled_path = os.path.join(base, 'Files_need_to_study', 'Filled_Output', '1 Base RMC _ 2026 February.xlsx')
unfilled_path = os.path.join(base, 'Files_need_to_study', 'Unfilled', '1 Base RMC _ 2026 February.xlsx')

# ═══════════════════════════════════════════════════════════════
# STEP 1: Read filled RMC Summary with pandas (values only)
# ═══════════════════════════════════════════════════════════════
print("Reading FILLED RMC Summary with pandas...", flush=True)
t0 = time.time()
df_filled = pd.read_excel(filled_path, sheet_name='RMC summary', header=None, engine='openpyxl')
print(f"  Done in {time.time()-t0:.1f}s. Shape: {df_filled.shape}", flush=True)

# Row 1 = headers/labels (0-indexed row 0)
# Row 2 = row 1 in 0-indexed... Let me check
# Print rows 0-4 (header area)
print("\n=== HEADER AREA (first 5 rows) ===", flush=True)
for i in range(min(5, len(df_filled))):
    row_vals = {}
    for c in range(min(88, df_filled.shape[1])):
        v = df_filled.iloc[i, c]
        if pd.notna(v):
            row_vals[c+1] = v
    print(f"  Row {i+1}: {row_vals}", flush=True)

# ═══════════════════════════════════════════════════════════════
# STEP 2: Extract column headers from row 1 (0-indexed row 0)
# ═══════════════════════════════════════════════════════════════
print("\n=== COLUMN HEADERS (looking for text headers) ===", flush=True)
# Check rows 0-4 for text headers
for check_row in range(5):
    text_cols = {}
    for c in range(min(88, df_filled.shape[1])):
        v = df_filled.iloc[check_row, c]
        if pd.notna(v) and isinstance(v, str):
            text_cols[c+1] = v
    if text_cols:
        print(f"  Row {check_row+1} text headers: {text_cols}", flush=True)

# ═══════════════════════════════════════════════════════════════
# STEP 3: Find data start and extract key columns
# ═══════════════════════════════════════════════════════════════
# Find first data row (where col B has an order number like L00xxx)
data_start = None
for i in range(len(df_filled)):
    v = df_filled.iloc[i, 1]  # Col B (0-indexed col 1)
    if pd.notna(v) and isinstance(v, str) and len(v) >= 5 and v[0].isalpha():
        # Looks like an order number
        if any(c.isdigit() for c in v):
            data_start = i
            break

print(f"\n  Data starts at 0-indexed row {data_start} (Excel row {data_start+1})", flush=True)

# ═══════════════════════════════════════════════════════════════
# STEP 4: Extract ALL data rows with their values
# ═══════════════════════════════════════════════════════════════
from openpyxl.utils import get_column_letter

data_rows = []
for i in range(data_start, len(df_filled)):
    order = df_filled.iloc[i, 1]  # Col B
    if pd.isna(order) or not str(order).strip():
        continue
    
    row = {}
    for c in range(min(88, df_filled.shape[1])):
        v = df_filled.iloc[i, c]
        if pd.notna(v):
            cl = get_column_letter(c+1)
            row[f'{cl}'] = round(float(v), 6) if isinstance(v, (int, float)) else str(v).strip()
    data_rows.append(row)

print(f"  Total data rows: {len(data_rows)}", flush=True)

# Print first 3 rows in detail
for i, row in enumerate(data_rows[:3]):
    print(f"\n  === Row {i+1} (Order={row.get('B', '?')}) ===", flush=True)
    for k in sorted(row.keys(), key=lambda x: (len(x), x)):
        print(f"    {k}: {row[k]}", flush=True)

# ═══════════════════════════════════════════════════════════════
# STEP 5: Column fill statistics
# ═══════════════════════════════════════════════════════════════
print(f"\n=== COLUMN FILL STATS (of {len(data_rows)} rows) ===", flush=True)
all_cols = set()
for row in data_rows:
    all_cols.update(row.keys())

for cl in sorted(all_cols, key=lambda x: (len(x), x)):
    filled = sum(1 for row in data_rows if cl in row)
    if filled > 0:
        # Get numeric values for stats
        nums = [row[cl] for row in data_rows if cl in row and isinstance(row[cl], (int, float))]
        if nums:
            print(f"  {cl}: {filled}/{len(data_rows)} filled, sum={sum(nums):.2f}, avg={sum(nums)/len(nums):.4f}", flush=True)
        else:
            print(f"  {cl}: {filled}/{len(data_rows)} filled (text)", flush=True)

# ═══════════════════════════════════════════════════════════════
# STEP 6: Now read the FORMULAS from the filled workbook
# ═══════════════════════════════════════════════════════════════
print("\n=== READING FORMULAS ===", flush=True)
import openpyxl
t0 = time.time()
wb_formulas = openpyxl.load_workbook(filled_path, data_only=False, read_only=True)
ws_f = wb_formulas['RMC summary']
print(f"  Loaded in {time.time()-t0:.1f}s", flush=True)

# Extract formulas from first 5 data rows
formula_row = data_start + 1  # Excel 1-indexed
print(f"\n  Formulas for Excel row {formula_row} (first data row):", flush=True)
for c in range(1, 89):
    v = ws_f.cell(row=formula_row, column=c).value
    if v is not None and isinstance(v, str) and v.startswith('='):
        cl = get_column_letter(c)
        print(f"    {cl}({c}): {v}", flush=True)

# Check a few more rows for formula patterns
for extra_row in [formula_row + 1, formula_row + 5, formula_row + 10]:
    if extra_row <= ws_f.max_row:
        print(f"\n  Formulas for Excel row {extra_row}:", flush=True)
        for c in range(1, 89):
            v = ws_f.cell(row=extra_row, column=c).value
            if v is not None and isinstance(v, str) and v.startswith('='):
                cl = get_column_letter(c)
                print(f"    {cl}({c}): {v}", flush=True)

wb_formulas.close()

# ═══════════════════════════════════════════════════════════════
# STEP 7: Save ground truth JSON
# ═══════════════════════════════════════════════════════════════
gt = {'data_rows': data_rows, 'total': len(data_rows)}
gt_path = os.path.join(base, 'rmc_summary_gt.json')
with open(gt_path, 'w') as f:
    json.dump(gt, f, indent=1, default=str)
print(f"\nGround truth saved: {gt_path} ({os.path.getsize(gt_path):,} bytes)", flush=True)
print("DONE!", flush=True)
