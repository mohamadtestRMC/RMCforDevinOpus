"""Quick check: Does the input Jobtrack already have rates filled?"""
import os, glob
base = r'c:\Users\mohamad.al\OneDrive - AL KHAYYAT INVESTMENTS\Desktop\IPP'

# Find Jobtrack input file
input_dir = os.path.join(base, 'input_files')
jt_files = glob.glob(os.path.join(input_dir, '*obtrack*'), recursive=True)
jt_files += glob.glob(os.path.join(input_dir, '*JT*'), recursive=True)
jt_files += glob.glob(os.path.join(input_dir, '**', '*obtrack*'), recursive=True)
print(f"Jobtrack files found: {jt_files}", flush=True)

# Also check the loaders to see what file is used
import importlib.util
spec = importlib.util.spec_from_file_location("loaders", 
    os.path.join(base, 'engine', 'base_rmc', 'loaders.py'))
print("\nChecking loaders.py for Jobtrack loading logic...", flush=True)
with open(os.path.join(base, 'engine', 'base_rmc', 'loaders.py'), 'r') as f:
    content = f.read()
    # Find jobtrack-related lines
    for i, line in enumerate(content.split('\n'), 1):
        if 'jobtrack' in line.lower() or 'jt' in line.lower().split('=')[0]:
            print(f"  L{i}: {line.strip()}", flush=True)

# Check if the Jobtrack has Rate columns filled
print("\n--- Checking Jobtrack data ---", flush=True)
if jt_files:
    import openpyxl
    jt_path = jt_files[0]
    print(f"Loading: {jt_path}", flush=True)
    wb = openpyxl.load_workbook(jt_path, data_only=True, read_only=True)
    ws = wb.active
    print(f"Sheet: {ws.title}, max_row={ws.max_row}", flush=True)
    
    # Check Film Rate (BC=55), Film Value (BD=56), Fresh1 Rate (CA=79), Adh Rate (CO=93)
    rate_cols = {55: 'Film Rate(BC)', 56: 'Film Value(BD)', 
                 79: 'Fresh1 Rate(CA)', 80: 'Fresh1 Value(CB)',
                 93: 'Adh Rate(CO)', 94: 'Adh Value(CP)'}
    
    for r in [4, 5, 6, 7, 8]:  # Header + first data rows
        vals = {}
        for c, label in rate_cols.items():
            v = ws.cell(row=r, column=c).value
            if v is not None:
                vals[label] = v
        if vals:
            print(f"  Row {r}: {vals}", flush=True)
    
    # Count non-zero rates
    film_rates = 0
    fresh1_rates = 0
    adh_rates = 0
    for r in range(5, min(ws.max_row + 1, 3500)):
        if ws.cell(row=r, column=55).value:
            try:
                if float(ws.cell(row=r, column=55).value) > 0:
                    film_rates += 1
            except: pass
        if ws.cell(row=r, column=79).value:
            try:
                if float(ws.cell(row=r, column=79).value) > 0:
                    fresh1_rates += 1
            except: pass
        if ws.cell(row=r, column=93).value:
            try:
                if float(ws.cell(row=r, column=93).value) > 0:
                    adh_rates += 1
            except: pass
    
    print(f"\n  Film rates filled: {film_rates}", flush=True)
    print(f"  Fresh1 rates filled: {fresh1_rates}", flush=True)
    print(f"  Adh rates filled: {adh_rates}", flush=True)
    wb.close()

print("\nDone!", flush=True)
