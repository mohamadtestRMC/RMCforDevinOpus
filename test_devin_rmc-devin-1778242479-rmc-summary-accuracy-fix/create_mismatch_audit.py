"""
Mismatch Audit Excel Generator
Creates a multi-sheet Excel file documenting all mismatches between
engine-calculated values and ground truth (With MRR) for Feb 2026 and Nov 2025.
"""
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io, sys, shutil, tempfile, os
sys.path.insert(0, '.')

from engine.fill_jobtrack import fill_jobtrack, COLS, DATA_START_ROW
from engine.mrr_lookup import load_stores_recordings, lookup_mrr_with_qty
from engine.rate_lookup import (
    load_purchase_register, lookup_film_rate_weighted,
    lookup_material_rate_for_month, _find_col, _get_rate_for_mrr
)

def safe_open_wb(path, data_only=True):
    try:
        return openpyxl.load_workbook(path, data_only=data_only)
    except PermissionError:
        tmp = tempfile.mktemp(suffix='.xlsx')
        shutil.copy2(path, tmp)
        wb = openpyxl.load_workbook(tmp, data_only=data_only)
        os.unlink(tmp)
        return wb

def sf(val):
    if val is None or val == '' or (isinstance(val, float) and pd.isna(val)):
        return 0.0
    if isinstance(val, str) and val.startswith('='):
        return 0.0
    try:
        return float(val)
    except:
        return 0.0

def ss(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ''
    s = str(val).strip()
    return '' if s.startswith('=') else s

COMPARE_COLS = {
    'Film_MR': 54, 'Film_Rate': 55, 'Film_Value': 56,
    'Fresh1_MR': 78, 'Fresh1_Rate': 79, 'Fresh1_Value': 80,
    'Fresh2_MR': 88, 'Fresh2_Rate': 89, 'Fresh2_Value': 90,
    'Adh_Rate': 93, 'Adh_Value': 94,
    'Hard_Rate': 97, 'Hard_Value': 98,
    'Sol_Rate': 101, 'Sol_Value': 102,
}

def collect_mismatches(dataset_name, jt_without, jt_with, stores, pr,
                       granules=None, megapack=None, prev_granules=None):
    """Run engine, compare, return list of mismatch dicts."""
    print(f"\n--- Processing {dataset_name} ---")
    
    with open(jt_without, 'rb') as f:
        jt_bytes = io.BytesIO(f.read())
    
    g_io = m_io = pg_io = None
    if granules:
        with open(granules, 'rb') as f: g_io = io.BytesIO(f.read())
    if megapack:
        with open(megapack, 'rb') as f: m_io = io.BytesIO(f.read())
    if prev_granules:
        with open(prev_granules, 'rb') as f: pg_io = io.BytesIO(f.read())
    
    filled_bytes, results_log, stats = fill_jobtrack(
        jt_bytes, stores, pr,
        granules_file=g_io, megapack_file=m_io, prev_granules_file=pg_io
    )
    
    filled_wb = openpyxl.load_workbook(filled_bytes, data_only=False)
    filled_ws = filled_wb.active
    gt_wb = safe_open_wb(jt_with, data_only=True)
    gt_ws = gt_wb.active
    max_row = min(filled_ws.max_row, gt_ws.max_row)
    
    # Also load stores/PR for diagnostic info
    stores_df = load_stores_recordings(stores)
    pr_df = load_purchase_register(pr)
    
    mismatches = []
    summary = {'match': 0, 'mismatch': 0, 'miss': 0, 'total': 0}
    
    for row in range(DATA_START_ROW, max_row + 1):
        process = ss(gt_ws.cell(row=row, column=COLS['Process']).value).upper()
        uid = ss(gt_ws.cell(row=row, column=COLS['UID']).value)
        order = ss(gt_ws.cell(row=row, column=COLS['Order_No']).value)
        if not process or not uid:
            continue
        
        for col_name, col_idx in COMPARE_COLS.items():
            is_film = col_name.startswith('Film_')
            is_fresh = col_name.startswith('Fresh')
            is_chem = col_name.startswith(('Adh_', 'Hard_', 'Sol_'))
            
            if process == 'PRINTING' and (is_fresh or is_chem): continue
            if process == 'LAM' and is_film: continue
            if process not in ('PRINTING', 'LAM'): continue
            
            gt_val = gt_ws.cell(row=row, column=col_idx).value
            eng_val = filled_ws.cell(row=row, column=col_idx).value
            gt_f = sf(gt_val); eng_f = sf(eng_val)
            gt_s = ss(gt_val); eng_s = ss(eng_val)
            
            is_mismatch = False
            mtype = ''
            
            if col_name.endswith('_MR'):
                gt_set = {s.strip() for s in gt_s.replace('/',',').split(',') if s.strip()} if gt_s else set()
                eng_set = {s.strip() for s in eng_s.replace('/',',').split(',') if s.strip()} if eng_s else set()
                if gt_set and not eng_set:
                    is_mismatch = True; mtype = 'MISSING MR#'
                elif gt_set and eng_set and gt_set != eng_set and not (gt_set & eng_set):
                    is_mismatch = True; mtype = 'DIFFERENT MR#'
                elif not gt_set and not eng_set:
                    pass  # both empty
                else:
                    summary['match'] += 1; summary['total'] += 1; continue
            else:
                if gt_f == 0 and eng_f == 0:
                    pass  # both empty
                elif gt_f > 0 and eng_f == 0:
                    is_mismatch = True; mtype = 'MISSING VALUE'
                elif gt_f == 0 and eng_f > 0:
                    pass  # engine-only
                elif abs(gt_f - eng_f) < 0.02 or (gt_f > 0 and abs(gt_f - eng_f)/gt_f < 0.005):
                    summary['match'] += 1; summary['total'] += 1; continue
                else:
                    pct = abs(gt_f - eng_f)/gt_f * 100 if gt_f > 0 else 999
                    is_mismatch = True; mtype = f'RATE DIFF ({pct:.1f}%)'
            
            if not is_mismatch:
                summary['total'] += 1
                continue
            
            summary['mismatch'] += 1; summary['total'] += 1
            
            # Get context
            mat = mic = size = ''
            if 'Film' in col_name:
                mat = ss(gt_ws.cell(row=row, column=COLS['Input_Name']).value)
                mic = sf(gt_ws.cell(row=row, column=COLS['Input_Mic']).value)
                size = sf(gt_ws.cell(row=row, column=COLS['Input_Size']).value)
            elif 'Fresh1' in col_name:
                mat = ss(gt_ws.cell(row=row, column=COLS['Fresh1_Name']).value)
                mic = sf(gt_ws.cell(row=row, column=COLS['Fresh1_Mic']).value)
                size = sf(gt_ws.cell(row=row, column=COLS['Fresh1_Size']).value)
            elif 'Fresh2' in col_name:
                mat = ss(gt_ws.cell(row=row, column=COLS['Fresh2_Name']).value)
                mic = sf(gt_ws.cell(row=row, column=COLS['Fresh2_Mic']).value)
                size = sf(gt_ws.cell(row=row, column=COLS['Fresh2_Size']).value)
            
            # Get MRR details from stores
            mrr_detail = ''
            if mat and ('Rate' in col_name or 'Value' in col_name or 'MR' in col_name):
                mrr_qty = lookup_mrr_with_qty(stores_df, mat, mic if mic else None,
                                               size if size else None, order, process)
                if not mrr_qty:
                    mrr_qty = lookup_mrr_with_qty(stores_df, mat, mic if mic else None, None, order, process)
                if not mrr_qty:
                    mrr_qty = lookup_mrr_with_qty(stores_df, mat, mic if mic else None, None, order)
                if mrr_qty:
                    parts = [f"MRR {m}: qty={q:.1f}" for m, q in sorted(mrr_qty.items(), key=lambda x: -x[1])]
                    mrr_detail = '; '.join(parts)
                else:
                    mrr_detail = 'NO MRR FOUND IN STORES'
            
            # Get PR rate details
            pr_detail = ''
            if mat and 'Rate' in col_name:
                tracking_col = _find_col(pr_df, 'tracking')
                material_col = _find_col(pr_df, 'material')
                rate_col = [c for c in pr_df.columns if str(c).strip().lower() == 'rate']
                rate_col = rate_col[0] if rate_col else 'Rate'
                
                if mrr_qty and tracking_col:
                    for mrr_num in list(mrr_qty.keys())[:5]:
                        try:
                            mrr_int = int(float(mrr_num))
                            mask = pd.to_numeric(pr_df[tracking_col], errors='coerce') == mrr_int
                            pr_rows = pr_df[mask]
                            if not pr_rows.empty:
                                for _, pr_row in pr_rows.iterrows():
                                    pr_mat = ss(pr_row.get(material_col, ''))
                                    pr_rate = sf(pr_row.get(rate_col, 0))
                                    pr_size = sf(pr_row.get(_find_col(pr_df, 'size') or '', 0))
                                    pr_detail += f"MRR {mrr_int}: {pr_mat} size={pr_size} rate={pr_rate:.4f}; "
                        except:
                            pass
            
            mismatches.append({
                'Dataset': dataset_name,
                'Row': row,
                'UID': uid,
                'Process': process,
                'Order': order,
                'Column': col_name,
                'Material': mat,
                'Mic': mic,
                'Size': size,
                'GT Value (Excel)': gt_s if col_name.endswith('_MR') else gt_f,
                'Engine Value': eng_s if col_name.endswith('_MR') else eng_f,
                'Difference': '' if col_name.endswith('_MR') else round(eng_f - gt_f, 4),
                'Mismatch Type': mtype,
                'MRR Detail (Stores)': mrr_detail,
                'PR Rate Detail': pr_detail,
            })
    
    gt_wb.close(); filled_wb.close()
    return mismatches, summary, stats


def style_header(ws, row, max_col, fill_color='1F4E79', font_color='FFFFFF'):
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    font = Font(bold=True, color=font_color, size=11)
    thin = Side(style='thin')
    border = Border(bottom=Side(style='medium'))
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = border

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value else ''
                max_len = max(max_len, len(val))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)


# ══════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════

T1 = "Template_Files"
T2 = "Template2"

mm_feb, sum_feb, stats_feb = collect_mismatches(
    "February 2026",
    f"{T1}/Jobtrack Feb Without MRR.xlsx",
    f"{T1}/Jobtrack Feb With MRR.xlsx",
    f"{T1}/Stores Recordings.xlsx",
    f"{T1}/Purchase Register - 2021 - 2026 _Feb 26.xlsx",
    granules=f"{T1}/Granules Recipe - February 2026.xlsx",
    megapack=f"{T1}/MEGA PACK.xlsx",
)

mm_nov, sum_nov, stats_nov = collect_mismatches(
    "November 2025",
    f"{T2}/Jobtrack Without MRR.xlsx",
    f"{T2}/Jobtrack With MRR.xlsx",
    f"{T2}/Stores Recordings.xlsx",
    f"{T2}/Purchase Register - 2021 - 2025 _Nov.xlsx",
    granules=f"{T2}/Granules Recipe -Nov_2025.xlsx",
)

all_mm = mm_feb + mm_nov

# ── Build Excel ──
wb = openpyxl.Workbook()

# ═══════════════════ Sheet 1: Summary ═══════════════════
ws1 = wb.active
ws1.title = "Summary"
ws1.append(["IPP Jobtrack MRR Engine — Mismatch Audit Report"])
ws1.merge_cells('A1:F1')
ws1.cell(1,1).font = Font(bold=True, size=16, color='1F4E79')
ws1.append([])

ws1.append(["Dataset", "Total Compared", "Match", "Mismatch", "Missing", "Accuracy %"])
style_header(ws1, 3, 6)

feb_acc = sum_feb['match']*100/(sum_feb['total'] or 1)
nov_acc = sum_nov['match']*100/(sum_nov['total'] or 1)
ws1.append(["February 2026", sum_feb['total'], sum_feb['match'], sum_feb['mismatch'], sum_feb.get('miss',0), f"{feb_acc:.1f}%"])
ws1.append(["November 2025", sum_nov['total'], sum_nov['match'], sum_nov['mismatch'], sum_nov.get('miss',0), f"{nov_acc:.1f}%"])
total_t = sum_feb['total'] + sum_nov['total']
total_m = sum_feb['match'] + sum_nov['match']
ws1.append(["TOTAL", total_t, total_m, sum_feb['mismatch']+sum_nov['mismatch'],
            sum_feb.get('miss',0)+sum_nov.get('miss',0), f"{total_m*100/(total_t or 1):.1f}%"])
ws1.cell(6,1).font = Font(bold=True)

ws1.append([])
ws1.append(["Engine Stats", "February 2026", "November 2025"])
style_header(ws1, 8, 3)
for key in ['total_rows','printing_rows','lam_rows','film_filled','fresh1_filled',
            'fresh2_filled','adh_filled','hard_filled','sol_filled','errors','skipped']:
    ws1.append([key, stats_feb.get(key,0), stats_nov.get(key,0)])
auto_width(ws1)

# ═══════════════════ Sheet 2: All Mismatches Detail ═══════════════════
ws2 = wb.create_sheet("All Mismatches")
headers = ['Dataset','Row','UID','Process','Order','Column','Material','Mic','Size',
           'GT Value (Excel)','Engine Value','Difference','Mismatch Type',
           'MRR Detail (Stores)','PR Rate Detail']
ws2.append(headers)
style_header(ws2, 1, len(headers))

red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

for mm in all_mm:
    r = ws2.max_row + 1
    ws2.append([mm[h] for h in headers])
    fill = red_fill if 'MISSING' in mm['Mismatch Type'] else yellow_fill
    for c in range(1, len(headers)+1):
        ws2.cell(r, c).fill = fill
auto_width(ws2)

# ═══════════════════ Sheet 3: Feb Mismatches ═══════════════════
ws3 = wb.create_sheet("February 2026")
ws3.append(headers)
style_header(ws3, 1, len(headers))
for mm in mm_feb:
    r = ws3.max_row + 1
    ws3.append([mm[h] for h in headers])
    for c in range(1, len(headers)+1):
        ws3.cell(r, c).fill = yellow_fill
auto_width(ws3)

# ═══════════════════ Sheet 4: Nov Mismatches ═══════════════════
ws4 = wb.create_sheet("November 2025")
ws4.append(headers)
style_header(ws4, 1, len(headers))
for mm in mm_nov:
    r = ws4.max_row + 1
    ws4.append([mm[h] for h in headers])
    fill = red_fill if 'MISSING' in mm['Mismatch Type'] else yellow_fill
    for c in range(1, len(headers)+1):
        ws4.cell(r, c).fill = fill
auto_width(ws4)

# ═══════════════════ Sheet 5: Root Cause & Rule Questions ═══════════════════
ws5 = wb.create_sheet("Root Cause & Questions")
ws5.append(["Case #", "Dataset", "Row", "Material", "Column", "Issue Description",
            "Root Cause Analysis", "Proposed Rule / Question"])
style_header(ws5, 1, 8)

# Manually categorized root causes from the mismatch data
cases = [
    # Feb 2026 cases
    [1, "Feb 2026", "42-43", "PET 12mic", "Fresh1_Rate",
     "GT has row-specific rates (4.22, 4.35) but engine gives single weighted-avg 4.3036",
     "Multiple MRRs exist with different rates. GT uses a SPECIFIC MRR rate per row, engine uses qty-weighted average across ALL MRRs for that material/mic.",
     "QUESTION: Should the rate be tied to the SPECIFIC MRR that was actually consumed for that row? Or is the weighted average across all MRRs correct? If per-MRR, how do we determine which MRR belongs to which row?"],
    [2, "Feb 2026", "46", "PET 12mic/913mm", "Film_Rate",
     "GT=4.5880, Engine=4.4604 (2.8% diff)",
     "PET with size 913 has multiple MRRs. The GT rate (4.5880) matches a SPECIFIC MRR's rate, while engine computes weighted average (4.4604). The size-913 MRR may have a distinct rate.",
     "QUESTION: Same as Case 1 — should the system assign per-MRR rate to specific rows? If so, the matching logic needs a row→MRR assignment rule (e.g., by date, by issue sequence)."],
    [3, "Feb 2026", "54", "TPE 100mic/893mm", "Fresh2_MR",
     "GT shows MRR 84080, Engine shows 85738/85775",
     "Stores has MRRs 85738 and 85775 for this material/order, but GT shows 84080 which may be from a different period or manually assigned.",
     "QUESTION: Is MRR 84080 manually assigned by the user? If so, how should the engine handle cases where the manual MRR differs from what Stores shows?"],
    # Nov 2025 cases
    [4, "Nov 2025", "7", "NYLON 15mic/1005mm", "Film_MR/Rate/Value",
     "GT has MRR=83005, Rate=7.72. Engine finds NOTHING (0).",
     "NYLON is not in the Stores material aliases. The engine cannot find any MRR because 'NYLON' doesn't match any sub-category in Stores Recordings.",
     "RULE NEEDED: Add 'NYLON' to MATERIAL_ALIASES in mrr_lookup.py. What are the Stores sub-categories for Nylon? (e.g., 'NYLON', 'NYLON PA', 'NY PA')"],
    [5, "Nov 2025", "17", "TPE 50mic/996mm", "Fresh1_Rate",
     "GT=4.1951, Engine=4.2968 (2.4% diff)",
     "Same pattern as Feb PET case. Multiple MRRs with different rates, engine uses weighted avg but GT uses a specific MRR rate.",
     "QUESTION: Same root cause as Cases 1-2. Per-MRR rate assignment vs weighted average."],
    [6, "Nov 2025", "48", "TPE 100mic/893mm", "Fresh2_Rate",
     "GT=4.8708, Engine=0.9000 (81.5% diff!). Engine's outlier detection kicked in and replaced 4.87 with month avg 0.90.",
     "The OUTLIER CHECK in fill_jobtrack.py line 498-501/707-712 replaces the MRR-derived rate with month average when >50% different. In this case the month avg (0.90) is WRONG — it's pulling a different material's rate. The MRR rate (4.87) was actually correct.",
     "RULE FIX: The outlier check should be REMOVED or made much more conservative. The per-MRR rate from PR is the source of truth. The 'month average' fallback is unreliable when the PR has mixed materials under the same category."],
]

for case in cases:
    r = ws5.max_row + 1
    ws5.append(case)
    if 'RULE FIX' in str(case[-1]):
        for c in range(1, 9):
            ws5.cell(r, c).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    elif 'RULE NEEDED' in str(case[-1]):
        for c in range(1, 9):
            ws5.cell(r, c).fill = red_fill

auto_width(ws5)

# ═══════════════════ Sheet 6: Calculation Method ═══════════════════
ws6 = wb.create_sheet("How We Calculate")
ws6.append(["Step", "Description", "Source File", "Logic"])
style_header(ws6, 1, 4)

steps = [
    [1, "Read Jobtrack row", "fill_jobtrack.py",
     "Read UID, Process, Order No, Material Name, Size, Mic, Qty from Jobtrack template"],
    [2, "Lookup MRR from Stores", "mrr_lookup.py",
     "Match material (with aliases), mic, order no, process in Stores Recordings → get {MRR: qty} dict"],
    [3, "Filter MRRs by PR", "rate_lookup.py → filter_mrr_by_pr()",
     "Remove MRRs that don't exist in Purchase Register (or don't match size)"],
    [4, "Check Supplier Override", "supplier_rates.py",
     "If supplier is BANDERA/CYM → use Granules Recipe rate. If MEGA PACK → use MEGA PACK rate."],
    [5, "Qty-Weighted Rate from PR", "rate_lookup.py → lookup_film_rate_weighted()",
     "For each MRR, get rate from PR. Compute: Σ(rate × stores_qty) / Σ(stores_qty)"],
    [6, "Outlier Check", "fill_jobtrack.py lines 496-501",
     "If MRR rate differs >50% from month average → replace with month avg. ⚠️ THIS CAUSES CASE 6 BUG"],
    [7, "Dominant MRR Display", "fill_jobtrack.py lines 503-509",
     "Only show MRRs with ≥10% of total qty in the MR# cell (display only, not rate calc)"],
    [8, "Write Value", "fill_jobtrack.py",
     "Value = Total Qty (manual: Qty + Balance) × Rate"],
    [9, "Chemical Rates (ADH/Hard/Sol)", "rate_lookup.py",
     "Lookup by material name + month in PR. Qty-weighted: Total Amount / Total Actual Quantity"],
]
for s in steps:
    ws6.append(s)
auto_width(ws6)

# Save
output_path = "Mismatch_Audit_Complete.xlsx"
wb.save(output_path)
print(f"\n✅ Saved: {output_path}")
print(f"   Feb mismatches: {len(mm_feb)}")
print(f"   Nov mismatches: {len(mm_nov)}")
print(f"   Total: {len(all_mm)}")
