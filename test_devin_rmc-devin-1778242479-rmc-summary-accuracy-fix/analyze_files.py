import openpyxl
import os
import sys

base = r'c:\Users\mohamad.al\OneDrive - AL KHAYYAT INVESTMENTS\Desktop\IPP\Files_need_to_study'

def analyze_workbook(path, label, max_rows=15, max_cols=25, formula_limit=50):
    print(f"\n{'='*80}")
    print(f"FILE: {label}")
    print(f"Path: {path}")
    print(f"{'='*80}")
    
    if not os.path.exists(path):
        print("  FILE NOT FOUND!")
        return
    
    try:
        wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    except Exception as e:
        print(f"  ERROR loading: {e}")
        return
        
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"\n--- Sheet: '{sn}' ---")
        
        row_count = 0
        formula_cells = []
        for ri, row in enumerate(ws.iter_rows(max_col=max_cols, values_only=False), 1):
            if ri <= max_rows:
                vals = []
                for cell in row:
                    v = cell.value
                    if v is not None:
                        vals.append(f"C{cell.column}={repr(v)[:80]}")
                if vals:
                    print(f"  Row {ri}: {vals}")
            
            for cell in row:
                v = cell.value
                if v and isinstance(v, str) and v.startswith('='):
                    if len(formula_cells) < formula_limit:
                        formula_cells.append(f"    {cell.coordinate}: {v[:120]}")
            
            row_count = ri
            if ri > 500:
                break
        
        print(f"  Total rows scanned: {row_count}")
        
        if formula_cells:
            print(f"\n  FORMULAS in '{sn}' ({len(formula_cells)} found):")
            for f in formula_cells:
                print(f)
    
    wb.close()
    print()

# 1. FILLED Base RMC
analyze_workbook(
    os.path.join(base, 'Filled_Output', '1 Base RMC _ 2026 February.xlsx'),
    "FILLED - 1 Base RMC _ 2026 February",
    max_rows=20, formula_limit=80
)

sys.stdout.flush()
print("\n\n===== PART 1 DONE =====\n")
sys.stdout.flush()
