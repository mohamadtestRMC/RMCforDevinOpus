"""Part 1: Sheet overview - dimensions for each sheet"""
import openpyxl
import time

FILE_PATH = r"Files_need_to_study\Filled_Output\1 Base RMC _ 2026 February.xlsx"

t0 = time.time()
print("Loading workbook (read_only, formulas)...")
wb = openpyxl.load_workbook(FILE_PATH, data_only=False, read_only=True)
print(f"Loaded in {time.time()-t0:.1f}s")

print(f"\nTotal sheets: {len(wb.sheetnames)}")

for name in wb.sheetnames:
    ws = wb[name]
    max_row = 0
    max_col = 0
    row_count = 0
    for row in ws.iter_rows():
        row_count += 1
        for cell in row:
            if cell.value is not None:
                if cell.row > max_row:
                    max_row = cell.row
                if cell.column > max_col:
                    max_col = cell.column
        if row_count > 5000:
            print(f"  '{name}': >5000 rows (stopped counting), max_col so far={max_col}")
            break
    else:
        print(f"  '{name}': {max_row} rows x {max_col} cols (data extent)")

wb.close()
print(f"\nDone in {time.time()-t0:.1f}s")
