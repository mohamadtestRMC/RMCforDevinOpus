"""Helper functions for meeting Excel generator."""
import pandas as pd
import openpyxl
import shutil, tempfile, os
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def safe_open(path, data_only=True):
    try:
        return openpyxl.load_workbook(path, data_only=data_only)
    except PermissionError:
        tmp = tempfile.mktemp(suffix='.xlsx')
        shutil.copy2(path, tmp)
        wb = openpyxl.load_workbook(tmp, data_only=data_only)
        os.unlink(tmp)
        return wb

def sf(val):
    if val is None or val == '' or (isinstance(val, float) and pd.isna(val)):
        return 0.0
    if isinstance(val, str) and val.startswith('='):
        return 0.0
    try:
        return float(val)
    except:
        return 0.0

def ss(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ''
    s = str(val).strip()
    return '' if s.startswith('=') else s

BLUE = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
GREEN = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
YELLOW = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
RED = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
LIGHT_BLUE = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
WHITE_FONT = Font(bold=True, color='FFFFFF', size=11)
BOLD = Font(bold=True, size=11)
WRAP = Alignment(horizontal='center', wrap_text=True, vertical='center')

def style_header(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = BLUE
        cell.font = WHITE_FONT
        cell.alignment = WRAP

def auto_width(ws):
    for col in ws.columns:
        mx = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                mx = max(mx, len(str(cell.value or '')))
            except:
                pass
        ws.column_dimensions[letter].width = min(mx + 3, 55)

def add_section_header(ws, row, text, max_col):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, size=13, color='1F4E79')
    cell.fill = PatternFill(start_color='E8EEF4', end_color='E8EEF4', fill_type='solid')
