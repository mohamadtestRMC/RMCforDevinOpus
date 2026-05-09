"""Quick check: what UIDs look like in Template2"""
import openpyxl
wb = openpyxl.load_workbook("Template2/Jobtrack Without MRR.xlsx", data_only=True)
ws = wb.active
for row in range(5, 30):
    uid = ws.cell(row=row, column=1).value
    date = ws.cell(row=row, column=4).value
    process = ws.cell(row=row, column=6).value
    if uid:
        print(f"Row {row}: UID='{uid}', Date='{date}', Process='{process}'")
wb.close()
