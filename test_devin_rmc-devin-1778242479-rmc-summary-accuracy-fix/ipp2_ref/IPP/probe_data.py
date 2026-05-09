"""
Probe script: inspect actual Excel structure and data for test design.
"""
import pandas as pd
import openpyxl
import json

BASE = "Template_Files"

print("=" * 80)
print("1. JOBTRACK STRUCTURE (With MRR — ground truth)")
print("=" * 80)

wb = openpyxl.load_workbook(f"{BASE}/Jobtrack Feb With MRR.xlsx", data_only=True)
ws = wb.active
print(f"Sheet: {ws.title}, Max row: {ws.max_row}, Max col: {ws.max_column}")

# Print header row (row 4)
print("\n--- Header Row 4 (key columns) ---")
key_cols = {
    1: 'UID(A)', 4: 'Date(D)', 6: 'Process(F)', 11: 'OrderNo(K)',
    47: 'InputName(AU)', 48: 'InputSize(AV)', 49: 'InputMic(AW)',
    51: 'InputQty(AY)', 52: 'BalQty(AZ)', 53: 'TotalInput(BA)',
    54: 'FilmMR(BB)', 55: 'FilmRate(BC)', 56: 'FilmValue(BD)',
    71: 'Fresh1Name(BS)', 72: 'Fresh1Size(BT)', 73: 'Fresh1Mic(BU)',
    75: 'Fresh1Qty(BW)', 76: 'Fresh1Bal(BX)', 77: 'TotalFresh1(BY)',
    78: 'Fresh1MR(BZ)', 79: 'Fresh1Rate(CA)', 80: 'Fresh1Value(CB)',
    81: 'Fresh2Name(CC)', 88: 'Fresh2MR(CJ)', 89: 'Fresh2Rate(CK)', 90: 'Fresh2Value(CL)',
    91: 'AdhName(CM)', 92: 'AdhKgs(CN)', 93: 'AdhRate(CO)', 94: 'AdhValue(CP)',
    96: 'HardKgs(CR)', 97: 'HardRate(CS)', 98: 'HardValue(CT)',
    100: 'SolQty(CV)', 101: 'SolRate(CW)', 102: 'SolValue(CX)',
}
for col_idx, label in sorted(key_cols.items()):
    val = ws.cell(row=4, column=col_idx).value
    print(f"  Col {col_idx:3d} ({label:20s}): {val}")

# Sample first 5 data rows
print("\n--- Sample Data Rows 5-9 ---")
for row in range(5, 10):
    uid = ws.cell(row=row, column=1).value
    process = ws.cell(row=row, column=6).value
    order = ws.cell(row=row, column=11).value
    film_mr = ws.cell(row=row, column=54).value
    film_rate = ws.cell(row=row, column=55).value
    film_val = ws.cell(row=row, column=56).value
    fresh1_mr = ws.cell(row=row, column=78).value
    adh_rate = ws.cell(row=row, column=93).value
    print(f"  Row {row}: UID={uid}, Process={process}, Order={order}, "
          f"FilmMR={film_mr}, FilmRate={film_rate}, FilmVal={film_val}, "
          f"Fresh1MR={fresh1_mr}, AdhRate={adh_rate}")
wb.close()

print("\n" + "=" * 80)
print("2. COUNTING ROWS BY PROCESS (ground truth)")
print("=" * 80)
wb = openpyxl.load_workbook(f"{BASE}/Jobtrack Feb With MRR.xlsx", data_only=True)
ws = wb.active
process_counts = {}
filled_film = 0
filled_fresh1 = 0
filled_fresh2 = 0
filled_adh = 0
filled_hard = 0
filled_sol = 0
total_rows = 0
for row in range(5, ws.max_row + 1):
    uid = ws.cell(row=row, column=1).value
    process = ws.cell(row=row, column=6).value
    if not uid or not process:
        continue
    total_rows += 1
    p = str(process).strip().upper()
    process_counts[p] = process_counts.get(p, 0) + 1
    
    if ws.cell(row=row, column=54).value is not None:  # Film MR
        filled_film += 1
    if ws.cell(row=row, column=78).value is not None:  # Fresh1 MR
        filled_fresh1 += 1
    if ws.cell(row=row, column=88).value is not None:  # Fresh2 MR
        filled_fresh2 += 1
    if ws.cell(row=row, column=93).value is not None:  # Adh Rate
        filled_adh += 1
    if ws.cell(row=row, column=97).value is not None:  # Hard Rate
        filled_hard += 1
    if ws.cell(row=row, column=101).value is not None:  # Sol Rate
        filled_sol += 1

print(f"Total data rows: {total_rows}")
print(f"Process counts: {json.dumps(process_counts, indent=2)}")
print(f"Ground truth fills: Film={filled_film}, Fresh1={filled_fresh1}, Fresh2={filled_fresh2}")
print(f"  Adh={filled_adh}, Hard={filled_hard}, Sol={filled_sol}")
wb.close()

print("\n" + "=" * 80)
print("3. STORES RECORDINGS STRUCTURE")
print("=" * 80)
df_stores = pd.read_excel(f"{BASE}/Stores Recordings.xlsx", sheet_name=0, header=1, nrows=5)
print(f"Columns: {list(df_stores.columns)}")
print(df_stores.head(3).to_string())

print("\n" + "=" * 80)
print("4. PURCHASE REGISTER STRUCTURE")
print("=" * 80)
df_pr = pd.read_excel(f"{BASE}/Purchase Register - 2021 - 2026 _Feb 26.xlsx", sheet_name=0, header=2, nrows=5)
print(f"Columns: {list(df_pr.columns)}")
print(df_pr.head(3).to_string())

print("\n" + "=" * 80)
print("5. GROUND TRUTH: Extract ALL filled values from 'With MRR' file")
print("=" * 80)
wb = openpyxl.load_workbook(f"{BASE}/Jobtrack Feb With MRR.xlsx", data_only=True)
ws = wb.active
ground_truth = []
for row in range(5, ws.max_row + 1):
    uid = ws.cell(row=row, column=1).value
    process = ws.cell(row=row, column=6).value
    if not uid or not process:
        continue
    p = str(process).strip().upper()
    entry = {
        'row': row,
        'uid': str(uid).strip(),
        'process': p,
        'order_no': str(ws.cell(row=row, column=11).value or '').strip(),
    }
    if p == 'PRINTING':
        entry['film_mr'] = ws.cell(row=row, column=54).value
        entry['film_rate'] = ws.cell(row=row, column=55).value
        entry['film_value'] = ws.cell(row=row, column=56).value
        entry['input_name'] = ws.cell(row=row, column=47).value
        entry['total_input'] = ws.cell(row=row, column=53).value
    elif p == 'LAM':
        entry['fresh1_mr'] = ws.cell(row=row, column=78).value
        entry['fresh1_rate'] = ws.cell(row=row, column=79).value
        entry['fresh1_value'] = ws.cell(row=row, column=80).value
        entry['fresh2_mr'] = ws.cell(row=row, column=88).value
        entry['fresh2_rate'] = ws.cell(row=row, column=89).value
        entry['fresh2_value'] = ws.cell(row=row, column=90).value
        entry['adh_rate'] = ws.cell(row=row, column=93).value
        entry['adh_value'] = ws.cell(row=row, column=94).value
        entry['hard_rate'] = ws.cell(row=row, column=97).value
        entry['hard_value'] = ws.cell(row=row, column=98).value
        entry['sol_rate'] = ws.cell(row=row, column=101).value
        entry['sol_value'] = ws.cell(row=row, column=102).value
    ground_truth.append(entry)

# Print summary
printing_rows = [e for e in ground_truth if e['process'] == 'PRINTING']
lam_rows = [e for e in ground_truth if e['process'] == 'LAM']
print(f"PRINTING rows: {len(printing_rows)}")
print(f"LAM rows: {len(lam_rows)}")

# Print a few printing examples
print("\nSample PRINTING ground truth:")
for e in printing_rows[:5]:
    print(f"  Row {e['row']}: UID={e['uid']}, Order={e['order_no']}, "
          f"MR={e.get('film_mr')}, Rate={e.get('film_rate')}, Value={e.get('film_value')}, "
          f"Input={e.get('input_name')}, TotalInput={e.get('total_input')}")

print("\nSample LAM ground truth:")
for e in lam_rows[:5]:
    print(f"  Row {e['row']}: UID={e['uid']}, Order={e['order_no']}, "
          f"F1MR={e.get('fresh1_mr')}, F1Rate={e.get('fresh1_rate')}, "
          f"AdhRate={e.get('adh_rate')}, HardRate={e.get('hard_rate')}, SolRate={e.get('sol_rate')}")

# Save ground truth for test comparison
import pickle
with open("ground_truth.pkl", "wb") as f:
    pickle.dump(ground_truth, f)
print(f"\nSaved {len(ground_truth)} ground truth entries to ground_truth.pkl")
wb.close()
