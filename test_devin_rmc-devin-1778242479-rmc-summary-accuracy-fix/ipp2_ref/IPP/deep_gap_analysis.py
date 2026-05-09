"""
Deep row-by-row gap analysis for ALL remaining mismatches.
Shows exactly what each row has, what's missing, and what question to ask.
"""
import pandas as pd
import openpyxl
import io, sys, shutil, tempfile, os
sys.path.insert(0, '.')

def open_wb(path):
    """Open workbook, copying to temp if locked."""
    try:
        return openpyxl.load_workbook(path, data_only=True)
    except PermissionError:
        tmp = tempfile.mktemp(suffix='.xlsx')
        shutil.copy2(path, tmp)
        wb = openpyxl.load_workbook(tmp, data_only=True)
        os.unlink(tmp)
        return wb

from engine.fill_jobtrack import COLS, DATA_START_ROW
from engine.mrr_lookup import load_stores_recordings, lookup_mrr_with_qty
from engine.rate_lookup import load_purchase_register, _find_col
from engine.supplier_rates import (build_mrr_supplier_map, get_supplier_for_mrrs,
                                    load_granules_rates, load_megapack_rates)

def sf(v):
    if v is None or v == '': return 0.0
    if isinstance(v, float) and pd.isna(v): return 0.0
    try: return float(v)
    except: return 0.0

def ss(v):
    if v is None or (isinstance(v, float) and pd.isna(v)): return ''
    s = str(v).strip()
    return '' if s.startswith('=') else s

def analyze_row(label, gt_ws, row, stores_df, pr_df, mrr_sup, gran_rates, mega_rates, report_month):
    """Deep analysis of a single mismatch row."""
    uid = ss(gt_ws.cell(row=row, column=COLS['UID']).value)
    process = ss(gt_ws.cell(row=row, column=COLS['Process']).value).upper()
    order = ss(gt_ws.cell(row=row, column=COLS['Order_No']).value)
    
    print(f"\n{'='*90}")
    print(f"  {label} | Row {row} | UID={uid} | Process={process} | Order={order}")
    print(f"{'='*90}")
    
    if process == 'PRINTING':
        mat = ss(gt_ws.cell(row=row, column=COLS['Input_Name']).value)
        mic = sf(gt_ws.cell(row=row, column=COLS['Input_Mic']).value)
        size = sf(gt_ws.cell(row=row, column=COLS['Input_Size']).value)
        gt_mr = ss(gt_ws.cell(row=row, column=COLS['Film_MR']).value)
        gt_rate = sf(gt_ws.cell(row=row, column=COLS['Film_Rate']).value)
        gt_val = sf(gt_ws.cell(row=row, column=COLS['Film_Value']).value)
        
        print(f"  Material: {mat} | Mic: {mic} | Size: {size}")
        print(f"  GT: MR#={gt_mr}, Rate={gt_rate:.4f}, Value={gt_val:.2f}")
        
        # Stores lookup
        mrr_qty = lookup_mrr_with_qty(stores_df, mat, mic, size, order, 'PRINTING')
        if not mrr_qty:
            mrr_qty = lookup_mrr_with_qty(stores_df, mat, mic, None, order, 'PRINTING')
        if not mrr_qty:
            mrr_qty = lookup_mrr_with_qty(stores_df, mat, mic, None, order)
        
        if mrr_qty:
            supplier = get_supplier_for_mrrs(mrr_sup, list(mrr_qty.keys())) or 'Standard'
            print(f"  Stores: {len(mrr_qty)} MRRs found, Supplier={supplier}")
            for m, q in mrr_qty.items():
                print(f"    MRR {m}: Qty={q:.1f}")
        else:
            print(f"  Stores: *** NO MRRs FOUND ***")
            # Try broader search
            for proc in ['PRINTING', 'LAMINATION', None]:
                for s in [size, None]:
                    test = lookup_mrr_with_qty(stores_df, mat, mic, s, order, proc)
                    if test:
                        print(f"    Found with proc={proc}, size={s}: {list(test.keys())}")
            print(f"  → QUESTION: Why is MRR {gt_mr} not found in Stores for {mat}/{mic}/{size}/{order}?")
        
        # Granules check
        ou = order.upper().strip()
        if gran_rates and ou in gran_rates:
            print(f"  Granules: WO# {ou} → Rate={gran_rates[ou]:.4f}")
        elif gran_rates:
            print(f"  Granules: WO# {ou} NOT found. Available: {list(gran_rates.keys())}")
    
    elif process == 'LAM':
        for prefix, name_col, mic_col, size_col, mr_col, rate_col, val_col in [
            ('Fresh1', COLS['Fresh1_Name'], COLS['Fresh1_Mic'], COLS['Fresh1_Size'],
             COLS['Fresh1_MR'], COLS['Fresh1_Rate'], COLS['Fresh1_Value']),
            ('Fresh2', COLS['Fresh2_Name'], COLS['Fresh2_Mic'], COLS['Fresh2_Size'],
             COLS['Fresh2_MR'], COLS['Fresh2_Rate'], COLS['Fresh2_Value']),
        ]:
            mat = ss(gt_ws.cell(row=row, column=name_col).value)
            if not mat: continue
            mic = sf(gt_ws.cell(row=row, column=mic_col).value)
            size = sf(gt_ws.cell(row=row, column=size_col).value)
            gt_mr = ss(gt_ws.cell(row=row, column=mr_col).value)
            gt_rate = sf(gt_ws.cell(row=row, column=rate_col).value)
            gt_val = sf(gt_ws.cell(row=row, column=val_col).value)
            
            if gt_rate == 0 and gt_mr == '': continue
            
            print(f"\n  --- {prefix} ---")
            print(f"  Material: {mat} | Mic: {mic} | Size: {size}")
            print(f"  GT: MR#={gt_mr}, Rate={gt_rate:.4f}, Value={gt_val:.2f}")
            
            mrr_qty = lookup_mrr_with_qty(stores_df, mat, mic, size, order, 'LAMINATION')
            if not mrr_qty:
                mrr_qty = lookup_mrr_with_qty(stores_df, mat, mic, None, order, 'LAMINATION')
            if not mrr_qty:
                mrr_qty = lookup_mrr_with_qty(stores_df, mat, mic, None, order)
            
            if mrr_qty:
                supplier = get_supplier_for_mrrs(mrr_sup, list(mrr_qty.keys())) or 'Standard'
                print(f"  Stores: {len(mrr_qty)} MRRs, Supplier={supplier}")
                for m, q in mrr_qty.items():
                    sup_m = mrr_sup.get(int(float(m)), 'Unknown') if mrr_sup else 'N/A'
                    print(f"    MRR {m}: Qty={q:.1f}, Supplier={sup_m}")
            else:
                print(f"  Stores: NO MRRs found")
            
            # Granules
            ou = order.upper().strip()
            if gran_rates:
                if ou in gran_rates:
                    print(f"  Granules: WO# {ou} → Rate={gran_rates[ou]:.4f}")
                else:
                    avg = sum(gran_rates.values()) / len(gran_rates)
                    print(f"  Granules: WO# {ou} NOT found. Avg={avg:.4f}")
                    print(f"    Available WOs: {list(gran_rates.keys())}")
            
            # Mega Pack
            if mega_rates:
                for key, rates in mega_rates.items():
                    mat_u = mat.upper()
                    r = rates.get('WPE', 0) if 'WPE' in mat_u else rates.get('TPE', 0)
                    print(f"  MEGA PACK [{key}]: {mat} → {r:.4f}")
            
            # Analysis
            if gt_mr == 'INH':
                if supplier in ('BANDERA', 'CYM'):
                    print(f"  ✅ RULE: BANDERA/CYM → INH is correct")
                    if gran_rates and ou not in gran_rates:
                        print(f"  ⚠️ GAP: WO# {ou} not in Granules. Need prev month file or manual rate")
                        print(f"  → Engine uses avg={avg:.4f}, GT={gt_rate:.4f}, Diff={abs(avg-gt_rate)/gt_rate*100:.1f}%")
                else:
                    print(f"  ❓ GT says INH but supplier={supplier}?")
            elif supplier in ('BANDERA', 'CYM') and gt_mr != 'INH':
                print(f"  ❓ QUESTION: Supplier=BANDERA/CYM but GT shows MR#={gt_mr} (not INH)")
                print(f"     Should this also be INH? Or is there a condition where BANDERA doesn't get INH?")
                print(f"     GT Rate={gt_rate:.4f}")

# ═══════════════════════════════════════════════════════
# ANALYZE BOTH DATASETS
# ═══════════════════════════════════════════════════════

print("╔══════════════════════════════════════════════════════════════════╗")
print("║          DEEP ROW-BY-ROW GAP ANALYSIS                          ║")
print("╚══════════════════════════════════════════════════════════════════╝")

# --- Feb 2026 ---
T1 = "Template_Files"
print(f"\n\n{'#'*90}")
print(f"  FEBRUARY 2026 — Mismatched Rows")
print(f"{'#'*90}")

gt_wb1 = open_wb(f"{T1}/Jobtrack Feb With MRR.xlsx")
gt_ws1 = gt_wb1.active
stores1 = load_stores_recordings(f"{T1}/Stores Recordings.xlsx")
pr1 = load_purchase_register(f"{T1}/Purchase Register - 2021 - 2026 _Feb 26.xlsx")
sup1 = build_mrr_supplier_map(stores1)
gran1 = load_granules_rates(f"{T1}/Granules Recipe - February 2026.xlsx")
mega1 = load_megapack_rates(f"{T1}/MEGA PACK.xlsx")

for row in [42, 43, 45, 46, 54, 66, 67]:
    analyze_row("Feb 2026", gt_ws1, row, stores1, pr1, sup1, gran1, mega1, "2-2026")
gt_wb1.close()

# --- Nov 2025 ---
T2 = "Template2"
print(f"\n\n{'#'*90}")
print(f"  NOVEMBER 2025 — Mismatched Rows")
print(f"{'#'*90}")

gt_wb2 = open_wb(f"{T2}/Jobtrack With MRR.xlsx")
gt_ws2 = gt_wb2.active
stores2 = load_stores_recordings(f"{T2}/Stores Recordings.xlsx")
pr2 = load_purchase_register(f"{T2}/Purchase Register - 2021 - 2025 _Nov.xlsx")
sup2 = build_mrr_supplier_map(stores2)
gran2 = load_granules_rates(f"{T2}/Granules Recipe -Nov_2025.xlsx")

for row in [7, 17, 48]:
    analyze_row("Nov 2025", gt_ws2, row, stores2, pr2, sup2, gran2, {}, "11-2025")
gt_wb2.close()

print(f"\n\n{'#'*90}")
print("  ANALYSIS COMPLETE")
print(f"{'#'*90}")
