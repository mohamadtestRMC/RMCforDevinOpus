"""Check if N00694 INH rate comes from the Granules formula 
when the order is NOT in the Granules but supplier IS BANDERA/CYM.
Theory: maybe the rule is simply BANDERA → "INH" MR# + Granules Recipe rate for THAT order.
If order not in Granules → STILL write INH but compute rate differently?

Let me check ALL the GT INH rows more carefully and see if their rates 
always match an entry in Granules Recipe."""
import pandas as pd
import openpyxl
import sys
sys.path.insert(0, '.')
from engine.supplier_rates import load_granules_rates

# Nov 2025 INH cases  
rates = load_granules_rates("Template2/Granules Recipe -Nov_2025.xlsx")
print("Nov 2025 Granules rates:", dict(rates))

wb = openpyxl.load_workbook("Template2/Jobtrack With MRR.xlsx", data_only=True)
ws = wb.active

print("\nAll INH rows in Nov 2025 GT:")
for row in range(5, ws.max_row + 1):
    process = str(ws.cell(row=row, column=6).value or '').strip().upper()
    if process != 'LAM': continue
    order = str(ws.cell(row=row, column=11).value or '').strip().upper()
    
    for prefix, mr_col, rate_col, name_col in [
        ('F1', 78, 79, 71), ('F2', 88, 89, 81)
    ]:
        mr = str(ws.cell(row=row, column=mr_col).value or '').strip()
        rate = float(ws.cell(row=row, column=rate_col).value or 0)
        mat = str(ws.cell(row=row, column=name_col).value or '').strip()
        
        if mr == 'INH':
            gran_rate = rates.get(order, None)
            match = "MATCH" if gran_rate and abs(gran_rate - rate) < 0.01 else "NO MATCH"
            gran_str = f"{gran_rate:.4f}" if gran_rate else "N/A"
            print(f"  Row {row} {prefix}: Order={order}, Mat={mat}, GT_Rate={rate:.4f}, "
                  f"Granules={gran_str}, {match}")

# Feb 2026 - check if there are INH rows
print("\nAll INH rows in Feb 2026 GT:")
wb2 = openpyxl.load_workbook("Template_Files/Jobtrack Feb With MRR.xlsx", data_only=True)
ws2 = wb2.active
inh_found = False
for row in range(5, ws2.max_row + 1):
    process = str(ws2.cell(row=row, column=6).value or '').strip().upper()
    if process != 'LAM': continue
    for mr_col in [78, 88]:
        mr = str(ws2.cell(row=row, column=mr_col).value or '').strip()
        if mr == 'INH':
            inh_found = True
            order = str(ws2.cell(row=row, column=11).value or '').strip()
            print(f"  Row {row}: Order={order}, MR=INH")
if not inh_found:
    print("  No INH rows in Feb 2026")

# For the MEGA PACK case — row 54 Feb, row 48 Nov
print("\n\nMEGA PACK cases:")
print("Row 54 Feb 2026: GT rate=4.7848, GT MR#=84080 (NOT INH)")
print("Row 48 Nov 2025: GT rate=4.8708, GT MR#=84080 (NOT INH)")

# Check if MEGA PACK file has these rates
print("\nChecking MEGA PACK file:")
try:
    from engine.supplier_rates import load_megapack_rates, lookup_megapack_rate
    mp_rates = load_megapack_rates("Template_Files/MEGA PACK.xlsx")
    print(f"MEGA PACK entries: {len(mp_rates)}")
    for k, v in mp_rates.items():
        print(f"  {k}: {v}")
    
    # Try looking up TPE
    r = lookup_megapack_rate(mp_rates, 'TPE', 2026, 2)
    print(f"\nTPE lookup Feb 2026: {r}")
    r = lookup_megapack_rate(mp_rates, 'TPE', 2025, 11)
    print(f"TPE lookup Nov 2025: {r}")
except Exception as e:
    print(f"Error: {e}")

wb.close()
wb2.close()
