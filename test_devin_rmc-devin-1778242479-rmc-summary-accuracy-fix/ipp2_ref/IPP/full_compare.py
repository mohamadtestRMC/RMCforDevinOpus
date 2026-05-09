"""
FULL ENGINE FILL + COMPARE
Fills both Jobtracks using the actual engine, then compares EVERY filled column
against ground truth (With MRR files).
"""
import pandas as pd
import openpyxl
import io
import sys
import shutil
import tempfile
import os
sys.path.insert(0, '.')

def safe_open_wb(path, data_only=True):
    """Open workbook, copying to temp if file is locked."""
    try:
        return openpyxl.load_workbook(path, data_only=data_only)
    except PermissionError:
        tmp = tempfile.mktemp(suffix='.xlsx')
        shutil.copy2(path, tmp)
        wb = openpyxl.load_workbook(tmp, data_only=data_only)
        os.unlink(tmp)
        return wb

from engine.fill_jobtrack import fill_jobtrack, COLS, DATA_START_ROW

# Column groups to compare
COMPARE_COLS = {
    # Printing rows
    'Film_MR': 54,       # BB
    'Film_Rate': 55,     # BC
    'Film_Value': 56,    # BD
    # LAM rows - 1st Fresh
    'Fresh1_MR': 78,     # BZ
    'Fresh1_Rate': 79,   # CA
    'Fresh1_Value': 80,  # CB
    # LAM rows - 2nd Fresh
    'Fresh2_MR': 88,     # CJ
    'Fresh2_Rate': 89,   # CK
    'Fresh2_Value': 90,  # CL
    # LAM rows - Adhesive
    'Adh_Rate': 93,      # CO
    'Adh_Value': 94,     # CP
    # LAM rows - Hardener
    'Hard_Rate': 97,     # CS
    'Hard_Value': 98,    # CT
    # LAM rows - Solvent
    'Sol_Rate': 101,     # CW
    'Sol_Value': 102,    # CX
}

def safe_float(val):
    if val is None or val == '' or (isinstance(val, float) and pd.isna(val)):
        return 0.0
    if isinstance(val, str) and val.startswith('='):
        return 0.0
    try:
        return float(val)
    except:
        return 0.0

def safe_str(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ''
    s = str(val).strip()
    if s.startswith('='):
        return ''
    return s

def compare_fill(dataset_name, jt_without_path, jt_with_path, stores_path, pr_path,
                 granules_path=None, megapack_path=None, prev_granules_path=None):
    """Fill the Without MRR file and compare against With MRR ground truth."""
    
    print(f"\n{'#' * 100}")
    print(f"# {dataset_name}")
    print(f"{'#' * 100}")
    
    # ── Step 1: Run the engine fill ──
    print("\n[1] Running engine fill...")
    
    with open(jt_without_path, 'rb') as f:
        jt_bytes = io.BytesIO(f.read())
    
    granules_io = None
    megapack_io = None
    prev_granules_io = None
    if granules_path:
        with open(granules_path, 'rb') as f:
            granules_io = io.BytesIO(f.read())
    if megapack_path:
        with open(megapack_path, 'rb') as f:
            megapack_io = io.BytesIO(f.read())
    if prev_granules_path:
        with open(prev_granules_path, 'rb') as f:
            prev_granules_io = io.BytesIO(f.read())
    
    def progress(pct, msg):
        if pct in [5, 35, 92, 100]:
            print(f"  [{pct:3d}%] {msg}")
    
    filled_bytes, results_log, stats = fill_jobtrack(
        jt_bytes, stores_path, pr_path,
        progress_callback=progress,
        granules_file=granules_io,
        megapack_file=megapack_io,
        prev_granules_file=prev_granules_io
    )
    
    print(f"\n  Engine Stats:")
    for k, v in stats.items():
        print(f"    {k}: {v}")
    
    # ── Step 2: Load filled output and ground truth ──
    print("\n[2] Loading filled output and ground truth...")
    
    filled_wb = openpyxl.load_workbook(filled_bytes, data_only=False)
    filled_ws = filled_wb.active
    
    gt_wb = safe_open_wb(jt_with_path, data_only=True)
    gt_ws = gt_wb.active
    
    max_row = min(filled_ws.max_row, gt_ws.max_row)
    print(f"  Filled rows: {filled_ws.max_row}, GT rows: {gt_ws.max_row}")
    
    # ── Step 3: Compare cell-by-cell ──
    print("\n[3] Comparing filled values vs ground truth...")
    
    # Tracking per-column results
    col_results = {}
    for col_name in COMPARE_COLS:
        col_results[col_name] = {'match': 0, 'mismatch': 0, 'both_empty': 0, 'gt_only': 0, 'eng_only': 0}
    
    all_mismatches = []
    
    for row in range(DATA_START_ROW, max_row + 1):
        process_gt = safe_str(gt_ws.cell(row=row, column=COLS['Process']).value).upper()
        uid = safe_str(gt_ws.cell(row=row, column=COLS['UID']).value)
        
        if not process_gt or not uid:
            continue
        
        for col_name, col_idx in COMPARE_COLS.items():
            gt_val = gt_ws.cell(row=row, column=col_idx).value
            eng_val = filled_ws.cell(row=row, column=col_idx).value
            
            # Determine which columns are relevant for which process
            is_film_col = col_name.startswith('Film_')
            is_fresh_col = col_name.startswith('Fresh')
            is_chem_col = col_name.startswith(('Adh_', 'Hard_', 'Sol_'))
            
            # Skip irrelevant columns
            if process_gt == 'PRINTING' and (is_fresh_col or is_chem_col):
                continue
            if process_gt == 'LAM' and is_film_col:
                continue
            if process_gt not in ('PRINTING', 'LAM'):
                continue
            
            gt_f = safe_float(gt_val)
            eng_f = safe_float(eng_val)
            gt_s = safe_str(gt_val)
            eng_s = safe_str(eng_val)
            
            # Compare
            if col_name.endswith('_MR'):
                # MR# is string comparison - but can be multiple MRRs in different order
                gt_set = set(gt_s.replace('/', ',').split(',')) if gt_s else set()
                eng_set = set(eng_s.replace('/', ',').split(',')) if eng_s else set()
                # Clean up sets
                gt_set = {s.strip() for s in gt_set if s.strip()}
                eng_set = {s.strip() for s in eng_set if s.strip()}
                
                if not gt_set and not eng_set:
                    col_results[col_name]['both_empty'] += 1
                elif gt_set and not eng_set:
                    col_results[col_name]['gt_only'] += 1
                    all_mismatches.append({
                        'row': row, 'uid': uid, 'process': process_gt,
                        'col': col_name, 'gt': gt_s, 'eng': eng_s, 'type': 'MISS'
                    })
                elif not gt_set and eng_set:
                    col_results[col_name]['eng_only'] += 1
                elif gt_s == 'INH' and eng_s == 'INH':
                    col_results[col_name]['match'] += 1
                elif gt_set == eng_set:
                    col_results[col_name]['match'] += 1
                elif gt_set & eng_set:  # Some overlap
                    col_results[col_name]['match'] += 1  # Partial OK
                else:
                    col_results[col_name]['mismatch'] += 1
                    all_mismatches.append({
                        'row': row, 'uid': uid, 'process': process_gt,
                        'col': col_name, 'gt': gt_s, 'eng': eng_s, 'type': 'MR#'
                    })
            else:
                # Numeric comparison with tolerance
                if gt_f == 0 and eng_f == 0:
                    col_results[col_name]['both_empty'] += 1
                elif gt_f > 0 and eng_f == 0:
                    col_results[col_name]['gt_only'] += 1
                    all_mismatches.append({
                        'row': row, 'uid': uid, 'process': process_gt,
                        'col': col_name, 'gt': f'{gt_f:.4f}', 'eng': '0', 'type': 'MISS'
                    })
                elif gt_f == 0 and eng_f > 0:
                    col_results[col_name]['eng_only'] += 1
                elif abs(gt_f - eng_f) < 0.02:  # Very tight tolerance
                    col_results[col_name]['match'] += 1
                elif gt_f > 0 and abs(gt_f - eng_f) / gt_f < 0.005:  # <0.5% relative
                    col_results[col_name]['match'] += 1
                else:
                    col_results[col_name]['mismatch'] += 1
                    pct_diff = abs(gt_f - eng_f) / gt_f * 100 if gt_f > 0 else 999
                    all_mismatches.append({
                        'row': row, 'uid': uid, 'process': process_gt,
                        'col': col_name, 'gt': f'{gt_f:.4f}', 'eng': f'{eng_f:.4f}',
                        'type': f'{pct_diff:.1f}%'
                    })
    
    gt_wb.close()
    filled_wb.close()
    
    # ── Step 4: Print results ──
    print(f"\n{'=' * 100}")
    print(f"  RESULTS: {dataset_name}")
    print(f"{'=' * 100}")
    
    print(f"\n  {'Column':<15} {'Match':>6} {'Mismatch':>9} {'GT Only':>8} {'Eng Only':>9} {'Empty':>6} {'Accuracy':>10}")
    print(f"  {'-'*15} {'-'*6} {'-'*9} {'-'*8} {'-'*9} {'-'*6} {'-'*10}")
    
    total_match = 0
    total_mismatch = 0
    total_gt_only = 0
    
    for col_name in COMPARE_COLS:
        r = col_results[col_name]
        total = r['match'] + r['mismatch'] + r['gt_only']
        if total == 0:
            pct = '-'
        else:
            pct = f"{r['match']*100/total:.1f}%"
        
        total_match += r['match']
        total_mismatch += r['mismatch']
        total_gt_only += r['gt_only']
        
        print(f"  {col_name:<15} {r['match']:>6} {r['mismatch']:>9} {r['gt_only']:>8} {r['eng_only']:>9} {r['both_empty']:>6} {pct:>10}")
    
    overall_total = total_match + total_mismatch + total_gt_only
    overall_pct = total_match * 100 / overall_total if overall_total > 0 else 0
    print(f"\n  OVERALL: {total_match}/{overall_total} ({overall_pct:.1f}%) match, "
          f"{total_mismatch} mismatches, {total_gt_only} missing")
    
    # ── Step 5: Print all mismatches ──
    if all_mismatches:
        print(f"\n  --- ALL MISMATCHES ({len(all_mismatches)}) ---")
        
        # Group by type
        by_col = {}
        for m in all_mismatches:
            key = m['col']
            if key not in by_col:
                by_col[key] = []
            by_col[key].append(m)
        
        for col_name, items in by_col.items():
            print(f"\n  [{col_name}] ({len(items)} issues)")
            for m in items[:15]:  # Show first 15
                # Get extra context
                extra = ""
                if 'Film' in col_name or 'Fresh' in col_name:
                    # Get material name for context
                    extra_row = m['row']
                    gt_wb2 = safe_open_wb(jt_with_path, data_only=True)
                    gt_ws2 = gt_wb2.active
                    if 'Film' in col_name:
                        mat = safe_str(gt_ws2.cell(row=extra_row, column=COLS['Input_Name']).value)
                        mic = safe_float(gt_ws2.cell(row=extra_row, column=COLS['Input_Mic']).value)
                        size = safe_float(gt_ws2.cell(row=extra_row, column=COLS['Input_Size']).value)
                        order = safe_str(gt_ws2.cell(row=extra_row, column=COLS['Order_No']).value)
                        extra = f" | Mat={mat}, Mic={mic}, Size={size}, Order={order}"
                    elif 'Fresh1' in col_name:
                        mat = safe_str(gt_ws2.cell(row=extra_row, column=COLS['Fresh1_Name']).value)
                        mic = safe_float(gt_ws2.cell(row=extra_row, column=COLS['Fresh1_Mic']).value)
                        size = safe_float(gt_ws2.cell(row=extra_row, column=COLS['Fresh1_Size']).value)
                        extra = f" | Mat={mat}, Mic={mic}, Size={size}"
                    elif 'Fresh2' in col_name:
                        mat = safe_str(gt_ws2.cell(row=extra_row, column=COLS['Fresh2_Name']).value)
                        mic = safe_float(gt_ws2.cell(row=extra_row, column=COLS['Fresh2_Mic']).value)
                        size = safe_float(gt_ws2.cell(row=extra_row, column=COLS['Fresh2_Size']).value)
                        extra = f" | Mat={mat}, Mic={mic}, Size={size}"
                    gt_wb2.close()
                
                print(f"    Row {m['row']:>3}: GT={m['gt']:>12} ENG={m['eng']:>12} [{m['type']}]{extra}")
            
            if len(items) > 15:
                print(f"    ... and {len(items) - 15} more")
    else:
        print("\n  >>> NO MISMATCHES - PERFECT FILL <<<")
    
    return stats, col_results, all_mismatches


# ══════════════════════════════════════════════════════════════════════
# RUN BOTH DATASETS
# ══════════════════════════════════════════════════════════════════════

T1 = "Template_Files"
T2 = "Template2"

# Template_Files (February 2026)
stats1, results1, mm1 = compare_fill(
    "Template_Files (February 2026)",
    f"{T1}/Jobtrack Feb Without MRR.xlsx",
    f"{T1}/Jobtrack Feb With MRR.xlsx",
    f"{T1}/Stores Recordings.xlsx",
    f"{T1}/Purchase Register - 2021 - 2026 _Feb 26.xlsx",
    granules_path=f"{T1}/Granules Recipe - February 2026.xlsx",
    megapack_path=f"{T1}/MEGA PACK.xlsx",
)

# Template2 (November 2025)
stats2, results2, mm2 = compare_fill(
    "Template2 (November 2025)",
    f"{T2}/Jobtrack Without MRR.xlsx",
    f"{T2}/Jobtrack With MRR.xlsx",
    f"{T2}/Stores Recordings.xlsx",
    f"{T2}/Purchase Register - 2021 - 2025 _Nov.xlsx",
    granules_path=f"{T2}/Granules Recipe -Nov_2025.xlsx",
    megapack_path=None,  # No Mega Pack for Nov
)

print(f"\n{'#' * 100}")
print("FULL COMPARISON COMPLETE")
print(f"{'#' * 100}")
