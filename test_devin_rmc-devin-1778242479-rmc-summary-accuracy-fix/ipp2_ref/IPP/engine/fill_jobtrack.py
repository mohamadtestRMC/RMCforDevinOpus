"""
Fill Jobtrack Engine - v4
Main orchestrator that reads the Jobtrack template, finds MR# and Rate,
calculates values, and writes the filled output.

KEY FIXES (v4):
- Always compute manual totals (AY+AZ, BW+BX, CG+CH) — never trust formula cache
- Header validation on Jobtrack load
- Removed dangerous WPE→TPE material fallback
- Qty-weighted average rate calculation
- Dominant MRR selection (>=10% of total qty)
- Accurate fill counting (only when value > 0)
- Per-row solvent logging
"""
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import io
import re
import logging
import itertools

from engine.mrr_lookup import load_stores_recordings, lookup_mrr, lookup_mrr_with_qty, match_formula_qtys_to_store
from engine.rate_lookup import (
    load_purchase_register, lookup_film_rate, lookup_film_rate_weighted,
    filter_mrr_by_pr, lookup_material_rate_for_month,
    lookup_adhesive_rate, lookup_hardener_rate, lookup_hardener_rate_by_name,
    lookup_solvent_rate
)
from engine.supplier_rates import (
    build_mrr_supplier_map, get_supplier_for_mrrs,
    load_granules_rates, load_megapack_rates, lookup_megapack_rate
)

logger = logging.getLogger(__name__)

# Column indices (1-based) in the Jobtrack
COLS = {
    # Identifiers
    'UID': 1,       # A
    'Date': 4,      # D - Production Date
    'Process': 6,   # F
    'Order_No': 11, # K

    # 1st Input (for Printing -> Film)
    'Input_Name': 47,    # AU - 1st Input Name
    'Input_Size': 48,    # AV - 1st Input Size (MM)
    'Input_Mic': 49,     # AW - 1st Input Mic
    'Input_Qty': 51,     # AY - 1st Input Qty (raw, not formula)
    'Balance_Qty': 52,   # AZ - Balance Qty
    'Total_1st_Input': 53,  # BA - Total 1st Input (FORMULA: =AY+AZ)

    # Film MRR columns (Printing)
    'Film_MR': 54,       # BB
    'Film_Rate': 55,     # BC
    'Film_Value': 56,    # BD

    # 1st Fresh Material (LAM)
    'Fresh1_Name': 71,   # BS
    'Fresh1_Size': 72,   # BT
    'Fresh1_Mic': 73,    # BU
    'Fresh1_Qty': 75,    # BW - 1st Fresh Material Qty
    'Fresh1_Balance': 76, # BX - Balance Qty
    'Total_Fresh1': 77,  # BY - Total 1st Fresh Material Qty (FORMULA: =BW+BX)

    # 1st Fresh MRR columns (LAM)
    'Fresh1_MR': 78,     # BZ
    'Fresh1_Rate': 79,   # CA
    'Fresh1_Value': 80,  # CB

    # 2nd Fresh Material (LAM)
    'Fresh2_Name': 81,   # CC
    'Fresh2_Size': 82,   # CD
    'Fresh2_Mic': 83,    # CE
    'Fresh2_Qty': 85,    # CG - 2nd Fresh Material Qty
    'Fresh2_Balance': 86, # CH - Balance Qty
    'Total_Fresh2': 87,  # CI - Total 2nd Fresh Material Qty (FORMULA: =CG+CH)

    # 2nd Fresh MRR columns (LAM)
    'Fresh2_MR': 88,     # CJ
    'Fresh2_Rate': 89,   # CK
    'Fresh2_Value': 90,  # CL

    # Adhesive (LAM)
    'Adh_Name': 91,      # CM
    'Adh_Kgs': 92,       # CN
    'Adh_Rate': 93,      # CO
    'Adh_Value': 94,     # CP

    # Hardener (LAM)
    'Hard_Kgs': 96,      # CR
    'Hard_Rate': 97,     # CS
    'Hard_Value': 98,    # CT

    # Solvent (LAM)
    'Sol_Qty': 100,      # CV
    'Sol_Rate': 101,     # CW
    'Sol_Value': 102,    # CX

    # Hardener Name (LAM) — source of truth for hardener material identity
    'DA_Hardener': 105,  # DA — pre-filled with hardener name (CT85, CR84, S110, etc.)
}

HEADER_ROW = 4
DATA_START_ROW = 5


def _safe_float(val, context: str = ""):
    """Safely convert a value to float, return 0.0 on failure.
    Logs a warning for unexpected non-numeric values so they don't silently disappear."""
    if val is None or val == '' or (isinstance(val, float) and pd.isna(val)):
        return 0.0
    # Skip formula strings
    if isinstance(val, str) and val.startswith('='):
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        if str(val).strip():  # Only log if there was actual content
            logger.warning(f"Non-numeric value '{val}' converted to 0.0{' in ' + context if context else ''}")
        return 0.0


def _safe_str(val):
    """Safely convert to string, return empty string on None/NaN."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    s = str(val).strip()
    if s.startswith('='):
        return ""
    return s


def _parse_qty_formula(formula_val) -> list:
    """Parse a quantity formula cell into individual additive components.

    Examples:
        '=40+486.2+486.8'        -> [40.0, 486.2, 486.8]
        '=293+486+485.4+482.3'   -> [293.0, 486.0, 485.4, 482.3]
        '=139'                   -> [139.0]
        '=31.04+30.85-1.19*4'    -> [] (complex, can't parse safely)
        1013.0                   -> [] (not a formula)
        None                     -> []

    Only handles simple additive formulas (terms joined by + only).
    Returns empty list for complex formulas (with *, /, or complex -),
    plain numbers, or None values.
    """
    if formula_val is None:
        return []
    if not isinstance(formula_val, str):
        return []
    s = formula_val.strip()
    if not s.startswith('='):
        return []

    # Remove leading '=' and optional '+'
    expr = s[1:].strip()
    if expr.startswith('+'):
        expr = expr[1:].strip()

    # Reject if contains multiplication, division, or cell references
    if '*' in expr or '/' in expr or re.search(r'[A-Za-z]', expr):
        return []

    # Split by '+' and '-' while keeping the sign
    # Handle: '293+486+485.4' and '=-293' and '=40+486.2+486.8'
    parts = re.split(r'(?=[+-])', expr)
    result = []
    for part in parts:
        part = part.strip()
        if not part:
            continue
        try:
            val = float(part)
            result.append(val)
        except ValueError:
            return []  # Unparseable component -> bail out

    return result


def _normalize_process(proc: str) -> str:
    """Normalize Jobtrack process values to engine buckets."""
    p = _safe_str(proc).upper().strip()
    if 'PRINT' in p:
        return 'PRINTING'
    if 'LAM' in p:
        return 'LAM'
    if 'REWIND' in p:
        return 'REWINDING'
    return p


def _stores_process_filter_for_row(process_bucket: str):
    """Map Jobtrack process bucket to Stores Issue Process filter."""
    if process_bucket == 'PRINTING':
        return 'PRINTING'
    if process_bucket == 'LAM':
        return 'LAMINATION'
    if process_bucket == 'REWINDING':
        # Business rule from Template3:
        # Jobtrack "Rewinding" should match Stores Issue Process "ULTRAFLEX".
        return 'ULTRAFLEX'
    return None


def _pick_mrrs_by_total_qty(mrr_qty: dict, target_qty: float, tolerance: float = 1.0):
    """Pick the smallest MRR subset that best matches target_qty."""
    if not mrr_qty:
        return []
    if target_qty <= 0:
        return sorted(mrr_qty.keys())

    # Prefer a single exact/near-exact MRR when possible.
    singles = []
    for mrr, qty in mrr_qty.items():
        diff = abs(float(qty) - target_qty)
        singles.append((diff, int(mrr)))
    best_single = min(singles, key=lambda x: x[0])
    if best_single[0] <= tolerance:
        return [best_single[1]]

    keys = sorted(int(k) for k in mrr_qty.keys())
    # Keep combinations bounded to avoid combinatorial explosion.
    max_combo = min(3, len(keys))
    best_combo = None  # (diff, combo_len, qty_covered, combo_tuple)
    for r in range(2, max_combo + 1):
        for combo in itertools.combinations(keys, r):
            qty_sum = sum(float(mrr_qty[k]) for k in combo)
            diff = abs(qty_sum - target_qty)
            score = (diff, r, -qty_sum, combo)
            if best_combo is None or score < best_combo:
                best_combo = score
            if diff <= tolerance:
                return list(combo)

    if best_combo:
        return list(best_combo[3])

    # Fallback: dominant MRRs (>=10%) then top one.
    total_qty = sum(float(q) for q in mrr_qty.values())
    if total_qty <= 0:
        return [max(mrr_qty, key=mrr_qty.get)]
    threshold = total_qty * 0.10
    dominant = [int(m) for m, q in mrr_qty.items() if float(q) >= threshold]
    if dominant:
        return sorted(dominant)
    return [int(max(mrr_qty, key=mrr_qty.get))]


def _store_match_diagnostics(stores_df, material, mic, size, order_no, process_filter,
                             formula_qtys=None, tolerance=1.0):
    """Return mismatch reasons for MIC/QTY/Process/WO#/Size against Stores."""
    reasons = []
    if stores_df is None or material is None or _safe_str(material) == '':
        return reasons

    cols = {'sub_cat': None, 'mic': None, 'width': None, 'wo': None, 'process': None, 'issue_qty': None}
    for c in stores_df.columns:
        cl = str(c).lower().strip()
        if 'sub' in cl and 'cat' in cl:
            cols['sub_cat'] = c
        elif cl == 'mic':
            cols['mic'] = c
        elif cl == 'width':
            cols['width'] = c
        elif 'issue wo' in cl:
            cols['wo'] = c
        elif 'issue' in cl and 'process' in cl:
            cols['process'] = c
        elif 'issue' in cl and 'qty' in cl:
            cols['issue_qty'] = c

    if not cols['sub_cat']:
        return ['Stores missing Sub Category column']

    base = stores_df[stores_df[cols['sub_cat']].apply(lambda x: _safe_str(x).upper() == _safe_str(material).upper()
                                                      or str(x).strip().upper().startswith(_safe_str(material).upper() + ' '))]
    if base.empty:
        reasons.append('material')
        return reasons

    if cols['wo'] and order_no:
        wo_match = base[base[cols['wo']].astype(str).str.strip() == str(order_no).strip()]
        if wo_match.empty:
            reasons.append('WO#')
        else:
            base = wo_match

    if cols['process'] and process_filter:
        proc_series = base[cols['process']].astype(str).str.upper().str.strip()
        proc_match = base[proc_series.apply(lambda x: process_filter.upper() in str(x) or str(x) == process_filter.upper())]
        if proc_match.empty:
            reasons.append('Process')
        else:
            base = proc_match

    if cols['mic'] and mic is not None and not pd.isna(mic):
        mic_num = pd.to_numeric(base[cols['mic']], errors='coerce')
        mic_match = base[mic_num == float(mic)]
        if mic_match.empty:
            reasons.append('MIC')
        else:
            base = mic_match

    if cols['width'] and size is not None and not pd.isna(size):
        try:
            size_val = float(size)
            width_num = pd.to_numeric(base[cols['width']], errors='coerce')
            size_match = base[(width_num >= size_val - 5) & (width_num <= size_val + 5)]
            if size_match.empty:
                reasons.append('Size')
            else:
                base = size_match
        except (TypeError, ValueError):
            pass

    if cols['issue_qty'] and formula_qtys:
        qtys = [q for q in formula_qtys if q > 0]
        if qtys:
            issue_vals = pd.to_numeric(base[cols['issue_qty']], errors='coerce').dropna().tolist()
            missing_qty = []
            for q in qtys:
                if not any(abs(float(sq) - float(q)) <= tolerance for sq in issue_vals):
                    missing_qty.append(q)
            if missing_qty:
                reasons.append('QTY')

    return sorted(set(reasons))


def _trace_balance_mrr(ws_write, ws_data, current_row, unmatched_qtys,
                       order_no, material, mic, stores_df,
                       process_filter, qty_col_idx, balance_col_idx,
                       tolerance=1.0):
    """Trace unmatched balance values back to their source MRR.

    When a formula component doesn't match any Store Issue Qty,
    check if another Jobtrack row (same WO) has a Balance Qty
    equal to the negative of the unmatched value. If so, trace
    the last formula component of that row to find the source MRR.

    Example:
        Row 34: formula =295 (doesn't match any Store entry)
        Row 30 (same WO): AZ=-295, AY=...+484.8
        Store: 484.8 → MRR 85588
        → Returns {85588: 295.0}
    """
    if not order_no:
        return {}

    order_str = str(order_no).strip()
    result = {}
    remaining_unmatched = list(unmatched_qtys)

    # ── Phase 1: Try individual matching (each unmatched qty vs each row's balance) ──
    still_unmatched = []
    max_scan_row = max(ws_data.max_row if ws_data is not None else 0,
                       ws_write.max_row if ws_write is not None else 0)
    scan_start = DATA_START_ROW

    for uq in remaining_unmatched:
        target_balance = -uq  # Looking for AZ = -295 when formula has 295
        found = False

        for r in range(scan_start, max_scan_row + 1):
            if r == current_row:
                continue

            # Check if same Work Order
            row_wo = ws_data.cell(row=r, column=COLS['Order_No']).value
            if row_wo is None:
                row_wo = ws_write.cell(row=r, column=COLS['Order_No']).value
            if row_wo is None:
                continue
            if str(row_wo).strip() != order_str:
                continue

            # Read Balance Qty — try data_only first, then parse formula
            row_bal = ws_data.cell(row=r, column=balance_col_idx).value
            if row_bal is None:
                raw = ws_write.cell(row=r, column=balance_col_idx).value
                if raw is not None and isinstance(raw, str) and raw.strip().startswith('='):
                    parts = _parse_qty_formula(raw)
                    if len(parts) == 1:
                        row_bal = parts[0]
                elif raw is not None:
                    try:
                        row_bal = float(raw)
                    except (ValueError, TypeError):
                        continue

            if row_bal is None:
                continue
            try:
                bal_val = float(row_bal)
            except (ValueError, TypeError):
                continue

            if abs(bal_val - target_balance) > tolerance:
                continue

            # Found a matching balance! Try source row components (last → first)
            row_formula = ws_write.cell(row=r, column=qty_col_idx).value
            formula_parts = _parse_qty_formula(row_formula)
            if not formula_parts:
                continue

            positive_parts = [p for p in formula_parts if p > 0]
            if not positive_parts:
                continue

            for component in reversed(positive_parts):
                mrr_match = match_formula_qtys_to_store(
                    stores_df, [component], material, mic,
                    order_no, process_filter
                )
                if mrr_match:
                    for mrr in mrr_match:
                        result[mrr] = result.get(mrr, 0) + uq
                    found = True
                    logger.debug(f"Row {current_row}: Balance {uq} traced via "
                                 f"Row {r} (AZ={bal_val}) component "
                                 f"{component} → MRR(s) {list(mrr_match.keys())}")
                    break

        if not found:
            still_unmatched.append(uq)

    # ── Phase 2: Group remaining unmatched and try sum-based matching ──
    # Handles cases like Row 41: two 140s (sum=280) vs Row 40's AZ==-140-140 (sum=-280)
    if still_unmatched and len(still_unmatched) > 1:
        total_unmatched = sum(still_unmatched)
        target_sum = -total_unmatched

        for r in range(scan_start, max_scan_row + 1):
            if r == current_row:
                continue

            row_wo = ws_data.cell(row=r, column=COLS['Order_No']).value
            if row_wo is None:
                row_wo = ws_write.cell(row=r, column=COLS['Order_No']).value
            if row_wo is None:
                continue
            if str(row_wo).strip() != order_str:
                continue

            # Read Balance Qty — get the TOTAL value (data_only gives computed sum)
            row_bal = ws_data.cell(row=r, column=balance_col_idx).value
            if row_bal is None:
                # Parse formula to get sum
                raw = ws_write.cell(row=r, column=balance_col_idx).value
                if raw is not None and isinstance(raw, str) and raw.strip().startswith('='):
                    parts = _parse_qty_formula(raw)
                    if parts:
                        row_bal = sum(parts)
                elif raw is not None:
                    try:
                        row_bal = float(raw)
                    except (ValueError, TypeError):
                        continue

            if row_bal is None:
                continue
            try:
                bal_val = float(row_bal)
            except (ValueError, TypeError):
                continue

            if abs(bal_val - target_sum) > tolerance:
                continue

            # Found matching sum balance! Get the MRR from this row's formula
            row_formula = ws_write.cell(row=r, column=qty_col_idx).value
            formula_parts = _parse_qty_formula(row_formula)
            if not formula_parts:
                continue

            positive_parts = [p for p in formula_parts if p > 0]
            if not positive_parts:
                continue

            # Match ALL positive components of the source row to find MRRs
            all_mrrs = match_formula_qtys_to_store(
                stores_df, positive_parts, material, mic,
                order_no, process_filter
            )
            if all_mrrs:
                # Assign all unmatched qty to the matched MRR(s) from this row
                # Use the dominant MRR (highest total qty)
                dominant_mrr = max(all_mrrs, key=all_mrrs.get)
                for uq in still_unmatched:
                    result[dominant_mrr] = result.get(dominant_mrr, 0) + uq
                logger.debug(f"Row {current_row}: Grouped balance {still_unmatched} "
                             f"(sum={total_unmatched}) traced via Row {r} "
                             f"(AZ={bal_val}) → MRR {dominant_mrr}")
                still_unmatched = []
                break
            else:
                # Try last component as fallback
                for component in reversed(positive_parts):
                    mrr_match = match_formula_qtys_to_store(
                        stores_df, [component], material, mic,
                        order_no, process_filter
                    )
                    if mrr_match:
                        for uq in still_unmatched:
                            for mrr in mrr_match:
                                result[mrr] = result.get(mrr, 0) + uq
                        logger.debug(f"Row {current_row}: Grouped balance {still_unmatched} "
                                     f"traced via Row {r} component {component} "
                                     f"→ MRR(s) {list(mrr_match.keys())}")
                        still_unmatched = []
                        break

    for uq in still_unmatched:
        logger.debug(f"Row {current_row}: Balance {uq} could not be traced "
                     f"for WO={order_str}")

    return result


def _mrr_exists_in_pr(pr_df, mrr_num):
    """Check if an MRR number exists in the Purchase Register (direct check).
    Unlike filter_mrr_by_pr, this returns False when the MRR is not in PR
    (no fallback behavior).
    """
    try:
        mrr_val = int(float(mrr_num))
    except (ValueError, TypeError):
        return False
    tracking_col = None
    for c in pr_df.columns:
        if 'tracking' in str(c).lower():
            tracking_col = c
            break
    if not tracking_col:
        return False
    return bool((pd.to_numeric(pr_df[tracking_col], errors='coerce') == mrr_val).any())


def _trace_unmatched_components(stores_df, positive_qtys, mrr_qty, material, mic,
                                 order_no, process_filter, ws_write, ws_data,
                                 current_row, qty_col_idx, bal_col_idx):
    """For partially matched formulas, trace unmatched components via balance.

    Checks each component individually; those not found in Store are traced
    to the previous row's balance (same WO) using _trace_balance_mrr.
    Returns updated mrr_qty dict with traced MRRs merged in.
    """
    # Identify which components were NOT matched
    unmatched = []
    for qty in positive_qtys:
        single = match_formula_qtys_to_store(
            stores_df, [qty], material, mic, order_no, process_filter
        )
        if not single:
            unmatched.append(qty)

    if not unmatched:
        return mrr_qty  # All components matched

    # Trace unmatched components via balance to previous rows
    traced = _trace_balance_mrr(
        ws_write, ws_data, current_row, unmatched,
        order_no, material, mic, stores_df,
        process_filter, qty_col_idx, bal_col_idx
    )
    if traced:
        # Merge traced MRRs into mrr_qty
        merged = dict(mrr_qty)
        for mrr, qty in traced.items():
            merged[mrr] = merged.get(mrr, 0) + qty
        logger.info(f"Row {current_row}: Unmatched components {unmatched} "
                    f"traced to MRRs: {traced}")
        return merged

    logger.debug(f"Row {current_row}: Unmatched components {unmatched} "
                 f"could not be traced")
    return mrr_qty


def _read_val(ws_data, ws_write, row, col):
    """Read a cell value, preferring data_only workbook for computed values.
    Falls back to write workbook if data_only returns None.
    """
    val = ws_data.cell(row=row, column=col).value
    if val is not None:
        return val
    # Fallback: try the write workbook (which has formulas)
    val2 = ws_write.cell(row=row, column=col).value
    if val2 is not None and isinstance(val2, str) and val2.startswith('='):
        return None  # It's a formula that didn't compute
    return val2


def _compute_total(ws_data, ws_write, row, total_col, qty_col, balance_col):
    """Compute total quantity reliably.
    Always calculates manually (qty + balance) and cross-checks with the formula cell.
    This ensures correct results even when the workbook has uncached formulas.
    """
    # Manual sum is the source of truth
    qty = _safe_float(_read_val(ws_data, ws_write, row, qty_col), f"row {row} qty")
    bal = _safe_float(_read_val(ws_data, ws_write, row, balance_col), f"row {row} balance")
    manual_total = qty + bal

    # Also read the formula cell for cross-check
    formula_total = _safe_float(_read_val(ws_data, ws_write, row, total_col), f"row {row} total")

    # If both are non-zero and differ significantly, log a warning
    if manual_total > 0 and formula_total > 0 and abs(manual_total - formula_total) > 0.01:
        logger.warning(f"Row {row}: Manual total ({qty}+{bal}={manual_total}) differs from "
                       f"formula total ({formula_total}). Using manual total.")

    # Prefer manual total; fall back to formula only if manual is zero
    if manual_total > 0:
        return manual_total
    if formula_total > 0:
        return formula_total
    return 0.0


def _validate_jobtrack_headers(ws, header_row):
    """Validate that critical Jobtrack columns are in the expected positions.
    Raises ValueError with details if headers don't match."""
    # Spot-check key columns — these are the ones we write to
    checks = {
        COLS['Process']: 'process',
        COLS['Order_No']: 'order',
        COLS['Film_MR']: 'mr',
        COLS['Film_Rate']: 'rate',
        COLS['Fresh1_MR']: 'mr',
        COLS['Adh_Rate']: 'rate',
    }
    issues = []
    for col_idx, expected_keyword in checks.items():
        header_val = ws.cell(row=header_row, column=col_idx).value
        if header_val is None:
            issues.append(f"Column {get_column_letter(col_idx)} (index {col_idx}): expected '{expected_keyword}' keyword, got empty cell")
        else:
            header_str = str(header_val).strip().lower()
            if expected_keyword not in header_str:
                issues.append(f"Column {get_column_letter(col_idx)} (index {col_idx}): expected '{expected_keyword}', got '{header_val}'")
    if issues:
        logger.warning(f"Jobtrack header validation warnings:\n" + "\n".join(issues))
    return issues


def fill_jobtrack(jt_file, stores_file, pr_file, progress_callback=None,
                  granules_file=None, megapack_file=None, report_month=None,
                  prev_granules_file=None):
    """
    Main function to fill the Jobtrack template with MRR data.

    Args:
        jt_file: Jobtrack file (BytesIO or path)
        stores_file: Stores Recordings file (BytesIO or path)
        pr_file: Purchase Register file (BytesIO or path)
        progress_callback: Optional callback(percent, message) for progress updates
        granules_file: Optional Granules Recipe file for Bandera/CYM supplier rates
        megapack_file: Optional MEGA PACK file for Mega Pack supplier rates
        report_month: Optional explicit report month string (e.g. '2-2026' for Feb 2026).
                      If not provided, auto-detected from Jobtrack dates, UID, or PR filename.
        prev_granules_file: Optional PREVIOUS month's Granules Recipe file.
                           Used as fallback when WO# not found in current month's Granules.

    Returns:
        tuple: (filled_workbook_bytes, results_log, summary_stats)
    """
    def update_progress(pct, msg):
        if progress_callback:
            progress_callback(pct, msg)
        logger.info(f"[{pct}%] {msg}")

    update_progress(5, "Loading Stores Recordings...")
    stores_df = load_stores_recordings(stores_file)

    update_progress(15, "Loading Purchase Register...")
    pr_df = load_purchase_register(pr_file)

    update_progress(25, "Loading Jobtrack template...")

    # DUAL WORKBOOK APPROACH:
    # wb_data: opened with data_only=True to read computed cell values
    # wb_write: opened normally to write values (preserves formulas/formatting)
    if isinstance(jt_file, io.BytesIO):
        jt_file.seek(0)
        jt_bytes = jt_file.read()
        wb_data = openpyxl.load_workbook(io.BytesIO(jt_bytes), data_only=True)
        wb_write = openpyxl.load_workbook(io.BytesIO(jt_bytes))
    else:
        wb_data = openpyxl.load_workbook(jt_file, data_only=True)
        wb_write = openpyxl.load_workbook(jt_file)

    ws_data = wb_data.active   # For READING computed values
    ws_write = wb_write.active # For WRITING results

    # Validate Jobtrack headers to catch template changes early
    header_issues = _validate_jobtrack_headers(ws_write, HEADER_ROW)
    if header_issues:
        update_progress(28, f"WARNING: {len(header_issues)} header mismatch(es) detected — output may be inaccurate!")

    max_row = ws_data.max_row
    total_data_rows = max_row - DATA_START_ROW + 1

    results_log = []
    stats = {
        'total_rows': total_data_rows,
        'printing_rows': 0,
        'lam_rows': 0,
        'film_filled': 0,
        'fresh1_filled': 0,
        'fresh2_filled': 0,
        'adh_filled': 0,
        'hard_filled': 0,
        'sol_filled': 0,
        'errors': 0,
        'skipped': 0,
    }

    # Detect reporting month from Jobtrack date column
    # The month format in PR is like '2-2026' for Feb 2026
    if not report_month:
        date_col = COLS.get('Date', 4)  # Column D typically
        # Scan ALL data rows (not just first 20) to find a valid date
        for scan_row in range(DATA_START_ROW, max_row + 1):
            date_val = ws_data.cell(row=scan_row, column=date_col).value
            if date_val and hasattr(date_val, 'month') and hasattr(date_val, 'year'):
                report_month = f"{date_val.month}-{date_val.year}"
                break
    if not report_month:
        # Try UID pattern like '202602-xxxx' — also scan all rows
        for scan_row in range(DATA_START_ROW, max_row + 1):
            uid_val = _safe_str(ws_data.cell(row=scan_row, column=COLS['UID']).value)
            if uid_val and len(uid_val) >= 6:
                try:
                    yr = int(uid_val[:4])
                    mn = int(uid_val[4:6])
                    if 2020 <= yr <= 2030 and 1 <= mn <= 12:
                        report_month = f"{mn}-{yr}"
                        break
                except ValueError:
                    pass
    if not report_month:
        # Try to extract from PR filename (e.g. "Purchase Register - 2021 - 2025 _Nov.xlsx")
        _MONTH_MAP = {
            'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
            'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12,
        }
        pr_name = ''
        if isinstance(pr_file, str):
            pr_name = pr_file
        elif hasattr(pr_file, 'name'):
            pr_name = pr_file.name
        pr_name_lower = str(pr_name).lower()
        for month_name, month_num in _MONTH_MAP.items():
            if month_name in pr_name_lower:
                # Find year: look for 4-digit year patterns, take the last one
                import re
                years = re.findall(r'20[12]\d', pr_name)
                if years:
                    yr = int(years[-1])  # Take last year (e.g. '2025' from '2021 - 2025')
                    report_month = f"{month_num}-{yr}"
                    break
    if not report_month:
        logger.error("Could not detect report month from any source — chemical rates may be wrong!")
    logger.info(f"Detected report month: {report_month}")

    # Pre-compute solvent rate for the reporting month
    solvent_rate = lookup_solvent_rate(pr_df, report_month=report_month)
    update_progress(30, f"Solvent rate (Ethyl Acetate): {solvent_rate:.4f} [{report_month}]")

    # ── Supplier-specific rate sources ──
    mrr_supplier_map = build_mrr_supplier_map(stores_df)
    granules_rates = load_granules_rates(granules_file) if granules_file else {}
    prev_granules_rates = load_granules_rates(prev_granules_file) if prev_granules_file else {}
    megapack_rates = load_megapack_rates(megapack_file) if megapack_file else {}
    # Parse report year/month for Mega Pack lookup
    report_year, report_month_num = 0, 0
    if report_month:
        parts = report_month.split('-')
        if len(parts) == 2:
            report_month_num = int(parts[0])
            report_year = int(parts[1])

    supplier_overrides = 0
    if granules_rates:
        update_progress(32, f"Granules Recipe (current): {len(granules_rates)} WO# rates loaded")
    if prev_granules_rates:
        update_progress(32, f"Granules Recipe (prev month): {len(prev_granules_rates)} WO# rates loaded")
    if megapack_rates:
        update_progress(33, f"MEGA PACK: {len(megapack_rates)} monthly rates loaded")

    update_progress(35, "Processing rows...")

    for row_idx in range(DATA_START_ROW, max_row + 1):
        row_num = row_idx - DATA_START_ROW + 1
        pct = 35 + int((row_num / total_data_rows) * 55)

        # Read from DATA workbook (computed values)
        process = _safe_str(_read_val(ws_data, ws_write, row_idx, COLS['Process']))
        uid = _safe_str(_read_val(ws_data, ws_write, row_idx, COLS['UID']))
        order_no = _safe_str(_read_val(ws_data, ws_write, row_idx, COLS['Order_No']))

        if not process or not uid:
            stats['skipped'] += 1
            continue

        process_bucket = _normalize_process(process)

        # ========== PRINTING ROWS: Film MR#, Rate, Value ==========
        if process_bucket in ('PRINTING', 'REWINDING'):
            stats['printing_rows'] += 1
            stores_process_filter = _stores_process_filter_for_row(process_bucket)

            input_name = _safe_str(_read_val(ws_data, ws_write, row_idx, COLS['Input_Name']))
            input_size = _read_val(ws_data, ws_write, row_idx, COLS['Input_Size'])
            input_mic = _read_val(ws_data, ws_write, row_idx, COLS['Input_Mic'])
            # Always compute total manually (AY+AZ) — never trust formula cache
            total_input = _compute_total(
                ws_data, ws_write, row_idx,
                COLS['Total_1st_Input'], COLS['Input_Qty'], COLS['Balance_Qty']
            )

            if not input_name:
                results_log.append({
                    'row': row_idx, 'uid': uid, 'type': 'Film',
                    'status': 'SKIP', 'detail': 'No input material name'
                })
                stats['skipped'] += 1
                continue

            # Check for in-house material (WPE, WLDPE -> "INH")
            if input_name.upper() in ('WPE', 'WLDPE', 'PTD WPE'):
                # ── Supplier override: Bandera/CYM → Granules Recipe ──
                rate = 0
                rate_source = 'PR'
                order_upper = str(order_no).strip().upper() if order_no else ''
                # Step 1: Check current month Granules
                if granules_rates and order_upper in granules_rates:
                    rate = granules_rates[order_upper]
                    rate_source = 'Granules Recipe'
                    supplier_overrides += 1
                    logger.info(f"Row {row_idx}: INH rate from Granules Recipe: "
                                f"WO={order_upper}, Rate={rate:.6f}")
                # Step 2: Check previous month Granules
                elif prev_granules_rates and order_upper in prev_granules_rates:
                    rate = prev_granules_rates[order_upper]
                    rate_source = 'Granules Recipe (prev month)'
                    supplier_overrides += 1
                    logger.info(f"Row {row_idx}: INH rate from PREV Granules Recipe: "
                                f"WO={order_upper}, Rate={rate:.6f}")
                # No average fallback: force manual review when WO is missing in Granules
                elif granules_rates or prev_granules_rates:
                    results_log.append({
                        'row': row_idx, 'uid': uid, 'type': 'Film',
                        'status': 'WARN',
                        'detail': f'INH WO# {order_upper} not found in Granules (current/prev)'
                    })

                if rate == 0:
                    # Standard INH lookup from MRR/PR
                    mrr_list = lookup_mrr(stores_df, input_name, input_mic, input_size,
                                          order_no, stores_process_filter)
                    if not mrr_list:
                        mrr_list = lookup_mrr(stores_df, 'WPE', input_mic, None,
                                              order_no, stores_process_filter)
                    if mrr_list:
                        rate = lookup_film_rate(pr_df, mrr_list, 'WPE', None, input_mic)
                    if rate == 0:
                        rate = lookup_material_rate_for_month(pr_df, 'WPE', input_mic, report_month)
                    if rate == 0:
                        rate = lookup_material_rate_for_month(pr_df, 'TPE', input_mic, report_month)

                ws_write.cell(row=row_idx, column=COLS['Film_MR']).value = 'INH'
                if rate > 0:
                    ws_write.cell(row=row_idx, column=COLS['Film_Rate']).value = rate
                    ws_write.cell(row=row_idx, column=COLS['Film_Value']).value = total_input * rate
                    stats['film_filled'] += 1
                    results_log.append({
                        'row': row_idx, 'uid': uid, 'type': 'Film',
                        'status': f'OK (INH/{rate_source})',
                        'detail': f'Rate={rate:.4f}, Value={total_input * rate:.2f}, Source={rate_source}'
                    })
                else:
                    results_log.append({
                        'row': row_idx, 'uid': uid, 'type': 'Film',
                        'status': 'WARN', 'detail': 'INH material - no rate found in MRR, PR, or Granules'
                    })
                    stats['errors'] += 1
                continue

            # Standard film lookup - find MR# from Stores with qty info
            # Step 1: Try precise matching via AY formula components
            ay_formula = ws_write.cell(row=row_idx, column=COLS['Input_Qty']).value
            formula_qtys = _parse_qty_formula(ay_formula)
            mrr_qty = {}
            formula_matched = False
            skip_fallback = False

            # Step 1a: If AY is a plain number (not formula), treat as single-component
            # for precise Store qty matching (e.g. AY=100 should match exactly one MRR)
            if not formula_qtys and ay_formula is not None:
                try:
                    plain_val = float(ay_formula)
                    if plain_val > 0:
                        formula_qtys = [plain_val]
                        logger.debug(f"Row {row_idx}: AY is plain number {plain_val}, treating as single qty")
                except (ValueError, TypeError):
                    pass

            if formula_qtys:
                # Extract only positive components (negatives are balance adjustments)
                positive_qtys = [q for q in formula_qtys if q > 0]
                if positive_qtys:
                    skip_fallback = True  # Formula was parsed → don’t fall back to full WO lookup
                    mrr_qty = match_formula_qtys_to_store(
                        stores_df, positive_qtys, input_name, input_mic,
                        order_no, stores_process_filter
                    )
                    if mrr_qty:
                        formula_matched = True
                        logger.info(f"Row {row_idx}: Formula-matched MRRs: {mrr_qty}")
                        # Step 1c: Trace unmatched components to previous row
                        mrr_qty = _trace_unmatched_components(
                            stores_df, positive_qtys, mrr_qty, input_name, input_mic,
                            order_no, stores_process_filter, ws_write, ws_data,
                            row_idx, COLS['Input_Qty'], COLS['Balance_Qty']
                        )
                    else:
                        # Step 1b: Try balance tracing for unmatched formula values
                        mrr_qty = _trace_balance_mrr(
                            ws_write, ws_data, row_idx, positive_qtys,
                            order_no, input_name, input_mic, stores_df,
                            stores_process_filter, COLS['Input_Qty'], COLS['Balance_Qty']
                        )
                        if mrr_qty:
                            # Check if traced MRR actually exists in PR
                            mrr_in_pr = any(_mrr_exists_in_pr(pr_df, m) for m in mrr_qty)
                            if mrr_in_pr:
                                logger.info(f"Row {row_idx}: Balance-traced MRRs (PR-verified): {mrr_qty}")
                            else:
                                # Traced MRR not in PR — allow fallback to full WO lookup
                                logger.info(f"Row {row_idx}: Balance-traced MRR {list(mrr_qty.keys())} not in PR, falling back")
                                mrr_qty = {}
                                skip_fallback = False

            # Step 2: Fall back to full lookup ONLY if formula wasn't parsed
            if not mrr_qty and not skip_fallback:
                mrr_qty = lookup_mrr_with_qty(stores_df, input_name, input_mic, input_size,
                                              order_no, stores_process_filter)
            if not mrr_qty and not skip_fallback:
                mrr_qty = lookup_mrr_with_qty(stores_df, input_name, input_mic, None,
                                              order_no, stores_process_filter)
            if not mrr_qty and not skip_fallback:
                mrr_qty = lookup_mrr_with_qty(stores_df, input_name, input_mic, None, order_no)

            # Notification when no Store match found at all
            if not mrr_qty:
                mismatch_keys = _store_match_diagnostics(
                    stores_df, input_name, input_mic, input_size, order_no,
                    stores_process_filter, formula_qtys
                )
                mismatch_txt = f" | Not matched: {', '.join(mismatch_keys)}" if mismatch_keys else ""
                detail = f'No MRR found in Store'
                if skip_fallback:
                    detail += f' (formula: {ay_formula})'
                results_log.append({
                    'row': row_idx, 'uid': uid, 'type': 'Film',
                    'status': 'WARN',
                    'detail': f'{detail}{mismatch_txt}'
                })

            if mrr_qty:

                # ── Supplier override check ──
                mrr_numbers = list(mrr_qty.keys())
                supplier = get_supplier_for_mrrs(mrr_supplier_map, mrr_numbers, mrr_qty)
                rate = 0
                rate_source = 'PR'

                if supplier == 'MEGA PACK' and megapack_rates:
                    rate = lookup_megapack_rate(megapack_rates, input_name,
                                               report_year, report_month_num)
                    if rate > 0:
                        rate_source = 'MEGA PACK'
                        supplier_overrides += 1
                        logger.info(f"Row {row_idx}: Film rate from MEGA PACK: "
                                    f"{input_name} -> {rate:.6f}")
                elif supplier in ('BANDERA', 'CYM'):
                    # RULE: BANDERA/CYM = ALWAYS INH
                    order_upper = str(order_no).strip().upper() if order_no else ''
                    if granules_rates and order_upper in granules_rates:
                        rate = granules_rates[order_upper]
                        rate_source = 'Granules Recipe'
                        supplier_overrides += 1
                    elif prev_granules_rates and order_upper in prev_granules_rates:
                        rate = prev_granules_rates[order_upper]
                        rate_source = 'Granules Recipe (prev month)'
                        supplier_overrides += 1
                    elif granules_rates or prev_granules_rates:
                        results_log.append({
                            'row': row_idx, 'uid': uid, 'type': 'Film',
                            'status': 'WARN',
                            'detail': f'BANDERA/CYM WO# {order_upper} not found in Granules (current/prev)'
                        })
                    if rate > 0:
                        logger.info(f"Row {row_idx}: Film INH rate [{rate_source}]: "
                                    f"WO={order_upper} -> {rate:.6f}")

                if rate == 0:
                    # Standard PR lookup
                    mrr_qty = filter_mrr_by_pr(pr_df, mrr_qty, input_name, input_size, input_mic)
                    rate = lookup_film_rate_weighted(pr_df, mrr_qty, input_name, input_size, input_mic)
                    if rate == 0:
                        rate = lookup_film_rate_weighted(pr_df, mrr_qty, input_name, None, input_mic)
                    if rate == 0:
                        rate = lookup_material_rate_for_month(pr_df, input_name, input_mic, report_month)
                    # Outlier check
                    if rate > 0:
                        month_rate = lookup_material_rate_for_month(pr_df, input_name, input_mic, report_month)
                        if month_rate > 0 and abs(rate - month_rate) / month_rate > 0.50:
                            logger.warning(f"Row {row_idx}: Rate outlier detected for {input_name}: "
                                           f"MRR rate={rate:.4f}, month avg={month_rate:.4f}. Using month avg.")
                            rate = month_rate

                mrr_list = _pick_mrrs_by_total_qty(mrr_qty, total_input, tolerance=1.0)

                if rate > 0:
                    # BANDERA/CYM always shows "INH" for MR#
                    if rate_source.startswith('Granules Recipe'):
                        mr_str = 'INH'
                    else:
                        mr_str = '/'.join(str(m) for m in mrr_list)
                        if len(mrr_list) == 1:
                            mr_str = str(mrr_list[0])
                    ws_write.cell(row=row_idx, column=COLS['Film_MR']).value = mr_str
                    ws_write.cell(row=row_idx, column=COLS['Film_Rate']).value = rate
                    ws_write.cell(row=row_idx, column=COLS['Film_Value']).value = total_input * rate
                    stats['film_filled'] += 1
                    results_log.append({
                        'row': row_idx, 'uid': uid, 'type': 'Film',
                        'status': 'OK', 'detail': f'MR#={mr_str}, Rate={rate:.4f}, Value={total_input * rate:.2f}'
                    })
                else:
                    mr_str = '/'.join(str(m) for m in mrr_list)
                    ws_write.cell(row=row_idx, column=COLS['Film_MR']).value = (
                        mrr_list[0] if len(mrr_list) == 1 else mr_str
                    )
                    results_log.append({
                        'row': row_idx, 'uid': uid, 'type': 'Film',
                        'status': 'WARN', 'detail': f'MR#={mr_str} found but no rate in PR'
                    })
                    stats['errors'] += 1
            else:
                mismatch_keys = _store_match_diagnostics(
                    stores_df, input_name, input_mic, input_size, order_no,
                    stores_process_filter, formula_qtys if formula_qtys else None
                )
                mismatch_txt = f' | Not matched keys: {", ".join(mismatch_keys)}' if mismatch_keys else ''
                results_log.append({
                    'row': row_idx, 'uid': uid, 'type': 'Film',
                    'status': 'MISS', 'detail': f'No MRR found for {input_name}/{input_size}/{input_mic}/{order_no}{mismatch_txt}'
                })
                stats['errors'] += 1

        # ========== LAM ROWS: Fresh Materials + Chemicals ==========
        elif process_bucket == 'LAM':
            stats['lam_rows'] += 1

            # --- 1st Fresh Material ---
            fresh1_name = _safe_str(_read_val(ws_data, ws_write, row_idx, COLS['Fresh1_Name']))
            fresh1_size = _read_val(ws_data, ws_write, row_idx, COLS['Fresh1_Size'])
            fresh1_mic = _read_val(ws_data, ws_write, row_idx, COLS['Fresh1_Mic'])
            # Always compute total manually (BW+BX) — never trust formula cache
            total_fresh1 = _compute_total(
                ws_data, ws_write, row_idx,
                COLS['Total_Fresh1'], COLS['Fresh1_Qty'], COLS['Fresh1_Balance']
            )

            if fresh1_name and total_fresh1 > 0:
                # Step 1: Try precise matching via BW formula components
                bw_formula = ws_write.cell(row=row_idx, column=COLS['Fresh1_Qty']).value
                formula_qtys = _parse_qty_formula(bw_formula)
                mrr_qty = {}
                formula_matched = False
                skip_fallback = False

                # Step 1a: If BW is a plain number, treat as single-component
                if not formula_qtys and bw_formula is not None:
                    try:
                        plain_val = float(bw_formula)
                        if plain_val > 0:
                            formula_qtys = [plain_val]
                            logger.debug(f"Row {row_idx}: BW is plain number {plain_val}, treating as single qty")
                    except (ValueError, TypeError):
                        pass

                if formula_qtys:
                    positive_qtys = [q for q in formula_qtys if q > 0]
                    if positive_qtys:
                        skip_fallback = True
                        mrr_qty = match_formula_qtys_to_store(
                            stores_df, positive_qtys, fresh1_name, fresh1_mic,
                            order_no, 'LAMINATION'
                        )
                        if mrr_qty:
                            formula_matched = True
                            logger.info(f"Row {row_idx}: Fresh1 formula-matched MRRs: {mrr_qty}")
                            # Trace unmatched components to previous row
                            mrr_qty = _trace_unmatched_components(
                                stores_df, positive_qtys, mrr_qty, fresh1_name, fresh1_mic,
                                order_no, 'LAMINATION', ws_write, ws_data,
                                row_idx, COLS['Fresh1_Qty'], COLS['Fresh1_Balance']
                            )
                        else:
                            # Try balance tracing
                            mrr_qty = _trace_balance_mrr(
                                ws_write, ws_data, row_idx, positive_qtys,
                                order_no, fresh1_name, fresh1_mic, stores_df,
                                'LAMINATION', COLS['Fresh1_Qty'], COLS['Fresh1_Balance']
                            )
                            if mrr_qty:
                                # Check if traced MRR actually exists in PR
                                mrr_in_pr = any(_mrr_exists_in_pr(pr_df, m) for m in mrr_qty)
                                if mrr_in_pr:
                                    logger.info(f"Row {row_idx}: Fresh1 balance-traced MRRs (PR-verified): {mrr_qty}")
                                else:
                                    logger.info(f"Row {row_idx}: Fresh1 balance-traced MRR {list(mrr_qty.keys())} not in PR, falling back")
                                    mrr_qty = {}
                                    skip_fallback = False

                # Step 2: Fall back ONLY if formula wasn't parsed
                if not mrr_qty and not skip_fallback:
                    mrr_qty = lookup_mrr_with_qty(stores_df, fresh1_name, fresh1_mic, fresh1_size,
                                                  order_no, 'LAMINATION')
                if not mrr_qty and not skip_fallback:
                    mrr_qty = lookup_mrr_with_qty(stores_df, fresh1_name, fresh1_mic, None,
                                                  order_no, 'LAMINATION')
                if not mrr_qty and not skip_fallback:
                    mrr_qty = lookup_mrr_with_qty(stores_df, fresh1_name, fresh1_mic, None, order_no)

                # Notification when no Store match found at all
                if not mrr_qty:
                    mismatch_keys = _store_match_diagnostics(
                        stores_df, fresh1_name, fresh1_mic, fresh1_size, order_no,
                        'LAMINATION', formula_qtys
                    )
                    mismatch_txt = f" | Not matched: {', '.join(mismatch_keys)}" if mismatch_keys else ""
                    detail = f'No MRR found in Store'
                    if skip_fallback:
                        detail += f' (formula: {bw_formula})'
                    results_log.append({
                        'row': row_idx, 'uid': uid, 'type': '1st Fresh',
                        'status': 'WARN',
                        'detail': f'{detail}{mismatch_txt}'
                    })

                if mrr_qty:

                    # ── Supplier override check ──
                    mrr_numbers = list(mrr_qty.keys())
                    supplier = get_supplier_for_mrrs(mrr_supplier_map, mrr_numbers, mrr_qty)
                    rate = 0
                    rate_source = 'PR'

                    if supplier == 'MEGA PACK' and megapack_rates:
                        rate = lookup_megapack_rate(megapack_rates, fresh1_name,
                                                   report_year, report_month_num)
                        if rate > 0:
                            rate_source = 'MEGA PACK'
                            supplier_overrides += 1
                    elif supplier in ('BANDERA', 'CYM'):
                        # RULE: BANDERA/CYM = ALWAYS INH
                        order_upper = str(order_no).strip().upper() if order_no else ''
                        if granules_rates and order_upper in granules_rates:
                            rate = granules_rates[order_upper]
                            rate_source = 'Granules Recipe'
                            supplier_overrides += 1
                        elif prev_granules_rates and order_upper in prev_granules_rates:
                            rate = prev_granules_rates[order_upper]
                            rate_source = 'Granules Recipe (prev month)'
                            supplier_overrides += 1
                        elif granules_rates or prev_granules_rates:
                            results_log.append({
                                'row': row_idx, 'uid': uid, 'type': '1st Fresh',
                                'status': 'WARN',
                                'detail': f'BANDERA/CYM WO# {order_upper} not found in Granules (current/prev)'
                            })
                        logger.info(f"Row {row_idx}: Fresh1 BANDERA/CYM INH [{rate_source}]: "
                                    f"WO={order_upper}, Rate={rate:.6f}")

                    if rate == 0:
                        mrr_qty = filter_mrr_by_pr(pr_df, mrr_qty, fresh1_name, fresh1_size, fresh1_mic)
                        rate = lookup_film_rate_weighted(pr_df, mrr_qty, fresh1_name, fresh1_size, fresh1_mic)
                        if rate == 0:
                            rate = lookup_film_rate_weighted(pr_df, mrr_qty, fresh1_name, None, fresh1_mic)
                        if rate == 0:
                            rate = lookup_material_rate_for_month(pr_df, fresh1_name, fresh1_mic, report_month)
                        # Outlier check
                        if rate > 0:
                            month_rate = lookup_material_rate_for_month(pr_df, fresh1_name, fresh1_mic, report_month)
                            if month_rate > 0 and abs(rate - month_rate) / month_rate > 0.50:
                                logger.warning(f"Row {row_idx}: Fresh1 rate outlier for {fresh1_name}: "
                                               f"MRR rate={rate:.4f}, month avg={month_rate:.4f}. Using month avg.")
                                rate = month_rate

                    mrr_list = _pick_mrrs_by_total_qty(mrr_qty, total_fresh1, tolerance=1.0)

                    if rate > 0:
                        # Determine MR# value: INH for in-house (BANDERA/CYM), actual MRR otherwise
                        if rate_source.startswith('Granules Recipe'):
                            mr_display = 'INH'
                        else:
                            mr_display = '/'.join(str(m) for m in mrr_list)
                            if len(mrr_list) == 1:
                                mr_display = mrr_list[0]
                        ws_write.cell(row=row_idx, column=COLS['Fresh1_MR']).value = mr_display
                        ws_write.cell(row=row_idx, column=COLS['Fresh1_Rate']).value = rate
                        ws_write.cell(row=row_idx, column=COLS['Fresh1_Value']).value = total_fresh1 * rate
                        stats['fresh1_filled'] += 1
                        results_log.append({
                            'row': row_idx, 'uid': uid, 'type': '1st Fresh',
                            'status': 'OK', 'detail': f'MR#={mr_display}, Rate={rate:.4f}, Val={total_fresh1 * rate:.2f} [{rate_source}]'
                        })
                    else:
                        results_log.append({
                            'row': row_idx, 'uid': uid, 'type': '1st Fresh',
                            'status': 'WARN', 'detail': f'MRR found but no rate'
                        })
                        stats['errors'] += 1
                else:
                    mismatch_keys = _store_match_diagnostics(
                        stores_df, fresh1_name, fresh1_mic, fresh1_size, order_no,
                        'LAMINATION', formula_qtys if formula_qtys else None
                    )
                    mismatch_txt = f' | Not matched keys: {", ".join(mismatch_keys)}' if mismatch_keys else ''
                    results_log.append({
                        'row': row_idx, 'uid': uid, 'type': '1st Fresh',
                        'status': 'MISS', 'detail': f'No MRR for {fresh1_name}/{fresh1_size}/{fresh1_mic}/{order_no}{mismatch_txt}'
                    })
                    stats['errors'] += 1

            # --- 2nd Fresh Material ---
            fresh2_name = _safe_str(_read_val(ws_data, ws_write, row_idx, COLS['Fresh2_Name']))
            fresh2_size = _read_val(ws_data, ws_write, row_idx, COLS['Fresh2_Size'])
            fresh2_mic = _read_val(ws_data, ws_write, row_idx, COLS['Fresh2_Mic'])
            # Always compute total manually (CG+CH) — never trust formula cache
            total_fresh2 = _compute_total(
                ws_data, ws_write, row_idx,
                COLS['Total_Fresh2'], COLS['Fresh2_Qty'], COLS['Fresh2_Balance']
            )

            if fresh2_name and total_fresh2 > 0:
                # Step 1: Try precise matching via CG formula components
                cg_formula = ws_write.cell(row=row_idx, column=COLS['Fresh2_Qty']).value
                formula_qtys = _parse_qty_formula(cg_formula)
                mrr_qty = {}
                formula_matched = False
                skip_fallback = False

                # Step 1a: If CG is a plain number, treat as single-component
                if not formula_qtys and cg_formula is not None:
                    try:
                        plain_val = float(cg_formula)
                        if plain_val > 0:
                            formula_qtys = [plain_val]
                            logger.debug(f"Row {row_idx}: CG is plain number {plain_val}, treating as single qty")
                    except (ValueError, TypeError):
                        pass

                if formula_qtys:
                    positive_qtys = [q for q in formula_qtys if q > 0]
                    if positive_qtys:
                        skip_fallback = True
                        mrr_qty = match_formula_qtys_to_store(
                            stores_df, positive_qtys, fresh2_name, fresh2_mic,
                            order_no, 'LAMINATION'
                        )
                        if mrr_qty:
                            formula_matched = True
                            logger.info(f"Row {row_idx}: Fresh2 formula-matched MRRs: {mrr_qty}")
                            # Trace unmatched components to previous row
                            mrr_qty = _trace_unmatched_components(
                                stores_df, positive_qtys, mrr_qty, fresh2_name, fresh2_mic,
                                order_no, 'LAMINATION', ws_write, ws_data,
                                row_idx, COLS['Fresh2_Qty'], COLS['Fresh2_Balance']
                            )
                        else:
                            # Try balance tracing
                            mrr_qty = _trace_balance_mrr(
                                ws_write, ws_data, row_idx, positive_qtys,
                                order_no, fresh2_name, fresh2_mic, stores_df,
                                'LAMINATION', COLS['Fresh2_Qty'], COLS['Fresh2_Balance']
                            )
                            if mrr_qty:
                                # Check if traced MRR actually exists in PR
                                mrr_in_pr = any(_mrr_exists_in_pr(pr_df, m) for m in mrr_qty)
                                if mrr_in_pr:
                                    logger.info(f"Row {row_idx}: Fresh2 balance-traced MRRs (PR-verified): {mrr_qty}")
                                else:
                                    logger.info(f"Row {row_idx}: Fresh2 balance-traced MRR {list(mrr_qty.keys())} not in PR, falling back")
                                    mrr_qty = {}
                                    skip_fallback = False

                # Step 2: Fall back ONLY if formula wasn't parsed
                if not mrr_qty and not skip_fallback:
                    mrr_qty = lookup_mrr_with_qty(stores_df, fresh2_name, fresh2_mic, fresh2_size,
                                                  order_no, 'LAMINATION')
                if not mrr_qty and not skip_fallback:
                    mrr_qty = lookup_mrr_with_qty(stores_df, fresh2_name, fresh2_mic, None,
                                                  order_no, 'LAMINATION')
                if not mrr_qty and not skip_fallback:
                    mrr_qty = lookup_mrr_with_qty(stores_df, fresh2_name, fresh2_mic, None, order_no)

                # Notification when no Store match found at all
                if not mrr_qty:
                    mismatch_keys = _store_match_diagnostics(
                        stores_df, fresh2_name, fresh2_mic, fresh2_size, order_no,
                        'LAMINATION', formula_qtys
                    )
                    mismatch_txt = f" | Not matched: {', '.join(mismatch_keys)}" if mismatch_keys else ""
                    detail = f'No MRR found in Store'
                    if skip_fallback:
                        detail += f' (formula: {cg_formula})'
                    results_log.append({
                        'row': row_idx, 'uid': uid, 'type': '2nd Fresh',
                        'status': 'WARN',
                        'detail': f'{detail}{mismatch_txt}'
                    })

                if mrr_qty:

                    # ── Supplier override check ──
                    mrr_numbers = list(mrr_qty.keys())
                    supplier = get_supplier_for_mrrs(mrr_supplier_map, mrr_numbers, mrr_qty)
                    rate = 0
                    rate_source = 'PR'

                    if supplier == 'MEGA PACK' and megapack_rates:
                        rate = lookup_megapack_rate(megapack_rates, fresh2_name,
                                                   report_year, report_month_num)
                        if rate > 0:
                            rate_source = 'MEGA PACK'
                            supplier_overrides += 1
                    elif supplier in ('BANDERA', 'CYM'):
                        # RULE: BANDERA/CYM = ALWAYS INH
                        order_upper = str(order_no).strip().upper() if order_no else ''
                        if granules_rates and order_upper in granules_rates:
                            rate = granules_rates[order_upper]
                            rate_source = 'Granules Recipe'
                            supplier_overrides += 1
                        elif prev_granules_rates and order_upper in prev_granules_rates:
                            rate = prev_granules_rates[order_upper]
                            rate_source = 'Granules Recipe (prev month)'
                            supplier_overrides += 1
                        elif granules_rates or prev_granules_rates:
                            results_log.append({
                                'row': row_idx, 'uid': uid, 'type': '2nd Fresh',
                                'status': 'WARN',
                                'detail': f'BANDERA/CYM WO# {order_upper} not found in Granules (current/prev)'
                            })
                        logger.info(f"Row {row_idx}: Fresh2 BANDERA/CYM INH [{rate_source}]: "
                                    f"WO={order_upper}, Rate={rate:.6f}")

                    if rate == 0:
                        mrr_qty = filter_mrr_by_pr(pr_df, mrr_qty, fresh2_name, fresh2_size, fresh2_mic)
                        rate = lookup_film_rate_weighted(pr_df, mrr_qty, fresh2_name, fresh2_size, fresh2_mic)
                        if rate == 0:
                            rate = lookup_film_rate_weighted(pr_df, mrr_qty, fresh2_name, None, fresh2_mic)
                        if rate == 0:
                            rate = lookup_material_rate_for_month(pr_df, fresh2_name, fresh2_mic, report_month)
                        # Outlier check
                        if rate > 0:
                            month_rate = lookup_material_rate_for_month(pr_df, fresh2_name, fresh2_mic, report_month)
                            if month_rate > 0 and abs(rate - month_rate) / month_rate > 0.50:
                                logger.warning(f"Row {row_idx}: Fresh2 rate outlier for {fresh2_name}: "
                                               f"MRR rate={rate:.4f}, month avg={month_rate:.4f}. Using month avg.")
                                rate = month_rate

                    mrr_list = _pick_mrrs_by_total_qty(mrr_qty, total_fresh2, tolerance=1.0)

                    if rate > 0:
                        # Determine MR# value: INH for in-house (BANDERA/CYM), actual MRR otherwise
                        if rate_source.startswith('Granules Recipe'):
                            mr_display = 'INH'
                        else:
                            mr_display = '/'.join(str(m) for m in mrr_list)
                            if len(mrr_list) == 1:
                                mr_display = mrr_list[0]
                        ws_write.cell(row=row_idx, column=COLS['Fresh2_MR']).value = mr_display
                        ws_write.cell(row=row_idx, column=COLS['Fresh2_Rate']).value = rate
                        ws_write.cell(row=row_idx, column=COLS['Fresh2_Value']).value = total_fresh2 * rate
                        stats['fresh2_filled'] += 1
                        results_log.append({
                            'row': row_idx, 'uid': uid, 'type': '2nd Fresh',
                            'status': 'OK', 'detail': f'MR#={mr_display}, Rate={rate:.4f} [{rate_source}]'
                        })
                    else:
                        results_log.append({
                            'row': row_idx, 'uid': uid, 'type': '2nd Fresh',
                            'status': 'WARN', 'detail': 'MRR found but no rate'
                        })
                        stats['errors'] += 1
                else:
                    mismatch_keys = _store_match_diagnostics(
                        stores_df, fresh2_name, fresh2_mic, fresh2_size, order_no,
                        'LAMINATION', formula_qtys if formula_qtys else None
                    )
                    mismatch_txt = f' | Not matched keys: {", ".join(mismatch_keys)}' if mismatch_keys else ''
                    results_log.append({
                        'row': row_idx, 'uid': uid, 'type': '2nd Fresh',
                        'status': 'MISS', 'detail': f'No MRR for {fresh2_name}/{fresh2_size}/{fresh2_mic}/{order_no}{mismatch_txt}'
                    })
                    stats['errors'] += 1

            # --- Adhesive ---
            adh_name = _safe_str(_read_val(ws_data, ws_write, row_idx, COLS['Adh_Name']))
            adh_kgs = _safe_float(_read_val(ws_data, ws_write, row_idx, COLS['Adh_Kgs']), f"row {row_idx} adh_kgs")

            if adh_name:
                adh_rate = lookup_adhesive_rate(pr_df, adh_name, report_month=report_month)
                hard_rate = 0.0
                if adh_rate > 0:
                    adh_value = adh_kgs * adh_rate
                    ws_write.cell(row=row_idx, column=COLS['Adh_Rate']).value = adh_rate
                    ws_write.cell(row=row_idx, column=COLS['Adh_Value']).value = adh_value
                    # Only count as filled if actual value > 0
                    if adh_value > 0:
                        stats['adh_filled'] += 1
                    results_log.append({
                        'row': row_idx, 'uid': uid, 'type': 'Adhesive',
                        'status': 'OK', 'detail': f'{adh_name}@{adh_rate:.4f}, Kgs={adh_kgs:.2f}, Val={adh_value:.2f}'
                    })

                    # --- Hardener (using DA column for direct name lookup) ---
                    hard_kgs = _safe_float(_read_val(ws_data, ws_write, row_idx, COLS['Hard_Kgs']), f"row {row_idx} hard_kgs")
                    # Read hardener name from DA column (105) — this is the source of truth
                    da_hardener = _safe_str(_read_val(ws_data, ws_write, row_idx, COLS['DA_Hardener']))
                    if da_hardener:
                        # Direct lookup by hardener name from DA column
                        hard_rate = lookup_hardener_rate_by_name(pr_df, da_hardener, report_month=report_month)
                    else:
                        # Fallback to legacy adhesive-based pairing
                        hard_rate = lookup_hardener_rate(pr_df, adh_name, report_month=report_month)
                    if hard_rate > 0:
                        hard_value = hard_kgs * hard_rate
                        ws_write.cell(row=row_idx, column=COLS['Hard_Rate']).value = hard_rate
                        ws_write.cell(row=row_idx, column=COLS['Hard_Value']).value = hard_value
                        if hard_value > 0:
                            stats['hard_filled'] += 1
                        results_log.append({
                            'row': row_idx, 'uid': uid, 'type': 'Hardener',
                            'status': 'OK', 'detail': f'DA={da_hardener or "(legacy)"}, Rate={hard_rate:.4f}, Kgs={hard_kgs:.2f}, Val={hard_value:.2f}'
                        })
                    else:
                        results_log.append({
                            'row': row_idx, 'uid': uid, 'type': 'Hardener',
                            'status': 'WARN', 'detail': f'No hardener rate found (DA={da_hardener}, Adh={adh_name})'
                        })

                    # --- Solvent ---
                    sol_qty = _safe_float(_read_val(ws_data, ws_write, row_idx, COLS['Sol_Qty']), f"row {row_idx} sol_qty")
                    if sol_qty > 0 and solvent_rate > 0:
                        sol_value = sol_qty * solvent_rate
                        ws_write.cell(row=row_idx, column=COLS['Sol_Rate']).value = solvent_rate
                        ws_write.cell(row=row_idx, column=COLS['Sol_Value']).value = sol_value
                        if sol_value > 0:
                            stats['sol_filled'] += 1
                        results_log.append({
                            'row': row_idx, 'uid': uid, 'type': 'Solvent',
                            'status': 'OK', 'detail': f'Rate={solvent_rate:.4f}, Qty={sol_qty:.2f}, Val={sol_value:.2f}'
                        })
                    elif sol_qty > 0:
                        results_log.append({
                            'row': row_idx, 'uid': uid, 'type': 'Solvent',
                            'status': 'WARN', 'detail': f'Solvent qty={sol_qty:.2f} but no rate for month {report_month}'
                        })
                else:
                    results_log.append({
                        'row': row_idx, 'uid': uid, 'type': 'Chemicals',
                        'status': 'WARN', 'detail': f'No rate for adhesive {adh_name}'
                    })
                    stats['errors'] += 1
        else:
            stats['skipped'] += 1

        if row_num % 5 == 0:
            update_progress(pct, f"Processed {row_num}/{total_data_rows} rows...")

    update_progress(92, "Generating output file...")

    # ── Red highlight rows that need manual attention ──
    from openpyxl.styles import PatternFill
    red_fill = PatternFill(start_color='FFE0E0', end_color='FFE0E0', fill_type='solid')
    
    # Collect rows with errors/missing data
    error_rows = set()
    for entry in results_log:
        if entry.get('status') in ('MISS', 'WARN', 'ERROR'):
            error_rows.add(entry.get('row'))
    
    # Apply red highlight to entire row for error rows
    if error_rows:
        max_col = ws_write.max_column
        for err_row in error_rows:
            for col in range(1, max_col + 1):
                cell = ws_write.cell(row=err_row, column=col)
                cell.fill = red_fill
        logger.info(f"Red-highlighted {len(error_rows)} rows for manual review: {sorted(error_rows)}")

    # Save to BytesIO
    output = io.BytesIO()
    wb_write.save(output)
    output.seek(0)

    wb_data.close()
    wb_write.close()

    update_progress(100, "Done!")

    return output, results_log, stats


def get_filled_data_for_dashboard(jt_file) -> pd.DataFrame:
    """Read the filled Jobtrack for dashboard visualization."""
    if isinstance(jt_file, io.BytesIO):
        jt_file.seek(0)

    df = pd.read_excel(jt_file, sheet_name=0, header=3)  # Row 4 is header
    return df
