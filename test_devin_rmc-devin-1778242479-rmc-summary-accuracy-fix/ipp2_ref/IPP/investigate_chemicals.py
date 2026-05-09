"""
Investigation script: ADHESIVE, HARDENER, SOLVENT rules
Compare Template2 data (Nov) with Template_Files data (Feb) to understand patterns.
"""
import pandas as pd
import openpyxl

T2 = "Template2"
T1 = "Template_Files"

print("=" * 80)
print("INVESTIGATION: ADHESIVE / HARDENER / SOLVENT RULES")
print("=" * 80)

# ── 1. Read the Jobtrack files (both months) ──
print("\n" + "═" * 80)
print("1. JOBTRACK ANALYSIS — Template2 (Nov 2025)")
print("═" * 80)

# Template2 - Jobtrack With MRR (Nov)
jt2_wb = openpyxl.load_workbook(f"{T2}/Jobtrack With MRR.xlsx", data_only=True)
jt2_ws = jt2_wb.active

# Print header row to find columns
print("\n--- Header Row (Row 4) for Chemical columns (col 91..102) ---")
for col in range(85, 110):
    val = jt2_ws.cell(row=4, column=col).value
    letter = openpyxl.utils.get_column_letter(col)
    if val:
        print(f"  Col {letter} ({col}): {val}")

# Read LAM rows and their chemical data
print("\n--- LAM rows with Adhesive/Hardener/Solvent data ---")
lam_data = []
for row in range(5, jt2_ws.max_row + 1):
    process = jt2_ws.cell(row=row, column=6).value  # F = Process
    if process and str(process).strip().upper() == 'LAM':
        uid = jt2_ws.cell(row=row, column=1).value
        adh_name = jt2_ws.cell(row=row, column=91).value  # CM
        adh_kgs = jt2_ws.cell(row=row, column=92).value    # CN
        adh_rate = jt2_ws.cell(row=row, column=93).value   # CO
        adh_val = jt2_ws.cell(row=row, column=94).value    # CP
        
        # What's in column CQ (95)? Let's check
        col_95 = jt2_ws.cell(row=row, column=95).value     # CQ
        
        hard_kgs = jt2_ws.cell(row=row, column=96).value   # CR
        hard_rate = jt2_ws.cell(row=row, column=97).value   # CS
        hard_val = jt2_ws.cell(row=row, column=98).value    # CT
        
        # What's in column CU (99)?
        col_99 = jt2_ws.cell(row=row, column=99).value     # CU
        
        sol_qty = jt2_ws.cell(row=row, column=100).value    # CV
        sol_rate = jt2_ws.cell(row=row, column=101).value   # CW
        sol_val = jt2_ws.cell(row=row, column=102).value    # CX
        
        if adh_name:
            lam_data.append({
                'row': row, 'uid': uid,
                'adh_name': adh_name,
                'adh_kgs': adh_kgs, 'adh_rate': adh_rate, 'adh_val': adh_val,
                'col_CQ': col_95,
                'hard_kgs': hard_kgs, 'hard_rate': hard_rate, 'hard_val': hard_val,
                'col_CU': col_99,
                'sol_qty': sol_qty, 'sol_rate': sol_rate, 'sol_val': sol_val,
            })

print(f"\nTotal LAM rows with adhesive: {len(lam_data)}")
print("\n--- Sample data (first 20) ---")
for d in lam_data[:20]:
    print(f"  Row {d['row']}: Adh={d['adh_name']}, Adh_Rate={d['adh_rate']}, "
          f"CQ={d['col_CQ']}, Hard_Rate={d['hard_rate']}, "
          f"CU={d['col_CU']}, Sol_Rate={d['sol_rate']}")

# Count unique adhesive names
adh_names = [str(d['adh_name']).strip().upper() for d in lam_data if d['adh_name']]
print(f"\n--- Unique Adhesive Names ---")
for name in sorted(set(adh_names)):
    count = adh_names.count(name)
    print(f"  {name}: {count} rows")

# Check all adhesive-hardener pairings
print("\n--- Adhesive → Hardener Pairing Analysis ---")
print("  (Checking what hardener name appears for each adhesive)")
# In the Jobtrack, there's usually a DA column for hardener name
# Let's check columns around 95-100 more carefully

# Print the full header from col 85 to 110
print("\n--- Full header row 4 (columns 85-110) ---")
for col in range(85, 115):
    val = jt2_ws.cell(row=4, column=col).value
    letter = openpyxl.utils.get_column_letter(col)
    print(f"  {letter} ({col}): {val}")

# Also check header row 3
print("\n--- Full header row 3 (columns 85-110) ---")
for col in range(85, 115):
    val = jt2_ws.cell(row=3, column=col).value
    letter = openpyxl.utils.get_column_letter(col)
    if val:
        print(f"  {letter} ({col}): {val}")

# Also header row 2
print("\n--- Full header row 2 (columns 85-110) ---")
for col in range(85, 115):
    val = jt2_ws.cell(row=3, column=col).value
    letter = openpyxl.utils.get_column_letter(col)
    if val:
        print(f"  {letter} ({col}): {val}")

jt2_wb.close()

# ── 2. Do the same for Jobtrack WITHOUT MRR (the template we need to fill) ──
print("\n" + "═" * 80)
print("2. JOBTRACK WITHOUT MRR — Template2 (Nov 2025)")
print("═" * 80)

jt2n_wb = openpyxl.load_workbook(f"{T2}/Jobtrack Without MRR.xlsx", data_only=True)
jt2n_ws = jt2n_wb.active

print("--- Header Row 4 for Chemical columns ---")
for col in range(85, 115):
    val = jt2n_ws.cell(row=4, column=col).value
    letter = openpyxl.utils.get_column_letter(col)
    if val:
        print(f"  {letter} ({col}): {val}")

# Check LAM row data in Without MRR to see what's pre-filled
print("\n--- Sample LAM rows (first 10 with adhesive name) ---")
count = 0
for row in range(5, jt2n_ws.max_row + 1):
    process = jt2n_ws.cell(row=row, column=6).value
    if process and str(process).strip().upper() == 'LAM':
        adh_name = jt2n_ws.cell(row=row, column=91).value
        if adh_name:
            count += 1
            if count <= 10:
                print(f"  Row {row}:")
                for col in range(91, 103):
                    val = jt2n_ws.cell(row=row, column=col).value
                    letter = openpyxl.utils.get_column_letter(col)
                    print(f"    {letter}({col}): {val}")

jt2n_wb.close()

# ── 3. Check the Stores Recordings for adhesive/hardener/solvent materials ──
print("\n" + "═" * 80)
print("3. STORES RECORDINGS — Template2 (Nov)")
print("═" * 80)

stores2 = pd.read_excel(f"{T2}/Stores Recordings.xlsx", sheet_name=0, header=2)
print(f"Stores columns: {list(stores2.columns)}")
print(f"Total rows: {len(stores2)}")

# Find category column
cat_col = None
mat_col = None
for c in stores2.columns:
    cl = str(c).strip().lower()
    if 'categ' in cl:
        cat_col = c
    if 'material' in cl and 'sub' not in cl:
        mat_col = c

print(f"\nCategory column: {cat_col}")
print(f"Material column: {mat_col}")

if cat_col:
    print("\n--- Unique categories ---")
    cats = stores2[cat_col].dropna().astype(str).str.strip().str.upper().unique()
    for c in sorted(cats):
        count = len(stores2[stores2[cat_col].astype(str).str.strip().str.upper() == c])
        print(f"  {c}: {count} rows")

    # Filter for adhesive/hardener/solvent
    for cat_kw in ['ADHESIVE', 'HARDENER', 'SOLVENT']:
        cat_rows = stores2[stores2[cat_col].astype(str).str.strip().str.upper().str.contains(cat_kw)]
        print(f"\n--- {cat_kw} entries in Stores ({len(cat_rows)} rows) ---")
        if mat_col and len(cat_rows) > 0:
            unique_mats = cat_rows[mat_col].dropna().astype(str).str.strip().unique()
            print(f"  Unique materials: {sorted(unique_mats)}")
            # Show sample
            sample = cat_rows.head(5)
            for _, r in sample.iterrows():
                print(f"  Material={r.get(mat_col, '?')}, Category={r.get(cat_col, '?')}")
                # print all non-null columns
                for col2 in stores2.columns:
                    v = r.get(col2)
                    if v is not None and not (isinstance(v, float) and pd.isna(v)):
                        print(f"    {col2}: {v}")

# ── 4. Check Purchase Register for adhesive/hardener/solvent ──
print("\n" + "═" * 80)
print("4. PURCHASE REGISTER — Template2 (Nov)")
print("═" * 80)

pr2 = pd.read_excel(f"{T2}/Purchase Register - 2021 - 2025 _Nov.xlsx", sheet_name=0, header=2)
print(f"PR columns: {list(pr2.columns)}")
print(f"Total rows: {len(pr2)}")

pr_cat_col = None
pr_mat_col = None
for c in pr2.columns:
    cl = str(c).strip().lower()
    if 'categ' in cl:
        pr_cat_col = c
    if 'material' in cl and 'sub' not in cl:
        pr_mat_col = c

if pr_cat_col:
    print("\n--- Unique categories in PR ---")
    cats = pr2[pr_cat_col].dropna().astype(str).str.strip().str.upper().unique()
    for c in sorted(cats):
        count = len(pr2[pr2[pr_cat_col].astype(str).str.strip().str.upper() == c])
        print(f"  {c}: {count} rows")

    for cat_kw in ['ADHESIVE', 'HARDENER', 'SOLVENT']:
        cat_rows = pr2[pr2[pr_cat_col].astype(str).str.strip().str.upper().str.contains(cat_kw)]
        print(f"\n--- {cat_kw} in PR ({len(cat_rows)} rows) ---")
        if pr_mat_col and len(cat_rows) > 0:
            unique_mats = cat_rows[pr_mat_col].dropna().astype(str).str.strip().unique()
            print(f"  Unique materials: {sorted(unique_mats)}")

# ── 5. Now check the column DA in the Jobtrack ──
# DA = column 105 (1-based). Let's check if hardener name is there
print("\n" + "═" * 80)
print("5. CHECKING 'DA' COLUMN (col 105) IN JOBTRACK WITH MRR")
print("═" * 80)

jt2_wb2 = openpyxl.load_workbook(f"{T2}/Jobtrack With MRR.xlsx", data_only=True)
jt2_ws2 = jt2_wb2.active

# Check exact column DA
da_col_idx = None
for col in range(1, 200):
    letter = openpyxl.utils.get_column_letter(col)
    if letter == 'DA':
        da_col_idx = col
        break

print(f"DA column index: {da_col_idx}")
if da_col_idx:
    # Print header
    for hdr_row in [2, 3, 4]:
        val = jt2_ws2.cell(row=hdr_row, column=da_col_idx).value
        print(f"  Header row {hdr_row}: {val}")
    
    # Print data for LAM rows
    print("\n--- DA column values for LAM rows (first 20) ---")
    count = 0
    for row in range(5, jt2_ws2.max_row + 1):
        process = jt2_ws2.cell(row=row, column=6).value
        if process and str(process).strip().upper() == 'LAM':
            da_val = jt2_ws2.cell(row=row, column=da_col_idx).value
            if da_val:
                count += 1
                if count <= 20:
                    adh_name = jt2_ws2.cell(row=row, column=91).value
                    print(f"  Row {row}: DA={da_val}, Adhesive(CM)={adh_name}")
    print(f"  Total LAM rows with DA value: {count}")

    # Collect all unique DA values  
    da_vals = set()
    for row in range(5, jt2_ws2.max_row + 1):
        process = jt2_ws2.cell(row=row, column=6).value
        if process and str(process).strip().upper() == 'LAM':
            da_val = jt2_ws2.cell(row=row, column=da_col_idx).value
            if da_val:
                da_vals.add(str(da_val).strip())
    print(f"\n  Unique DA values: {sorted(da_vals)}")

# Also check what other columns between 91-DA have
print("\n--- All columns from CM(91) to beyond DA ---")
for col in range(91, da_col_idx + 5 if da_col_idx else 115):
    letter = openpyxl.utils.get_column_letter(col)
    h2 = jt2_ws2.cell(row=2, column=col).value
    h3 = jt2_ws2.cell(row=3, column=col).value
    h4 = jt2_ws2.cell(row=4, column=col).value
    # Get first non-empty data value
    sample_val = None
    for r in range(5, min(50, jt2_ws2.max_row + 1)):
        process = jt2_ws2.cell(row=r, column=6).value
        if process and str(process).strip().upper() == 'LAM':
            v = jt2_ws2.cell(row=r, column=col).value
            if v is not None:
                sample_val = v
                break
    print(f"  {letter}({col}): h2={h2}, h3={h3}, h4={h4}, sample={sample_val}")

jt2_wb2.close()

print("\n" + "═" * 80)
print("INVESTIGATION COMPLETE")
print("═" * 80)
