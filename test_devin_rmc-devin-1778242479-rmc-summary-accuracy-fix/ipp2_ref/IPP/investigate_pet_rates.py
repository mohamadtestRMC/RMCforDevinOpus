"""
Investigate PET MRR rate selection: which specific MRR rate does the GT pick?
"""
import sys, shutil, tempfile, os
sys.path.insert(0, '.')
import openpyxl
from engine.fill_jobtrack import COLS, DATA_START_ROW
from engine.mrr_lookup import load_stores_recordings, lookup_mrr_with_qty
from engine.rate_lookup import load_purchase_register, lookup_film_rate_weighted, _find_col

def open_wb(path):
    try: return openpyxl.load_workbook(path, data_only=True)
    except PermissionError:
        tmp = tempfile.mktemp(suffix='.xlsx')
        shutil.copy2(path, tmp)
        wb = openpyxl.load_workbook(tmp, data_only=True)
        os.unlink(tmp)
        return wb

T1 = "Template_Files"
stores = load_stores_recordings(f"{T1}/Stores Recordings.xlsx")
pr = load_purchase_register(f"{T1}/Purchase Register - 2021 - 2026 _Feb 26.xlsx")
gt_wb = open_wb(f"{T1}/Jobtrack Feb With MRR.xlsx")
gt_ws = gt_wb.active

# PR columns
mrr_col = 'Tracking N o.'
rate_col = 'Rate'
mat_col = 'Material'
qty_col = 'Actual Quantity'
amt_col = 'Amount'
mic_col = 'Mic'
size_col = 'Size'
month_col = 'month'

print(f"PR: {len(pr)} rows")

def get_mrr_rate_from_pr(mrr_num):
    """Look up a specific MRR's rate and details from PR."""
    results = []
    for _, row in pr.iterrows():
        m = row.get(mrr_col)
        if m is not None:
            try:
                if int(float(m)) == int(float(mrr_num)):
                    r = float(row.get(rate_col, 0)) if row.get(rate_col) is not None else 0
                    q = float(row.get(qty_col, 0)) if row.get(qty_col) is not None else 0
                    a = float(row.get(amt_col, 0)) if row.get(amt_col) is not None else 0
                    mat = str(row.get(mat_col, '')).strip()
                    results.append({'mrr': int(float(m)), 'rate': r, 'qty': q, 'amt': a, 'mat': mat})
            except: pass
    return results

def sf(v):
    if v is None: return 0.0
    try: return float(v)
    except: return 0.0
def ss(v):
    if v is None: return ''
    return str(v).strip()

# Analyze each mismatch row
for row in [42, 43, 45, 46]:
    uid = ss(gt_ws.cell(row=row, column=COLS['UID']).value)
    order = ss(gt_ws.cell(row=row, column=COLS['Order_No']).value)
    process = ss(gt_ws.cell(row=row, column=COLS['Process']).value).upper()
    
    if process == 'PRINTING':
        mat = ss(gt_ws.cell(row=row, column=COLS['Input_Name']).value)
        mic = sf(gt_ws.cell(row=row, column=COLS['Input_Mic']).value)
        size = sf(gt_ws.cell(row=row, column=COLS['Input_Size']).value)
        gt_mr = ss(gt_ws.cell(row=row, column=COLS['Film_MR']).value)
        gt_rate = sf(gt_ws.cell(row=row, column=COLS['Film_Rate']).value)
        col_type = 'Film'
    else:
        mat = ss(gt_ws.cell(row=row, column=COLS['Fresh1_Name']).value)
        mic = sf(gt_ws.cell(row=row, column=COLS['Fresh1_Mic']).value)
        size = sf(gt_ws.cell(row=row, column=COLS['Fresh1_Size']).value)
        gt_mr = ss(gt_ws.cell(row=row, column=COLS['Fresh1_MR']).value)
        gt_rate = sf(gt_ws.cell(row=row, column=COLS['Fresh1_Rate']).value)
        col_type = 'Fresh1'
    
    print(f"\n{'='*100}")
    print(f"Row {row} | {col_type} | {mat} {mic}mic {size}w | Order={order} | GT MR#={gt_mr} | GT Rate={gt_rate:.4f}")
    print(f"{'='*100}")
    
    # Get all MRRs from Stores
    mrr_qty = lookup_mrr_with_qty(stores, mat, mic, size, order, 'LAMINATION' if process=='LAM' else 'PRINTING')
    if not mrr_qty:
        mrr_qty = lookup_mrr_with_qty(stores, mat, mic, None, order, 'LAMINATION' if process=='LAM' else 'PRINTING')
    
    # Look up each MRR's rate from PR
    total_q = 0
    total_v = 0
    print(f"\n  {'MRR':>8} {'Stores Qty':>12} {'PR Rate':>10} {'PR Qty':>10} {'PR Amt':>12} {'Material'}")
    print(f"  {'-'*8} {'-'*12} {'-'*10} {'-'*10} {'-'*12} {'-'*30}")
    
    mrr_rates = {}
    for mrr, sq in sorted(mrr_qty.items(), key=lambda x: x[1], reverse=True):
        pr_entries = get_mrr_rate_from_pr(mrr)
        if pr_entries:
            for e in pr_entries:
                print(f"  {mrr:>8} {sq:>12.1f} {e['rate']:>10.4f} {e['qty']:>10.1f} {e['amt']:>12.2f} {e['mat'][:30]}")
                mrr_rates[mrr] = e['rate']
                total_q += e['qty']
                total_v += e['amt']
        else:
            print(f"  {mrr:>8} {sq:>12.1f} {'N/A':>10} {'N/A':>10} {'N/A':>12}")
    
    if total_q > 0:
        wavg = total_v / total_q
        print(f"\n  Weighted avg (PR qty): {wavg:.4f}")
    
    # Try different selection strategies
    print(f"\n  --- Rate Selection Strategies ---")
    
    # Strategy 1: Weighted avg by Stores qty
    wavg_stores = 0
    tq = sum(mrr_qty.values())
    for mrr, sq in mrr_qty.items():
        if mrr in mrr_rates:
            wavg_stores += mrr_rates[mrr] * sq / tq
    print(f"  1. Weighted avg (Stores qty): {wavg_stores:.4f}  {'✅ MATCH' if abs(wavg_stores - gt_rate) < 0.01 else f'❌ diff={abs(wavg_stores-gt_rate):.4f}'}")
    
    # Strategy 2: Rate from GT's MRR only
    gt_mrrs = [m.strip() for m in gt_mr.replace('/', ',').split(',') if m.strip()]
    if gt_mrrs:
        gt_rates = [mrr_rates.get(int(m), 0) for m in gt_mrrs if m.isdigit()]
        gt_rates_valid = [r for r in gt_rates if r > 0]
        if gt_rates_valid:
            avg_gt_mrrs = sum(gt_rates_valid) / len(gt_rates_valid)
            # Also try qty-weighted for GT MRRs only
            gt_tq = sum(mrr_qty.get(int(m), 0) for m in gt_mrrs if m.isdigit())
            gt_wavg = sum(mrr_rates.get(int(m), 0) * mrr_qty.get(int(m), 0) for m in gt_mrrs if m.isdigit()) / gt_tq if gt_tq > 0 else 0
            print(f"  2. Simple avg of GT MRRs ({gt_mrrs}): {avg_gt_mrrs:.4f}  {'✅ MATCH' if abs(avg_gt_mrrs - gt_rate) < 0.01 else f'❌ diff={abs(avg_gt_mrrs-gt_rate):.4f}'}")
            print(f"  3. Weighted avg of GT MRRs only: {gt_wavg:.4f}  {'✅ MATCH' if abs(gt_wavg - gt_rate) < 0.01 else f'❌ diff={abs(gt_wavg-gt_rate):.4f}'}")
    
    # Strategy 3: Largest MRR rate
    if mrr_qty and mrr_rates:
        largest_mrr = max(mrr_qty, key=mrr_qty.get)
        largest_rate = mrr_rates.get(largest_mrr, 0)
        print(f"  4. Largest MRR ({largest_mrr}) rate: {largest_rate:.4f}  {'✅ MATCH' if abs(largest_rate - gt_rate) < 0.01 else f'❌ diff={abs(largest_rate-gt_rate):.4f}'}")
    
    # Strategy 4: Each individual MRR rate
    print(f"\n  --- Individual MRR rates ---")
    for mrr in sorted(mrr_rates.keys()):
        r = mrr_rates[mrr]
        match = '✅ MATCH' if abs(r - gt_rate) < 0.01 else ''
        print(f"    MRR {mrr}: rate={r:.4f} {match}")

    # Strategy 5: PR amt/qty for this specific month
    print(f"\n  GT Rate = {gt_rate:.4f}")

gt_wb.close()
