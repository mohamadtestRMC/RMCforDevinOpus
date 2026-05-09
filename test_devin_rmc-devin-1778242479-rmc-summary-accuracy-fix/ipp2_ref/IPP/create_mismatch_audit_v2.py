"""
Mismatch Audit v2 - With full calculation breakdown
Shows: MRRs found, qty per MRR, rate per MRR, weighted avg calc, etc.
"""
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io, sys, shutil, tempfile, os
sys.path.insert(0, '.')

from engine.fill_jobtrack import fill_jobtrack, COLS, DATA_START_ROW
from engine.mrr_lookup import load_stores_recordings, lookup_mrr_with_qty
from engine.rate_lookup import (
    load_purchase_register, _find_col, _get_rate_for_mrr
)

def safe_open_wb(path, data_only=True):
    try: return openpyxl.load_workbook(path, data_only=data_only)
    except PermissionError:
        tmp = tempfile.mktemp(suffix='.xlsx'); shutil.copy2(path, tmp)
        wb = openpyxl.load_workbook(tmp, data_only=data_only); os.unlink(tmp); return wb

def sf(val):
    if val is None or val == '' or (isinstance(val, float) and pd.isna(val)): return 0.0
    if isinstance(val, str) and val.startswith('='): return 0.0
    try: return float(val)
    except: return 0.0

def ss(val):
    if val is None or (isinstance(val, float) and pd.isna(val)): return ''
    s = str(val).strip()
    return '' if s.startswith('=') else s

COMPARE_COLS = {
    'Film_MR': 54, 'Film_Rate': 55, 'Film_Value': 56,
    'Fresh1_MR': 78, 'Fresh1_Rate': 79, 'Fresh1_Value': 80,
    'Fresh2_MR': 88, 'Fresh2_Rate': 89, 'Fresh2_Value': 90,
    'Adh_Rate': 93, 'Adh_Value': 94, 'Hard_Rate': 97, 'Hard_Value': 98,
    'Sol_Rate': 101, 'Sol_Value': 102,
}

def get_pr_details_for_mrr(pr_df, mrr_num, material=None, size=None, mic=None):
    """Get ALL PR rows for a given MRR with their details."""
    tracking_col = _find_col(pr_df, 'tracking')
    material_col = _find_col(pr_df, 'material')
    size_col = _find_col(pr_df, 'size')
    mic_col = _find_col(pr_df, 'mic')
    rate_col = [c for c in pr_df.columns if str(c).strip().lower() == 'rate']
    rate_col = rate_col[0] if rate_col else 'Rate'
    amt_col = None; qty_col = None
    for c in pr_df.columns:
        cl = str(c).strip().lower()
        if cl == 'amount': amt_col = c
        if cl == 'actual quantity': qty_col = c
    
    if not tracking_col: return []
    mask = pd.to_numeric(pr_df[tracking_col], errors='coerce') == mrr_num
    rows = pr_df[mask]
    details = []
    for _, r in rows.iterrows():
        details.append({
            'MRR': mrr_num,
            'PR_Material': ss(r.get(material_col, '')),
            'PR_Size': sf(r.get(size_col, 0)) if size_col else '',
            'PR_Mic': sf(r.get(mic_col, 0)) if mic_col else '',
            'PR_Rate': sf(r.get(rate_col, 0)),
            'PR_Amount': sf(r.get(amt_col, 0)) if amt_col else '',
            'PR_Qty': sf(r.get(qty_col, 0)) if qty_col else '',
        })
    return details

def collect_mismatches(dataset_name, jt_without, jt_with, stores_path, pr_path,
                       granules=None, megapack=None):
    print(f"Processing {dataset_name}...")
    with open(jt_without, 'rb') as f: jt_bytes = io.BytesIO(f.read())
    g_io = m_io = None
    if granules:
        with open(granules, 'rb') as f: g_io = io.BytesIO(f.read())
    if megapack:
        with open(megapack, 'rb') as f: m_io = io.BytesIO(f.read())
    
    filled_bytes, _, stats = fill_jobtrack(jt_bytes, stores_path, pr_path,
                                            granules_file=g_io, megapack_file=m_io)
    filled_wb = openpyxl.load_workbook(filled_bytes, data_only=False)
    filled_ws = filled_wb.active
    gt_wb = safe_open_wb(jt_with, data_only=True)
    gt_ws = gt_wb.active
    max_row = min(filled_ws.max_row, gt_ws.max_row)
    
    stores_df = load_stores_recordings(stores_path)
    pr_df = load_purchase_register(pr_path)
    
    mismatches = []
    calc_details = []  # Detailed calculation breakdown
    
    for row in range(DATA_START_ROW, max_row + 1):
        process = ss(gt_ws.cell(row=row, column=COLS['Process']).value).upper()
        uid = ss(gt_ws.cell(row=row, column=COLS['UID']).value)
        order = ss(gt_ws.cell(row=row, column=COLS['Order_No']).value)
        if not process or not uid: continue
        
        for col_name, col_idx in COMPARE_COLS.items():
            is_film = col_name.startswith('Film_')
            is_fresh = col_name.startswith('Fresh')
            is_chem = col_name.startswith(('Adh_', 'Hard_', 'Sol_'))
            if process == 'PRINTING' and (is_fresh or is_chem): continue
            if process == 'LAM' and is_film: continue
            if process not in ('PRINTING', 'LAM'): continue
            
            gt_val = gt_ws.cell(row=row, column=col_idx).value
            eng_val = filled_ws.cell(row=row, column=col_idx).value
            gt_f = sf(gt_val); eng_f = sf(eng_val)
            gt_s = ss(gt_val); eng_s = ss(eng_val)
            
            is_mm = False; mtype = ''
            if col_name.endswith('_MR'):
                gt_set = {s.strip() for s in gt_s.replace('/',',').split(',') if s.strip()} if gt_s else set()
                eng_set = {s.strip() for s in eng_s.replace('/',',').split(',') if s.strip()} if eng_s else set()
                if gt_set and not eng_set: is_mm = True; mtype = 'MISSING'
                elif gt_set and eng_set and gt_set != eng_set and not (gt_set & eng_set):
                    is_mm = True; mtype = 'DIFFERENT MR#'
            else:
                if gt_f == 0 and eng_f == 0: continue
                elif gt_f > 0 and eng_f == 0: is_mm = True; mtype = 'MISSING'
                elif abs(gt_f - eng_f) < 0.02: continue
                elif gt_f > 0 and abs(gt_f - eng_f)/gt_f < 0.005: continue
                else:
                    pct = abs(gt_f - eng_f)/gt_f * 100 if gt_f > 0 else 999
                    is_mm = True; mtype = f'DIFF {pct:.1f}%'
            
            if not is_mm: continue
            
            # Get material context
            mat = mic = size = ''
            if 'Film' in col_name:
                mat = ss(gt_ws.cell(row=row, column=COLS['Input_Name']).value)
                mic = sf(gt_ws.cell(row=row, column=COLS['Input_Mic']).value)
                size = sf(gt_ws.cell(row=row, column=COLS['Input_Size']).value)
            elif 'Fresh1' in col_name:
                mat = ss(gt_ws.cell(row=row, column=COLS['Fresh1_Name']).value)
                mic = sf(gt_ws.cell(row=row, column=COLS['Fresh1_Mic']).value)
                size = sf(gt_ws.cell(row=row, column=COLS['Fresh1_Size']).value)
            elif 'Fresh2' in col_name:
                mat = ss(gt_ws.cell(row=row, column=COLS['Fresh2_Name']).value)
                mic = sf(gt_ws.cell(row=row, column=COLS['Fresh2_Mic']).value)
                size = sf(gt_ws.cell(row=row, column=COLS['Fresh2_Size']).value)
            
            # Get total qty used for Value calc
            total_qty = 0
            if 'Film' in col_name:
                q1 = sf(gt_ws.cell(row=row, column=COLS['Input_Qty']).value)
                q2 = sf(gt_ws.cell(row=row, column=COLS['Balance_Qty']).value)
                total_qty = q1 + q2
            elif 'Fresh1' in col_name:
                q1 = sf(gt_ws.cell(row=row, column=COLS['Fresh1_Qty']).value)
                q2 = sf(gt_ws.cell(row=row, column=COLS['Fresh1_Balance']).value)
                total_qty = q1 + q2
            elif 'Fresh2' in col_name:
                q1 = sf(gt_ws.cell(row=row, column=COLS['Fresh2_Qty']).value)
                q2 = sf(gt_ws.cell(row=row, column=COLS['Fresh2_Balance']).value)
                total_qty = q1 + q2
            
            # === DETAILED CALC BREAKDOWN ===
            mrr_qty = {}
            if mat:
                mrr_qty = lookup_mrr_with_qty(stores_df, mat, mic if mic else None,
                                               size if size else None, order, process)
                if not mrr_qty:
                    mrr_qty = lookup_mrr_with_qty(stores_df, mat, mic if mic else None, None, order, process)
                if not mrr_qty:
                    mrr_qty = lookup_mrr_with_qty(stores_df, mat, mic if mic else None, None, order)
            
            # For each MRR found in stores, get its PR rate details
            mrr_breakdown = []
            weighted_sum = 0; total_stores_qty = 0
            for mrr_num, s_qty in sorted(mrr_qty.items(), key=lambda x: -x[1]):
                pr_rows = get_pr_details_for_mrr(pr_df, mrr_num, mat, size, mic)
                pr_rate = pr_rows[0]['PR_Rate'] if pr_rows else 0
                pr_mat = pr_rows[0]['PR_Material'] if pr_rows else 'NOT FOUND'
                pr_size = pr_rows[0]['PR_Size'] if pr_rows else ''
                weighted_sum += pr_rate * s_qty
                total_stores_qty += s_qty
                mrr_breakdown.append({
                    'MRR': mrr_num, 'Stores_Qty': s_qty,
                    'PR_Rate': pr_rate, 'PR_Material': pr_mat, 'PR_Size': pr_size
                })
            
            calc_rate = weighted_sum / total_stores_qty if total_stores_qty > 0 else 0
            
            # Build readable strings
            mrr_list_str = ', '.join([str(m['MRR']) for m in mrr_breakdown]) if mrr_breakdown else 'NONE FOUND'
            stores_qty_str = ', '.join([f"{m['MRR']}={m['Stores_Qty']:.1f}" for m in mrr_breakdown])
            pr_rates_str = ', '.join([f"{m['MRR']}={m['PR_Rate']:.4f}" for m in mrr_breakdown])
            calc_formula = ''
            if len(mrr_breakdown) > 1:
                parts = [f"({m['PR_Rate']:.4f} x {m['Stores_Qty']:.1f})" for m in mrr_breakdown]
                calc_formula = f"({' + '.join(parts)}) / {total_stores_qty:.1f} = {calc_rate:.4f}"
            elif len(mrr_breakdown) == 1:
                calc_formula = f"Single MRR rate = {calc_rate:.4f}"
            
            mismatches.append({
                'Dataset': dataset_name,
                'Row': row,
                'UID': uid,
                'Process': process,
                'Order No': order,
                'Column': col_name,
                'Material': mat,
                'Mic': mic,
                'Size': size,
                'Total Qty (Jobtrack)': round(total_qty, 2),
                'Manual Value (With MRR)': gt_s if col_name.endswith('_MR') else round(gt_f, 4),
                'Calculated Value (Engine)': eng_s if col_name.endswith('_MR') else round(eng_f, 4),
                'Difference': '' if col_name.endswith('_MR') else round(eng_f - gt_f, 4),
                'Mismatch Type': mtype,
                'MRRs Found (Stores)': mrr_list_str,
                'Stores Qty per MRR': stores_qty_str,
                'PR Rate per MRR': pr_rates_str,
                'Weighted Avg Calculation': calc_formula,
            })
            
            # Add per-MRR detail rows for the calc sheet
            for m in mrr_breakdown:
                calc_details.append({
                    'Dataset': dataset_name, 'Row': row, 'UID': uid,
                    'Material': mat, 'Mic': mic, 'Size': size,
                    'MRR Number': m['MRR'],
                    'Stores Issue Qty': m['Stores_Qty'],
                    'PR Material Name': m['PR_Material'],
                    'PR Size': m['PR_Size'],
                    'PR Rate': m['PR_Rate'],
                    'Rate x Qty': round(m['PR_Rate'] * m['Stores_Qty'], 4),
                    'Total Stores Qty': total_stores_qty,
                    'Weighted Avg Rate': round(calc_rate, 4),
                    'Manual Rate (GT)': gt_f if 'Rate' in col_name else '',
                    'Match?': 'YES' if abs(calc_rate - gt_f) < 0.02 else 'NO',
                })
    
    gt_wb.close(); filled_wb.close()
    return mismatches, calc_details, stats


def style_header(ws, row, max_col, color='1F4E79'):
    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    font = Font(bold=True, color='FFFFFF', size=10)
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill; cell.font = font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

def auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0; col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try: max_len = max(max_len, len(str(cell.value or '')))
            except: pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 55)


# ══════════════════════════════════════════════════════════════
T1, T2 = "Template_Files", "Template2"

mm1, cd1, st1 = collect_mismatches("February 2026",
    f"{T1}/Jobtrack Feb Without MRR.xlsx", f"{T1}/Jobtrack Feb With MRR.xlsx",
    f"{T1}/Stores Recordings.xlsx", f"{T1}/Purchase Register - 2021 - 2026 _Feb 26.xlsx",
    granules=f"{T1}/Granules Recipe - February 2026.xlsx", megapack=f"{T1}/MEGA PACK.xlsx")

mm2, cd2, st2 = collect_mismatches("November 2025",
    f"{T2}/Jobtrack Without MRR.xlsx", f"{T2}/Jobtrack With MRR.xlsx",
    f"{T2}/Stores Recordings.xlsx", f"{T2}/Purchase Register - 2021 - 2025 _Nov.xlsx",
    granules=f"{T2}/Granules Recipe -Nov_2025.xlsx")

all_mm = mm1 + mm2
all_cd = cd1 + cd2

# ── Build Excel ──
wb = openpyxl.Workbook()
red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
yellow = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

# === Sheet 1: All Mismatches with Calc Breakdown ===
ws1 = wb.active; ws1.title = "Mismatches Detail"
h1 = list(all_mm[0].keys()) if all_mm else []
ws1.append(h1); style_header(ws1, 1, len(h1))
for mm in all_mm:
    r = ws1.max_row + 1
    ws1.append([mm[h] for h in h1])
    fill = red if 'MISSING' in str(mm['Mismatch Type']) else yellow
    for c in range(1, len(h1)+1): ws1.cell(r, c).fill = fill
auto_width(ws1)

# === Sheet 2: Per-MRR Calculation Breakdown ===
ws2 = wb.create_sheet("Calculation Breakdown")
h2 = list(all_cd[0].keys()) if all_cd else []
ws2.append(h2); style_header(ws2, 1, len(h2))
for cd in all_cd:
    r = ws2.max_row + 1
    ws2.append([cd[h] for h in h2])
    fill = green if cd.get('Match?') == 'YES' else red
    for c in range(1, len(h2)+1): ws2.cell(r, c).fill = fill
auto_width(ws2)

# === Sheet 3: Root Cause & Questions ===
ws3 = wb.create_sheet("Root Cause & Questions")
h3 = ["Case", "Dataset", "Rows", "Material", "Issue", "Root Cause", "Question / Rule"]
ws3.append(h3); style_header(ws3, 1, len(h3))
cases = [
    [1, "Feb 2026", "42-43", "PET 12mic",
     "GT=4.22/4.35 per row, Engine=4.3036 (single avg)",
     "GT uses SPECIFIC MRR rate per row. Engine averages ALL MRRs.",
     "Q: Should rate be per-MRR or weighted avg? If per-MRR, how to assign MRR to row?"],
    [2, "Feb 2026", "46", "PET 12mic/913mm",
     "GT=4.5880, Engine=4.4604 (2.8% diff)",
     "Same: size-913 MRR has unique rate, engine averages all.",
     "Q: Same as Case 1"],
    [3, "Feb 2026", "54", "TPE 100mic/893mm",
     "GT MRR=84080, Engine=85738/85775",
     "Stores has 85738/85775, GT shows 84080 (manual?)",
     "Q: Was 84080 manually entered? How to handle?"],
    [4, "Nov 2025", "7", "NYLON 15mic/1005mm",
     "Engine finds NOTHING - complete miss",
     "NYLON not in MATERIAL_ALIASES",
     "RULE: Add NYLON aliases. What Stores sub-categories?"],
    [5, "Nov 2025", "17", "TPE 50mic/996mm",
     "GT=4.1951, Engine=4.2968 (2.4%)",
     "Same weighted-avg vs per-MRR pattern",
     "Q: Same as Case 1"],
    [6, "Nov 2025", "48", "TPE 100mic/893mm",
     "GT=4.8708, Engine=0.9000 (81.5% WRONG!)",
     "Outlier check replaced correct 4.87 with wrong month avg 0.90",
     "FIX: Remove outlier check - it's harmful here"],
]
for case in cases:
    r = ws3.max_row + 1
    ws3.append(case)
    if 'FIX' in str(case[-1]): fill_c = green
    elif 'RULE' in str(case[-1]): fill_c = red
    else: fill_c = yellow
    for c in range(1, len(h3)+1): ws3.cell(r, c).fill = fill_c
auto_width(ws3)

# === Sheet 4: How We Calculate ===
ws4 = wb.create_sheet("How We Calculate")
h4 = ["Step", "What", "Source", "Detail"]
ws4.append(h4); style_header(ws4, 1, 4)
for s in [
    [1, "Read Jobtrack Row", "Jobtrack Without MRR", "UID, Process, Order, Material, Size, Mic, Qty + Balance"],
    [2, "Find MRRs in Stores", "Stores Recordings", "Match: Material (with aliases) + Mic + Order No + Process -> {MRR: issue_qty}"],
    [3, "Get Rate per MRR from PR", "Purchase Register", "For each MRR: match Tracking No -> get Rate column value"],
    [4, "Weighted Average Rate", "Calculation", "Sum(Rate_i * StoresQty_i) / Sum(StoresQty_i) for all MRRs"],
    [5, "Check Supplier Override", "Granules/MEGA PACK", "BANDERA/CYM -> Granules Recipe rate. MEGA PACK -> monthly rate"],
    [6, "Outlier Check (PROBLEM)", "fill_jobtrack.py", "If rate >50% diff from month avg -> replace. THIS IS CASE 6 BUG"],
    [7, "Calc Value", "Calculation", "Value = (Qty + Balance) * Rate"],
    [8, "Chemical Rates", "Purchase Register", "ADH/Hardener/Solvent: Total Amount / Total Actual Qty for the month"],
]: ws4.append(s)
auto_width(ws4)

out = "Mismatch_Audit_v2.xlsx"
wb.save(out)
print(f"Saved: {out} | Feb={len(mm1)} Nov={len(mm2)} Total={len(all_mm)} mismatches")
print(f"Calc breakdown rows: {len(all_cd)}")
