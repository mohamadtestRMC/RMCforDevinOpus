"""
Root cause analysis for each failure.
"""
import pandas as pd
import openpyxl
from engine.mrr_lookup import load_stores_recordings, lookup_mrr_with_qty, _find_stores_columns
from engine.rate_lookup import load_purchase_register, lookup_film_rate_weighted, filter_mrr_by_pr, _find_col

BASE = "Template_Files"
stores_df = load_stores_recordings(f"{BASE}/Stores Recordings.xlsx")
pr_df = load_purchase_register(f"{BASE}/Purchase Register - 2021 - 2026 _Feb 26.xlsx")
wb = openpyxl.load_workbook(f"{BASE}/Jobtrack Feb With MRR.xlsx", data_only=True)
ws = wb.active

print("=" * 80)
print("FAILURE 1: Row 9 — WPE/INH material, no rate found")
print("=" * 80)
# Row 9: WPE material, expected INH with rate 4.699...
# Our engine sets MR=INH but rate=None after removing TPE fallback
# The ground truth rate is 4.6992526305327065 — what material is this rate from?
print(f"Input name: {ws.cell(row=9, column=47).value}")
print(f"Order: {ws.cell(row=9, column=11).value}")
print(f"Expected MR: {ws.cell(row=9, column=54).value}")
print(f"Expected Rate: {ws.cell(row=9, column=55).value}")

# Check what MRRs exist for WPE in this order
from engine.mrr_lookup import lookup_mrr
order = str(ws.cell(row=9, column=11).value).strip()
mic = ws.cell(row=9, column=49).value
mrr_list = lookup_mrr(stores_df, 'WPE', mic, None, order, 'PRINTING')
print(f"MRRs for WPE/{order}: {mrr_list}")
if not mrr_list:
    mrr_list = lookup_mrr(stores_df, 'WPE', mic, None, order)
    print(f"MRRs for WPE/{order} (no process filter): {mrr_list}")

# Check PR for these MRRs
tracking_col = _find_col(pr_df, 'tracking')
material_col = _find_col(pr_df, 'material')
rate_col = [c for c in pr_df.columns if str(c).strip().lower() == 'rate'][0]

# What materials exist for these MRRs in PR?
if mrr_list:
    for m in mrr_list:
        mask = pd.to_numeric(pr_df[tracking_col], errors='coerce') == m
        entries = pr_df[mask]
        if not entries.empty:
            for _, row_data in entries.iterrows():
                print(f"  MRR {m}: Material={row_data[material_col]}, Rate={row_data[rate_col]}, Size={row_data.get('Size')}")
        else:
            print(f"  MRR {m}: NOT IN PR")

# Also search PR for all WPE-like entries
print("\nAll WPE/PE WHITE entries in PR:")
for _, row_data in pr_df.iterrows():
    mat = str(row_data[material_col]).strip().upper() if pd.notna(row_data[material_col]) else ''
    if mat in ('WPE', 'PE WHITE', 'WLDPE', 'TPE'):
        print(f"  Tracking={row_data[tracking_col]}, Material={mat}, Rate={row_data[rate_col]}")

print("\n" + "=" * 80)
print("FAILURE 2: Row 54 — Fresh2 rate 0.9 vs expected 4.78")
print("=" * 80)
print(f"Fresh2 name: {ws.cell(row=54, column=81).value}")
print(f"Fresh2 mic: {ws.cell(row=54, column=83).value}")
print(f"Fresh2 size: {ws.cell(row=54, column=82).value}")
print(f"Order: {ws.cell(row=54, column=11).value}")
f2_name = str(ws.cell(row=54, column=81).value).strip()
f2_mic = ws.cell(row=54, column=83).value
f2_size = ws.cell(row=54, column=82).value
f2_order = str(ws.cell(row=54, column=11).value).strip()

mrr_qty = lookup_mrr_with_qty(stores_df, f2_name, f2_mic, f2_size, f2_order, 'LAMINATION')
print(f"MRR lookup (with size): {mrr_qty}")
if not mrr_qty:
    mrr_qty = lookup_mrr_with_qty(stores_df, f2_name, f2_mic, None, f2_order, 'LAMINATION')
    print(f"MRR lookup (no size): {mrr_qty}")
if not mrr_qty:
    mrr_qty = lookup_mrr_with_qty(stores_df, f2_name, f2_mic, None, f2_order)
    print(f"MRR lookup (no process): {mrr_qty}")

if mrr_qty:
    filtered = filter_mrr_by_pr(pr_df, mrr_qty, f2_name, f2_size, f2_mic)
    print(f"After PR filter: {filtered}")
    rate = lookup_film_rate_weighted(pr_df, filtered, f2_name, f2_size, f2_mic)
    print(f"Weighted rate with size: {rate}")
    rate2 = lookup_film_rate_weighted(pr_df, filtered, f2_name, None, f2_mic)
    print(f"Weighted rate no size: {rate2}")
    # Check what's in PR for these MRRs
    for m in mrr_qty:
        mask = pd.to_numeric(pr_df[tracking_col], errors='coerce') == int(float(m))
        entries = pr_df[mask]
        for _, r in entries.iterrows():
            print(f"  MRR {m}: Mat={r[material_col]}, Rate={r[rate_col]}, Size={r.get('Size')}, Mic={r.get('Mic')}")

print("\n" + "=" * 80)
print("FAILURE 3: Row 40-41 — Film rate 4.587 vs 4.576 (0.23% diff)")
print("=" * 80)
print(f"Input name: {ws.cell(row=40, column=47).value}")
print(f"Order: {ws.cell(row=40, column=11).value}")
print(f"Expected MR: {ws.cell(row=40, column=54).value}")
print(f"Expected Rate: {ws.cell(row=40, column=55).value}")
# This is a multi-MRR case — the ground truth has 85526, we have 85526/85573/85587
# The rate difference comes from including extra MRRs in the weighted average
f_name = str(ws.cell(row=40, column=47).value).strip()
f_mic = ws.cell(row=40, column=49).value
f_order = str(ws.cell(row=40, column=11).value).strip()
mqty = lookup_mrr_with_qty(stores_df, f_name, f_mic, None, f_order, 'PRINTING')
print(f"All MRRs found: {mqty}")
# Show rates for each
for m, q in mqty.items():
    mask = pd.to_numeric(pr_df[tracking_col], errors='coerce') == int(float(m))
    entries = pr_df[mask]
    for _, r in entries.iterrows():
        mat = str(r[material_col]).strip()
        if 'PET' in mat.upper():
            print(f"  MRR {m}: qty={q}, Mat={mat}, Rate={r[rate_col]}, Size={r.get('Size')}")

wb.close()
