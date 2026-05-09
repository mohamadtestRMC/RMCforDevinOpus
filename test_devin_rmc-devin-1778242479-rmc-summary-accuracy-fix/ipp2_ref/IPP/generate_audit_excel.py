"""
Generate detailed Excel audit report for ALL mismatches across both datasets.
Shows: how engine computed, how GT computed, difference, and questions needed.
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys, io
sys.path.insert(0, '.')

from engine.fill_jobtrack import fill_jobtrack, COLS, DATA_START_ROW, _safe_str, _safe_float
from engine.mrr_lookup import load_stores_recordings, lookup_mrr_with_qty
from engine.rate_lookup import (load_purchase_register, lookup_film_rate_weighted,
                                 filter_mrr_by_pr, _find_col, lookup_adhesive_rate,
                                 lookup_material_rate_for_month)
from engine.supplier_rates import (build_mrr_supplier_map, get_supplier_for_mrrs,
                                    load_granules_rates, load_megapack_rates, lookup_megapack_rate)

def sf(v):
    try: return float(v) if v and not (isinstance(v,float) and pd.isna(v)) else 0.0
    except: return 0.0
def ss(v):
    if v is None or (isinstance(v,float) and pd.isna(v)): return ''
    s = str(v).strip()
    return '' if s.startswith('=') else s

COMPARE_COLS = {
    'Film_MR': 54, 'Film_Rate': 55, 'Film_Value': 56,
    'Fresh1_MR': 78, 'Fresh1_Rate': 79, 'Fresh1_Value': 80,
    'Fresh2_MR': 88, 'Fresh2_Rate': 89, 'Fresh2_Value': 90,
    'Adh_Rate': 93, 'Adh_Value': 94,
    'Hard_Rate': 97, 'Hard_Value': 98,
    'Sol_Rate': 101, 'Sol_Value': 102,
}

COL_F1 = {'name':71,'size':72,'mic':73}
COL_F2 = {'name':81,'size':82,'mic':83}
COL_FILM = {'name':47,'size':48,'mic':49}

rows_data = []

for ds_name, jt_without, jt_with, stores_path, pr_path, gran_path, mega_path in [
    ("Feb 2026",
     "Template_Files/Jobtrack Feb Without MRR.xlsx",
     "Template_Files/Jobtrack Feb With MRR.xlsx",
     "Template_Files/Stores Recordings.xlsx",
     "Template_Files/Purchase Register - 2021 - 2026 _Feb 26.xlsx",
     "Template_Files/Granules Recipe - February 2026.xlsx",
     "Template_Files/MEGA PACK.xlsx"),
    ("Nov 2025",
     "Template2/Jobtrack Without MRR.xlsx",
     "Template2/Jobtrack With MRR.xlsx",
     "Template2/Stores Recordings.xlsx",
     "Template2/Purchase Register - 2021 - 2025 _Nov.xlsx",
     "Template2/Granules Recipe -Nov_2025.xlsx",
     None),
]:
    print(f"\nProcessing {ds_name}...")

    # Fill the file
    with open(jt_without, 'rb') as f:
        jt_bytes = io.BytesIO(f.read())
    gran_io = None
    mega_io = None
    if gran_path:
        with open(gran_path, 'rb') as f: gran_io = io.BytesIO(f.read())
    if mega_path:
        with open(mega_path, 'rb') as f: mega_io = io.BytesIO(f.read())

    filled_bytes, _, _ = fill_jobtrack(jt_bytes, stores_path, pr_path,
                                        granules_file=gran_io, megapack_file=mega_io)

    # Load filled + GT
    filled_wb = openpyxl.load_workbook(filled_bytes, data_only=False)
    filled_ws = filled_wb.active
    gt_wb = openpyxl.load_workbook(jt_with, data_only=True)
    gt_ws = gt_wb.active

    # Load reference data
    stores = load_stores_recordings(stores_path)
    pr = load_purchase_register(pr_path)
    supplier_map = build_mrr_supplier_map(stores)
    granules = load_granules_rates(gran_path) if gran_path else {}
    megapack = load_megapack_rates(mega_path) if mega_path else {}

    tracking_col = _find_col(pr, 'tracking')
    mat_col_pr = _find_col(pr, 'material')
    rate_col_pr = [c for c in pr.columns if str(c).strip().lower() == 'rate']
    rate_col_pr = rate_col_pr[0] if rate_col_pr else None
    qty_col_pr = [c for c in pr.columns if str(c).strip().lower() == 'actual quantity']
    qty_col_pr = qty_col_pr[0] if qty_col_pr else None
    amt_col_pr = [c for c in pr.columns if str(c).strip().lower() == 'amount']
    amt_col_pr = amt_col_pr[0] if amt_col_pr else None

    max_row = min(filled_ws.max_row, gt_ws.max_row)

    for row in range(DATA_START_ROW, max_row + 1):
        process = ss(gt_ws.cell(row=row, column=COLS['Process']).value).upper()
        uid = ss(gt_ws.cell(row=row, column=COLS['UID']).value)
        order = ss(gt_ws.cell(row=row, column=COLS['Order_No']).value)
        if not process or not uid: continue

        for col_name, col_idx in COMPARE_COLS.items():
            is_film = col_name.startswith('Film_')
            is_fresh = col_name.startswith('Fresh')
            is_chem = col_name.startswith(('Adh_', 'Hard_', 'Sol_'))
            if process == 'PRINTING' and (is_fresh or is_chem): continue
            if process == 'LAM' and is_film: continue
            if process not in ('PRINTING', 'LAM'): continue

            gt_val = gt_ws.cell(row=row, column=col_idx).value
            eng_val = filled_ws.cell(row=row, column=col_idx).value
            gt_f = sf(gt_val); eng_f = sf(eng_val)
            gt_s = ss(gt_val); eng_s = ss(eng_val)

            is_mismatch = False
            if col_name.endswith('_MR'):
                gt_set = {s.strip() for s in gt_s.replace('/',',').split(',') if s.strip()} if gt_s else set()
                eng_set = {s.strip() for s in eng_s.replace('/',',').split(',') if s.strip()} if eng_s else set()
                if gt_set and not eng_set: is_mismatch = True
                elif gt_set and eng_set and gt_s != 'INH' and gt_set != eng_set:
                    if not (gt_set & eng_set): is_mismatch = True
                elif gt_s == 'INH' and eng_s != 'INH': is_mismatch = True
                elif gt_s != 'INH' and eng_s == 'INH': is_mismatch = True
            else:
                if gt_f == 0 and eng_f == 0: pass
                elif gt_f > 0 and eng_f == 0: is_mismatch = True
                elif gt_f > 0 and abs(gt_f - eng_f) >= 0.02 and abs(gt_f - eng_f)/gt_f >= 0.005:
                    is_mismatch = True

            if not is_mismatch: continue

            # ── Build explanation ──
            # Determine material context
            if 'Film' in col_name:
                mat = ss(gt_ws.cell(row=row, column=COL_FILM['name']).value)
                mic = sf(gt_ws.cell(row=row, column=COL_FILM['mic']).value)
                size = sf(gt_ws.cell(row=row, column=COL_FILM['size']).value)
                lookup_proc = 'PRINTING'
            elif 'Fresh1' in col_name:
                mat = ss(gt_ws.cell(row=row, column=COL_F1['name']).value)
                mic = sf(gt_ws.cell(row=row, column=COL_F1['mic']).value)
                size = sf(gt_ws.cell(row=row, column=COL_F1['size']).value)
                lookup_proc = 'LAMINATION'
            elif 'Fresh2' in col_name:
                mat = ss(gt_ws.cell(row=row, column=COL_F2['name']).value)
                mic = sf(gt_ws.cell(row=row, column=COL_F2['mic']).value)
                size = sf(gt_ws.cell(row=row, column=COL_F2['size']).value)
                lookup_proc = 'LAMINATION'
            else:
                mat = col_name.split('_')[0]
                mic = 0; size = 0; lookup_proc = ''

            # Engine calculation breakdown
            eng_explain = ""
            gt_explain = ""
            question = ""
            diff_pct = ""

            if col_name.endswith('_MR'):
                diff_pct = f"GT={gt_s}, ENG={eng_s}"
            elif gt_f > 0:
                diff_abs = eng_f - gt_f
                diff_pct = f"{abs(diff_abs)/gt_f*100:.2f}%"
            elif gt_f > 0 and eng_f == 0:
                diff_pct = "MISS (100%)"

            # Film/Fresh MRR lookup
            if 'Film' in col_name or 'Fresh' in col_name:
                mrr_qty = lookup_mrr_with_qty(stores, mat, mic, size, order, lookup_proc)
                if not mrr_qty:
                    mrr_qty = lookup_mrr_with_qty(stores, mat, mic, None, order, lookup_proc)
                if not mrr_qty:
                    mrr_qty = lookup_mrr_with_qty(stores, mat, mic, None, order)

                supplier = get_supplier_for_mrrs(supplier_map, list(mrr_qty.keys())) if mrr_qty else None

                if mrr_qty:
                    mrr_f = filter_mrr_by_pr(pr, mrr_qty, mat, size, mic)
                    eng_rate = lookup_film_rate_weighted(pr, mrr_f, mat, size, mic)

                    # Build per-MRR breakdown
                    mrr_details = []
                    for m, q in mrr_qty.items():
                        sup = supplier_map.get(m, 'UNKNOWN')
                        # Get PR rate for this MRR
                        try:
                            mv = int(float(m))
                            mask = pd.to_numeric(pr[tracking_col], errors='coerce') == mv
                            pr_rows = pr[mask]
                            if mat_col_pr:
                                pr_rows = pr_rows[pr_rows[mat_col_pr].astype(str).str.upper().str.strip() == mat.upper()]
                            pr_rate = sf(pr_rows[rate_col_pr].iloc[0]) if len(pr_rows) > 0 else 0
                            pr_qty = sf(pr_rows[qty_col_pr].iloc[0]) if len(pr_rows) > 0 else 0
                        except:
                            pr_rate = 0; pr_qty = 0
                        mrr_details.append(f"MR#{m}: StoresQty={q:.1f}, PR_Rate={pr_rate:.4f}, Supplier={sup}")

                    eng_explain = f"Engine found {len(mrr_qty)} MRRs in Stores.\n"
                    eng_explain += "\n".join(mrr_details)
                    eng_explain += f"\nWeighted avg rate = {eng_rate:.4f}"
                    if supplier:
                        eng_explain += f"\nSupplier override: {supplier}"
                        if supplier in ('BANDERA','CYM'):
                            order_up = order.upper()
                            if order_up in granules:
                                eng_explain += f"\nGranules Recipe rate for {order_up} = {granules[order_up]:.4f} -> MR#=INH"
                            else:
                                eng_explain += f"\nOrder {order_up} NOT in Granules Recipe ({list(granules.keys())})"
                        elif supplier == 'MEGA PACK':
                            eng_explain += f"\nMEGA PACK rates available: {list(megapack.keys())}"
                else:
                    eng_explain = f"No MRRs found in Stores for {mat}/{mic}/{size}/{order}/{lookup_proc}"

                # GT explanation
                if gt_s == 'INH':
                    gt_explain = f"GT marked as INH (In-House production). Rate={gt_f:.4f}"
                    if order.upper() in granules:
                        gt_explain += f"\nOrder {order.upper()} IS in Granules Recipe = {granules[order.upper()]:.4f}"
                    else:
                        gt_explain += f"\nOrder {order.upper()} NOT in Granules. Rate appears to be manual."
                elif gt_s and not col_name.endswith('_Value'):
                    gt_mrrs = [s.strip() for s in gt_s.replace('/',',').split(',')]
                    gt_details = []
                    for gm in gt_mrrs:
                        try:
                            gmv = int(float(gm))
                            mask = pd.to_numeric(pr[tracking_col], errors='coerce') == gmv
                            pr_rows = pr[mask]
                            if mat_col_pr:
                                pr_rows = pr_rows[pr_rows[mat_col_pr].astype(str).str.upper().str.strip() == mat.upper()]
                            for _, r in pr_rows.iterrows():
                                gt_details.append(f"MR#{gm}: Rate={sf(r.get(rate_col_pr,0)):.4f}, Qty={sf(r.get(qty_col_pr,0)):.1f}")
                        except: pass
                    gt_explain = f"GT uses MR#={gt_s}\n" + "\n".join(gt_details)
                elif col_name.endswith('_Value'):
                    gt_explain = f"GT Value = Qty x Rate"
                elif gt_f > 0 and eng_f == 0:
                    gt_explain = f"GT has value {gt_f:.4f} but engine produced 0 (MISS)"

                # Questions
                if eng_f == 0 and gt_f > 0:
                    question = f"Why is MRR {gt_s} not found in Stores for {mat}/{order}? Is the data missing or under a different name?"
                elif gt_s == 'INH' and eng_s != 'INH':
                    question = f"Should ALL BANDERA orders use INH, or only those in Granules Recipe? Order {order} is not in Granules but GT says INH."
                elif abs(gt_f - eng_f) > 0 and gt_f > 0:
                    question = f"GT uses specific MR#={gt_s} (rate={gt_f:.4f}). Should the engine pick the same specific MRRs, or use ALL matching MRRs weighted avg ({eng_f:.4f})?"

            rows_data.append({
                'Dataset': ds_name,
                'Row': row,
                'UID': uid,
                'Order': order,
                'Process': process,
                'Column': col_name,
                'Material': mat,
                'Micron': mic,
                'Size': size,
                'GT_Value': gt_s if col_name.endswith('_MR') else gt_f,
                'Engine_Value': eng_s if col_name.endswith('_MR') else eng_f,
                'Difference': diff_pct,
                'Engine_Calculation': eng_explain,
                'GT_Calculation': gt_explain,
                'Question_To_Resolve': question,
            })

    filled_wb.close()
    gt_wb.close()

# ═══════════════════════════════════════════════════════
# Write Excel
# ═══════════════════════════════════════════════════════
print(f"\nTotal mismatches: {len(rows_data)}")

df = pd.DataFrame(rows_data)

output_path = "Mismatch_Audit_Report.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Mismatches Detail"

# Headers
headers = list(df.columns)
header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
header_font = Font(name='Calibri', bold=True, color="FFFFFF", size=11)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Data rows
miss_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
warn_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
data_font = Font(name='Calibri', size=10)
wrap_align = Alignment(vertical='top', wrap_text=True)

for r_idx, row_data in enumerate(rows_data, 2):
    for c_idx, h in enumerate(headers, 1):
        val = row_data[h]
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        cell.font = data_font
        cell.alignment = wrap_align
        cell.border = thin_border

        # Color coding
        if h == 'Difference' and val and ('MISS' in str(val) or '78' in str(val) or '81' in str(val)):
            cell.fill = miss_fill
        elif h == 'Difference' and val:
            cell.fill = warn_fill

# Column widths
col_widths = {
    'Dataset': 12, 'Row': 6, 'UID': 14, 'Order': 10, 'Process': 10,
    'Column': 14, 'Material': 10, 'Micron': 8, 'Size': 8,
    'GT_Value': 18, 'Engine_Value': 18, 'Difference': 14,
    'Engine_Calculation': 60, 'GT_Calculation': 45, 'Question_To_Resolve': 55,
}
for c, h in enumerate(headers, 1):
    ws.column_dimensions[get_column_letter(c)].width = col_widths.get(h, 15)

# Row heights for wrapped text
for r in range(2, len(rows_data) + 2):
    ws.row_dimensions[r].height = 80

# Freeze panes
ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions

# ── Summary Sheet ──
ws2 = wb.create_sheet("Summary")
summary_data = [
    ["IPP Engine Mismatch Audit Report", "", "", ""],
    ["", "", "", ""],
    ["Dataset", "Total Compared", "Matched", "Accuracy"],
    ["Feb 2026", 195, 186, "95.4%"],
    ["Nov 2025", 88, 80, "90.9%"],
    ["COMBINED", 283, 266, "94.0%"],
    ["", "", "", ""],
    ["Root Cause Category", "Count", "Fixable?", "Action Needed"],
    ["PET per-row MRR selection (manual GT)", 8, "No", "GT uses different MRR subset per row - manual decision"],
    ["NYLON missing from Stores", 3, "No", "MRR 83005 for N00694 not in stores data file"],
    ["INH for order not in Granules", 3, "Ask", "N00694 BANDERA: INH in Nov, not in Feb. Which is correct?"],
    ["MEGA PACK missing Nov data", 2, "No", "MEGA PACK file only has Feb 2026 data"],
    ["TPE MR# display order", 1, "Minor", "Engine shows dominant MRRs, GT shows specific one"],
    ["", "", "", ""],
    ["CHEMICAL RATES (Adh/Hard/Sol)", "", "", ""],
    ["Adhesive Rate", "100%", "31/31", "PERFECT across both datasets"],
    ["Hardener Rate", "100%", "31/31", "PERFECT across both datasets"],
    ["Solvent Rate", "100%", "7/7", "PERFECT across both datasets"],
]

for r_idx, row in enumerate(summary_data, 1):
    for c_idx, val in enumerate(row, 1):
        cell = ws2.cell(row=r_idx, column=c_idx, value=val)
        cell.font = Font(name='Calibri', size=11)
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        if r_idx == 1:
            cell.font = Font(name='Calibri', bold=True, size=14, color="1E293B")
        elif r_idx in (3, 8, 15):
            cell.font = Font(name='Calibri', bold=True, size=11)
            cell.fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
        elif r_idx in (16, 17, 18):
            cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")

ws2.column_dimensions['A'].width = 40
ws2.column_dimensions['B'].width = 18
ws2.column_dimensions['C'].width = 18
ws2.column_dimensions['D'].width = 55

# ── Questions Sheet ──
ws3 = wb.create_sheet("Questions To Resolve")
questions = [
    ["#", "Question", "Context", "Impact", "Affects"],
    [1, "Should ALL BANDERA/CYM TPE/WPE orders use INH, or only orders listed in Granules Recipe?",
     "In Nov 2025, order N00694 (BANDERA) is marked INH in GT but is NOT in Granules Recipe.\nIn Feb 2026, order G00418 (BANDERA) is NOT marked INH and uses regular MR#.\nThis is inconsistent between months.",
     "3 cells (MR#, Rate, Value) in Nov 2025 Row 17",
     "Nov 2025"],
    [2, "For PET 12mic: should the engine use ALL matching MRRs (weighted average) or select specific MRRs per row?",
     "GT file uses different MRR subsets for different rows of the same material.\nRow 42: GT picks MR#85547/85588 (rate=4.22)\nRow 43: GT picks ALL 5 MRRs but shows rate 4.35 (doesn't match any weighted avg)\nThis suggests manual rate assignment per row.",
     "8 cells across 4 rows in Feb 2026",
     "Feb 2026"],
    [3, "Is MRR 83005 for NYLON/N00694 missing from the Nov Stores Recordings file, or recorded under a different material name?",
     "Stores has N00694 entries only for material 'PE', not NYLON.\nThe GT shows MR#83005 with rate 7.72.",
     "3 cells (MR#, Rate, Value) in Nov 2025 Row 7",
     "Nov 2025"],
    [4, "Can you provide the MEGA PACK file for November 2025?",
     "Current MEGA PACK.xlsx only has February 2026 rates (TPE=4.7848, WPE=4.9568).\nRow 48 needs Nov 2025 TPE rate (GT=4.8708) which is missing.",
     "2 cells (Rate, Value) in Nov 2025 Row 48",
     "Nov 2025"],
    [5, "For TPE MEGA PACK: should the engine show MR#84080 specifically, or the dominant MRRs by quantity?",
     "Row 54 Feb: Engine finds 3 MRRs (85738, 85775, 84080) and shows the 2 with highest qty.\nGT shows only 84080. Rate is correct (MEGA PACK override).",
     "1 cell (MR# display only) in Feb 2026 Row 54",
     "Feb 2026"],
]

for r_idx, row in enumerate(questions, 1):
    for c_idx, val in enumerate(row, 1):
        cell = ws3.cell(row=r_idx, column=c_idx, value=val)
        cell.font = Font(name='Calibri', size=11)
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        cell.border = thin_border
        if r_idx == 1:
            cell.font = Font(name='Calibri', bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")

ws3.column_dimensions['A'].width = 5
ws3.column_dimensions['B'].width = 55
ws3.column_dimensions['C'].width = 60
ws3.column_dimensions['D'].width = 35
ws3.column_dimensions['E'].width = 12
for r in range(2, len(questions) + 1):
    ws3.row_dimensions[r].height = 90

wb.save(output_path)
print(f"\nSaved to: {output_path}")
