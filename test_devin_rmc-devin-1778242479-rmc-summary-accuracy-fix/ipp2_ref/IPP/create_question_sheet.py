"""
Create a simple Excel sheet for the user to show their friend
the contradiction in MRR selection across rows.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Question for You"

# Styles
title_font = Font(name='Calibri', size=16, bold=True, color='1E293B')
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='4F46E5')
data_font = Font(name='Calibri', size=11)
highlight_fill = PatternFill('solid', fgColor='FEF3C7')  # Yellow
green_fill = PatternFill('solid', fgColor='D1FAE5')
red_fill = PatternFill('solid', fgColor='FEE2E2')
blue_fill = PatternFill('solid', fgColor='DBEAFE')
gray_fill = PatternFill('solid', fgColor='F1F5F9')
question_font = Font(name='Calibri', size=13, bold=True, color='DC2626')
thin_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1'),
)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ═══════════════════════════════════════════════════
# TITLE
# ═══════════════════════════════════════════════════
ws.merge_cells('A1:H1')
ws['A1'] = 'Question: How do you decide which MRRs to use for the rate?'
ws['A1'].font = title_font
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 35

ws.merge_cells('A2:H2')
ws['A2'] = 'Below are 4 rows from February Jobtrack. Same material (PET), same MRR pool — but different selections. Why?'
ws['A2'].font = Font(name='Calibri', size=11, color='64748B')
ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[2].height = 25

# ═══════════════════════════════════════════════════
# SECTION 1: Available MRRs for Order L00335
# ═══════════════════════════════════════════════════
r = 4
ws.merge_cells(f'A{r}:H{r}')
ws[f'A{r}'] = 'CASE 1: Order L00335 — PET Material — These MRRs exist in Stores:'
ws[f'A{r}'].font = Font(name='Calibri', size=12, bold=True, color='4F46E5')
ws.row_dimensions[r].height = 25

r = 5
headers = ['MRR #', 'Supplier', 'Month', 'Rate (AED)', 'Qty (kg)']
for i, h in enumerate(headers):
    c = ws.cell(row=r, column=i+1, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = thin_border

mrrs = [
    (85547, 'JBF RAK', 'Feb 2026', 4.22, 1810),
    (85588, 'JBF RAK', 'Feb 2026', 4.22, 1858),
    (85572, 'FLEX', 'Feb 2026', 4.588, 991),
    (85157, 'FLEX', 'Jan 2026', 4.404, 498),
    (85226, 'JBF RAK', 'Jan 2026', 4.22, 305),
]

for i, (mrr, sup, month, rate, qty) in enumerate(mrrs):
    r = 6 + i
    row_fill = green_fill if 'Feb' in month else gray_fill
    for j, val in enumerate([mrr, sup, month, rate, qty]):
        c = ws.cell(row=r, column=j+1, value=val)
        c.font = data_font
        c.fill = row_fill
        c.alignment = center
        c.border = thin_border

# ═══════════════════════════════════════════════════
# SECTION 2: Row 42 — What you chose
# ═══════════════════════════════════════════════════
r = 12
ws.merge_cells(f'A{r}:H{r}')
ws[f'A{r}'] = 'ROW 42 (UID: 202602-1120-L) — Your Selection:'
ws[f'A{r}'].font = Font(name='Calibri', size=12, bold=True, color='059669')
ws.row_dimensions[r].height = 22

r = 13
for i, h in enumerate(['You Used', 'MRR #', 'Rate', '', 'Result', '', '', '']):
    c = ws.cell(row=r, column=i+1, value=h)
    c.font = Font(name='Calibri', size=10, bold=True, color='64748B')
    c.alignment = center

r = 14
data_42 = [
    ('✓', 85547, 4.22, '', '', '', '', ''),
    ('✓', 85588, 4.22, '', '', '', '', ''),
    ('✗', '85572 (skipped)', '', '', '', '', '', ''),
    ('✗', '85157 (skipped)', '', '', '', '', '', ''),
    ('✗', '85226 (skipped)', '', '', '', '', '', ''),
]
for i, row_data in enumerate(data_42):
    for j, val in enumerate(row_data):
        c = ws.cell(row=r+i, column=j+1, value=val)
        c.font = data_font
        c.alignment = center
        c.border = thin_border
        if val == '✓':
            c.fill = green_fill
            c.font = Font(name='Calibri', size=11, bold=True, color='059669')
        elif val == '✗':
            c.fill = red_fill
            c.font = Font(name='Calibri', size=11, bold=True, color='DC2626')

ws.cell(row=14, column=5, value='Your Rate:').font = Font(name='Calibri', size=11, bold=True)
ws.cell(row=14, column=6, value=4.22).font = Font(name='Calibri', size=14, bold=True, color='059669')
ws.cell(row=14, column=6).fill = green_fill
ws.cell(row=15, column=5, value='Method:').font = Font(name='Calibri', size=10, color='64748B')
ws.cell(row=15, column=6, value='Only JBF RAK, Feb only').font = Font(name='Calibri', size=10, color='64748B')

# ═══════════════════════════════════════════════════
# SECTION 3: Row 45 — What you chose (DIFFERENT!)
# ═══════════════════════════════════════════════════
r = 20
ws.merge_cells(f'A{r}:H{r}')
ws[f'A{r}'] = 'CASE 2: Order G00418 — PET Material — These MRRs exist in Stores:'
ws[f'A{r}'].font = Font(name='Calibri', size=12, bold=True, color='4F46E5')

r = 21
headers2 = ['MRR #', 'Supplier', 'Month', 'Rate (AED)', 'Qty (kg)']
for i, h in enumerate(headers2):
    c = ws.cell(row=r, column=i+1, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = thin_border

mrrs2 = [
    (85573, 'FLEX', 'Feb 2026', 4.588, 2621),
    (85330, 'JBF RAK', 'Jan 2026', 4.22, 253),
    (85157, 'FLEX', 'Jan 2026', 4.404, 180),
]
for i, (mrr, sup, month, rate, qty) in enumerate(mrrs2):
    r = 22 + i
    row_fill = green_fill if 'Feb' in month else gray_fill
    for j, val in enumerate([mrr, sup, month, rate, qty]):
        c = ws.cell(row=r, column=j+1, value=val)
        c.font = data_font
        c.fill = row_fill
        c.alignment = center
        c.border = thin_border

r = 26
ws.merge_cells(f'A{r}:H{r}')
ws[f'A{r}'] = 'ROW 45 (UID: 202602-1360-P) — Your Selection:'
ws[f'A{r}'].font = Font(name='Calibri', size=12, bold=True, color='059669')

r = 27
for i, h in enumerate(['You Used', 'MRR #', 'Rate', '', 'Result', '', '', '']):
    c = ws.cell(row=r, column=i+1, value=h)
    c.font = Font(name='Calibri', size=10, bold=True, color='64748B')
    c.alignment = center

data_45 = [
    ('✓', 85573, 4.588),
    ('✓', 85330, 4.22),
    ('✓', 85157, 4.404),
]
for i, (used, mrr, rate) in enumerate(data_45):
    r = 28 + i
    ws.cell(row=r, column=1, value=used).font = Font(name='Calibri', size=11, bold=True, color='059669')
    ws.cell(row=r, column=1).fill = green_fill
    ws.cell(row=r, column=1).alignment = center
    ws.cell(row=r, column=1).border = thin_border
    ws.cell(row=r, column=2, value=mrr).font = data_font
    ws.cell(row=r, column=2).alignment = center
    ws.cell(row=r, column=2).border = thin_border
    ws.cell(row=r, column=3, value=rate).font = data_font
    ws.cell(row=r, column=3).alignment = center
    ws.cell(row=r, column=3).border = thin_border

ws.cell(row=28, column=5, value='Your Rate:').font = Font(name='Calibri', size=11, bold=True)
ws.cell(row=28, column=6, value=4.5467).font = Font(name='Calibri', size=14, bold=True, color='059669')
ws.cell(row=28, column=6).fill = green_fill
ws.cell(row=29, column=5, value='Method:').font = Font(name='Calibri', size=10, color='64748B')
ws.cell(row=29, column=6, value='ALL MRRs, ALL months used').font = Font(name='Calibri', size=10, color='64748B')

# ═══════════════════════════════════════════════════
# SECTION 4: THE QUESTION
# ═══════════════════════════════════════════════════
r = 32
ws.merge_cells(f'A{r}:H{r}')
c = ws[f'A{r}']
c.value = '⬇  THE QUESTION  ⬇'
c.font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
c.fill = PatternFill('solid', fgColor='DC2626')
c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[r].height = 30

r = 33
ws.merge_cells(f'A{r}:H{r}')
ws[f'A{r}'] = 'In Row 42: You used ONLY Feb MRRs from JBF RAK (skipped FLEX and Jan MRRs)'
ws[f'A{r}'].font = Font(name='Calibri', size=12, bold=True, color='1E293B')
ws[f'A{r}'].fill = highlight_fill
ws[f'A{r}'].alignment = left
ws.row_dimensions[r].height = 22

r = 34
ws.merge_cells(f'A{r}:H{r}')
ws[f'A{r}'] = 'In Row 45: You used ALL MRRs including Jan MRRs and mixed suppliers'
ws[f'A{r}'].font = Font(name='Calibri', size=12, bold=True, color='1E293B')
ws[f'A{r}'].fill = highlight_fill
ws[f'A{r}'].alignment = left
ws.row_dimensions[r].height = 22

r = 35
ws.merge_cells(f'A{r}:H{r}')
ws[f'A{r}'] = ''
ws.row_dimensions[r].height = 10

r = 36
ws.merge_cells(f'A{r}:H{r}')
ws[f'A{r}'] = 'Question 1: When do you include old-month MRRs (Jan) and when do you skip them?'
ws[f'A{r}'].font = question_font
ws[f'A{r}'].alignment = left
ws.row_dimensions[r].height = 25

r = 37
ws.merge_cells(f'A{r}:H{r}')
ws[f'A{r}'] = 'Question 2: When do you filter by supplier (only JBF RAK) vs. use all suppliers?'
ws[f'A{r}'].font = question_font
ws[f'A{r}'].alignment = left
ws.row_dimensions[r].height = 25

r = 38
ws.merge_cells(f'A{r}:H{r}')
ws[f'A{r}'] = 'Question 3: What is the general rule you follow to select MRRs for the rate?'
ws[f'A{r}'].font = question_font
ws[f'A{r}'].alignment = left
ws.row_dimensions[r].height = 25

# ═══════════════════════════════════════════════════
# SECTION 5: Row 46 example (simple — just for reference)
# ═══════════════════════════════════════════════════
r = 40
ws.merge_cells(f'A{r}:H{r}')
ws[f'A{r}'] = 'BONUS: Row 46 (N00945) — You used only 85572 (Feb) and skipped 85226 (Jan). Same as Row 42 logic.'
ws[f'A{r}'].font = Font(name='Calibri', size=10, italic=True, color='64748B')
ws[f'A{r}'].alignment = left

r = 41
ws.merge_cells(f'A{r}:H{r}')
ws[f'A{r}'] = 'BONUS: Row 43 (L00335) — You used ALL 5 MRRs but got rate 4.35 which we cannot reproduce. How did you calculate it?'
ws[f'A{r}'].font = Font(name='Calibri', size=10, italic=True, color='DC2626')
ws[f'A{r}'].alignment = left

# Column widths
widths = [12, 22, 14, 14, 14, 28, 14, 14]
for i, w in enumerate(widths):
    ws.column_dimensions[get_column_letter(i+1)].width = w

# Save
path = "Template_Files/Question_MRR_Selection.xlsx"
wb.save(path)
print(f"Saved to: {path}")
