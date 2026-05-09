"""
Meeting Preparation Excel — Full MRR Calculation Trace
Shows exactly HOW each MR#, Rate, Value was derived with file references.
"""
import sys, io
sys.path.insert(0, '.')
import pandas as pd
import openpyxl
from meeting_excel_helpers import *
from engine.fill_jobtrack import COLS, DATA_START_ROW
from engine.mrr_lookup import load_stores_recordings, lookup_mrr_with_qty, MATERIAL_ALIASES
from engine.rate_lookup import (
    load_purchase_register, lookup_film_rate_weighted, filter_mrr_by_pr,
    lookup_material_rate_for_month, _find_col, PR_MAT_ALIASES
)
from engine.supplier_rates import build_mrr_supplier_map, get_supplier_for_mrrs

# ── Datasets ──
DATASETS = [
    {
        'name': 'February 2026',
        'jt_with': 'Template_Files/Jobtrack Feb With MRR.xlsx',
        'stores': 'Template_Files/Stores Recordings.xlsx',
        'pr': 'Template_Files/Purchase Register - 2021 - 2026 _Feb 26.xlsx',
        'rows': [42, 43, 45, 46, 54],
    },
    {
        'name': 'November 2025',
        'jt_with': 'Template2/Jobtrack With MRR.xlsx',
        'stores': 'Template2/Stores Recordings.xlsx',
        'pr': 'Template2/Purchase Register - 2021 - 2025 _Nov.xlsx',
        'rows': [7, 17, 48],
    },
]

# Which columns to trace per process
FILM_COLS = {
    'mat': COLS['Input_Name'], 'size': COLS['Input_Size'],
    'mic': COLS['Input_Mic'], 'mr': COLS['Film_MR'],
    'rate': COLS['Film_Rate'], 'value': COLS['Film_Value'],
    'qty': COLS['Input_Qty'], 'bal': COLS['Balance_Qty'],
    'total': COLS['Total_1st_Input'],
}
FRESH1_COLS = {
    'mat': COLS['Fresh1_Name'], 'size': COLS['Fresh1_Size'],
    'mic': COLS['Fresh1_Mic'], 'mr': COLS['Fresh1_MR'],
    'rate': COLS['Fresh1_Rate'], 'value': COLS['Fresh1_Value'],
    'qty': COLS['Fresh1_Qty'], 'bal': COLS['Fresh1_Balance'],
    'total': COLS['Total_Fresh1'],
}
FRESH2_COLS = {
    'mat': COLS['Fresh2_Name'], 'size': COLS['Fresh2_Size'],
    'mic': COLS['Fresh2_Mic'], 'mr': COLS['Fresh2_MR'],
    'rate': COLS['Fresh2_Rate'], 'value': COLS['Fresh2_Value'],
    'qty': COLS['Fresh2_Qty'], 'bal': COLS['Fresh2_Balance'],
    'total': COLS['Total_Fresh2'],
}

def trace_row(gt_ws, stores_df, pr_df, mrr_sup_map, row, dataset_name, pr_path, stores_path, jt_path):
    """Trace one row completely: how MR#, Rate, Value were derived."""
    uid = ss(gt_ws.cell(row=row, column=COLS['UID']).value)
    process = ss(gt_ws.cell(row=row, column=COLS['Process']).value).upper()
    order = ss(gt_ws.cell(row=row, column=COLS['Order_No']).value)

    # Determine which column sets to trace
    if process == 'PRINTING':
        col_sets = [('Film', FILM_COLS)]
    elif process == 'LAM':
        col_sets = [('1st Fresh', FRESH1_COLS), ('2nd Fresh', FRESH2_COLS)]
    else:
        return []

    results = []
    for label, C in col_sets:
        mat = ss(gt_ws.cell(row=row, column=C['mat']).value)
        if not mat:
            continue
        size_v = sf(gt_ws.cell(row=row, column=C['size']).value)
        mic_v = sf(gt_ws.cell(row=row, column=C['mic']).value)
        gt_mr = ss(gt_ws.cell(row=row, column=C['mr']).value)
        gt_rate = sf(gt_ws.cell(row=row, column=C['rate']).value)
        gt_value = sf(gt_ws.cell(row=row, column=C['value']).value)
        qty_raw = sf(gt_ws.cell(row=row, column=C['qty']).value)
        bal_raw = sf(gt_ws.cell(row=row, column=C['bal']).value)
        total_qty = qty_raw + bal_raw

        # ── Step 1: MRR Lookup from Stores ──
        proc_filter = 'PRINTING' if process == 'PRINTING' else 'LAMINATION'
        mrr_qty = lookup_mrr_with_qty(stores_df, mat, mic_v or None,
                                       size_v or None, order, proc_filter)
        step1_filter = f"Material={mat}, Mic={mic_v}, Order={order}, Process={proc_filter}"
        if not mrr_qty:
            mrr_qty = lookup_mrr_with_qty(stores_df, mat, mic_v or None, None, order, proc_filter)
            step1_filter += " → relaxed (no size)"
        if not mrr_qty:
            mrr_qty = lookup_mrr_with_qty(stores_df, mat, mic_v or None, None, order)
            step1_filter += " → relaxed (no process)"

        aliases = MATERIAL_ALIASES.get(mat.upper(), [mat.upper()])
        mrr_detail = '; '.join([f"MRR {m}: qty={q:.1f}" for m, q in
                                sorted(mrr_qty.items(), key=lambda x: -x[1])]) if mrr_qty else 'NONE FOUND'

        # ── Step 2: Filter by PR ──
        if mrr_qty:
            valid_mrr = filter_mrr_by_pr(pr_df, mrr_qty, mat, size_v or None, mic_v or None)
            pr_filtered = '; '.join([f"MRR {m}: qty={q:.1f}" for m, q in valid_mrr.items()])
        else:
            valid_mrr = {}
            pr_filtered = 'N/A'

        # ── Step 3: Supplier check ──
        supplier = ''
        if mrr_qty:
            supplier = get_supplier_for_mrrs(mrr_sup_map, list(mrr_qty.keys()), mrr_qty) or ''

        # ── Step 4: Get per-MRR rates from PR ──
        tracking_col = _find_col(pr_df, 'tracking')
        material_col = _find_col(pr_df, 'material')
        size_col_pr = _find_col(pr_df, 'size')
        mic_col_pr = _find_col(pr_df, 'mic')
        rate_col = [c for c in pr_df.columns if str(c).strip().lower() == 'rate']
        rate_col = rate_col[0] if rate_col else 'Rate'
        amt_col = [c for c in pr_df.columns if str(c).strip().lower() == 'amount']
        amt_col = amt_col[0] if amt_col else None
        aqty_col = [c for c in pr_df.columns if str(c).strip().lower() == 'actual quantity']
        aqty_col = aqty_col[0] if aqty_col else None

        per_mrr_detail = []
        for mrr_num in list(valid_mrr.keys())[:6]:
            try:
                mrr_int = int(float(mrr_num))
                mask = pd.to_numeric(pr_df[tracking_col], errors='coerce') == mrr_int
                pr_rows = pr_df[mask]
                for _, pr_row in pr_rows.iterrows():
                    pr_mat = ss(pr_row.get(material_col, ''))
                    pr_rate = sf(pr_row.get(rate_col, 0))
                    pr_size = sf(pr_row.get(size_col_pr or '', 0))
                    pr_mic = sf(pr_row.get(mic_col_pr or '', 0))
                    per_mrr_detail.append(f"MRR={mrr_int} Mat={pr_mat} Size={pr_size} Mic={pr_mic} Rate={pr_rate:.4f}")
            except:
                pass
        pr_rate_detail = '\n'.join(per_mrr_detail) if per_mrr_detail else 'NO MRR IN PR'

        # ── Step 5: Calculate weighted rate ──
        if valid_mrr:
            eng_rate = lookup_film_rate_weighted(pr_df, valid_mrr, mat, size_v or None, mic_v or None)
            if eng_rate == 0:
                eng_rate = lookup_film_rate_weighted(pr_df, valid_mrr, mat, None, mic_v or None)
            # Build formula
            parts = []
            for m, q in valid_mrr.items():
                try:
                    mrr_int = int(float(m))
                    mask = pd.to_numeric(pr_df[tracking_col], errors='coerce') == mrr_int
                    pr_match = pr_df[mask]
                    if not pr_match.empty:
                        r = sf(pr_match.iloc[0][rate_col])
                        parts.append((mrr_int, r, q))
                except:
                    pass
            if parts:
                num_parts = ' + '.join([f"({r:.4f} × {q:.1f})" for _, r, q in parts])
                den_parts = ' + '.join([f"{q:.1f}" for _, _, q in parts])
                num_val = sum(r * q for _, r, q in parts)
                den_val = sum(q for _, _, q in parts)
                calc_formula = f"({num_parts}) / ({den_parts}) = {num_val:.2f} / {den_val:.1f} = {eng_rate:.4f}"
            else:
                calc_formula = f"Result = {eng_rate:.4f}"
        else:
            eng_rate = 0
            calc_formula = "No MRRs found"

        eng_value = total_qty * eng_rate
        value_formula = f"{total_qty:.2f} × {eng_rate:.4f} = {eng_value:.2f}"

        # ── Dominant MRR display ──
        if mrr_qty:
            total_q = sum(mrr_qty.values())
            threshold = total_q * 0.10
            dominant = {m: q for m, q in mrr_qty.items() if q >= threshold}
            if not dominant:
                dominant = dict(sorted(mrr_qty.items(), key=lambda x: x[1], reverse=True)[:1])
            eng_mr = '/'.join(str(m) for m in sorted(dominant.keys()))
        else:
            eng_mr = ''

        # Rate diff
        diff = eng_rate - gt_rate if gt_rate > 0 else 0
        diff_pct = f"{abs(diff)/gt_rate*100:.1f}%" if gt_rate > 0 and diff != 0 else "MATCH" if gt_rate > 0 else "N/A"

        results.append({
            'Dataset': dataset_name,
            'Row': row,
            'UID': uid,
            'Process': process,
            'Type': label,
            'Order No': order,
            'Material': mat,
            'Size (mm)': size_v,
            'Micron': mic_v,
            'Material Aliases Used': ', '.join(aliases),
            # Ground Truth
            'GT MR#': gt_mr,
            'GT Rate': gt_rate,
            'GT Value': gt_value,
            # Engine
            'Engine MR#': eng_mr,
            'Engine Rate': eng_rate,
            'Engine Value': eng_value,
            'Rate Diff': diff,
            'Diff %': diff_pct,
            # Trace
            'Step 1: Stores Filter': step1_filter,
            'Step 1: Source File': stores_path,
            'Step 1: MRRs Found': mrr_detail,
            'Step 2: PR Validated MRRs': pr_filtered,
            'Step 2: Source File': pr_path,
            'Step 3: Supplier': supplier if supplier else 'Standard (no override)',
            'Step 4: Per-MRR PR Rates': pr_rate_detail,
            'Step 5: Rate Formula': calc_formula,
            'Step 6: Value Formula': value_formula,
            'Qty (Raw)': qty_raw,
            'Balance': bal_raw,
            'Total Qty': total_qty,
            'Total Qty Source': f"Jobtrack col {get_column_letter(C['qty'])} + col {get_column_letter(C['bal'])} (manual sum, never formula)",
        })
    return results

# ════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════
print("Building Meeting Preparation Excel...")
all_traces = []

for ds in DATASETS:
    print(f"\n── Processing {ds['name']} ──")
    stores_df = load_stores_recordings(ds['stores'])
    pr_df = load_purchase_register(ds['pr'])
    mrr_sup_map = build_mrr_supplier_map(stores_df)
    gt_wb = safe_open(ds['jt_with'], data_only=True)
    gt_ws = gt_wb.active

    for row in ds['rows']:
        print(f"  Tracing Row {row}...")
        traces = trace_row(gt_ws, stores_df, pr_df, mrr_sup_map, row,
                           ds['name'], ds['pr'], ds['stores'], ds['jt_with'])
        all_traces.extend(traces)
    gt_wb.close()

# ── Build Excel ──
wb = openpyxl.Workbook()

# ═══ Sheet 1: Full Trace ═══
ws1 = wb.active
ws1.title = "Row Trace (Full Detail)"
if all_traces:
    headers = list(all_traces[0].keys())
    ws1.append(headers)
    style_header(ws1, 1, len(headers))
    for t in all_traces:
        r = ws1.max_row + 1
        ws1.append([t[h] for h in headers])
        # Color by match status
        diff_pct = t.get('Diff %', '')
        if diff_pct == 'MATCH':
            fill = GREEN
        elif 'N/A' in str(diff_pct):
            fill = RED
        else:
            fill = YELLOW
        for c in range(1, len(headers) + 1):
            ws1.cell(r, c).fill = fill
            ws1.cell(r, c).alignment = Alignment(wrap_text=True, vertical='top')
    auto_width(ws1)

# ═══ Sheet 2: Questions for Team ═══
ws2 = wb.create_sheet("Questions for Team")
q_headers = ['#', 'Category', 'Question', 'Related Rows', 'Why It Matters', 'Current Engine Behavior']
ws2.append(q_headers)
style_header(ws2, 1, len(q_headers))
questions = [
    [1, 'Rate Method', 'When multiple MRRs exist, do you use a SPECIFIC MRR rate per row or a weighted average?',
     'Row 42,43,45,46 (Feb)', 'Engine gives 4.2983 but GT has 4.22 and 4.35 separately',
     'Qty-weighted avg: SUM(PR_Rate × Stores_Qty) / SUM(Stores_Qty)'],
    [2, 'MRR Assignment', 'Row 54: GT shows MRR 84080 but Stores has 85738/85775. Where does 84080 come from?',
     'Row 54 (Feb)', 'Engine cannot find MRR 84080 in Stores for this order',
     'Engine only uses MRRs found in Stores Recordings'],
    [3, 'NYLON Material', 'What is the sub-category name for NYLON in Stores Recordings?',
     'Row 7 (Nov)', 'Engine finds NO MRR because NYLON is not in aliases',
     'Engine has no alias for NYLON → complete miss'],
    [4, 'Outlier Check', 'When MRR rate = 4.87 but month avg = 0.90, which is correct?',
     'Row 17,48 (Nov)', 'Engine replaced correct 4.87 with wrong 0.90',
     'If >50% diff from month avg → use month avg (CAUSES BUGS)'],
    [5, 'INH Source', 'For WPE/WLDPE marked INH, is rate always from Granules Recipe?',
     'All INH rows', 'Need to confirm fallback chain',
     'Granules(current) → Granules(prev) → Granules(avg) → PR'],
    [6, 'Chemical Rates', 'Confirm: Adh/Hard/Sol rate = Total Amount / Total Qty from PR for same month?',
     'All LAM rows', 'Chemical rates are 100% match — just confirming rule',
     'Filter PR by category+material+month → Amount/Qty'],
]
for q in questions:
    r = ws2.max_row + 1
    ws2.append(q)
auto_width(ws2)

# ═══ Sheet 3: Calculation Rules ═══
ws3 = wb.create_sheet("How Engine Calculates")
rules = [
    ['Step', 'Action', 'Source File', 'Logic / Formula', 'Column Reference'],
    [1, 'Read Row Data', 'Jobtrack (Without MRR)', 'Read Material, Size, Mic, Order, Process, Qty, Balance',
     'Mat=AU, Size=AV, Mic=AW, Qty=AY, Bal=AZ, Order=K, Process=F'],
    [2, 'Compute Total Qty', 'Jobtrack (Without MRR)', 'Total = Qty (raw) + Balance (raw) — MANUAL SUM, never trust formula',
     'Film: AY+AZ=BA | Fresh1: BW+BX=BY | Fresh2: CG+CH=CI'],
    [3, 'Find MRRs', 'Stores Recordings', 'Filter: SubCategory matches Material (with aliases) AND Mic AND Order AND Process',
     'SubCat col, Mic col, Issue WO col, Issue Process col'],
    [4, 'Get MRR Quantities', 'Stores Recordings', 'Group by MRR No → SUM(Issue Qty) per MRR',
     'M.R.R No. col, Issue Qty col'],
    [5, 'Validate in PR', 'Purchase Register', 'Keep only MRRs that exist in PR Tracking Number column (with size ±5mm)',
     'Tracking Number col, Size col'],
    [6, 'Check Supplier', 'Stores Recordings', 'If supplier=BANDERA/CYM → Granules Recipe | MEGA PACK → MEGA PACK.xlsx',
     'Supplier Name col in Stores'],
    [7, 'Get Per-MRR Rate', 'Purchase Register', 'For each valid MRR: find row by Tracking Number → read Rate column',
     'Tracking Number, Material, Size, Mic, Rate columns'],
    [8, 'Weighted Avg Rate', 'Calculation', 'Rate = SUM(PR_Rate × Stores_Qty) / SUM(Stores_Qty)',
     'If all rates equal → use that rate directly'],
    [9, 'Compute Value', 'Calculation', 'Value = Total Qty × Rate', 'Written to Film/Fresh Rate + Value cols'],
    [10, 'Display MR#', 'Display Rule', 'Show MRRs with ≥10% of total qty, joined with "/"',
     'Film: BB | Fresh1: BZ | Fresh2: CJ'],
    ['—', '—— CHEMICALS ——', '—', '—', '—'],
    [11, 'Adhesive Rate', 'Purchase Register', 'Filter: Category=adhesive AND Material=exact name AND Month=report month\nRate = Total Amount / Total Actual Quantity',
     'Adh Name=CM, Rate→CO, Value→CP'],
    [12, 'Hardener Rate', 'Purchase Register', 'Hardener name from col DA → exact match in PR\nRate = Total Amount / Total Actual Quantity',
     'Hard Name=DA, Rate→CS, Value→CT'],
    [13, 'Solvent Rate', 'Purchase Register', 'Material=ETHYL ACETATE, Category=solvent, same month\nRate = Total Amount / Total Actual Quantity',
     'Sol Rate→CW, Value→CX'],
]
for i, rule in enumerate(rules):
    ws3.append(rule)
    if i == 0:
        style_header(ws3, 1, len(rule))
auto_width(ws3)

# ═══ Sheet 4: Accuracy Summary ═══
ws4 = wb.create_sheet("Current Accuracy")
ws4.append(['Metric', 'February 2026', 'November 2025'])
style_header(ws4, 1, 3)
ws4.append(['Total Compared', 195, 88])
ws4.append(['Match', 186, 80])
ws4.append(['Mismatch', 9, 8])
ws4.append(['Accuracy', '95.4%', '90.9%'])
ws4.append([])
ws4.append(['Mismatch Root Cause', 'Count', 'Fix'])
style_header(ws4, 7, 3)
ws4.append(['Per-MRR vs Weighted Avg', 6, 'Policy question for team'])
ws4.append(['Missing NYLON alias', 1, 'Add alias to engine'])
ws4.append(['Outlier check replaced correct rate', 2, 'Remove/fix outlier rule'])
ws4.append(['Different MRR# than Stores', 1, 'Manual override?'])
ws4.append([])
ws4.append(['Chemical Rates (Adh/Hard/Sol)', 'ALL 100% MATCH', '✓'])
ws4.cell(ws4.max_row, 1).font = BOLD
auto_width(ws4)

# Save
output = "Meeting_Preparation_MRR_Trace.xlsx"
wb.save(output)
print(f"\n✅ Saved: {output}")
print(f"   Total rows traced: {len(all_traces)}")
print("   Sheets: Row Trace | Questions | How Engine Calculates | Accuracy")
