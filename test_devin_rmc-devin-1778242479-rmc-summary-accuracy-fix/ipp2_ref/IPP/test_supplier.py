"""Quick test: Run engine with supplier files and check rows 9 and 54."""
import io
import shutil
import os
from engine.fill_jobtrack import fill_jobtrack
import openpyxl

BASE = "Template_Files"

with open(f"{BASE}/Jobtrack Feb Without MRR.xlsx", "rb") as f:
    jt = io.BytesIO(f.read())
with open(f"{BASE}/Stores Recordings.xlsx", "rb") as f:
    stores = io.BytesIO(f.read())
with open(f"{BASE}/Purchase Register - 2021 - 2026 _Feb 26.xlsx", "rb") as f:
    pr = io.BytesIO(f.read())
with open(f"{BASE}/Granules Recipe - February 2026.xlsx", "rb") as f:
    granules = io.BytesIO(f.read())

# MEGA PACK may be locked — copy first
src = f"{BASE}/MEGA PACK.xlsx"
tmp = f"{BASE}/MEGA_PACK_copy.xlsx"
shutil.copy2(src, tmp)
with open(tmp, "rb") as f:
    megapack = io.BytesIO(f.read())
os.remove(tmp)

output_bytes, results_log, fill_stats = fill_jobtrack(
    jt, stores, pr, granules_file=granules, megapack_file=megapack
)

print(f"Stats: {fill_stats}")

# Check rows 9 and 54
wb = openpyxl.load_workbook(io.BytesIO(output_bytes) if isinstance(output_bytes, bytes) else output_bytes, data_only=True)
ws = wb.active

for row in [9, 54]:
    uid = ws.cell(row=row, column=1).value
    for label, mr_col, rate_col, name_col in [
        ("Film", 54, 55, 47),
        ("Fresh1", 78, 79, 71),
        ("Fresh2", 88, 89, 81),
    ]:
        mr = ws.cell(row=row, column=mr_col).value
        rate = ws.cell(row=row, column=rate_col).value
        name = ws.cell(row=row, column=name_col).value
        if rate:
            print(f"Row {row} ({uid}): {label} = MR:{mr}, Rate:{rate}, Material:{name}")

# Expected:
# Row 9: Film Rate = 4.6993 (Granules Recipe)
# Row 54: Fresh2 Rate = 4.7848 (MEGA PACK)
print("\nExpected:")
print("  Row 9 Film Rate = 4.699253 (Granules Recipe for G00411)")
print("  Row 54 Fresh2 Rate = 4.784833 (MEGA PACK TPE for Feb/26)")

wb.close()

# Check log for supplier overrides
for entry in results_log:
    if 'Granules' in entry.get('detail', '') or 'MEGA' in entry.get('detail', ''):
        print(f"\nLog: {entry}")
