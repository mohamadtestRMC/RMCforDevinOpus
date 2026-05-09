"""
Detailed Mismatch Audit Excel — shows engine calc vs GT calc with questions.
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io, sys
sys.path.insert(0, '.')

from engine.fill_jobtrack import fill_jobtrack, COLS, DATA_START_ROW
from engine.rate_lookup import load_purchase_register, _find_col, lookup_adhesive_rate
from engine.mrr_lookup import load_stores_recordings, lookup_mrr_with_qty
from engine.supplier_rates import build_mrr_supplier_map, get_supplier_for_mrrs, load_granules_rates

def safe_float(val):
    if val is None or val == '' or (isinstance(val, float) and pd.isna(val)):
        return 0.0
    if isinstance(val, str) and val.startswith('='):
        return 0.0
    try: return float(val)
    except: return 0.0

def safe_str(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ''
    s = str(val).strip()
    return '' if s.startswith('=') else s

def get_pr_details(pr_df, material, report_month):
    """Get PR entries for a material in a given month."""
    mat_col = _find_col(pr_df, 'material')
    rate_col = _find_col(pr_df, 'rate')
    month_col = _find_col(pr_df, 'month')
    qty_col = _find_col(pr_df, 'qty') or _find_col(pr_df, 'quantity')
    amt_col = _find_col(pr_df, 'amount') or _find_col(pr_df, 'value')
    
    if not mat_col: return "No material column found"
    
    mat_upper = material.upper().strip()
    mask = pr_df[mat_col].astype(str).str.upper().str.strip().str.contains(mat_upper, na=False)
    if report_month and month_col:
        mask = mask & (pr_df[month_col].astype(str) == str(report_month))
    
    matches = pr_df[mask]
    if matches.empty:
        return f"No PR entries for '{material}' in month {report_month}"
    
    details = []
    for _, row in matches.head(10).iterrows():
        r = safe_float(row.get(rate_col)) if rate_col else 0
        q = safe_float(row.get(qty_col)) if qty_col else 0
        a = safe_float(row.get(amt_col)) if amt_col else 0
        details.append(f"Rate={r:.4f}, Qty={q:.1f}, Amt={a:.2f}")
    return "; ".join(details)

def get_stores_details(stores_df, mat, mic, size, order):
    """Get stores lookup results."""
    mrr_qty = lookup_mrr_with_qty(stores_df, mat, mic, size, order, 'PRINTING')
    if not mrr_qty:
        mrr_qty = lookup_mrr_with_qty(stores_df, mat, mic, size, order, 'LAMINATION')
    if not mrr_qty:
        mrr_qty = lookup_mrr_with_qty(stores_df, mat, mic, None, order)
    if mrr_qty:
        return ", ".join([f"MRR {m}: Qty={q:.1f}" for m, q in mrr_qty.items()])
    return f"No MRR found for {mat}/{mic}/{size}/{order}"

def collect_mismatches(dataset_name, jt_without, jt_with, stores_path, pr_path,
                       granules_path=None, megapack_path=None):
    """Run engine fill and collect all mismatches with details."""
    
    COMPARE_COLS = {
        'Film_MR': 54, 'Film_Rate': 55, 'Film_Value': 56,
        'Fresh1_MR': 78, 'Fresh1_Rate': 79, 'Fresh1_Value': 80,
        'Fresh2_MR': 88, 'Fresh2_Rate': 89, 'Fresh2_Value': 90,
        'Adh_Rate': 93, 'Adh_Value': 94,
        'Hard_Rate': 97, 'Hard_Value': 98,
        'Sol_Rate': 101, 'Sol_Value': 102,
    }
    
    # Run engine
    with open(jt_without, 'rb') as f:
        jt_bytes = io.BytesIO(f.read())
    
    gran_io = None
    mega_io = None
    if granules_path:
        with open(granules_path, 'rb') as f: gran_io = io.BytesIO(f.read())
    if megapack_path:
        with open(megapack_path, 'rb') as f: mega_io = io.BytesIO(f.read())
    
    filled_bytes, results_log, stats = fill_jobtrack(
        jt_bytes, stores_path, pr_path,
        granules_file=gran_io, megapack_file=mega_io
    )
    
    # Load workbooks
    filled_wb = openpyxl.load_workbook(filled_bytes)
    filled_ws = filled_wb.active
    gt_wb = openpyxl.load_workbook(jt_with, data_only=True)
    gt_ws = gt_wb.active
    
    # Load reference data
    stores_df = load_stores_recordings(stores_path)
    pr_df = load_purchase_register(pr_path)
    mrr_sup_map = build_mrr_supplier_map(stores_df)
    gran_rates = load_granules_rates(granules_path) if granules_path else {}
    
    # Detect report month from engine log
    report_month = None
    for scan_row in range(DATA_START_ROW, min(gt_ws.max_row + 1, DATA_START_ROW + 50)):
        dv = gt_ws.cell(row=scan_row, column=4).value
        if dv and hasattr(dv, 'month'):
            report_month = f"{dv.month}-{dv.year}"
            break
    if not report_month:
        # From PR filename
        import re
        _MONTHS = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
                    'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}
        for mn, num in _MONTHS.items():
            if mn in pr_path.lower():
                years = re.findall(r'20[12]\d', pr_path)
                if years:
                    report_month = f"{num}-{years[-1]}"
                    break
    
    max_row = min(filled_ws.max_row, gt_ws.max_row)
    mismatches = []
    
    for row in range(DATA_START_ROW, max_row + 1):
        process = safe_str(gt_ws.cell(row=row, column=COLS['Process']).value).upper()
        uid = safe_str(gt_ws.cell(row=row, column=COLS['UID']).value)
        order = safe_str(gt_ws.cell(row=row, column=COLS['Order_No']).value)
        if not process or not uid: continue
        
        for col_name, col_idx in COMPARE_COLS.items():
            is_film = col_name.startswith('Film_')
            is_fresh = col_name.startswith('Fresh')
            is_chem = col_name.startswith(('Adh_', 'Hard_', 'Sol_'))
            if process == 'PRINTING' and (is_fresh or is_chem): continue
            if process == 'LAM' and is_film: continue
            if process not in ('PRINTING', 'LAM'): continue
            
            gt_val = gt_ws.cell(row=row, column=col_idx).value
            eng_val = filled_ws.cell(row=row, column=col_idx).value
            gt_f = safe_float(gt_val); eng_f = safe_float(eng_val)
            gt_s = safe_str(gt_val); eng_s = safe_str(eng_val)
            
            is_mismatch = False
            if col_name.endswith('_MR'):
                gt_set = {s.strip() for s in gt_s.replace('/',',').split(',') if s.strip()} if gt_s else set()
                eng_set = {s.strip() for s in eng_s.replace('/',',').split(',') if s.strip()} if eng_s else set()
                if gt_set and not eng_set: is_mismatch = True
                elif gt_set and eng_set and gt_s != 'INH' and gt_set != eng_set and not (gt_set & eng_set):
                    is_mismatch = True
                elif gt_s == 'INH' and eng_s != 'INH': is_mismatch = True
            else:
                if gt_f == 0 and eng_f == 0: continue
                if gt_f > 0 and eng_f == 0: is_mismatch = True
                elif gt_f > 0 and abs(gt_f - eng_f) > 0.02 and abs(gt_f - eng_f)/gt_f > 0.005:
                    is_mismatch = True
            
            if not is_mismatch: continue
            
            # Collect details
            mat, mic, size = '', 0, 0
            if 'Film' in col_name:
                mat = safe_str(gt_ws.cell(row=row, column=COLS['Input_Name']).value)
                mic = safe_float(gt_ws.cell(row=row, column=COLS['Input_Mic']).value)
                size = safe_float(gt_ws.cell(row=row, column=COLS['Input_Size']).value)
            elif 'Fresh1' in col_name:
                mat = safe_str(gt_ws.cell(row=row, column=COLS['Fresh1_Name']).value)
                mic = safe_float(gt_ws.cell(row=row, column=COLS['Fresh1_Mic']).value)
                size = safe_float(gt_ws.cell(row=row, column=COLS['Fresh1_Size']).value)
            elif 'Fresh2' in col_name:
                mat = safe_str(gt_ws.cell(row=row, column=COLS['Fresh2_Name']).value)
                mic = safe_float(gt_ws.cell(row=row, column=COLS['Fresh2_Mic']).value)
                size = safe_float(gt_ws.cell(row=row, column=COLS['Fresh2_Size']).value)
            
            # Supplier info
            stores_info = get_stores_details(stores_df, mat, mic, size, order) if mat else "N/A"
            supplier = "N/A"
            mrr_qty = lookup_mrr_with_qty(stores_df, mat, mic, size, order, 'LAMINATION') if mat else {}
            if not mrr_qty and mat:
                mrr_qty = lookup_mrr_with_qty(stores_df, mat, mic, size, order, 'PRINTING')
            if not mrr_qty and mat:
                mrr_qty = lookup_mrr_with_qty(stores_df, mat, mic, None, order)
            if mrr_qty:
                supplier = get_supplier_for_mrrs(mrr_sup_map, list(mrr_qty.keys())) or "Standard"
            
            # Granules info
            gran_info = "N/A"
            if gran_rates:
                order_upper = order.upper().strip()
                if order_upper in gran_rates:
                    gran_info = f"WO# {order_upper} found: Rate={gran_rates[order_upper]:.4f}"
                else:
                    gran_info = f"WO# {order_upper} NOT in Granules. Available: {', '.join(gran_rates.keys())}"
            
            # PR details
            pr_info = get_pr_details(pr_df, mat, report_month) if mat else "N/A"
            
            # Diff calculation
            if col_name.endswith('_MR'):
                diff_str = "MR# mismatch"
                pct_diff = ""
            elif gt_f > 0:
                pct_diff = f"{abs(gt_f - eng_f)/gt_f*100:.1f}%"
                diff_str = f"{eng_f - gt_f:+.4f}"
            else:
                pct_diff = "GT=0"
                diff_str = f"{eng_f:.4f}"
            
            # Engine explanation
            if supplier in ('BANDERA', 'CYM') and 'Fresh' in col_name:
                eng_explain = (f"Supplier={supplier}. Looked for WO# {order.upper()} in Granules Recipe. "
                              f"{gran_info}. Since not found, fell through to PR lookup → got rate {eng_f}")
            elif mat.upper() == 'NYLON' and eng_f == 0:
                eng_explain = f"Searched Stores for NYLON/{mic}/{size}/{order}. {stores_info}. No MRR found → rate=0"
            elif 'PET' in mat.upper() and abs(gt_f - eng_f) / max(gt_f, 0.001) < 0.05:
                eng_explain = (f"Found MRRs in Stores: {stores_info}. Computed qty-weighted avg rate from PR = {eng_f:.4f}. "
                              f"GT uses a different per-row MRR selection = {gt_f:.4f}")
            elif supplier == 'MEGA PACK':
                eng_explain = f"Supplier=MEGA PACK. No MEGA PACK file for this month → fell through to PR → rate={eng_f}"
            else:
                eng_explain = f"Stores: {stores_info}. PR rate for {report_month}: {eng_f:.4f}"
            
            # GT explanation
            if gt_s == 'INH':
                gt_explain = f"Marked as INH (in-house). Supplier={supplier}. Rate={gt_f:.4f} from Granules Recipe"
            elif gt_f > 0 and mat.upper() == 'NYLON':
                gt_explain = f"MRR 83005 manually entered. Rate 7.72 from PR for NYLON"
            elif supplier == 'MEGA PACK' and gt_f > 4:
                gt_explain = f"Rate {gt_f:.4f} from MEGA PACK file (converted rate for {mat})"
            else:
                gt_explain = f"GT value = {gt_f:.4f}. Likely uses specific MRR rate rather than weighted avg"
            
            # Question
            if supplier in ('BANDERA', 'CYM') and gt_s == 'INH':
                question = (f"WO# {order.upper()} not in current month Granules. "
                           f"Should we check PREVIOUS month's Granules Recipe? "
                           f"If so, which file? The rate 4.1951 - is it from Oct 2025 Granules?")
            elif mat.upper() == 'NYLON':
                question = "MRR 83005 not in Stores file. Is this a data entry gap? Should NYLON use a different lookup?"
            elif supplier == 'MEGA PACK':
                question = f"No MEGA PACK file for Nov 2025. Is there one? Or should we use previous month's MEGA PACK rates?"
            elif 'PET' in mat.upper():
                question = (f"Engine uses weighted avg of ALL MRRs = {eng_f:.4f}. "
                           f"GT uses specific MRR rate = {gt_f:.4f}. "
                           f"Which MRR selection rule does the GT follow?")
            elif gt_s and eng_s and gt_s != eng_s:
                question = f"GT shows MR#={gt_s}, Engine shows MR#={eng_s}. Which MRR selection is correct?"
            else:
                question = "How should the engine handle this case?"
            
            mismatches.append({
                'Dataset': dataset_name,
                'Row': row,
                'UID': uid,
                'Order': order,
                'Column': col_name,
                'Material': mat,
                'Micron': mic,
                'Size': size,
                'Supplier': supplier,
                'GT_Value': gt_s if col_name.endswith('_MR') else f"{gt_f:.4f}" if gt_f else "0",
                'Engine_Value': eng_s if col_name.endswith('_MR') else f"{eng_f:.4f}" if eng_f else "0",
                'Difference': diff_str,
                'Pct_Diff': pct_diff,
                'Engine_Calculation': eng_explain,
                'GT_Calculation': gt_explain,
                'Stores_Lookup': stores_info,
                'PR_Details': pr_info,
                'Granules_Info': gran_info,
                'Question': question,
            })
    
    gt_wb.close()
    filled_wb.close()
    return mismatches

# ═══════════════════════════════════════════════════════
# COLLECT ALL MISMATCHES
# ═══════════════════════════════════════════════════════
print("Collecting Feb 2026 mismatches...")
T1 = "Template_Files"
mm1 = collect_mismatches(
    "Feb 2026", f"{T1}/Jobtrack Feb Without MRR.xlsx", f"{T1}/Jobtrack Feb With MRR.xlsx",
    f"{T1}/Stores Recordings.xlsx", f"{T1}/Purchase Register - 2021 - 2026 _Feb 26.xlsx",
    f"{T1}/Granules Recipe - February 2026.xlsx", f"{T1}/MEGA PACK.xlsx"
)

print("Collecting Nov 2025 mismatches...")
T2 = "Template2"
mm2 = collect_mismatches(
    "Nov 2025", f"{T2}/Jobtrack Without MRR.xlsx", f"{T2}/Jobtrack With MRR.xlsx",
    f"{T2}/Stores Recordings.xlsx", f"{T2}/Purchase Register - 2021 - 2025 _Nov.xlsx",
    f"{T2}/Granules Recipe -Nov_2025.xlsx", None
)

all_mm = mm1 + mm2
print(f"\nTotal mismatches: {len(all_mm)}")

# ═══════════════════════════════════════════════════════
# CREATE EXCEL
# ═══════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ── Sheet 1: Detailed Mismatches ──
ws1 = wb.active
ws1.title = "Mismatch Details"

headers = ['#', 'Dataset', 'Row', 'UID', 'Order', 'Column', 'Material', 'Micron', 'Size',
           'Supplier', 'GT Value', 'Engine Value', 'Difference', '% Diff',
           'How Engine Calculated', 'How GT Got Its Value',
           'Stores Lookup Result', 'PR Data Details', 'Granules Recipe Info',
           'Question To Resolve']

# Styles
hdr_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
hdr_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
gt_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
eng_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
q_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
thin = Side(style='thin', color='D9D9D9')
border = Border(top=thin, bottom=thin, left=thin, right=thin)

for c, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=c, value=h)
    cell.fill = hdr_fill
    cell.font = hdr_font
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

for i, mm in enumerate(all_mm, 1):
    r = i + 1
    vals = [i, mm['Dataset'], mm['Row'], mm['UID'], mm['Order'], mm['Column'],
            mm['Material'], mm['Micron'], mm['Size'], mm['Supplier'],
            mm['GT_Value'], mm['Engine_Value'], mm['Difference'], mm['Pct_Diff'],
            mm['Engine_Calculation'], mm['GT_Calculation'],
            mm['Stores_Lookup'], mm['PR_Details'], mm['Granules_Info'],
            mm['Question']]
    for c, v in enumerate(vals, 1):
        cell = ws1.cell(row=r, column=c, value=v)
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.font = Font(name='Calibri', size=10)
        if c == 11: cell.fill = gt_fill
        elif c == 12: cell.fill = eng_fill
        elif c == 20: cell.fill = q_fill

# Column widths
widths = [4, 10, 5, 15, 10, 14, 10, 7, 7, 12, 14, 14, 10, 7, 50, 40, 40, 40, 35, 50]
for c, w in enumerate(widths, 1):
    ws1.column_dimensions[get_column_letter(c)].width = w

ws1.auto_filter.ref = f"A1:T{len(all_mm)+1}"
ws1.freeze_panes = 'A2'

# ── Sheet 2: Summary by Root Cause ──
ws2 = wb.create_sheet("Summary by Root Cause")

causes = [
    ["PET Per-Row MRR Selection", "Feb 2026", "Rows 42-46", 8,
     "Engine computes qty-weighted average rate across ALL MRRs. GT manually selects specific MRR per row, giving slightly different rates (1-3% variance).",
     "Need to understand GT's MRR selection rule: does it always pick the closest MRR? The most recent? A specific one per order?"],
    ["INH Not Detected (N00694)", "Nov 2025", "Row 17", 3,
     "Supplier=BANDERA → should be INH. But WO# N00694 not in current month Granules Recipe. Engine fell through to PR lookup (rate=0.9).",
     "CONFIRMED RULE: BANDERA/CYM = ALWAYS INH. If WO# not in current Granules, check PREVIOUS month's Granules Recipe file. Need Oct 2025 Granules file."],
    ["NYLON MRR Missing", "Nov 2025", "Row 7", 3,
     "MRR 83005 for NYLON/15mic/1005/N00694 does not exist in the Nov Stores Recordings file. Complete data gap.",
     "Is MRR 83005 missing from the Stores file by mistake? Or should NYLON use a different lookup method?"],
    ["MEGA PACK Missing Nov Data", "Nov 2025", "Row 48", 2,
     "Supplier=MEGA PACK, material=TPE. No MEGA PACK file provided for Nov 2025. Engine fell through to PR (rate=0.9).",
     "Is there a MEGA PACK file for Nov 2025? Or should we check previous month's MEGA PACK rates?"],
    ["Fresh2 MR# Display", "Feb 2026", "Row 54", 1,
     "GT shows MR#=84080, Engine shows 85738/85775. Both are valid MRRs but GT picked a different one.",
     "For MEGA PACK supplier, should we always show MR# from the MEGA PACK file instead of from Stores?"],
]

cause_headers = ['Root Cause', 'Dataset', 'Affected Rows', 'Mismatch Count',
                 'Technical Explanation', 'Question / Action Needed']
for c, h in enumerate(cause_headers, 1):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.fill = hdr_fill
    cell.font = hdr_font
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

for i, cause in enumerate(causes, 2):
    for c, v in enumerate(cause, 1):
        cell = ws2.cell(row=i, column=c, value=v)
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.font = Font(name='Calibri', size=10)
        if c == 6: cell.fill = q_fill

cause_widths = [25, 12, 15, 10, 60, 60]
for c, w in enumerate(cause_widths, 1):
    ws2.column_dimensions[get_column_letter(c)].width = w

# ── Sheet 3: INH Rule (New confirmed rule) ──
ws3 = wb.create_sheet("INH Rule - CONFIRMED")

rule_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
confirm_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

rules = [
    ["RULE", "DETAIL"],
    ["When to mark INH", "When Supplier = BANDERA or CYM → ALWAYS mark MR# = 'INH'"],
    ["Rate source (Step 1)", "Look for WO# in CURRENT month's Granules Recipe file"],
    ["Rate source (Step 2)", "If WO# NOT found → check PREVIOUS month's Granules Recipe file"],
    ["Rate source (Step 3)", "If still not found → ???  (Need clarification: use average? error?)"],
    ["", ""],
    ["EVIDENCE", "DETAIL"],
    ["Nov Row 23 (B00942)", "Supplier=BANDERA, WO# in Nov Granules → Rate=4.4020 ✅ MATCH"],
    ["Nov Row 28 (G00340)", "Supplier=BANDERA, WO# in Nov Granules → Rate=4.0373 ✅ MATCH"],
    ["Nov Row 50 (B00969)", "Supplier=BANDERA, WO# in Nov Granules → Rate=4.4512 ✅ MATCH"],
    ["Nov Row 17 (N00694)", "Supplier=BANDERA, WO# NOT in Nov Granules → GT=4.1951 ❌ Need Oct Granules"],
    ["Feb Row 66 (G00418)", "Supplier=BANDERA, WO# NOT in Feb Granules → GT=4.9000 (uses actual MRR?)"],
    ["", ""],
    ["OPEN QUESTION", "DETAIL"],
    ["Q1: Oct 2025 Granules", "Can you provide the Granules Recipe file for October 2025?"],
    ["Q2: Feb G00418 case", "Feb Row 66: BANDERA but GT=MR#85854 (not INH). Is this an exception or should it also be INH?"],
    ["Q3: Final fallback", "If WO# not in current OR previous month Granules, what rate to use?"],
]

for i, row in enumerate(rules, 1):
    for c, v in enumerate(row, 1):
        cell = ws3.cell(row=i, column=c, value=v)
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.font = Font(name='Calibri', size=11)
        if i == 1:
            cell.fill = hdr_fill
            cell.font = hdr_font
        elif row[0] in ('RULE', 'EVIDENCE', 'OPEN QUESTION'):
            cell.fill = rule_fill
            cell.font = Font(name='Calibri', size=11, bold=True)
        elif '✅' in str(v):
            cell.fill = confirm_fill

ws3.column_dimensions['A'].width = 25
ws3.column_dimensions['B'].width = 80

# Save
output_path = "Detailed_Mismatch_Audit.xlsx"
wb.save(output_path)
print(f"\n✅ Saved: {output_path}")
print(f"   Sheet 1: {len(all_mm)} mismatches with full details")
print(f"   Sheet 2: 5 root causes with explanations")
print(f"   Sheet 3: INH Rule confirmed + open questions")
