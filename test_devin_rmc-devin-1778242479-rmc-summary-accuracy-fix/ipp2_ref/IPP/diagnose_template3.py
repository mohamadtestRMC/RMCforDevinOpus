"""
Diagnostic script for Template3 issues reported by user.
Investigates specific lines and data issues.
"""
import pandas as pd
import openpyxl
import os
import sys

# Fix encoding for Windows console
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

BASE = r"c:\Users\mohamad.al\OneDrive - AL KHAYYAT INVESTMENTS\Desktop\IPP\Template3"

print("=" * 80)
print("LOADING DATA FILES")
print("=" * 80)

stores_path = os.path.join(BASE, "Stores Recordings.xlsx")
stores_df = pd.read_excel(stores_path, sheet_name=0, header=1)
col_map = {c: str(c).strip().replace('\n', ' ').replace('\r', '') for c in stores_df.columns}
stores_df.rename(columns=col_map, inplace=True)
print(f"Stores: {len(stores_df)} rows")

stores_cols = {}
for c in stores_df.columns:
    cl = str(c).lower().strip()
    if 'sub' in cl and 'cat' in cl:
        stores_cols['sub_cat'] = c
    elif cl == 'mic':
        stores_cols['mic'] = c
    elif cl == 'width':
        stores_cols['width'] = c
    elif 'm.r.r' in cl and 'no' in cl:
        stores_cols['mrr'] = c
    elif 'issue wo' in cl:
        stores_cols['wo'] = c
    elif 'issue' in cl and 'process' in cl:
        stores_cols['process'] = c
    elif 'issue' in cl and 'qty' in cl:
        stores_cols['issue_qty'] = c

pr_path = os.path.join(BASE, "Purchase Register - 2021 - 2026 _Feb 26.xlsx")
pr_df = pd.read_excel(pr_path, sheet_name=0, header=2)
pr_col_map = {c: str(c).strip().replace('\n', ' ').replace('\r', '') for c in pr_df.columns}
pr_df.rename(columns=pr_col_map, inplace=True)
print(f"PR: {len(pr_df)} rows")

tracking_col = None
for c in pr_df.columns:
    if 'tracking' in str(c).lower():
        tracking_col = c
        break
print(f"PR tracking col: {tracking_col}")

jt_path = os.path.join(BASE, "Job Track Feb 26.xlsx")
wb = openpyxl.load_workbook(jt_path, data_only=True)
ws = wb.active
wb_formulas = openpyxl.load_workbook(jt_path)
ws_f = wb_formulas.active
print(f"Jobtrack: {ws.max_row} rows")

filled_path = os.path.join(BASE, "Jobtrack_Filled_MRR_20260429_1929.xlsx")
if os.path.exists(filled_path):
    wb_filled = openpyxl.load_workbook(filled_path, data_only=True)
    ws_filled = wb_filled.active
    print(f"Filled output: {ws_filled.max_row} rows")
else:
    ws_filled = None
    print("No filled output found")

# Also check the Comparison_Result.xlsx
comp_path = os.path.join(BASE, "Comparison_Result.xlsx")
if os.path.exists(comp_path):
    comp_df = pd.read_excel(comp_path, sheet_name=0)
    print(f"Comparison Result: {len(comp_df)} rows")
else:
    comp_df = None

print("\n" + "=" * 80)
print("ISSUE 1: Rewinding -> ULTRAFLEX mapping")
print("=" * 80)

proc_col = stores_cols['process']
proc_values = stores_df[proc_col].dropna().astype(str).str.upper().str.strip().unique()
print(f"Unique Stores Issue Process values: {sorted(proc_values)}")

ultraflex_rows = stores_df[stores_df[proc_col].astype(str).str.upper().str.strip().str.contains('ULTRAFLEX', na=False)]
print(f"Rows with ULTRAFLEX process: {len(ultraflex_rows)}")
if not ultraflex_rows.empty:
    display_cols = [stores_cols.get(k) for k in ['sub_cat', 'mic', 'width', 'wo', 'process', 'issue_qty', 'mrr'] if k in stores_cols]
    display_cols = [c for c in display_cols if c is not None]
    print(ultraflex_rows[display_cols].head(10).to_string())

# Check code already handles this
print("\nCode check: _stores_process_filter_for_row('REWINDING') returns 'ULTRAFLEX'")
print("This is already implemented at line 201-204 of fill_jobtrack.py")

PROCESS_COL = 6
print("\nSample Jobtrack Rewinding rows:")
count = 0
for r in range(5, ws.max_row + 1):
    proc = ws.cell(row=r, column=PROCESS_COL).value
    if proc and 'rewind' in str(proc).lower():
        count += 1
        if count <= 5:
            order = ws.cell(row=r, column=11).value
            uid = ws.cell(row=r, column=1).value
            mat = ws.cell(row=r, column=47).value
            print(f"  Row {r}: UID={uid}, WO={order}, Material={mat}")
print(f"Total Rewinding rows: {count}")

print("\n" + "=" * 80)
print("ISSUE 2: WO B01077 -> Line 625 (84526/85460)")
print("=" * 80)

row = 625
print(f"Row {row} in original Jobtrack:")
print(f"  UID (A): {ws.cell(row=row, column=1).value}")
print(f"  Process (F): {ws.cell(row=row, column=6).value}")
print(f"  Order No (K): {ws.cell(row=row, column=11).value}")
print(f"  Input Name (AU): {ws.cell(row=row, column=47).value}")
print(f"  Input Size (AV): {ws.cell(row=row, column=48).value}")
print(f"  Input Mic (AW): {ws.cell(row=row, column=49).value}")
print(f"  Input Qty formula (AY): {ws_f.cell(row=row, column=51).value}")
print(f"  Input Qty value (AY): {ws.cell(row=row, column=51).value}")
print(f"  Balance Qty formula (AZ): {ws_f.cell(row=row, column=52).value}")
print(f"  Balance Qty value (AZ): {ws.cell(row=row, column=52).value}")
print(f"  Total 1st Input (BA): {ws.cell(row=row, column=53).value}")

if ws_filled:
    print(f"\nRow {row} in FILLED output:")
    print(f"  Film MR# (BB): {ws_filled.cell(row=row, column=54).value}")
    print(f"  Film Rate (BC): {ws_filled.cell(row=row, column=55).value}")

wo = 'B01077'
stores_wo = stores_df[stores_df[stores_cols['wo']].astype(str).str.strip() == wo]
print(f"\nStores entries for WO={wo}: {len(stores_wo)} rows")
if not stores_wo.empty:
    display_cols = [stores_cols.get(k) for k in ['sub_cat', 'mic', 'width', 'wo', 'process', 'issue_qty', 'mrr'] if k in stores_cols]
    display_cols = [c for c in display_cols if c is not None]
    print(stores_wo[display_cols].to_string())

for mrr in [84526, 85460]:
    if tracking_col:
        pr_mrr = pr_df[pd.to_numeric(pr_df[tracking_col], errors='coerce') == mrr]
        print(f"\nPR entries for MRR {mrr}: {len(pr_mrr)} rows")
        if not pr_mrr.empty:
            rate_cols = [c for c in pr_df.columns if str(c).strip().lower() == 'rate']
            mat_cols = [c for c in pr_df.columns if 'material' in str(c).lower()]
            cols_show = [tracking_col]
            if mat_cols: cols_show.append(mat_cols[0])
            if rate_cols: cols_show.append(rate_cols[0])
            qty_cols = [c for c in pr_df.columns if 'actual' in str(c).lower() and 'qty' in str(c).lower()]
            if qty_cols: cols_show.append(qty_cols[0])
            print(pr_mrr[cols_show].to_string())

# Check nearby rows for same WO
wo_625 = str(ws.cell(row=625, column=11).value).strip() if ws.cell(row=625, column=11).value else ''
print(f"\nNearby Jobtrack rows with same WO={wo_625}:")
for r in range(max(5, 625 - 30), min(ws.max_row + 1, 625 + 30)):
    r_wo = ws.cell(row=r, column=11).value
    if r_wo and str(r_wo).strip() == wo_625:
        ay_formula = ws_f.cell(row=r, column=51).value
        az_val = ws.cell(row=r, column=52).value
        ba_val = ws.cell(row=r, column=53).value
        print(f"  Row {r}: AY_formula={ay_formula}, AZ={az_val}, BA={ba_val}")

print("\n" + "=" * 80)
print("ISSUE 3: Line 747 - WO same, QTY=365")
print("=" * 80)

row = 747
print(f"Row {row}:")
print(f"  UID (A): {ws.cell(row=row, column=1).value}")
print(f"  Process (F): {ws.cell(row=row, column=6).value}")
print(f"  Order No (K): {ws.cell(row=row, column=11).value}")
print(f"  Input Name (AU): {ws.cell(row=row, column=47).value}")
print(f"  Input Mic (AW): {ws.cell(row=row, column=49).value}")
print(f"  Input Qty formula (AY): {ws_f.cell(row=row, column=51).value}")
print(f"  Input Qty value (AY): {ws.cell(row=row, column=51).value}")
print(f"  Balance Qty formula (AZ): {ws_f.cell(row=row, column=52).value}")
print(f"  Balance Qty value (AZ): {ws.cell(row=row, column=52).value}")

if ws_filled:
    print(f"  FILLED Film MR# (BB): {ws_filled.cell(row=row, column=54).value}")

wo_747 = str(ws.cell(row=row, column=11).value).strip() if ws.cell(row=row, column=11).value else ''
print(f"\nNearby rows with same WO={wo_747}:")
for r in range(max(5, row - 30), min(ws.max_row + 1, row + 30)):
    r_wo = ws.cell(row=r, column=11).value
    if r_wo and str(r_wo).strip() == wo_747:
        ay_formula = ws_f.cell(row=r, column=51).value
        az_formula = ws_f.cell(row=r, column=52).value
        ay_val = ws.cell(row=r, column=51).value
        az_val = ws.cell(row=r, column=52).value
        filled_mrr = ws_filled.cell(row=r, column=54).value if ws_filled else 'N/A'
        print(f"  Row {r}: AY_f={ay_formula} (val={ay_val}), AZ_f={az_formula} (val={az_val}), MRR={filled_mrr}")

stores_wo747 = stores_df[stores_df[stores_cols['wo']].astype(str).str.strip() == wo_747]
print(f"\nStores for WO={wo_747}: {len(stores_wo747)} rows")
if not stores_wo747.empty:
    display_cols = [stores_cols.get(k) for k in ['sub_cat', 'mic', 'width', 'wo', 'process', 'issue_qty', 'mrr'] if k in stores_cols]
    display_cols = [c for c in display_cols if c is not None]
    print(stores_wo747[display_cols].to_string())

print("\n" + "=" * 80)
print("ISSUE 4: Lines 907, 1025")
print("=" * 80)

for row in [907, 1025]:
    print(f"\nRow {row}:")
    print(f"  UID (A): {ws.cell(row=row, column=1).value}")
    print(f"  Process (F): {ws.cell(row=row, column=6).value}")
    print(f"  Order No (K): {ws.cell(row=row, column=11).value}")
    print(f"  Input Name (AU): {ws.cell(row=row, column=47).value}")
    print(f"  Input Mic (AW): {ws.cell(row=row, column=49).value}")
    print(f"  Input Qty formula (AY): {ws_f.cell(row=row, column=51).value}")
    print(f"  Input Qty value (AY): {ws.cell(row=row, column=51).value}")
    print(f"  Balance Qty formula (AZ): {ws_f.cell(row=row, column=52).value}")
    print(f"  Balance Qty value (AZ): {ws.cell(row=row, column=52).value}")
    if ws_filled:
        print(f"  FILLED Film MR# (BB): {ws_filled.cell(row=row, column=54).value}")
    
    wo_r = str(ws.cell(row=row, column=11).value).strip() if ws.cell(row=row, column=11).value else ''
    print(f"  Nearby rows with same WO={wo_r}:")
    for r in range(max(5, row - 20), min(ws.max_row + 1, row + 20)):
        r_wo = ws.cell(row=r, column=11).value
        if r_wo and str(r_wo).strip() == wo_r:
            ay_f = ws_f.cell(row=r, column=51).value
            az_f = ws_f.cell(row=r, column=52).value
            filled_mrr = ws_filled.cell(row=r, column=54).value if ws_filled else 'N/A'
            print(f"    Row {r}: AY={ay_f}, AZ={az_f}, MRR={filled_mrr}")

print("\n" + "=" * 80)
print("ISSUE 5: Line 111 - Rate 5.075... and J00877")
print("=" * 80)

row = 111
print(f"Row {row}:")
print(f"  UID (A): {ws.cell(row=row, column=1).value}")
print(f"  Process (F): {ws.cell(row=row, column=6).value}")
print(f"  Order No (K): {ws.cell(row=row, column=11).value}")
print(f"  Input Name (AU): {ws.cell(row=row, column=47).value}")
print(f"  Input Size (AV): {ws.cell(row=row, column=48).value}")
print(f"  Input Mic (AW): {ws.cell(row=row, column=49).value}")
print(f"  Input Qty formula (AY): {ws_f.cell(row=row, column=51).value}")
print(f"  Total 1st Input (BA): {ws.cell(row=row, column=53).value}")

if ws_filled:
    print(f"\n  FILLED Film MR# (BB): {ws_filled.cell(row=row, column=54).value}")
    print(f"  FILLED Film Rate (BC): {ws_filled.cell(row=row, column=55).value}")
    print(f"  FILLED Film Value (BD): {ws_filled.cell(row=row, column=56).value}")

wo_111 = str(ws.cell(row=row, column=11).value).strip() if ws.cell(row=row, column=11).value else ''
print(f"\nWO for row 111: '{wo_111}'")

stores_wo111 = stores_df[stores_df[stores_cols['wo']].astype(str).str.strip() == wo_111]
print(f"Stores entries for WO={wo_111}: {len(stores_wo111)} rows")
if not stores_wo111.empty:
    display_cols = [stores_cols.get(k) for k in ['sub_cat', 'mic', 'width', 'wo', 'process', 'issue_qty', 'mrr'] if k in stores_cols]
    display_cols = [c for c in display_cols if c is not None]
    print(stores_wo111[display_cols].to_string())

# Search for J00877 everywhere
print(f"\nSearching for 'J00877' in Stores WO column:")
stores_j = stores_df[stores_df[stores_cols['wo']].astype(str).str.strip().str.contains('J00877', na=False)]
print(f"  Found: {len(stores_j)} rows")

print(f"\nSearching for 'J00877' in Jobtrack Order column:")
found_j = False
for r in range(5, ws.max_row + 1):
    wo_val = ws.cell(row=r, column=11).value
    if wo_val and 'J00877' in str(wo_val):
        print(f"  Found at row {r}: {wo_val}")
        found_j = True
if not found_j:
    print("  NOT FOUND in Jobtrack!")

print(f"\nSearching for 'J00877' in PR:")
if tracking_col:
    pr_j = pr_df[pr_df[tracking_col].astype(str).str.strip().str.contains('J00877', na=False)]
    print(f"  Found in tracking: {len(pr_j)} rows")
# Also search party/material columns
for c in pr_df.columns:
    matches = pr_df[pr_df[c].astype(str).str.contains('J00877', na=False)]
    if len(matches) > 0:
        print(f"  Found 'J00877' in PR column '{c}': {len(matches)} rows")

print("\n" + "=" * 80)
print("ISSUE 6: Line 2350 - 81732/85226 but QTY=100 -> should be 1 MRR")
print("=" * 80)

row = 2350
if row <= ws.max_row:
    print(f"Row {row}:")
    print(f"  UID (A): {ws.cell(row=row, column=1).value}")
    print(f"  Process (F): {ws.cell(row=row, column=6).value}")
    print(f"  Order No (K): {ws.cell(row=row, column=11).value}")
    print(f"  Input Name (AU): {ws.cell(row=row, column=47).value}")
    print(f"  Input Size (AV): {ws.cell(row=row, column=48).value}")
    print(f"  Input Mic (AW): {ws.cell(row=row, column=49).value}")
    print(f"  Input Qty formula (AY): {ws_f.cell(row=row, column=51).value}")
    print(f"  Input Qty value (AY): {ws.cell(row=row, column=51).value}")
    print(f"  Balance Qty formula (AZ): {ws_f.cell(row=row, column=52).value}")
    print(f"  Balance Qty value (AZ): {ws.cell(row=row, column=52).value}")
    print(f"  Total (BA): {ws.cell(row=row, column=53).value}")
    
    # Also check LAM columns
    print(f"  Fresh1 Name (BS): {ws.cell(row=row, column=71).value}")
    print(f"  Fresh1 Qty formula (BW): {ws_f.cell(row=row, column=75).value}")
    print(f"  Fresh2 Name (CC): {ws.cell(row=row, column=81).value}")
    
    if ws_filled:
        print(f"\n  FILLED Film MR# (BB): {ws_filled.cell(row=row, column=54).value}")
        print(f"  FILLED Fresh1 MR# (BZ): {ws_filled.cell(row=row, column=78).value}")
        print(f"  FILLED Fresh2 MR# (CJ): {ws_filled.cell(row=row, column=88).value}")

    wo_2350 = str(ws.cell(row=row, column=11).value).strip() if ws.cell(row=row, column=11).value else ''
    mat_2350 = str(ws.cell(row=row, column=47).value).strip() if ws.cell(row=row, column=47).value else ''
    mic_2350 = ws.cell(row=row, column=49).value
    
    stores_wo2350 = stores_df[stores_df[stores_cols['wo']].astype(str).str.strip() == wo_2350]
    print(f"\nStores for WO={wo_2350}: {len(stores_wo2350)} rows")
    if not stores_wo2350.empty:
        display_cols = [stores_cols.get(k) for k in ['sub_cat', 'mic', 'width', 'wo', 'process', 'issue_qty', 'mrr'] if k in stores_cols]
        display_cols = [c for c in display_cols if c is not None]
        print(stores_wo2350[display_cols].to_string())
    
    for mrr in [81732, 85226]:
        if tracking_col:
            pr_mrr = pr_df[pd.to_numeric(pr_df[tracking_col], errors='coerce') == mrr]
            rate_cols = [c for c in pr_df.columns if str(c).strip().lower() == 'rate']
            mat_cols = [c for c in pr_df.columns if 'material' in str(c).lower()]
            print(f"\nPR for MRR {mrr}: {len(pr_mrr)} rows")
            if not pr_mrr.empty:
                cols_show = [tracking_col]
                if mat_cols: cols_show.append(mat_cols[0])
                if rate_cols: cols_show.append(rate_cols[0])
                print(pr_mrr[cols_show].to_string())
else:
    print(f"Row {row} exceeds max {ws.max_row}")

print("\n" + "=" * 80)
print("CHECKING COMPARISON RESULT FILE")
print("=" * 80)
if comp_df is not None:
    print(f"Columns: {list(comp_df.columns)}")
    print(f"Shape: {comp_df.shape}")
    print(f"\nFirst few rows:")
    print(comp_df.head(3).to_string())

wb.close()
wb_formulas.close()
if ws_filled:
    wb_filled.close()

print("\n\nDONE!")
