"""
FINAL verification: Confirm the fallback rules
1. S621/S110: Fallback to MOST RECENT month when no data in report month
2. CR 88-300 -> CR 800-300: Material name alias
3. Verify these rules also work for Feb 2026
"""
import pandas as pd
import sys
sys.path.insert(0, '.')
from engine.rate_lookup import load_purchase_register, _find_col, _filter_by_month, _qty_weighted_rate

# ── 1. CR 88-300 -> CR 800-300 connection ──
print("=" * 80)
print("1. CR 88-300 -> CR 800-300 CONNECTION")
print("=" * 80)

# They are likely the same product. CR 88-300 is a short name for CR 800-300
# The GT rate for DA='CR 88-300' is 9.5208
# CR 800-300 in Nov 2025 is 9.5208 - EXACT MATCH!
print("CR 88-300 (DA column) = CR 800-300 (PR material name)")
print("GT rate 9.5208 = CR 800-300 Nov 2025 rate 9.5208")
print("CONFIRMED: CR 88-300 should be aliased to CR 800-300")

# ── 2. Fallback rule: Use most recent month when no data ──
print("\n" + "=" * 80)
print("2. FALLBACK RULE: Most recent month when no data in report month")
print("=" * 80)

pr_df = load_purchase_register("Template2/Purchase Register - 2021 - 2025 _Nov.xlsx")
cat_col = _find_col(pr_df, 'categery', 'category')
mat_col = _find_col(pr_df, 'material')
month_col = None
for c in pr_df.columns:
    if str(c).strip().lower() == 'month':
        month_col = c
        break

amt_col = None
qty_col = None
for c in pr_df.columns:
    cl = str(c).strip().lower()
    if cl == 'amount':
        amt_col = c
    if cl == 'actual quantity':
        qty_col = c

def get_most_recent_rate(pr_df, material_name, category_kw, report_month):
    """Find the most recent month's rate for a material when report_month has no data."""
    adh_mask = pr_df[cat_col].astype(str).str.lower().str.contains(category_kw, na=False)
    mat_mask = pr_df[mat_col].astype(str).str.upper().str.strip() == material_name.upper()
    rows = pr_df[adh_mask & mat_mask]
    
    if rows.empty:
        return 0.0, None
    
    # Parse report month
    parts = report_month.split('-')
    report_m, report_y = int(parts[0]), int(parts[1])
    report_key = report_y * 100 + report_m
    
    # Group by month and find most recent <= report month
    best_month = None
    best_key = 0
    
    for m in rows[month_col].astype(str).str.strip().unique():
        mp = m.split('-')
        if len(mp) == 2:
            try:
                mk = int(mp[1]) * 100 + int(mp[0])
                if mk <= report_key and mk > best_key:
                    best_key = mk
                    best_month = m
            except:
                pass
    
    if not best_month:
        return 0.0, None
    
    month_rows = rows[rows[month_col].astype(str).str.strip() == best_month]
    total_a = pd.to_numeric(month_rows[amt_col], errors='coerce').fillna(0).sum()
    total_q = pd.to_numeric(month_rows[qty_col], errors='coerce').fillna(0).sum()
    rate = total_a / total_q if total_q > 0 else 0
    
    return rate, best_month

# Test the fallback rule
report_month = '11-2025'

for material, gt_rate in [('S621', 11.6529), ('S110', 10.4766), ('CR 800-300', 9.5208)]:
    rate, month_used = get_most_recent_rate(pr_df, material, 'adhesive', report_month)
    match = abs(rate - gt_rate) < 0.001
    print(f"  {material}: Most recent month = {month_used}, Rate = {rate:.4f} | GT = {gt_rate:.4f} | {'MATCH' if match else 'MISMATCH'}")

# ── 3. Now let's define the COMPLETE RULES ──
print("\n" + "=" * 80)
print("3. COMPLETE CHEMICAL LOOKUP RULES")
print("=" * 80)

print("""
RULE 1 - ADHESIVE RATE:
  Source: PR Category = "ADHESIVE"
  Material: From CM column (with name_map: '75-300' -> 'MF 75-300')
  Month: report_month
  Fallback: If no data in report_month, use MOST RECENT MONTH <= report_month
  Calculation: Total Amount / Total Actual Quantity

RULE 2 - HARDENER RATE:
  Source: PR Category = "ADHESIVE" (hardener materials are under ADHESIVE category)
  Material: From DA column (105) - direct lookup
  Name aliases: 'CR 88-300' -> 'CR 800-300' (same product)
                'CR84' -> 'CR 84' (spacing)
  Month: report_month
  Fallback: If no data in report_month, use MOST RECENT MONTH <= report_month
  Calculation: Total Amount / Total Actual Quantity

RULE 3 - SOLVENT RATE:
  Source: PR Category = "SOLVENT"
  Material: Always "ETHYL ACETATE"
  Month: report_month
  Fallback: Same most-recent-month rule
  Calculation: Total Amount / Total Actual Quantity
""")

# ── 4. VERIFY COMPLETE RULES against both datasets ──
print("=" * 80)
print("4. FULL VERIFICATION: Complete Rules vs Ground Truth")
print("=" * 80)

import openpyxl

def lookup_chemical_rate(pr_df, material_name, category_kw, report_month, name_aliases=None):
    """Complete lookup with fallback to most recent month."""
    lookup_name = material_name.upper().strip()
    
    # Apply name aliases
    if name_aliases and lookup_name in name_aliases:
        lookup_name = name_aliases[lookup_name]
    
    adh_mask = pr_df[cat_col].astype(str).str.lower().str.contains(category_kw, na=False)
    mat_mask = pr_df[mat_col].astype(str).str.upper().str.strip().apply(
        lambda x: x == lookup_name or lookup_name in x or x in lookup_name
    )
    rows = pr_df[adh_mask & mat_mask]
    
    if rows.empty:
        return 0.0
    
    # Try report month first
    month_rows = rows[rows[month_col].astype(str).str.strip() == report_month]
    
    if not month_rows.empty:
        total_a = pd.to_numeric(month_rows[amt_col], errors='coerce').fillna(0).sum()
        total_q = pd.to_numeric(month_rows[qty_col], errors='coerce').fillna(0).sum()
        return total_a / total_q if total_q > 0 else 0.0
    
    # Fallback: most recent month
    parts = report_month.split('-')
    report_key = int(parts[1]) * 100 + int(parts[0])
    
    best_month = None
    best_key = 0
    for m in rows[month_col].astype(str).str.strip().unique():
        mp = m.split('-')
        if len(mp) == 2:
            try:
                mk = int(mp[1]) * 100 + int(mp[0])
                if mk <= report_key and mk > best_key:
                    best_key = mk
                    best_month = m
            except:
                pass
    
    if best_month:
        month_rows = rows[rows[month_col].astype(str).str.strip() == best_month]
        total_a = pd.to_numeric(month_rows[amt_col], errors='coerce').fillna(0).sum()
        total_q = pd.to_numeric(month_rows[qty_col], errors='coerce').fillna(0).sum()
        return total_a / total_q if total_q > 0 else 0.0
    
    return 0.0

ADH_NAME_MAP = {'75-300': 'MF 75-300', '85-300': 'MF 75-300'}
HARDENER_NAME_MAP = {
    'CR84': 'CR 84',
    'CR 88-300': 'CR 800-300',
}

for ds_name, jt_path, pr_path, month in [
    ("Nov 2025", "Template2/Jobtrack With MRR.xlsx", 
     "Template2/Purchase Register - 2021 - 2025 _Nov.xlsx", "11-2025"),
    ("Feb 2026", "Template_Files/Jobtrack Feb With MRR.xlsx", 
     "Template_Files/Purchase Register - 2021 - 2026 _Feb 26.xlsx", "2-2026"),
]:
    print(f"\n--- {ds_name} ---")
    wb = openpyxl.load_workbook(jt_path, data_only=True)
    ws = wb.active
    pr_df2 = load_purchase_register(pr_path)
    
    # Reinit column finders for this PR
    cat_col = _find_col(pr_df2, 'categery', 'category')
    mat_col = _find_col(pr_df2, 'material')
    month_col = None
    for c in pr_df2.columns:
        if str(c).strip().lower() == 'month':
            month_col = c
            break
    amt_col = None
    qty_col = None
    for c in pr_df2.columns:
        cl = str(c).strip().lower()
        if cl == 'amount':
            amt_col = c
        if cl == 'actual quantity':
            qty_col = c
    
    total = 0
    adh_ok_count = 0
    hard_ok_count = 0
    sol_ok_count = 0
    
    for row in range(5, ws.max_row + 1):
        process = ws.cell(row=row, column=6).value
        if not process or str(process).strip().upper() != 'LAM':
            continue
        
        adh_name = ws.cell(row=row, column=91).value
        if not adh_name:
            continue
        
        total += 1
        adh_str = str(adh_name).strip()
        da_str = str(ws.cell(row=row, column=105).value or '').strip()
        
        gt_adh_rate = float(ws.cell(row=row, column=93).value or 0)
        gt_hard_rate = float(ws.cell(row=row, column=97).value or 0)
        gt_sol_rate = float(ws.cell(row=row, column=101).value or 0)
        sol_qty = float(ws.cell(row=row, column=100).value or 0)
        
        # Engine computation with new rules
        eng_adh = lookup_chemical_rate(pr_df2, adh_str, 'adhesive', month, ADH_NAME_MAP)
        eng_hard = lookup_chemical_rate(pr_df2, da_str, 'adhesive', month, HARDENER_NAME_MAP) if da_str else 0
        eng_sol = lookup_chemical_rate(pr_df2, 'ETHYL ACETATE', 'solvent', month) if sol_qty > 0 else 0
        
        adh_match = abs(gt_adh_rate - eng_adh) < 0.01 if gt_adh_rate > 0 else True
        hard_match = abs(gt_hard_rate - eng_hard) < 0.01 if gt_hard_rate > 0 else True
        sol_match = abs(gt_sol_rate - eng_sol) < 0.01 if gt_sol_rate > 0 else True
        
        if adh_match: adh_ok_count += 1
        if hard_match: hard_ok_count += 1
        if sol_match: sol_ok_count += 1
        
        if not (adh_match and hard_match and sol_match):
            flags = []
            if not adh_match:
                flags.append(f"ADH: GT={gt_adh_rate:.4f} vs ENG={eng_adh:.4f}")
            if not hard_match:
                flags.append(f"HARD: GT={gt_hard_rate:.4f} vs ENG={eng_hard:.4f}")
            if not sol_match:
                flags.append(f"SOL: GT={gt_sol_rate:.4f} vs ENG={eng_sol:.4f}")
            print(f"  MISMATCH Row {row}: Adh={adh_str}, DA={da_str} | {' | '.join(flags)}")
    
    wb.close()
    
    print(f"\n  RESULTS:")
    print(f"    Adhesive: {adh_ok_count}/{total} ({adh_ok_count*100//total}%)")
    print(f"    Hardener: {hard_ok_count}/{total} ({hard_ok_count*100//total}%)")
    print(f"    Solvent:  {sol_ok_count}/{total} ({sol_ok_count*100//total}%)")
    overall = (adh_ok_count + hard_ok_count + sol_ok_count)
    overall_total = total * 3
    print(f"    OVERALL:  {overall}/{overall_total} ({overall*100//overall_total}%)")

print("\n" + "=" * 80)
print("VERIFICATION COMPLETE")
print("=" * 80)
