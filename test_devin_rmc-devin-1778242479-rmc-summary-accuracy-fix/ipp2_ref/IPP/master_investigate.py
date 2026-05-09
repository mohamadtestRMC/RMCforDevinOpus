"""
MASTER RULE INVESTIGATION
Investigate EVERY mismatch to extract the exact rules for 100% accuracy.
"""
import pandas as pd
import openpyxl
import sys
sys.path.insert(0, '.')
from engine.mrr_lookup import load_stores_recordings, lookup_mrr, lookup_mrr_with_qty
from engine.rate_lookup import (
    load_purchase_register, _find_col, _filter_by_month, _qty_weighted_rate,
    lookup_film_rate_weighted, lookup_material_rate_for_month, filter_mrr_by_pr
)

def sf(val):
    if val is None or val == '' or (isinstance(val, float) and pd.isna(val)):
        return 0.0
    try: return float(val)
    except: return 0.0

def ss(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ''
    s = str(val).strip()
    return '' if s.startswith('=') else s

print("=" * 100)
print("MASTER RULE INVESTIGATION - Finding rules for 100% accuracy")
print("=" * 100)

# ══════════════════════════════════════════════════════════════
# INVESTIGATION 1: Feb 2026 Film/Fresh rate mismatches (PET)
# ══════════════════════════════════════════════════════════════
print("\n" + "#" * 100)
print("# INVESTIGATION 1: Feb 2026 — PET Film/Fresh rate mismatches")
print("#" * 100)

T1 = "Template_Files"
gt1_wb = openpyxl.load_workbook(f"{T1}/Jobtrack Feb With MRR.xlsx", data_only=True)
gt1_ws = gt1_wb.active
stores1 = load_stores_recordings(f"{T1}/Stores Recordings.xlsx")
pr1 = load_purchase_register(f"{T1}/Purchase Register - 2021 - 2026 _Feb 26.xlsx")

# Rows 42, 43 — Fresh1_Rate PET 12mic 1197mm
# Row 45 — Film_Rate PET 12mic 1063mm 
# Row 46 — Film_Rate PET 12mic 913mm
# Row 54 — Fresh2_MR TPE 100mic 893mm

for row in [42, 43, 45, 46, 54]:
    process = ss(gt1_ws.cell(row=row, column=6).value)
    uid = ss(gt1_ws.cell(row=row, column=1).value)
    order = ss(gt1_ws.cell(row=row, column=11).value)
    
    print(f"\n{'='*80}")
    print(f"Row {row}: Process={process}, UID={uid}, Order={order}")
    
    if process.upper() == 'PRINTING':
        mat = ss(gt1_ws.cell(row=row, column=47).value)  # AU - Input Name
        size = sf(gt1_ws.cell(row=row, column=48).value)  # AV
        mic = sf(gt1_ws.cell(row=row, column=49).value)   # AW
        gt_mr = ss(gt1_ws.cell(row=row, column=54).value)  # BB
        gt_rate = sf(gt1_ws.cell(row=row, column=55).value)  # BC
        
        print(f"  Film: Mat={mat}, Size={size}, Mic={mic}")
        print(f"  GT: MR#={gt_mr}, Rate={gt_rate:.4f}")
        
        # What MRRs does engine find?
        mrr_qty = lookup_mrr_with_qty(stores1, mat, mic, size, order, 'PRINTING')
        if not mrr_qty:
            mrr_qty = lookup_mrr_with_qty(stores1, mat, mic, None, order, 'PRINTING')
        if not mrr_qty:
            mrr_qty = lookup_mrr_with_qty(stores1, mat, mic, None, order)
        print(f"  Engine MRRs: {mrr_qty}")
        
        # Filter by PR
        if mrr_qty:
            mrr_qty_filtered = filter_mrr_by_pr(pr1, mrr_qty, mat, size, mic)
            print(f"  After PR filter: {mrr_qty_filtered}")
            
            eng_rate = lookup_film_rate_weighted(pr1, mrr_qty_filtered, mat, size, mic)
            eng_rate2 = lookup_film_rate_weighted(pr1, mrr_qty_filtered, mat, None, mic)
            month_rate = lookup_material_rate_for_month(pr1, mat, mic, "2-2026")
            print(f"  Engine rates: with_size={eng_rate:.4f}, no_size={eng_rate2:.4f}, month_avg={month_rate:.4f}")
        
        # Look up GT MR# in PR to see what rate it gets
        if gt_mr:
            tracking_col = _find_col(pr1, 'tracking')
            mat_col = _find_col(pr1, 'material')
            rate_col2 = [c for c in pr1.columns if str(c).strip().lower() == 'rate']
            rate_col2 = rate_col2[0] if rate_col2 else 'Rate'
            amt_col = [c for c in pr1.columns if str(c).strip().lower() == 'amount']
            amt_col = amt_col[0] if amt_col else None
            qty_col = [c for c in pr1.columns if str(c).strip().lower() == 'actual quantity']
            qty_col = qty_col[0] if qty_col else None
            
            for mr_s in gt_mr.split('/'):
                mr_s = mr_s.strip()
                try:
                    mr_val = int(float(mr_s))
                    mask = pd.to_numeric(pr1[tracking_col], errors='coerce') == mr_val
                    rows = pr1[mask]
                    print(f"\n  PR entries for MR#={mr_val}: {len(rows)}")
                    for _, r in rows.iterrows():
                        print(f"    Mat={r.get(mat_col,'?')}, Size={r.get('Size','?')}, Mic={r.get('Mic','?')}, "
                              f"Rate={r.get(rate_col2,'?')}, Qty={r.get(qty_col,'?')}, Amt={r.get(amt_col,'?')}")
                except:
                    pass
    
    elif process.upper() == 'LAM':
        # Check Fresh1 and Fresh2
        for prefix, name_col, size_col, mic_col, mr_col, rate_col, val_col in [
            ('Fresh1', 71, 72, 73, 78, 79, 80),
            ('Fresh2', 81, 82, 83, 88, 89, 90),
        ]:
            mat = ss(gt1_ws.cell(row=row, column=name_col).value)
            size = sf(gt1_ws.cell(row=row, column=size_col).value)
            mic = sf(gt1_ws.cell(row=row, column=mic_col).value)
            gt_mr = ss(gt1_ws.cell(row=row, column=mr_col).value)
            gt_rate = sf(gt1_ws.cell(row=row, column=rate_col).value)
            
            if mat and gt_rate > 0:
                print(f"\n  {prefix}: Mat={mat}, Size={size}, Mic={mic}")
                print(f"  GT: MR#={gt_mr}, Rate={gt_rate:.4f}")
                
                mrr_qty = lookup_mrr_with_qty(stores1, mat, mic, size, order, 'LAMINATION')
                if not mrr_qty:
                    mrr_qty = lookup_mrr_with_qty(stores1, mat, mic, None, order, 'LAMINATION')
                if not mrr_qty:
                    mrr_qty = lookup_mrr_with_qty(stores1, mat, mic, None, order)
                print(f"  Engine MRRs: {mrr_qty}")
                
                if mrr_qty:
                    mrr_qty_f = filter_mrr_by_pr(pr1, mrr_qty, mat, size, mic)
                    eng_rate = lookup_film_rate_weighted(pr1, mrr_qty_f, mat, size, mic)
                    print(f"  Engine rate: {eng_rate:.4f}")
                
                # Check GT MR# in PR
                if gt_mr and gt_mr != 'INH':
                    tracking_col = _find_col(pr1, 'tracking')
                    rate_col3 = [c for c in pr1.columns if str(c).strip().lower() == 'rate']
                    rate_col3 = rate_col3[0] if rate_col3 else 'Rate'
                    mat_col2 = _find_col(pr1, 'material')
                    
                    for mr_s in gt_mr.split('/'):
                        try:
                            mr_val = int(float(mr_s.strip()))
                            mask = pd.to_numeric(pr1[tracking_col], errors='coerce') == mr_val
                            rows = pr1[mask]
                            print(f"  PR for MR#={mr_val}: {len(rows)} rows")
                            for _, r in rows.iterrows():
                                print(f"    Mat={r.get(mat_col2,'?')}, Size={r.get('Size','?')}, "
                                      f"Mic={r.get('Mic','?')}, Rate={r.get(rate_col3,'?')}")
                        except:
                            pass

gt1_wb.close()

# ══════════════════════════════════════════════════════════════
# INVESTIGATION 2: Nov 2025 — INH detection & NYLON miss
# ══════════════════════════════════════════════════════════════
print("\n\n" + "#" * 100)
print("# INVESTIGATION 2: Nov 2025 — INH detection, NYLON miss, TPE rates")
print("#" * 100)

T2 = "Template2"
gt2_wb = openpyxl.load_workbook(f"{T2}/Jobtrack With MRR.xlsx", data_only=True)
gt2_ws = gt2_wb.active
stores2 = load_stores_recordings(f"{T2}/Stores Recordings.xlsx")
pr2 = load_purchase_register(f"{T2}/Purchase Register - 2021 - 2025 _Nov.xlsx")

# Row 7: NYLON Film miss
print(f"\n{'='*80}")
print("Row 7: NYLON Film — MISS")
row = 7
mat = ss(gt2_ws.cell(row=row, column=47).value)
size = sf(gt2_ws.cell(row=row, column=48).value)
mic = sf(gt2_ws.cell(row=row, column=49).value)
order = ss(gt2_ws.cell(row=row, column=11).value)
gt_mr = ss(gt2_ws.cell(row=row, column=54).value)
gt_rate = sf(gt2_ws.cell(row=row, column=55).value)
print(f"  Mat={mat}, Size={size}, Mic={mic}, Order={order}")
print(f"  GT: MR#={gt_mr}, Rate={gt_rate:.4f}")

# Check stores for NYLON
mrr_qty = lookup_mrr_with_qty(stores2, mat, mic, size, order, 'PRINTING')
print(f"  Stores lookup (PRINTING, size={size}): {mrr_qty}")
mrr_qty = lookup_mrr_with_qty(stores2, mat, mic, None, order, 'PRINTING')
print(f"  Stores lookup (PRINTING, no size): {mrr_qty}")
mrr_qty = lookup_mrr_with_qty(stores2, mat, mic, None, order)
print(f"  Stores lookup (any process): {mrr_qty}")

# Check what categories contain "NYLON" in stores
cat_col = None
mat_col_s = None
for c in stores2.columns:
    cl = str(c).strip().lower()
    if 'categ' in cl or 'sub' in cl:
        cat_col = c
    if c == 'Main Group' or cl == 'main group':
        mat_col_s = c

print(f"\n  Stores columns for filtering: cat={cat_col}, main_group={mat_col_s}")

if cat_col:
    nylon_stores = stores2[stores2[cat_col].astype(str).str.upper().str.contains('NYLON')]
    print(f"  NYLON entries in stores: {len(nylon_stores)}")
    if len(nylon_stores) > 0:
        # Check if any match our order
        for c in stores2.columns:
            cl = str(c).strip().lower()
            if 'wo' in cl or 'order' in cl or 'intimation' in cl:
                order_vals = nylon_stores[c].astype(str).str.strip().unique()
                matching = [v for v in order_vals if order.upper() in v.upper() or v.upper() in order.upper()]
                if matching:
                    print(f"    Column '{c}' has matching orders: {matching}")

# Search directly for GT MR# 83005 in stores
print(f"\n  Direct search for MR#={gt_mr} in stores:")
for c in stores2.columns:
    cl = str(c).strip().lower()
    if 'mr' in cl or 'tracking' in cl:
        matches = stores2[stores2[c].astype(str).str.strip() == str(gt_mr)]
        if len(matches) > 0:
            print(f"    Found in column '{c}': {len(matches)} rows")
            for _, r in matches.head(3).iterrows():
                print(f"      {dict(r)}")

# Rows 17, 28, 50: Fresh1 INH detection (TPE/WPE)
for row in [17, 28, 50]:
    print(f"\n{'='*80}")
    print(f"Row {row}: Fresh1 — INH detection")
    
    mat = ss(gt2_ws.cell(row=row, column=71).value)  # BS - Fresh1 Name
    size = sf(gt2_ws.cell(row=row, column=72).value)  # BT
    mic = sf(gt2_ws.cell(row=row, column=73).value)   # BU
    order = ss(gt2_ws.cell(row=row, column=11).value)
    gt_mr = ss(gt2_ws.cell(row=row, column=78).value)  # BZ
    gt_rate = sf(gt2_ws.cell(row=row, column=79).value) # CA
    
    print(f"  Mat={mat}, Size={size}, Mic={mic}, Order={order}")
    print(f"  GT: MR#={gt_mr}, Rate={gt_rate:.4f}")
    
    # Check: is this INH? GT says INH
    print(f"  GT says MR#='{gt_mr}' — this means IN-HOUSE material")
    
    # What does the engine currently do?
    # The engine checks if input_name is in ('WPE', 'WLDPE', 'PTD WPE') for PRINTING
    # But for LAM Fresh materials, there's no INH check!
    print(f"  Engine issue: LAM Fresh does NOT have INH detection for TPE/WPE")
    
    # How does the GT determine INH?
    # TPE and WPE in LAM context = in-house produced films
    # The material is produced by the factory's own BFL (Blown Film) process
    
    # What rate does GT use for INH TPE?
    if gt_rate > 0:
        print(f"  GT INH rate: {gt_rate:.4f}")
        # Search PR for this rate
        tracking_col = _find_col(pr2, 'tracking')
        mat_col2 = _find_col(pr2, 'material')
        rate_col3 = [c for c in pr2.columns if str(c).strip().lower() == 'rate']
        rate_col3 = rate_col3[0] if rate_col3 else 'Rate'
        
        # Is this a Granules Recipe rate? Check order
        print(f"  Order: {order}")

# Row 23: Fresh2 INH (TPE)
print(f"\n{'='*80}")
print("Row 23: Fresh2 — INH detection")
row = 23
mat = ss(gt2_ws.cell(row=row, column=81).value)  # CC - Fresh2 Name
size = sf(gt2_ws.cell(row=row, column=82).value)  # CD
mic = sf(gt2_ws.cell(row=row, column=83).value)   # CE
gt_mr = ss(gt2_ws.cell(row=row, column=88).value)  # CJ
gt_rate = sf(gt2_ws.cell(row=row, column=89).value) # CK
print(f"  Mat={mat}, Size={size}, Mic={mic}")
print(f"  GT: MR#={gt_mr}, Rate={gt_rate:.4f}")

# Row 17: Fresh1 TPE rate wrong (0.9 vs 4.1951)
# Row 48: Fresh2 TPE rate wrong (0.9 vs 4.8708)
print(f"\n{'='*80}")
print("TPE RATE INVESTIGATION")
print("Row 17 Fresh1: TPE Mic=50, GT=4.1951, Eng=0.9000")
print("Row 48 Fresh2: TPE Mic=100, GT=4.8708, Eng=0.9000")

# Where does 0.9 come from? Check month-level lookup
month_rate_tpe50 = lookup_material_rate_for_month(pr2, 'TPE', 50, "11-2025")
month_rate_tpe100 = lookup_material_rate_for_month(pr2, 'TPE', 100, "11-2025")
print(f"\n  Month rate TPE mic=50: {month_rate_tpe50:.4f}")
print(f"  Month rate TPE mic=100: {month_rate_tpe100:.4f}")

# Check what TPE looks like in PR
cat_col_pr = _find_col(pr2, 'categery', 'category')
mat_col_pr = _find_col(pr2, 'material')
month_col_pr = None
for c in pr2.columns:
    if str(c).strip().lower() == 'month':
        month_col_pr = c

tpe_mask = pr2[mat_col_pr].astype(str).str.upper().str.strip().str.contains('TPE')
tpe_rows = pr2[tpe_mask]
print(f"\n  All TPE entries in PR: {len(tpe_rows)}")
for _, r in tpe_rows.iterrows():
    print(f"    Cat={r.get(cat_col_pr,'?')}, Mat={r.get(mat_col_pr,'?')}, "
          f"Size={r.get('Size','?')}, Mic={r.get('Mic','?')}, "
          f"Rate={r.get('Rate','?')}, Month={r.get(month_col_pr,'?')}")

# Where does GT 4.1951 and 4.8708 come from?
# These are INH materials — maybe from Granules Recipe?
print("\n  Checking Granules Recipe for TPE rates...")
try:
    gran = pd.read_excel(f"{T2}/Granules Recipe -Nov_2025.xlsx", sheet_name=0)
    print(f"  Granules Recipe columns: {list(gran.columns)[:10]}")
    print(f"  Granules Recipe rows: {len(gran)}")
    # Show first few rows
    for i, r in gran.head(10).iterrows():
        print(f"    {dict(r)}")
except Exception as e:
    print(f"  Error: {e}")

# Check what orders the INH rows have
print("\n\n  INH Rows and their orders:")
for row in [17, 28, 50]:
    order = ss(gt2_ws.cell(row=row, column=11).value)
    mat = ss(gt2_ws.cell(row=row, column=71).value)
    gt_rate = sf(gt2_ws.cell(row=row, column=79).value)
    print(f"  Row {row}: Order={order}, Mat={mat}, GT_Rate={gt_rate:.4f}")

# Row 48 Fresh2
order48 = ss(gt2_ws.cell(row=48, column=11).value)
mat48 = ss(gt2_ws.cell(row=48, column=81).value)
gt_rate48 = sf(gt2_ws.cell(row=48, column=89).value)
print(f"  Row 48: Order={order48}, Mat={mat48}, GT_Rate={gt_rate48:.4f}")

gt2_wb.close()

# ══════════════════════════════════════════════════════════════
# INVESTIGATION 3: Check the Feb data for same INH patterns
# ══════════════════════════════════════════════════════════════
print("\n\n" + "#" * 100)
print("# INVESTIGATION 3: Feb 2026 — Do INH patterns exist?")
print("#" * 100)

gt1_wb2 = openpyxl.load_workbook(f"{T1}/Jobtrack Feb With MRR.xlsx", data_only=True)
gt1_ws2 = gt1_wb2.active

print("\nAll LAM rows with Fresh materials in Feb 2026:")
for row in range(5, gt1_ws2.max_row + 1):
    process = ss(gt1_ws2.cell(row=row, column=6).value)
    if process.upper() != 'LAM':
        continue
    
    for prefix, name_col, mr_col, rate_col in [
        ('F1', 71, 78, 79),
        ('F2', 81, 88, 89),
    ]:
        mat = ss(gt1_ws2.cell(row=row, column=name_col).value)
        mr = ss(gt1_ws2.cell(row=row, column=mr_col).value)
        rate = sf(gt1_ws2.cell(row=row, column=rate_col).value)
        
        if mat and (mr == 'INH' or mat.upper() in ('TPE', 'WPE', 'WLDPE', 'PTD WPE')):
            order = ss(gt1_ws2.cell(row=row, column=11).value)
            mic = sf(gt1_ws2.cell(row=row, column=name_col+2).value)
            print(f"  Row {row} {prefix}: Mat={mat}, MR={mr}, Rate={rate:.4f}, Order={order}, Mic={mic}")

gt1_wb2.close()

print("\n" + "=" * 100)
print("INVESTIGATION COMPLETE")
print("=" * 100)
