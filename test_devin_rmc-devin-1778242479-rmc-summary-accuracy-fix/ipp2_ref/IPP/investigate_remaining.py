"""
DEEP INVESTIGATION: Find universal rules for ALL remaining mismatches.
Covers BOTH datasets to find ONE common rule set.
"""
import pandas as pd
import openpyxl
import sys
sys.path.insert(0, '.')
from engine.mrr_lookup import load_stores_recordings, lookup_mrr_with_qty
from engine.rate_lookup import load_purchase_register, _find_col, lookup_film_rate_weighted, filter_mrr_by_pr
from engine.supplier_rates import build_mrr_supplier_map, get_supplier_for_mrrs

def sf(v):
    try: return float(v) if v and not (isinstance(v,float) and pd.isna(v)) else 0.0
    except: return 0.0
def ss(v):
    if v is None or (isinstance(v,float) and pd.isna(v)): return ''
    s = str(v).strip()
    return '' if s.startswith('=') else s

COLS_F1 = {'name':71,'size':72,'mic':73,'mr':78,'rate':79,'val':80}
COLS_F2 = {'name':81,'size':82,'mic':83,'mr':88,'rate':89,'val':90}
COLS_FILM = {'name':47,'size':48,'mic':49,'mr':54,'rate':55,'val':56}

print("=" * 100)
print("INVESTIGATION: Common rules for 100% accuracy across both datasets")
print("=" * 100)

# ══════════════════════════════════════════════════════════════════
# PART 1: Analyze ALL TPE/WPE rows across BOTH datasets
# ══════════════════════════════════════════════════════════════════
print("\n" + "#" * 100)
print("# PART 1: TPE/WPE behavior across BOTH datasets")
print("#" * 100)

for ds_name, jt_path, stores_path, pr_path, gran_path in [
    ("Feb 2026", "Template_Files/Jobtrack Feb With MRR.xlsx",
     "Template_Files/Stores Recordings.xlsx",
     "Template_Files/Purchase Register - 2021 - 2026 _Feb 26.xlsx",
     "Template_Files/Granules Recipe - February 2026.xlsx"),
    ("Nov 2025", "Template2/Jobtrack With MRR.xlsx",
     "Template2/Stores Recordings.xlsx",
     "Template2/Purchase Register - 2021 - 2025 _Nov.xlsx",
     "Template2/Granules Recipe -Nov_2025.xlsx"),
]:
    print(f"\n{'='*80}")
    print(f"  {ds_name}")
    print(f"{'='*80}")
    
    wb = openpyxl.load_workbook(jt_path, data_only=True)
    ws = wb.active
    stores = load_stores_recordings(stores_path)
    pr = load_purchase_register(pr_path)
    supplier_map = build_mrr_supplier_map(stores)
    
    for row in range(5, ws.max_row + 1):
        process = ss(ws.cell(row=row, column=6).value).upper()
        if process != 'LAM': continue
        order = ss(ws.cell(row=row, column=11).value)
        
        for prefix, cols in [('F1', COLS_F1), ('F2', COLS_F2)]:
            mat = ss(ws.cell(row=row, column=cols['name']).value).upper()
            if mat not in ('TPE', 'WPE', 'WLDPE', 'PTD WPE'): continue
            
            size = sf(ws.cell(row=row, column=cols['size']).value)
            mic = sf(ws.cell(row=row, column=cols['mic']).value)
            gt_mr = ss(ws.cell(row=row, column=cols['mr']).value)
            gt_rate = sf(ws.cell(row=row, column=cols['rate']).value)
            
            # Check supplier
            mrr_qty = lookup_mrr_with_qty(stores, mat, mic, size, order, 'LAMINATION')
            if not mrr_qty:
                mrr_qty = lookup_mrr_with_qty(stores, mat, mic, None, order, 'LAMINATION')
            if not mrr_qty:
                mrr_qty = lookup_mrr_with_qty(stores, mat, mic, None, order)
            
            supplier = 'NONE'
            if mrr_qty:
                supplier = get_supplier_for_mrrs(supplier_map, list(mrr_qty.keys())) or 'UNKNOWN'
            
            is_inh = (gt_mr == 'INH')
            
            print(f"  Row {row:>3} {prefix}: Mat={mat:<5} Mic={mic:>5.0f} Size={size:>6.0f} "
                  f"Order={order:<7} GT_MR={gt_mr:<15} GT_Rate={gt_rate:>8.4f} "
                  f"Supplier={supplier:<12} INH={is_inh} "
                  f"Engine_MRRs={list(mrr_qty.keys()) if mrr_qty else '[]'}")
    
    wb.close()

# ══════════════════════════════════════════════════════════════════
# PART 2: Analyze the supplier for ALL MRRs of TPE/WPE
# ══════════════════════════════════════════════════════════════════
print("\n\n" + "#" * 100)
print("# PART 2: Supplier analysis for TPE/WPE MRRs")
print("#" * 100)

for ds_name, stores_path in [
    ("Feb 2026", "Template_Files/Stores Recordings.xlsx"),
    ("Nov 2025", "Template2/Stores Recordings.xlsx"),
]:
    print(f"\n{'='*80}")
    print(f"  {ds_name}")
    print(f"{'='*80}")
    
    stores = load_stores_recordings(stores_path)
    supplier_map = build_mrr_supplier_map(stores)
    
    # Find all TPE/WPE related columns
    cat_col = mat_col = mrr_col = supplier_col = None
    for c in stores.columns:
        cl = str(c).strip().lower()
        if 'categ' in cl or 'sub' in cl: cat_col = c
        if cl == 'main group' or cl == 'material' and not mat_col: mat_col = c
        if 'supplier' in cl: supplier_col = c
    
    if cat_col:
        tpe_mask = stores[cat_col].astype(str).str.upper().str.contains('TPE|WPE|WLDPE', na=False)
        tpe_rows = stores[tpe_mask]
        print(f"  TPE/WPE entries in stores: {len(tpe_rows)}")
        if supplier_col and len(tpe_rows) > 0:
            suppliers = tpe_rows[supplier_col].astype(str).str.strip().unique()
            print(f"  Suppliers: {suppliers}")

# ══════════════════════════════════════════════════════════════════
# PART 3: Check Granules Recipe structure in both datasets
# ══════════════════════════════════════════════════════════════════
print("\n\n" + "#" * 100)
print("# PART 3: Granules Recipe file structure")
print("#" * 100)

from engine.supplier_rates import load_granules_rates

for ds_name, gran_path in [
    ("Feb 2026", "Template_Files/Granules Recipe - February 2026.xlsx"),
    ("Nov 2025", "Template2/Granules Recipe -Nov_2025.xlsx"),
]:
    print(f"\n{'='*80}")
    print(f"  {ds_name}")
    print(f"{'='*80}")
    
    rates = load_granules_rates(gran_path)
    print(f"  Granules rates loaded: {len(rates)} entries")
    for wo, rate in sorted(rates.items())[:20]:
        print(f"    WO#{wo} = {rate:.4f}")

# ══════════════════════════════════════════════════════════════════
# PART 4: Check Feb PET Film/Fresh mismatches (rows 42,43,45,46,54)
# ══════════════════════════════════════════════════════════════════
print("\n\n" + "#" * 100)
print("# PART 4: Feb 2026 PET Film/Fresh mismatches — GT MRR selection")
print("#" * 100)

wb = openpyxl.load_workbook("Template_Files/Jobtrack Feb With MRR.xlsx", data_only=True)
ws = wb.active
stores = load_stores_recordings("Template_Files/Stores Recordings.xlsx")
pr = load_purchase_register("Template_Files/Purchase Register - 2021 - 2026 _Feb 26.xlsx")
supplier_map = build_mrr_supplier_map(stores)

tracking_col = _find_col(pr, 'tracking')
mat_col = _find_col(pr, 'material')
rate_col_name = [c for c in pr.columns if str(c).strip().lower() == 'rate']
rate_col_name = rate_col_name[0] if rate_col_name else 'Rate'
qty_col_name = [c for c in pr.columns if str(c).strip().lower() == 'actual quantity']
qty_col_name = qty_col_name[0] if qty_col_name else None

for row in [42, 43, 45, 46, 54]:
    process = ss(ws.cell(row=row, column=6).value).upper()
    order = ss(ws.cell(row=row, column=11).value)
    
    if process == 'PRINTING':
        cols = COLS_FILM
        ctx = 'Film'
    else:
        # Check which fresh column has data
        if ss(ws.cell(row=row, column=COLS_F2['name']).value):
            cols = COLS_F2
            ctx = 'Fresh2'
        else:
            cols = COLS_F1
            ctx = 'Fresh1'
    
    mat = ss(ws.cell(row=row, column=cols['name']).value)
    size = sf(ws.cell(row=row, column=cols['size']).value)
    mic = sf(ws.cell(row=row, column=cols['mic']).value)
    gt_mr = ss(ws.cell(row=row, column=cols['mr']).value)
    gt_rate = sf(ws.cell(row=row, column=cols['rate']).value)
    
    print(f"\n  Row {row} ({ctx}): Mat={mat}, Size={size}, Mic={mic}, Order={order}")
    print(f"  GT: MR#={gt_mr}, Rate={gt_rate:.4f}")
    
    # What engine finds
    proc = 'PRINTING' if process == 'PRINTING' else 'LAMINATION'
    mrr_qty = lookup_mrr_with_qty(stores, mat, mic, size, order, proc)
    if not mrr_qty:
        mrr_qty = lookup_mrr_with_qty(stores, mat, mic, None, order, proc)
    if not mrr_qty:
        mrr_qty = lookup_mrr_with_qty(stores, mat, mic, None, order)
    
    supplier = get_supplier_for_mrrs(supplier_map, list(mrr_qty.keys())) if mrr_qty else 'NONE'
    supplier = supplier or 'UNKNOWN'
    
    mrr_qty_f = filter_mrr_by_pr(pr, mrr_qty, mat, size, mic) if mrr_qty else {}
    eng_rate = lookup_film_rate_weighted(pr, mrr_qty_f, mat, size, mic) if mrr_qty_f else 0
    
    print(f"  Engine: MRRs={dict(mrr_qty_f)}, Rate={eng_rate:.4f}, Supplier={supplier}")
    
    # Show per-MRR rates from GT
    if gt_mr:
        gt_mrrs = [m.strip() for m in gt_mr.replace('/',',').split(',')]
        print(f"  GT MRRs: {gt_mrrs}")
        
        # Compute GT's weighted rate from only GT's MRRs
        gt_total_amt = 0
        gt_total_qty = 0
        for m in gt_mrrs:
            try:
                mv = int(float(m))
                mask = pd.to_numeric(pr[tracking_col], errors='coerce') == mv
                rows = pr[mask]
                if mat_col:
                    rows = rows[rows[mat_col].astype(str).str.upper().str.strip() == mat.upper()]
                for _, r in rows.iterrows():
                    pr_size = sf(r.get('Size', 0))
                    if abs(pr_size - size) <= 5 or size == 0:
                        rate_v = sf(r.get(rate_col_name, 0))
                        qty_v = sf(r.get(qty_col_name, 0))
                        print(f"    MR#{mv}: Size={pr_size}, Rate={rate_v:.4f}, Qty={qty_v:.1f}")
                        gt_total_amt += sf(r.get('Amount', 0))
                        gt_total_qty += qty_v
            except: pass
        
        if gt_total_qty > 0:
            computed_rate = gt_total_amt / gt_total_qty
            print(f"  Computed from GT MRRs: {computed_rate:.4f} (vs GT={gt_rate:.4f})")

# Row 54 special: TPE Fresh2 
print(f"\n  --- Row 54: TPE Fresh2 (MR# mismatch) ---")
row = 54
mat = ss(ws.cell(row=row, column=COLS_F2['name']).value)
size = sf(ws.cell(row=row, column=COLS_F2['size']).value)
mic = sf(ws.cell(row=row, column=COLS_F2['mic']).value)
order = ss(ws.cell(row=row, column=11).value)
gt_mr = ss(ws.cell(row=row, column=COLS_F2['mr']).value)
gt_rate = sf(ws.cell(row=row, column=COLS_F2['rate']).value)
print(f"  Mat={mat}, Size={size}, Mic={mic}, Order={order}")
print(f"  GT: MR#={gt_mr}, Rate={gt_rate:.4f}")

# What supplier is GT MR# 84080?
supplier_84080 = get_supplier_for_mrrs(supplier_map, [84080])
print(f"  Supplier for MR#84080: {supplier_84080}")

# What supplier are engine MRRs?
mrr_qty = lookup_mrr_with_qty(stores, mat, mic, size, order, 'LAMINATION')
if not mrr_qty:
    mrr_qty = lookup_mrr_with_qty(stores, mat, mic, None, order, 'LAMINATION')
if not mrr_qty:
    mrr_qty = lookup_mrr_with_qty(stores, mat, mic, None, order)
for m, q in mrr_qty.items():
    s = get_supplier_for_mrrs(supplier_map, [m])
    print(f"  MR#{m}: Qty={q}, Supplier={s}")

wb.close()

# ══════════════════════════════════════════════════════════════════
# PART 5: NYLON issue in Nov 2025
# ══════════════════════════════════════════════════════════════════
print("\n\n" + "#" * 100)
print("# PART 5: NYLON miss in Nov 2025")
print("#" * 100)

stores2 = load_stores_recordings("Template2/Stores Recordings.xlsx")
# Try different lookup strategies
for proc in ['PRINTING', 'LAMINATION', None]:
    mrr = lookup_mrr_with_qty(stores2, 'NYLON', 15, 1005, 'N00694', proc)
    print(f"  NYLON lookup (process={proc}): {mrr}")
    mrr = lookup_mrr_with_qty(stores2, 'NYLON', 15, None, 'N00694', proc)
    print(f"  NYLON lookup no size (process={proc}): {mrr}")

# Check stores columns
print(f"\n  Stores columns: {list(stores2.columns)[:15]}")

# Look for NYLON directly
for c in stores2.columns:
    vals = stores2[c].astype(str).str.upper()
    if vals.str.contains('NYLON').any():
        nylon_count = vals.str.contains('NYLON').sum()
        print(f"  Column '{c}' has {nylon_count} NYLON entries")
        if nylon_count <= 5:
            nylon_rows = stores2[vals.str.contains('NYLON')]
            for _, r in nylon_rows.iterrows():
                print(f"    {dict(list(r.items())[:10])}")

# Check for MR# 83005
for c in stores2.columns:
    cl = str(c).strip().lower()
    if 'tracking' in cl or 'mr' in cl or 'receipt' in cl:
        matches = stores2[stores2[c].astype(str).str.strip() == '83005']
        if len(matches) > 0:
            print(f"\n  MR#83005 found in column '{c}': {len(matches)} rows")

print("\n" + "=" * 100)
print("INVESTIGATION COMPLETE")
print("=" * 100)
