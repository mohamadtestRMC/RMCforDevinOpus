"""
REVERSE ENGINEERING: Trace every ground truth rate back to its source file.
For each row in Jobtrack With MRR, determine:
1. Which MRR was used
2. What supplier that MRR comes from (Stores Recordings col 6)
3. If Bandera/CYM -> check Granules Recipe for matching WO# rate
4. If Mega Pack -> check MEGA PACK.xlsx for matching rate
5. Otherwise -> check Purchase Register rate
"""
import openpyxl
import pandas as pd
from collections import defaultdict

BASE = "Template_Files"

# ═══════════════════════════════════════════════════════════
# LOAD ALL DATA SOURCES
# ═══════════════════════════════════════════════════════════

# 1. Stores Recordings — build MRR→Supplier map
print("Loading Stores Recordings...")
wb_st = openpyxl.load_workbook(f"{BASE}/Stores Recordings.xlsx", data_only=True)
ws_st = wb_st.active

mrr_supplier = {}  # mrr_number -> supplier name
mrr_wo = {}  # mrr_number -> WO#
for r in range(3, ws_st.max_row + 1):
    mrr = ws_st.cell(row=r, column=16).value  # MRR No
    supplier = str(ws_st.cell(row=r, column=6).value or '').strip()
    wo = ws_st.cell(row=r, column=3).value  # WO NO
    if mrr and supplier:
        try:
            mrr_num = int(float(mrr))
            mrr_supplier[mrr_num] = supplier.upper()
            if wo:
                mrr_wo[mrr_num] = str(wo).strip()
        except (ValueError, TypeError):
            pass
wb_st.close()
print(f"  Loaded {len(mrr_supplier)} MRR->Supplier mappings")

# 2. Granules Recipe — build WO#→Rate AED map
print("Loading Granules Recipe...")
wb_gr = openpyxl.load_workbook(f"{BASE}/Granules Recipe - February 2026.xlsx", data_only=True)

granules_rates = {}  # wo_number -> rate_aed (from first/most recent sheet)
# Use the first sheet (most recent month = February 2026)
ws_gr = wb_gr[wb_gr.sheetnames[0]]
print(f"  Using sheet: '{wb_gr.sheetnames[0]}' ({ws_gr.max_row} rows x {ws_gr.max_column} cols)")

# Find header row and Rate AED column
header_row_gr = None
wo_col_gr = None
rate_col_gr = None
for r in range(1, 12):
    for c in range(1, 10):
        v = ws_gr.cell(row=r, column=c).value
        if v and 'WO' in str(v).upper():
            header_row_gr = r
            wo_col_gr = c
            break
    if header_row_gr:
        break

# Find "Rate AED" column
for c in range(1, ws_gr.max_column + 1):
    for r in [header_row_gr, max(1, header_row_gr - 2), 5]:
        v = ws_gr.cell(row=r, column=c).value
        if v and isinstance(v, str) and 'rate' in v.lower() and 'aed' in v.lower():
            rate_col_gr = c
            break
    if rate_col_gr:
        break

print(f"  Header row: {header_row_gr}, WO# col: {wo_col_gr}, Rate AED col: {rate_col_gr}")

# Build WO# → Rate map
if header_row_gr and wo_col_gr and rate_col_gr:
    for r in range(header_row_gr + 1, ws_gr.max_row + 1):
        wo = ws_gr.cell(row=r, column=wo_col_gr).value
        rate = ws_gr.cell(row=r, column=rate_col_gr).value
        mat = ws_gr.cell(row=r, column=wo_col_gr + 7).value  # Material column
        if wo and rate and isinstance(rate, (int, float)):
            granules_rates[str(wo).strip().upper()] = {
                'rate': rate, 'material': mat, 'row': r
            }
    print(f"  Loaded {len(granules_rates)} WO#->Rate mappings")
    for wo, info in granules_rates.items():
        print(f"    WO={wo}: Rate={info['rate']:.6f}, Material={info['material']}")

wb_gr.close()

# 3. MEGA PACK rates
print("Loading MEGA PACK.xlsx...")
wb_mp = openpyxl.load_workbook(f"{BASE}/MEGA PACK.xlsx", data_only=True)
ws_mp = wb_mp.active
mega_rates = {}  # month_label -> {TPE: rate, WPE: rate}
for r in range(5, ws_mp.max_row + 1):
    month_label = ws_mp.cell(row=r, column=4).value  # e.g. "Feb/26"
    date_val = ws_mp.cell(row=r, column=2).value
    tpe_raw = ws_mp.cell(row=r, column=5).value
    wpe_raw = ws_mp.cell(row=r, column=6).value
    tpe_conv = ws_mp.cell(row=r, column=7).value  # TPE with conversion
    wpe_conv = ws_mp.cell(row=r, column=8).value  # WPE with conversion
    if month_label:
        mega_rates[str(month_label).strip()] = {
            'TPE_raw': tpe_raw, 'WPE_raw': wpe_raw,
            'TPE_conv': tpe_conv, 'WPE_conv': wpe_conv,
            'date': date_val
        }
        print(f"  {month_label}: TPE_raw={tpe_raw}, WPE_raw={wpe_raw}, "
              f"TPE_conv={tpe_conv}, WPE_conv={wpe_conv}")
wb_mp.close()

# ═══════════════════════════════════════════════════════════
# TRACE EVERY GROUND TRUTH RATE
# ═══════════════════════════════════════════════════════════
print("\n" + "=" * 80)
print("TRACING EVERY GROUND TRUTH RATE TO ITS SOURCE")
print("=" * 80)

wb_gt = openpyxl.load_workbook(f"{BASE}/Jobtrack Feb With MRR.xlsx", data_only=True)
ws_gt = wb_gt.active

# Column indices
COLS = {
    'UID': 1, 'Process': 6, 'Order_No': 11,
    'Input_Name': 47, 'Film_MR': 54, 'Film_Rate': 55, 'Film_Value': 56,
    'Fresh1_Name': 71, 'Fresh1_MR': 78, 'Fresh1_Rate': 79,
    'Fresh2_Name': 81, 'Fresh2_MR': 88, 'Fresh2_Rate': 89,
}

for r in range(5, ws_gt.max_row + 1):
    uid = ws_gt.cell(row=r, column=COLS['UID']).value
    if not uid:
        continue
    
    order = ws_gt.cell(row=r, column=COLS['Order_No']).value
    process = ws_gt.cell(row=r, column=COLS['Process']).value
    
    # Check each material type
    checks = []
    if process and 'PRINT' in str(process).upper():
        film_mr = ws_gt.cell(row=r, column=COLS['Film_MR']).value
        film_rate = ws_gt.cell(row=r, column=COLS['Film_Rate']).value
        input_name = ws_gt.cell(row=r, column=COLS['Input_Name']).value
        if film_mr and film_rate:
            checks.append(('Film', film_mr, film_rate, input_name))
    
    if process and 'LAM' in str(process).upper():
        for prefix, mr_col, rate_col, name_col in [
            ('Fresh1', 'Fresh1_MR', 'Fresh1_Rate', 'Fresh1_Name'),
            ('Fresh2', 'Fresh2_MR', 'Fresh2_Rate', 'Fresh2_Name'),
        ]:
            mr = ws_gt.cell(row=r, column=COLS[mr_col]).value
            rate = ws_gt.cell(row=r, column=COLS[rate_col]).value
            name = ws_gt.cell(row=r, column=COLS[name_col]).value
            if mr and rate:
                checks.append((prefix, mr, rate, name))
    
    for mat_type, mr_val, gt_rate, mat_name in checks:
        # Parse MR numbers
        if str(mr_val) == 'INH':
            # INH = in-house, check if order is in Granules Recipe
            order_str = str(order).strip().upper() if order else ''
            if order_str in granules_rates:
                gr_rate = granules_rates[order_str]['rate']
                match = abs(gt_rate - gr_rate) < 0.001
                if match:
                    print(f"  Row {r}: {mat_type} INH, Order={order}, Rate={gt_rate:.6f} "
                          f"← GRANULES RECIPE ({gr_rate:.6f}) ✓")
                else:
                    print(f"  Row {r}: {mat_type} INH, Order={order}, Rate={gt_rate:.6f} "
                          f"vs Granules={gr_rate:.6f} MISMATCH!")
            continue
        
        mr_strs = str(mr_val).split('/')
        suppliers = set()
        for ms in mr_strs:
            try:
                mrr_num = int(float(ms.strip()))
                sup = mrr_supplier.get(mrr_num, 'UNKNOWN')
                suppliers.add(sup)
            except (ValueError, TypeError):
                pass
        
        # Check if any MRR is from a special supplier
        special = [s for s in suppliers if any(k in s for k in ['BANDERA', 'CYM', 'MEGA'])]
        
        if special:
            supplier = special[0]
            if 'MEGA' in supplier:
                # Check MEGA PACK rates
                for label, rates in mega_rates.items():
                    mat_upper = str(mat_name).strip().upper() if mat_name else ''
                    if 'TPE' in mat_upper:
                        mp_rate = rates.get('TPE_conv')
                    elif 'WPE' in mat_upper or 'WLDPE' in mat_upper:
                        mp_rate = rates.get('WPE_conv')
                    else:
                        mp_rate = rates.get('TPE_conv')  # default to TPE
                    
                    if mp_rate and abs(gt_rate - mp_rate) < 0.001:
                        print(f"  Row {r}: {mat_type} MR={mr_val}, Order={order}, Material={mat_name}, "
                              f"Supplier=MEGA PACK, Rate={gt_rate:.6f} "
                              f"← MEGA PACK {label} ({mp_rate:.6f}) ✓")
                        break
                else:
                    print(f"  Row {r}: {mat_type} MR={mr_val}, Order={order}, Material={mat_name}, "
                          f"Supplier=MEGA PACK, Rate={gt_rate:.6f} — NOT FOUND IN MEGA PACK!")
            
            elif 'BANDERA' in supplier or 'CYM' in supplier:
                order_str = str(order).strip().upper() if order else ''
                if order_str in granules_rates:
                    gr_rate = granules_rates[order_str]['rate']
                    match = abs(gt_rate - gr_rate) < 0.001
                    status = 'MATCH' if match else 'MISMATCH!'
                    print(f"  Row {r}: {mat_type} MR={mr_val}, Order={order}, Material={mat_name}, "
                          f"Supplier={supplier}, Rate={gt_rate:.6f} "
                          f"← GRANULES RECIPE ({gr_rate:.6f}) {status}")
                else:
                    print(f"  Row {r}: {mat_type} MR={mr_val}, Order={order}, Material={mat_name}, "
                          f"Supplier={supplier}, Rate={gt_rate:.6f} — WO# NOT in Granules Recipe!")

wb_gt.close()

# ═══════════════════════════════════════════════════════════
# NOW CHECK: Does every "other" supplier row match PR rate?
# ═══════════════════════════════════════════════════════════
print("\n" + "=" * 80)
print("ALL REMAINING MISMATCHES (non-special suppliers)")
print("=" * 80)

# Reload ground truth and engine results
import io, json
from engine.fill_jobtrack import fill_jobtrack

with open(f"{BASE}/Jobtrack Feb Without MRR.xlsx", "rb") as f:
    jt = io.BytesIO(f.read())
with open(f"{BASE}/Stores Recordings.xlsx", "rb") as f:
    stores = io.BytesIO(f.read())
with open(f"{BASE}/Purchase Register - 2021 - 2026 _Feb 26.xlsx", "rb") as f:
    pr = io.BytesIO(f.read())

output_bytes, results_log, fill_stats = fill_jobtrack(jt, stores, pr)

# Load the test results
files = [f for f in __import__('os').listdir('.') if f.startswith('test_results_')]
files.sort()
if files:
    with open(files[-1]) as f:
        test_data = json.load(f)
    
    fails = [d for d in test_data['details'] if d[0] == 'FAIL' and 'Rate' in d[1]]
    for status, name, detail in fails:
        print(f"  {name}: {detail}")
        # Extract row number
        row_match = name.split('Row ')[1].split(' ')[0] if 'Row' in name else None
        if row_match:
            row_num = int(row_match)
            # Check supplier for this row's MRRs
            for entry in results_log:
                if entry.get('row') == row_num:
                    detail_str = entry.get('detail', '')
                    if 'MR#' in detail_str:
                        mr_part = detail_str.split('MR#=')[1].split(',')[0] if 'MR#=' in detail_str else ''
                        mrrs = mr_part.split('/')
                        for m in mrrs:
                            try:
                                mn = int(float(m.strip()))
                                sup = mrr_supplier.get(mn, 'N/A')
                                print(f"    MRR {mn} -> Supplier: {sup}")
                            except:
                                pass
