"""
Fast Excel reader utilities. Uses openpyxl iter_rows for speed.
"""
from __future__ import annotations

import io
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook


def safe_float(v) -> float:
    if v is None:
        return 0.0
    try:
        f = float(v)
        return 0.0 if np.isnan(f) or np.isinf(f) else f
    except (ValueError, TypeError):
        return 0.0


def safe_str(v) -> str:
    return str(v).strip() if v is not None else ""


def open_workbook(source: Any, data_only: bool = True, read_only: bool = True):
    """Open a workbook from Path or BytesIO."""
    if isinstance(source, (str, Path)):
        return load_workbook(str(source), read_only=read_only, data_only=data_only)
    buf = source
    if hasattr(buf, "seek"):
        buf.seek(0)
    return load_workbook(buf, read_only=read_only, data_only=data_only)


_GHOST_DIM_THRESHOLD = 100_000


def read_sheet_fast(
    wb, sheet_name: str, header_row: int = 1, max_row: Optional[int] = None
) -> Tuple[List[str], List[tuple]]:
    """Read sheet using fast iter_rows. Returns (headers, data_rows).

    When ``max_row`` is None (default), the sheet's actual ``ws.max_row`` is used so
    no rows are silently truncated when monthly inputs grow beyond hardcoded caps.
    Callers may still pass an explicit ``max_row`` to bound very large sheets, in
    which case ``min(max_row, ws.max_row)`` is honoured.

    Ghost-dimension protection: some Excel files report a worksheet ``<dimension>``
    of e.g. ``A1:L1048576`` even when actual data only spans a few hundred rows
    (Feb 2026 OPN_WIP / CLS_WIP do this). To avoid iterating millions of empty
    rows, when ``ws.max_row`` exceeds 100k we probe forward from ``header_row`` and
    stop after 200 consecutive empty rows, returning only the populated range.
    """
    if sheet_name not in wb.sheetnames:
        return [], []
    ws = wb[sheet_name]
    actual_max = ws.max_row or header_row
    if actual_max > _GHOST_DIM_THRESHOLD and max_row is None:
        # Ghost dimension — find real last data row by scanning until we see
        # 200 consecutive empty rows, which signals end of populated data.
        last_data_row = header_row
        empty_streak = 0
        for r_idx, row in enumerate(
            ws.iter_rows(min_row=header_row, values_only=True), start=header_row
        ):
            if any(c is not None for c in row):
                last_data_row = r_idx
                empty_streak = 0
            else:
                empty_streak += 1
                if empty_streak >= 200:
                    break
        effective_max = max(last_data_row, header_row)
    else:
        effective_max = actual_max if max_row is None else min(max_row, actual_max)
    if effective_max < header_row:
        effective_max = header_row
    all_rows = list(ws.iter_rows(
        min_row=header_row, max_row=effective_max, values_only=True
    ))
    if not all_rows:
        return [], []
    raw_headers = all_rows[0]
    headers: List[str] = []
    seen: Dict[str, int] = {}
    for h in raw_headers:
        name = str(h).strip().replace("\n", " ") if h else "_blank"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        headers.append(name)
    data_rows = [r for r in all_rows[1:] if any(c is not None for c in r)]
    return headers, data_rows


def sheet_to_dataframe(
    source: Any,
    sheet_name: str,
    header_row: int = 1,
    max_row: int = 5000,
) -> pd.DataFrame:
    """Read a sheet into a DataFrame with deduped headers."""
    wb = open_workbook(source, data_only=True, read_only=True)
    try:
        headers, rows = read_sheet_fast(wb, sheet_name, header_row, max_row)
        if not headers:
            return pd.DataFrame()
        df = pd.DataFrame(rows, columns=headers)
        return df
    finally:
        wb.close()


class OrderIndex:
    """Fast order-based SUMIF/VLOOKUP lookup built from rows + column headers."""

    def __init__(self, headers: List[str], rows: List[tuple], order_col_idx: int):
        self.headers = headers
        self.all_rows = rows
        self._col_map = {h: i for i, h in enumerate(headers)}
        self._by_order: Dict[str, List[tuple]] = {}
        for row in rows:
            if order_col_idx < len(row):
                o = safe_str(row[order_col_idx])
                if o:
                    self._by_order.setdefault(o, []).append(row)

    def _ci(self, col_name: str) -> int:
        return self._col_map.get(col_name, -1)

    def sumif(self, order: str, col_name: str) -> float:
        ci = self._ci(col_name)
        if ci < 0:
            return 0.0
        total = 0.0
        for row in self._by_order.get(order, []):
            if ci < len(row):
                total += safe_float(row[ci])
        return total

    def vlookup(self, order: str, col_name: str) -> float:
        ci = self._ci(col_name)
        if ci < 0:
            return 0.0
        entries = self._by_order.get(order, [])
        if entries and ci < len(entries[0]):
            return safe_float(entries[0][ci])
        return 0.0

    def vlookup_str(self, order: str, col_name: str) -> str:
        ci = self._ci(col_name)
        if ci < 0:
            return ""
        entries = self._by_order.get(order, [])
        if entries and ci < len(entries[0]):
            return safe_str(entries[0][ci])
        return ""

    def orders(self) -> List[str]:
        return list(self._by_order.keys())

    @classmethod
    def from_sheet(
        cls, wb, sheet_name: str, header_row: int, max_row: Optional[int], order_col: str
    ) -> Optional["OrderIndex"]:
        """Build an OrderIndex from a sheet. Pass ``max_row=None`` to use the sheet's
        actual ``ws.max_row`` (avoids silent truncation when monthly inputs grow)."""
        if sheet_name not in wb.sheetnames:
            return None
        headers, rows = read_sheet_fast(wb, sheet_name, header_row, max_row)
        if not headers:
            return None
        oci = headers.index(order_col) if order_col in headers else 0
        return cls(headers, rows, oci)
