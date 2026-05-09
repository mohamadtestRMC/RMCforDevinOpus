"""
Fast test: skip fill_jobtrack, read Jobtrack from filled Base RMC directly.
This tests the CORE computation logic (process sheet building + RMC summary).
"""
import sys, logging, time, io
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent))
logging.basicConfig(level=logging.INFO, format='%(message)s')

from rmc_engine.data_reader import open_workbook, read_sheet_fast
from rmc_engine.jobtrack_processor import build_all_from_jobtrack
from rmc_engine.rmc_compute import compute_rmc_summary, validate_rmc
from rmc_engine.process_builder import build_indexes_from_filled_reference

BASE = Path(r"c:\Users\mohamad.al\OneDrive - AL KHAYYAT INVESTMENTS\Desktop\IPP\Files_need_to_study")
UNFILLED = BASE / "Unfilled"
FILLED = BASE / "Filled_Output"

FILLED_RMC = FILLED / "1 Base RMC _ 2026 February.xlsx"
UNFILLED_RMC = UNFILLED / "1 Base RMC _ 2026 February.xlsx"
INK_FILE = UNFILLED / "5 Ink Consumption February 2026.xlsx"
COMP_FILE = UNFILLED / "12 Components Consumptions Dispensed Details.xlsx"

t0 = time.time()

# Step 1: Extract Jobtrack bytes from the FILLED file (simulating enriched output)
print("Step 1: Extracting enriched Jobtrack from filled file...")
from openpyxl import load_workbook
wb = load_workbook(str(FILLED_RMC), read_only=True, data_only=True)
ws = wb["Jobtrack"]
buf = io.BytesIO()

from openpyxl import Workbook
out_wb = Workbook()
out_ws = out_wb.active
out_ws.title = "Job Track"
for row in ws.iter_rows(values_only=False):
    out_ws.append([c.value for c in row])
wb.close()
out_wb.save(buf)
jt_bytes = buf.getvalue()
print(f"  Extracted Jobtrack: {len(jt_bytes)} bytes in {time.time()-t0:.1f}s")

# Step 2: Build from Jobtrack with all fixes
print("\nStep 2: Building process sheets from Jobtrack...")
t1 = time.time()
idx, rmc_orders, offsets, transfers, other_film, combined = build_all_from_jobtrack(
    jt_bytes,
    opening_wip_source=None,
    closing_wip_source=None,
    ink_consumption_source=INK_FILE,
    components_source=COMP_FILE,
    unfilled_rmc_template=UNFILLED_RMC,
    prev_month_rmc=FILLED_RMC,
)
print(f"  Built in {time.time()-t1:.1f}s")
print(f"  Indexes: {list(idx.keys())}")
print(f"  Orders: {len(rmc_orders)}")
print(f"  Transfers: {len(transfers)}")
print(f"  Other Film: {len(other_film)}")
print(f"  Combined: {len(combined)}")

# Step 3: Compute RMC summary
print("\nStep 3: Computing RMC summary...")
rmc_rows = compute_rmc_summary(
    rmc_orders, idx, offsets, transfers, other_film, combined,
)
print(f"  Computed {len(rmc_rows)} rows")

# Step 4: Validate against reference
print("\nStep 4: Validating against filled reference...")
_, rmc_ref_rows, offsets_ref = build_indexes_from_filled_reference(FILLED_RMC)
val = validate_rmc(rmc_ref_rows, rmc_rows)

print(f"\n{'='*60}")
print(f"ACCURACY: {val['accuracy_pct']}%")
print(f"Exact: {val['exact_matches']} / {val['total_checks']}")
print(f"Close (<1): {val['close_lt1']}")
print(f"Mismatches (>1): {val['mismatches_gt1']}")
print(f"Total time: {time.time()-t0:.1f}s")

if val['top_mismatches']:
    print(f"\nTop 15 mismatches:")
    for m in val['top_mismatches'][:15]:
        print(f"  {m['order']:10s} | {m['col']:35s} | comp={m['computed']:12,.2f} | ref={m['reference']:12,.2f} | diff={m['diff']:+12,.2f}")

    by_col = {}
    for m in val['top_mismatches']:
        c = m['col']
        by_col[c] = by_col.get(c, 0) + 1
    print(f"\nMismatches by column:")
    for c, cnt in sorted(by_col.items(), key=lambda x: -x[1]):
        print(f"  {c:40s}: {cnt}")
