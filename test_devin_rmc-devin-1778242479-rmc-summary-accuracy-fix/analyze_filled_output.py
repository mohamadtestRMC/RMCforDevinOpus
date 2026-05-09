import openpyxl
import os

files = [
    r"c:\Users\mohamad.al\OneDrive - AL KHAYYAT INVESTMENTS\Desktop\IPP\Files_need_to_study\Filled_Output\4 Granules Recipe - February 2026.xlsx",
    r"c:\Users\mohamad.al\OneDrive - AL KHAYYAT INVESTMENTS\Desktop\IPP\Files_need_to_study\Filled_Output\5 Ink Consumption February 2026.xlsx",
    r"c:\Users\mohamad.al\OneDrive - AL KHAYYAT INVESTMENTS\Desktop\IPP\Files_need_to_study\Filled_Output\Base RMC Documents.xlsx",
    r"c:\Users\mohamad.al\OneDrive - AL KHAYYAT INVESTMENTS\Desktop\IPP\Files_need_to_study\Filled_Output\Material Names.xlsx",
]

for fpath in files:
    fname = os.path.basename(fpath)
    print("=" * 120)
    print(f"FILE: {fname}")
    print("=" * 120)

    if not os.path.exists(fpath):
        print(f"  *** FILE NOT FOUND ***\n")
        continue

    # Load with data_only=False to see formulas
    wb_formula = openpyxl.load_workbook(fpath, data_only=False)
    # Load with data_only=True to see cached values
    wb_data = openpyxl.load_workbook(fpath, data_only=True)

    print(f"Sheet names: {wb_formula.sheetnames}\n")

    for sheet_name in wb_formula.sheetnames:
        ws_f = wb_formula[sheet_name]
        ws_d = wb_data[sheet_name]

        print("-" * 100)
        print(f"  SHEET: '{sheet_name}'")
        print(f"  Dimensions: {ws_f.dimensions}")
        print(f"  Max row: {ws_f.max_row}, Max col: {ws_f.max_column}")
        print(f"  Merged cells: {list(ws_f.merged_cells.ranges)[:20]}")
        print()

        # Show headers (first 3 rows to capture multi-row headers)
        print(f"  --- HEADER ROWS (first 3 rows) ---")
        for row_idx in range(1, min(4, ws_f.max_row + 1)):
            row_vals = []
            for col_idx in range(1, min(ws_f.max_column + 1, 30)):
                cell = ws_d.cell(row=row_idx, column=col_idx)
                val = cell.value
                if val is not None:
                    row_vals.append(f"[{col_idx}]{val}")
            if row_vals:
                print(f"    Row {row_idx}: {' | '.join(row_vals)}")
        print()

        # Show first 10 data rows (rows 4-13)
        print(f"  --- DATA ROWS (rows 4-13) ---")
        for row_idx in range(4, min(14, ws_f.max_row + 1)):
            row_vals = []
            for col_idx in range(1, min(ws_f.max_column + 1, 30)):
                cell = ws_d.cell(row=row_idx, column=col_idx)
                val = cell.value
                if val is not None:
                    row_vals.append(f"[{col_idx}]{val}")
            if row_vals:
                print(f"    Row {row_idx}: {' | '.join(row_vals)}")
        print()

        # Show last 5 rows (might have totals)
        if ws_f.max_row > 13:
            print(f"  --- LAST 5 ROWS (rows {max(14, ws_f.max_row-4)}-{ws_f.max_row}) ---")
            for row_idx in range(max(14, ws_f.max_row - 4), ws_f.max_row + 1):
                row_vals = []
                for col_idx in range(1, min(ws_f.max_column + 1, 30)):
                    cell = ws_d.cell(row=row_idx, column=col_idx)
                    val = cell.value
                    if val is not None:
                        row_vals.append(f"[{col_idx}]{val}")
                if row_vals:
                    print(f"    Row {row_idx}: {' | '.join(row_vals)}")
            print()

        # Find ALL formulas
        formulas = []
        for row in ws_f.iter_rows(min_row=1, max_row=ws_f.max_row,
                                   min_col=1, max_col=ws_f.max_column):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append(f"    {cell.coordinate}: {cell.value}")

        if formulas:
            print(f"  --- FORMULAS ({len(formulas)} found) ---")
            for f in formulas[:50]:
                print(f)
            if len(formulas) > 50:
                print(f"    ... and {len(formulas) - 50} more formulas")
        else:
            print(f"  --- NO FORMULAS FOUND ---")
        print()

    wb_formula.close()
    wb_data.close()
    print()

print("=" * 120)
print("ANALYSIS COMPLETE")
print("=" * 120)
