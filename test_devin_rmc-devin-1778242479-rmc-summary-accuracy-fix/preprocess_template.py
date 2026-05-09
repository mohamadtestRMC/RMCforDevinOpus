"""Preprocess the template to strip empty rows from sheets with 1M rows.
This makes openpyxl loading fast (from 15+ min to ~30 sec)."""
import openpyxl, os, time, shutil

base = r'c:\Users\mohamad.al\OneDrive - AL KHAYYAT INVESTMENTS\Desktop\IPP'
src = os.path.join(base, 'Files_need_to_study', 'Unfilled', '1 Base RMC _ 2026 February.xlsx')
dst = os.path.join(base, 'output', 'template_stripped.xlsx')
os.makedirs(os.path.dirname(dst), exist_ok=True)

# First copy the file
shutil.copy2(src, dst)

print(f"Opening template... ({os.path.getsize(dst):,} bytes)")
t0 = time.time()

# Use read-only to quickly scan actual data ranges
wb_ro = openpyxl.load_workbook(dst, read_only=True)
actual_rows = {}
for sn in wb_ro.sheetnames:
    ws = wb_ro[sn]
    max_r = 0
    for row in ws.iter_rows(min_row=1, max_row=100, max_col=1):
        for cell in row:
            if cell.value is not None:
                max_r = cell.row
    # Also check a broader range for data sheets
    for row in ws.iter_rows(min_row=1, max_col=30):
        for cell in row:
            if cell.value is not None:
                if cell.row > max_r:
                    max_r = cell.row
    actual_rows[sn] = max_r
    print(f"  {sn}: actual data up to row {max_r} (reported max={ws.max_row})")
wb_ro.close()

print(f"\nScan complete in {time.time()-t0:.1f}s")
print(f"\nSheets with 1M+ reported rows that need stripping:")
for sn, actual in actual_rows.items():
    print(f"  {sn}: {actual} actual rows")
