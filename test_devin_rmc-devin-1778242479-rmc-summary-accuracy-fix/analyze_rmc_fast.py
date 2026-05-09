"""Fast analysis using read_only mode - RMC summary full dump"""
import openpyxl
from openpyxl.utils import get_column_letter
import time, re, sys

FILE_PATH = r"Files_need_to_study\Filled_Output\1 Base RMC _ 2026 February.xlsx"

t0 = time.time()
print("Loading workbook (read_only mode)...", flush=True)
wb = openpyxl.load_workbook(FILE_PATH, data_only=False, read_only=True)
print(f"Loaded in {time.time()-t0:.1f}s", flush=True)

ws = wb['RMC summary']
print(f"\n{'='*80}", flush=True)
print("FULL DUMP OF 'RMC summary' SHEET", flush=True)
print(f"{'='*80}", flush=True)

formulas = []
row_count = 0
current_row = 0
for row in ws.iter_rows():
    current_row += 1
    row_cells = []
    row_num = current_row
    for cell in row:
        val = cell.value
        if val is not None:
            try:
                col_idx = cell.column
                col_letter = get_column_letter(col_idx)
                coord = f"{col_letter}{row_num}"
            except:
                coord = f"?{row_num}"
            row_cells.append(f"    {coord}: {repr(val)}")
            if isinstance(val, str) and val.startswith('='):
                formulas.append((coord, val))
    if row_cells:
        row_count += 1
        print(f"\n  ROW {row_num}:", flush=True)
        for rc in row_cells:
            print(rc, flush=True)

print(f"\nTotal rows with data: {row_count}", flush=True)
print(f"Total rows scanned: {current_row}", flush=True)

print(f"\n{'='*80}", flush=True)
print(f"ALL FORMULAS IN 'RMC summary' ({len(formulas)} total)", flush=True)
print(f"{'='*80}", flush=True)
for coord, formula in formulas:
    print(f"  {coord}: {formula}", flush=True)

unique_patterns = {}
for coord, f in formulas:
    pattern = re.sub(r"'[^']*'", 'SHEET', f)
    pattern = re.sub(r'[A-Z]{1,3}\d+', 'REF', pattern)
    pattern = re.sub(r'\d+\.?\d*', 'N', pattern)
    if pattern not in unique_patterns:
        unique_patterns[pattern] = coord
print(f"\nUNIQUE FORMULA PATTERNS ({len(unique_patterns)}):", flush=True)
for p, example_coord in sorted(unique_patterns.items()):
    print(f"  Pattern: {p}  (example: {example_coord})", flush=True)

wb.close()
print(f"\nDone in {time.time()-t0:.1f}s", flush=True)
