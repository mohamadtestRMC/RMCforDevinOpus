"""Analyze all sheets except RMC summary - headers, sample data, formulas"""
import openpyxl
from openpyxl.utils import get_column_letter
import time, re

FILE_PATH = r"Files_need_to_study\Filled_Output\1 Base RMC _ 2026 February.xlsx"

t0 = time.time()
print("Loading workbook (read_only)...", flush=True)
wb = openpyxl.load_workbook(FILE_PATH, data_only=False, read_only=True)
print(f"Loaded in {time.time()-t0:.1f}s\n", flush=True)

for sheet_name in wb.sheetnames:
    if sheet_name == 'RMC summary':
        continue

    ws = wb[sheet_name]
    print(f"\n{'#'*80}", flush=True)
    print(f"SHEET: '{sheet_name}'", flush=True)
    print(f"{'#'*80}", flush=True)

    formulas = []
    all_rows = []
    row_count = 0
    for row in ws.iter_rows():
        row_count += 1
        row_data = {}
        for cell in row:
            val = cell.value
            if val is not None:
                try:
                    col_letter = get_column_letter(cell.column)
                    coord = f"{col_letter}{row_count}"
                except:
                    coord = f"?{row_count}"
                row_data[coord] = val
                if isinstance(val, str) and val.startswith('='):
                    formulas.append((coord, val))
        if row_data:
            all_rows.append((row_count, row_data))

    print(f"  Total rows with data: {len(all_rows)}", flush=True)
    print(f"  Total rows scanned: {row_count}", flush=True)
    print(f"  Total formulas: {len(formulas)}", flush=True)

    show_rows = min(15, len(all_rows))
    print(f"\n  --- FIRST {show_rows} DATA ROWS ---", flush=True)
    for i in range(show_rows):
        rn, rd = all_rows[i]
        print(f"  ROW {rn}:", flush=True)
        for coord, val in rd.items():
            print(f"    {coord}: {repr(val)}", flush=True)

    if len(all_rows) > 15:
        print(f"\n  --- LAST 3 DATA ROWS ---", flush=True)
        for i in range(max(len(all_rows)-3, 15), len(all_rows)):
            rn, rd = all_rows[i]
            print(f"  ROW {rn}:", flush=True)
            for coord, val in rd.items():
                print(f"    {coord}: {repr(val)}", flush=True)

    if formulas:
        print(f"\n  --- FORMULA SAMPLES (first 20) ---", flush=True)
        for coord, f in formulas[:20]:
            print(f"    {coord}: {f}", flush=True)
        if len(formulas) > 20:
            print(f"    ... {len(formulas)-20} more formulas", flush=True)

        unique_patterns = {}
        for coord, f in formulas:
            pattern = re.sub(r"'[^']*'", 'SHEET', f)
            pattern = re.sub(r'[A-Z]{1,3}\d+', 'REF', pattern)
            pattern = re.sub(r'\d+\.?\d*', 'N', pattern)
            if pattern not in unique_patterns:
                unique_patterns[pattern] = coord
        print(f"\n  --- UNIQUE FORMULA PATTERNS ({len(unique_patterns)}) ---", flush=True)
        for p, ex in sorted(unique_patterns.items()):
            print(f"    {p}  (ex: {ex})", flush=True)

wb.close()
print(f"\nDone in {time.time()-t0:.1f}s", flush=True)
