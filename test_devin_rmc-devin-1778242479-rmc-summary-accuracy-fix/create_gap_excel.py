"""
Create final Excel showing: for each row, exactly how friend calculated,
what our engine does differently, and what the fix is.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Styles
tf = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
sf = Font(name='Calibri', size=12, bold=True, color='1E293B')
hf = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
df = Font(name='Calibri', size=10)
bf = Font(name='Calibri', size=10, bold=True)
nf = Font(name='Calibri', size=11, bold=True, color='059669')
rf = Font(name='Calibri', size=11, bold=True, color='DC2626')
hfill = PatternFill('solid', fgColor='4F46E5')
gfill = PatternFill('solid', fgColor='D1FAE5')
rfill = PatternFill('solid', fgColor='FEE2E2')
yfill = PatternFill('solid', fgColor='FEF3C7')
bfill = PatternFill('solid', fgColor='DBEAFE')
dark = PatternFill('solid', fgColor='1E293B')
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left = Alignment(horizontal='left', vertical='center', wrap_text=True)
border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1'),
)

def style_cell(ws, r, c, val, font=df, fill=None, align=center):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = font
    if fill: cell.fill = fill
    cell.alignment = align
    cell.border = border
    return cell

# ══════════════════════════════════════════════════════════
# SHEET 1: Summary
# ══════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Summary"

ws.merge_cells('A1:H1')
c = ws.cell(row=1, column=1, value="GAP ANALYSIS: How Friend Calculated vs Our Engine")
c.font = tf; c.fill = dark; c.alignment = center
ws.row_dimensions[1].height = 35

r = 3
headers = ['Row', 'UID', 'Order', 'Material', 'Friend Rate', 'Engine Rate', 'Diff %', 'Root Cause']
for i, h in enumerate(headers):
    style_cell(ws, r, i+1, h, hf, hfill)

data = [
    (42, '202602-1120-L', 'L00335', 'PET Fresh1', 4.22, 4.298, '1.86%',
     'Friend used 2 MRRs (JBF RAK only)\nEngine used 3 MRRs (10% filter kept FLEX)'),
    (43, '202602-1232-L', 'L00335', 'PET Fresh1', 4.35, 4.298, '1.19%',
     'Friend used 2 MRRs (85547+85572)\nwith slightly different quantities'),
    (45, '202602-1360-P', 'G00418', 'PET Film', 4.547, 4.588, '0.91%',
     'BUG: 10% filter drops 85330 and 85157\nFriend used ALL 3 MRRs → FIX AVAILABLE'),
    (46, '202602-1474-P', 'N00945', 'PET Film', 4.588, 4.460, '2.78%',
     'Friend used 1 MRR (85572 only)\nEngine uses 2 MRRs (both pass 10% filter)'),
]
for i, d in enumerate(data):
    r = 4 + i
    for j, v in enumerate(d):
        fill = gfill if j == 4 else (rfill if j == 5 else None)
        style_cell(ws, r, j+1, v, bf if j in (4,5) else df, fill, left if j == 7 else center)
    ws.row_dimensions[r].height = 40

r = 9
ws.merge_cells(f'A{r}:H{r}')
c = ws.cell(row=r, column=1, value="KEY FINDING: The 10% dominant MRR filter in our engine is the main bug")
c.font = Font(name='Calibri', size=12, bold=True, color='DC2626'); c.fill = rfill; c.alignment = left

r = 10
ws.merge_cells(f'A{r}:H{r}')
c = ws.cell(row=r, column=1, value="Row 45: Removing the filter will FIX this row (all 3 MRRs → weighted avg = 4.547 = exact match)")
c.font = Font(name='Calibri', size=11, bold=True, color='059669'); c.fill = gfill; c.alignment = left

r = 11
ws.merge_cells(f'A{r}:H{r}')
c = ws.cell(row=r, column=1, value="Rows 42, 43, 46: Different MRR subset selection — friend picked specific MRRs manually")
c.font = Font(name='Calibri', size=11, color='64748B'); c.alignment = left

widths_s = [8, 18, 10, 12, 12, 12, 10, 55]
for i, w in enumerate(widths_s):
    ws.column_dimensions[get_column_letter(i+1)].width = w

# ══════════════════════════════════════════════════════════
# SHEET 2-5: One per row with full calculation trace
# ══════════════════════════════════════════════════════════
row_details = [
    {
        'row': 42, 'uid': '202602-1120-L', 'order': 'L00335',
        'mat': 'PET', 'size': 1197, 'mic': 12, 'slot': 'Fresh1',
        'gt_mr': '85547/85588', 'gt_rate': 4.22, 'eng_rate': 4.298302,
        'mrrs': [
            (85547, 1809.5, 4.22, 'JBF RAK', '2-2026', True),
            (85588, 1857.7, 4.22, 'JBF RAK', '2-2026', True),
            (85572, 991.2, 4.588, 'FLEX', '2-2026', False),
            (85157, 497.8, 4.404, 'FLEX', '1-2026', False),
            (85226, 305.0, 4.22, 'JBF RAK', '1-2026', False),
        ],
        'friend_used': [85547, 85588],
        'explanation': [
            "Friend picked only JBF RAK supplier MRRs from current month (Feb 2026)",
            "Both have rate 4.22 → weighted avg = 4.22",
            "Engine includes FLEX MRRs (85572) which has higher rate (4.588)",
            "This pulls our average UP to 4.298",
        ],
        'fix': "Friend manually selected supplier subset — hard to automate without a rule"
    },
    {
        'row': 43, 'uid': '202602-1232-L', 'order': 'L00335',
        'mat': 'PET', 'size': 1197, 'mic': 12, 'slot': 'Fresh1',
        'gt_mr': '85226/85157/85547/85572/85588', 'gt_rate': 4.349972, 'eng_rate': 4.298302,
        'mrrs': [
            (85547, 1809.5, 4.22, 'JBF RAK', '2-2026', True),
            (85588, 1857.7, 4.22, 'JBF RAK', '2-2026', True),
            (85572, 991.2, 4.588, 'FLEX', '2-2026', True),
            (85157, 497.8, 4.404, 'FLEX', '1-2026', True),
            (85226, 305.0, 4.22, 'JBF RAK', '1-2026', True),
        ],
        'friend_used': [85547, 85572],
        'explanation': [
            "Friend listed ALL 5 MRRs in the MR# column",
            "BUT the rate 4.3500 matches weighted avg of only 85547+85572",
            "85547: 1814 kg × 4.22 = 7655.1   (friend may have used slightly different qty)",
            "85572: 990.5 kg × 4.588 = 4544.6",
            "Total: 12199.7 / 2804.5 = 4.3500",
            "Likely: friend calculated rate from 2 MRRs but listed all 5 in display",
        ],
        'fix': "Friend used different quantities than what Stores shows — possible manual entry"
    },
    {
        'row': 45, 'uid': '202602-1360-P', 'order': 'G00418',
        'mat': 'PET', 'size': 1063, 'mic': 12, 'slot': 'Film',
        'gt_mr': '85157/85573/85330', 'gt_rate': 4.546673, 'eng_rate': 4.588,
        'mrrs': [
            (85573, 2621.3, 4.588, 'FLEX', '2-2026', True),
            (85330, 253.0, 4.22, 'JBF RAK', '1-2026', True),
            (85157, 180.0, 4.404, 'FLEX', '1-2026', True),
        ],
        'friend_used': [85573, 85330, 85157],
        'explanation': [
            "Friend used ALL 3 MRRs — weighted average",
            "85573: 2621.3 × 4.588 = 12,026.52",
            "85330: 253.0 × 4.22  = 1,067.66",
            "85157: 180.0 × 4.404 = 792.72",
            "Total: 13,886.90 / 3,054.3 = 4.5467 ← EXACT MATCH",
            "",
            "OUR ENGINE BUG: 10% filter drops 85330 (8.3%) and 85157 (5.9%)",
            "Engine uses only 85573 → rate = 4.588 (wrong!)",
            "FIX: Remove 10% filter → rate = 4.5467 (correct!)",
        ],
        'fix': "REMOVE 10% DOMINANT MRR FILTER → This row will be FIXED!"
    },
    {
        'row': 46, 'uid': '202602-1474-P', 'order': 'N00945',
        'mat': 'PET', 'size': 913, 'mic': 12, 'slot': 'Film',
        'gt_mr': '85572', 'gt_rate': 4.588, 'eng_rate': 4.460399,
        'mrrs': [
            (85572, 376.8, 4.588, 'FLEX', '2-2026', True),
            (85226, 200.0, 4.22, 'JBF RAK', '1-2026', False),
        ],
        'friend_used': [85572],
        'explanation': [
            "Friend used ONLY MRR 85572 → rate = 4.588",
            "MRR 85226 has PR month 1-2026 (January purchase)",
            "Friend may have excluded it because it's from a different purchase batch",
            "Engine includes both MRRs → weighted avg = 4.460 (lower)",
            "",
            "Note: Friend said dates don't matter, but here he only used Feb MRR",
            "Possible: this was a manual selection based on work order context",
        ],
        'fix': "Friend excluded old-purchase MRR — inconsistent with stated rule"
    },
]

for rd in row_details:
    ws2 = wb.create_sheet(f"Row {rd['row']}")
    
    # Title
    ws2.merge_cells('A1:I1')
    c = ws2.cell(row=1, column=1, 
                 value=f"Row {rd['row']} — {rd['uid']} — Order {rd['order']} — "
                       f"{rd['slot']}={rd['mat']} Size={rd['size']} Mic={rd['mic']}")
    c.font = tf; c.fill = dark; c.alignment = left
    ws2.row_dimensions[1].height = 30
    
    # Rates comparison
    style_cell(ws2, 3, 1, "Friend's Rate:", bf)
    style_cell(ws2, 3, 2, rd['gt_rate'], nf, gfill)
    style_cell(ws2, 3, 3, "Our Engine Rate:", bf)
    style_cell(ws2, 3, 4, rd['eng_rate'], rf, rfill)
    style_cell(ws2, 3, 5, "GT MR#:", bf)
    style_cell(ws2, 3, 6, rd['gt_mr'], bf)
    
    # All MRRs table
    r = 5
    style_cell(ws2, r, 1, "ALL AVAILABLE MRRs FROM STORES:", sf)
    r = 6
    for i, h in enumerate(['MRR #', 'Qty (kg)', 'Rate (AED)', 'Supplier', 'PR Month', 
                            'Friend Used?', 'Engine Used?', 'Qty × Rate']):
        style_cell(ws2, r, i+1, h, hf, hfill)
    
    for mrr, qty, rate, sup, month, in_gt in rd['mrrs']:
        r += 1
        in_friend = '✓ YES' if mrr in rd['friend_used'] else '✗ NO'
        # Engine uses all that pass 10% filter
        total_q = sum(m[1] for m in rd['mrrs'])
        in_engine = '✓' if qty >= total_q * 0.10 else '✗ (< 10%)'
        
        vals = [mrr, qty, rate, sup, month, in_friend, in_engine, round(qty * rate, 2)]
        for i, v in enumerate(vals):
            fill = gfill if (i == 5 and in_friend == '✓ YES') else (rfill if (i == 5 and in_friend == '✗ NO') else None)
            style_cell(ws2, r, i+1, v, df, fill)
    
    # Friend's calculation
    r += 2
    style_cell(ws2, r, 1, "HOW FRIEND CALCULATED:", Font(name='Calibri', size=12, bold=True, color='4F46E5'), bfill, left)
    r += 1
    
    for i, h in enumerate(['Step', 'MRR', 'Qty (kg)', '×', 'Rate', '=', 'Amount']):
        style_cell(ws2, r, i+1, h, hf, hfill)
    r += 1
    
    total_num = 0
    total_den = 0
    step = 1
    for mrr, qty, rate, sup, month, _ in rd['mrrs']:
        if mrr in rd['friend_used']:
            amount = qty * rate
            total_num += amount
            total_den += qty
            vals = [f"Step {step}", mrr, qty, '×', rate, '=', round(amount, 2)]
            for i, v in enumerate(vals):
                style_cell(ws2, r, i+1, v, df)
            r += 1
            step += 1
    
    # Total
    vals = ['TOTAL', '', total_den, '', '', '=', round(total_num, 2)]
    for i, v in enumerate(vals):
        style_cell(ws2, r, i+1, v, bf, yfill)
    r += 1
    
    # Result
    result = total_num / total_den if total_den > 0 else 0
    style_cell(ws2, r, 1, "RESULT:", Font(name='Calibri', size=13, bold=True))
    style_cell(ws2, r, 2, f"{total_num:.2f} ÷ {total_den:.1f}", bf)
    style_cell(ws2, r, 3, "=", bf)
    style_cell(ws2, r, 4, round(result, 6), Font(name='Calibri', size=14, bold=True, color='059669'), gfill)
    style_cell(ws2, r, 5, f"Friend got: {rd['gt_rate']}", bf)
    match = abs(result - rd['gt_rate']) < 0.01
    style_cell(ws2, r, 6, "✓ MATCH" if match else f"Close (diff={abs(result - rd['gt_rate']):.6f})", 
               nf if match else rf, gfill if match else yfill)
    
    r += 2
    # Explanation
    style_cell(ws2, r, 1, "WHY DIFFERENT FROM ENGINE:", Font(name='Calibri', size=12, bold=True, color='DC2626'), rfill, left)
    r += 1
    for line in rd['explanation']:
        if line:
            ws2.merge_cells(f'A{r}:I{r}')
            c = ws2.cell(row=r, column=1, value=line)
            c.font = df if not line.startswith('FIX') and not line.startswith('OUR ENGINE BUG') else Font(name='Calibri', size=10, bold=True, color='DC2626')
            c.alignment = left
            if 'FIX' in line or 'BUG' in line:
                c.fill = yfill
        r += 1
    
    r += 1
    ws2.merge_cells(f'A{r}:I{r}')
    c = ws2.cell(row=r, column=1, value=f"CONCLUSION: {rd['fix']}")
    c.font = Font(name='Calibri', size=12, bold=True, color='4F46E5')
    c.fill = bfill; c.alignment = left
    
    # Column widths
    widths = [12, 12, 12, 5, 12, 14, 14, 14, 14]
    for i, w in enumerate(widths):
        ws2.column_dimensions[get_column_letter(i+1)].width = w

path = "Template_Files/Gap_Analysis_Detail.xlsx"
wb.save(path)
print(f"Saved to: {path}")
