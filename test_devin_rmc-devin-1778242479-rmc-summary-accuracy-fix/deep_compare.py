"""
Deep compare: Engine output vs Filled Reference for each process sheet.
Identifies exactly which cells differ and why.
"""
import sys, os
sys.path.insert(0, r'c:\Users\mohamad.al\OneDrive - AL KHAYYAT INVESTMENTS\Desktop\IPP')
import openpyxl

base = r'c:\Users\mohamad.al\OneDrive - AL KHAYYAT INVESTMENTS\Desktop\IPP'
engine_path = os.path.join(base, 'output', 'Base_RMC_Feb2026_FILLED.xlsx')
ref_path = os.path.join(base, 'Files_need_to_study', 'Filled_Output', '1 Base RMC _ 2026 February.xlsx')

def _sf(v):
    if v is None: return 0.0
    try: return float(v)
    except: return 0.0

print("Loading files...", flush=True)
wb_eng = openpyxl.load_workbook(engine_path, data_only=True, read_only=True)
wb_ref = openpyxl.load_workbook(ref_path, data_only=True, read_only=True)

# Compare specific process sheets
for sheet_name in ['Print', 'Lam', 'FG', 'CLS_WIP']:
    if sheet_name not in wb_eng.sheetnames or sheet_name not in wb_ref.sheetnames:
        print(f"\n{sheet_name}: MISSING from one file", flush=True)
        continue
    
    ws_eng = wb_eng[sheet_name]
    ws_ref = wb_ref[sheet_name]
    
    print(f"\n{'='*70}", flush=True)
    print(f"SHEET: {sheet_name}", flush=True)
    print(f"  Engine: {ws_eng.max_row} rows x {ws_eng.max_column} cols", flush=True)
    print(f"  Ref:    {ws_ref.max_row} rows x {ws_ref.max_column} cols", flush=True)
    
    # Build order→row indexes for both
    eng_orders = {}
    for r in range(5, min(ws_eng.max_row + 1, 1000)):
        o = ws_eng.cell(row=r, column=2).value
        if o and str(o).strip():
            ou = str(o).strip().upper()
            if ou not in eng_orders:
                eng_orders[ou] = []
            eng_orders[ou].append(r)
    
    ref_orders = {}
    for r in range(5, min(ws_ref.max_row + 1, 1000)):
        o = ws_ref.cell(row=r, column=2).value
        if o and str(o).strip():
            ou = str(o).strip().upper()
            if ou not in ref_orders:
                ref_orders[ou] = []
            ref_orders[ou].append(r)
    
    common = set(eng_orders.keys()) & set(ref_orders.keys())
    eng_only = set(eng_orders.keys()) - set(ref_orders.keys())
    ref_only = set(ref_orders.keys()) - set(eng_orders.keys())
    
    print(f"  Orders: Engine={len(eng_orders)}, Ref={len(ref_orders)}, Common={len(common)}", flush=True)
    if eng_only:
        print(f"  Engine-only orders (first 5): {list(eng_only)[:5]}", flush=True)
    if ref_only:
        print(f"  Ref-only orders (first 5): {list(ref_only)[:5]}", flush=True)
    
    # For common orders, compare specific columns
    if sheet_name == 'Print':
        # Check key columns: G=7(Film Input), H=8(Dry Ink), I=9(Total), J=10(Film Val), K=11(Ink Val)
        check_cols = {7: 'Film Input Kgs(G)', 8: 'Dry Ink(H)', 9: 'Total Input(I)', 
                      10: 'Film Value(J)', 11: 'Ink Value(K)', 13: 'Output Kgs(M)',
                      17: 'Wastage Kgs(Q)', 18: 'Wastage Val(R)'}
        print(f"\n  --- Column comparison (first 5 common orders) ---", flush=True)
        sample = sorted(common)[:5]
        for ou in sample:
            er = eng_orders[ou][0]
            rr = ref_orders[ou][0]
            print(f"\n  Order {ou} (eng row {er}, ref row {rr}):", flush=True)
            for c, label in check_cols.items():
                ev = _sf(ws_eng.cell(row=er, column=c).value)
                rv = _sf(ws_ref.cell(row=rr, column=c).value)
                match = "✅" if abs(ev - rv) < 0.5 else "❌"
                print(f"    {match} {label}: eng={ev:.2f}, ref={rv:.2f}", flush=True)
    
    elif sheet_name == 'Lam':
        # Check: AY=51, AZ=52, BA=53, BB=54, Z=26(ptd_qty), AB=28(ptd_val), AF=32(fresh1_qty), AH=34(fresh1_val)
        check_cols = {26: 'Ptd Qty(Z)', 27: 'Ptd Rate(AA)', 28: 'Ptd Val(AB)',
                      32: 'Fresh1 Qty(AF)', 33: 'Fresh1 Rate(AG)', 34: 'Fresh1 Val(AH)',
                      38: 'Adh Qty(AL)', 40: 'Adh Rate(AN)', 41: 'Adh Val(AO)',
                      51: 'Fresh Mat Qty(AY)', 52: 'Fresh Mat Val(AZ)',
                      53: 'AdhHard Solids(BA)', 54: 'AdhHardSolv Val(BB)'}
        print(f"\n  --- Column comparison (first 5 common orders) ---", flush=True)
        sample = sorted(common)[:5]
        for ou in sample:
            er = eng_orders[ou][0]
            rr = ref_orders[ou][0]
            print(f"\n  Order {ou} (eng row {er}, ref row {rr}):", flush=True)
            for c, label in check_cols.items():
                ev = _sf(ws_eng.cell(row=er, column=c).value)
                rv = _sf(ws_ref.cell(row=rr, column=c).value)
                match = "✅" if abs(ev - rv) < 0.5 else "❌"
                if ev != 0 or rv != 0:
                    print(f"    {match} {label}: eng={ev:.2f}, ref={rv:.2f}", flush=True)
    
    elif sheet_name == 'FG':
        # Check: A=1(order), G=7(FG output)
        check_cols = {1: 'Order(A)', 7: 'FG Output(G)'}
        print(f"\n  --- FG structure ---", flush=True)
        # Check FG by column A (not B)
        fg_eng = {}
        for r in range(3, min(ws_eng.max_row + 1, 700)):
            o = ws_eng.cell(row=r, column=1).value
            if o: fg_eng[str(o).strip().upper()] = r
        fg_ref = {}
        for r in range(3, min(ws_ref.max_row + 1, 700)):
            o = ws_ref.cell(row=r, column=1).value
            if o: fg_ref[str(o).strip().upper()] = r
        
        common_fg = set(fg_eng.keys()) & set(fg_ref.keys())
        print(f"  FG: Engine={len(fg_eng)}, Ref={len(fg_ref)}, Common={len(common_fg)}", flush=True)
        
        # Sample compare
        sample = sorted(common_fg)[:5]
        for ou in sample:
            er = fg_eng[ou]
            rr = fg_ref[ou]
            ev_g = _sf(ws_eng.cell(row=er, column=7).value)
            rv_g = _sf(ws_ref.cell(row=rr, column=7).value)
            # Also check all cols
            eng_row = {c: ws_eng.cell(row=er, column=c).value for c in range(1, 10) if ws_eng.cell(row=er, column=c).value}
            ref_row = {c: ws_ref.cell(row=rr, column=c).value for c in range(1, 10) if ws_ref.cell(row=rr, column=c).value}
            match = "✅" if abs(ev_g - rv_g) < 0.5 else "❌"
            print(f"  {match} {ou}: eng G={ev_g:.2f}, ref G={rv_g:.2f}", flush=True)
            print(f"      eng: {eng_row}", flush=True)
            print(f"      ref: {ref_row}", flush=True)
    
    elif sheet_name == 'CLS_WIP':
        check_cols = {8: 'Qty(H)', 10: 'Value(J)'}
        print(f"\n  --- CLS_WIP comparison ---", flush=True)
        sample = sorted(common)[:5]
        for ou in sample:
            er = eng_orders[ou][0]
            rr = ref_orders[ou][0]
            print(f"  Order {ou}:", flush=True)
            for c, label in check_cols.items():
                ev = _sf(ws_eng.cell(row=er, column=c).value)
                rv = _sf(ws_ref.cell(row=rr, column=c).value)
                match = "✅" if abs(ev - rv) < 0.5 else "❌"
                print(f"    {match} {label}: eng={ev:.2f}, ref={rv:.2f}", flush=True)

wb_eng.close()
wb_ref.close()
print("\nDone!", flush=True)
