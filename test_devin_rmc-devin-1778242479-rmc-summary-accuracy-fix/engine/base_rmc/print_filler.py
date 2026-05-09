"""
Print Filler — Fills the Print sheet FROM Jobtrack data.

CORRECTED column mapping based on actual template headers:
  Row 6: headers
  Row 7+: data rows
  B=Order No, C=Design Name, D=Date, E=M/c, F=Material, G=Structure,
  H=Mic, I=Input Name,
  J=Film Input (Kgs), K=Dry Ink (Kgs), L=Total Input,
  M=Film Value, N=Ink Value, O=Total Value,
  P=Output Kgs, Q=RMC/Kg, R=Output Meters, S=Output Sq Mtrs,
  T=Wastage Qty, U=Wastage Value, V=Wastage %, W=Input RMC,
  X=Waste Qty (Log), Y=Avg Ink GSM, Z=Check

NOTE: Columns D (Date), E (M/c), H (Mic), S (Output Sq Mtrs),
Y (Avg Ink GSM), Z (Check) were previously MISSING and are now filled.
"""
from __future__ import annotations
import logging
import pandas as pd
from engine.base_rmc.context import RMCContext

logger = logging.getLogger(__name__)

# FULL column mapping — includes ALL metadata columns from Jobtrack
COL = {
    'order_no': 2,         # B
    'design': 3,           # C
    'date': 4,             # D = Production Date (FROM Jobtrack)
    'machine': 5,          # E = M/c Machine (FROM Jobtrack)
    'material': 6,         # F
    'structure': 7,        # G
    'mic': 8,              # H = Input Micron (FROM Jobtrack)
    'input_name': 9,       # I
    'film_input_kgs': 10,  # J = Film Input (Kgs) = total_input - dry_ink
    'dry_ink_kgs': 11,     # K = Dry Ink (Kgs)
    'total_input': 12,     # L = Total Input
    'film_value': 13,      # M = Film Value
    'ink_value': 14,       # N = Ink Value
    'total_value': 15,     # O = Total Value = Film + Ink
    'output_kgs': 16,      # P = Output Kgs
    'output_rmc_kg': 17,   # Q = RMC/Kg = Total Value / Output Kgs
    'output_mtrs': 18,     # R = Output Meters
    'output_sq_mtrs': 19,  # S = Output Sq Mtrs (FROM Jobtrack)
    'wastage_kgs': 20,     # T = Wastage Qty
    'wastage_val': 21,     # U = Wastage Value
    'wastage_pct': 22,     # V = Wastage %
    'input_rmc_kg': 23,    # W = Input RMC
    'waste_log': 24,       # X = Waste Qty (Log Sheet)
    'avg_ink_gsm': 25,     # Y = Average Ink GSM (computed)
    'check': 26,           # Z = Check (= Total Input - Output - Wastage)
}

DATA_START = 7
SUBTOTAL_ROW = 5


def _sf(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return 0.0
    try: return float(val)
    except: return 0.0


def _ss(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    return str(val).strip()


def fill_print(ctx: RMCContext) -> None:
    """Fill Print sheet from Jobtrack PRINTING rows with ALL metadata columns."""
    ctx._log("Filling Print sheet (with metadata: Date, M/c, Mic, Ink GSM)...")

    ws = ctx.wb['Print'] if 'Print' in ctx.wb.sheetnames else None
    if ws is None:
        ctx._error("  Print sheet not found!")
        return

    if ctx.jobtrack_df is None or ctx.jobtrack_df.empty:
        ctx._error("  No Jobtrack data for Print")
        return

    df = ctx.jobtrack_df
    proc_col = None
    for c in df.columns:
        if str(c).strip().lower() == 'process':
            proc_col = c
            break

    print_df = df[df[proc_col].astype(str).str.strip().str.upper() == 'PRINTING'].copy()
    ctx._log(f"  Print rows from Jobtrack: {len(print_df)}")

    if print_df.empty:
        return

    def fc(name):
        for c in df.columns:
            if str(c).strip() == name:
                return c
        return None

    ink_costs = ctx.ink_rate_cache

    filled_count = 0
    meta_filled = {'date': 0, 'machine': 0, 'mic': 0, 'ink_gsm': 0}
    print_cache = {}
    row_idx = DATA_START

    for _, jt in print_df.iterrows():
        order_no = _ss(jt.get(fc('Order No'), ''))
        if not order_no:
            continue

        order_upper = order_no.upper()
        design = _ss(jt.get(fc('Design Name'), ''))
        material = _ss(jt.get(fc('Material'), ''))
        structure = _ss(jt.get(fc('Structure'), ''))
        input_name = _ss(jt.get(fc('1st Input Name'), ''))

        # ── NEW: Metadata columns from Jobtrack ──
        date_val = jt.get(fc('Date'), None)
        machine = _ss(jt.get(fc('Machine'), ''))
        mic = _sf(jt.get(fc('1st Input Mic'), 0))
        output_sq_mtrs = _sf(jt.get(fc('Output (Sq. Mtrs)'), 0))
        if output_sq_mtrs == 0:
            output_sq_mtrs = _sf(jt.get(fc('Prod Sq Mtr'), 0))

        # Total Input
        total_input = _sf(jt.get(fc('TOTAL INPUT'), 0))
        if total_input == 0:
            qty = _sf(jt.get(fc('1st Input  Qty'), 0))
            bal = _sf(jt.get(fc('Balance Qty'), 0))
            total_input = qty + bal

        output_kgs = _sf(jt.get(fc('Net Wt. (Kgs-Output)'), 0))
        output_mtrs = _sf(jt.get(fc('Output (Meters)'), 0))
        dry_ink = _sf(jt.get(fc('DRY INK QTY'), 0))
        waste = _sf(jt.get(fc('Waste'), 0))
        waste_log = _sf(jt.get(fc('Total Wastage'), 0))
        if waste_log == 0:
            waste_log = waste

        # Film Input Kgs = Total Input - Dry Ink
        film_input_kgs = total_input - dry_ink if total_input > dry_ink else total_input

        # Film value from Jobtrack enrichment
        film_rate = _sf(jt.get(fc('Rate'), 0))
        film_value = _sf(jt.get(fc('Film Value'), 0))
        if film_value == 0 and film_rate > 0 and film_input_kgs > 0:
            film_value = film_input_kgs * film_rate

        # Ink Value from Ink Consumption Summary
        ink_value = ink_costs.get(order_upper, ink_costs.get(order_no, 0.0))

        # Computed values
        total_value = film_value + ink_value
        input_rmc_kg = film_value / film_input_kgs if film_input_kgs > 0 else 0
        output_rmc_kg = total_value / output_kgs if output_kgs > 0 else 0
        wastage_kgs = total_input - output_kgs if total_input > 0 else waste
        wastage_val = wastage_kgs * input_rmc_kg
        wastage_pct = wastage_kgs / total_input if total_input > 0 else 0

        # ── NEW: Avg Ink GSM = (Dry Ink Kgs / Output Sq Mtrs) * 1000 ──
        avg_ink_gsm = 0.0
        if output_sq_mtrs > 0 and dry_ink > 0:
            avg_ink_gsm = (dry_ink / output_sq_mtrs) * 1000

        # ── NEW: Check = Total Input - Output - Wastage (should ≈ 0) ──
        check_val = total_input - output_kgs - wastage_kgs

        # Write to Print sheet — ALL column positions including metadata
        ws.cell(row=row_idx, column=COL['order_no'], value=order_no)
        ws.cell(row=row_idx, column=COL['design'], value=design or None)

        # ── NEW: Date, Machine, Mic ──
        if date_val is not None and not (isinstance(date_val, float) and pd.isna(date_val)):
            ws.cell(row=row_idx, column=COL['date'], value=date_val)
            meta_filled['date'] += 1
        if machine:
            ws.cell(row=row_idx, column=COL['machine'], value=machine)
            meta_filled['machine'] += 1
        ws.cell(row=row_idx, column=COL['material'], value=material or None)
        ws.cell(row=row_idx, column=COL['structure'], value=structure or None)
        if mic > 0:
            ws.cell(row=row_idx, column=COL['mic'], value=mic)
            meta_filled['mic'] += 1
        ws.cell(row=row_idx, column=COL['input_name'], value=input_name or None)

        ws.cell(row=row_idx, column=COL['film_input_kgs'], value=film_input_kgs)
        ws.cell(row=row_idx, column=COL['dry_ink_kgs'], value=dry_ink)
        ws.cell(row=row_idx, column=COL['total_input'], value=total_input)
        ws.cell(row=row_idx, column=COL['film_value'], value=film_value)
        ws.cell(row=row_idx, column=COL['ink_value'], value=ink_value)
        ws.cell(row=row_idx, column=COL['total_value'], value=total_value)
        ws.cell(row=row_idx, column=COL['output_kgs'], value=output_kgs)
        ws.cell(row=row_idx, column=COL['output_rmc_kg'], value=output_rmc_kg)
        ws.cell(row=row_idx, column=COL['output_mtrs'], value=output_mtrs)

        # ── NEW: Output Sq Mtrs ──
        if output_sq_mtrs > 0:
            ws.cell(row=row_idx, column=COL['output_sq_mtrs'], value=output_sq_mtrs)

        ws.cell(row=row_idx, column=COL['wastage_kgs'], value=wastage_kgs)
        ws.cell(row=row_idx, column=COL['wastage_val'], value=wastage_val)
        ws.cell(row=row_idx, column=COL['wastage_pct'], value=wastage_pct)
        ws.cell(row=row_idx, column=COL['input_rmc_kg'], value=input_rmc_kg)
        ws.cell(row=row_idx, column=COL['waste_log'], value=waste_log)

        # ── NEW: Avg Ink GSM + Check ──
        if avg_ink_gsm > 0:
            ws.cell(row=row_idx, column=COL['avg_ink_gsm'], value=avg_ink_gsm)
            meta_filled['ink_gsm'] += 1
        ws.cell(row=row_idx, column=COL['check'], value=round(check_val, 4))

        if output_rmc_kg > 0:
            filled_count += 1

        ctx.print_rate_cache[order_upper] = output_rmc_kg

        if order_upper not in print_cache:
            print_cache[order_upper] = {
                'input_kgs': 0, 'output_kgs': 0, 'film_value': 0,
                'ink_value': 0, 'ink_kgs': 0, 'wastage_kgs': 0, 'wastage_val': 0,
            }
        print_cache[order_upper]['input_kgs'] += total_input
        print_cache[order_upper]['output_kgs'] += output_kgs
        print_cache[order_upper]['film_value'] += film_value
        print_cache[order_upper]['ink_value'] += ink_value
        print_cache[order_upper]['ink_kgs'] += dry_ink
        print_cache[order_upper]['wastage_kgs'] += wastage_kgs
        print_cache[order_upper]['wastage_val'] += wastage_val

        row_idx += 1

    total_rows = row_idx - DATA_START

    # Subtotals
    if total_rows > 0:
        last = DATA_START + total_rows - 1
        from openpyxl.utils import get_column_letter
        for key in ('total_input', 'output_kgs', 'film_value', 'ink_value',
                     'total_value', 'wastage_kgs', 'wastage_val'):
            cn = COL.get(key)
            if cn:
                cl = get_column_letter(cn)
                ws.cell(row=SUBTOTAL_ROW, column=cn,
                        value=f"=SUBTOTAL(9,{cl}{DATA_START}:{cl}{last})")

    ctx.print_by_order = print_cache
    ctx._log(f"  Print filled: {filled_count}/{total_rows} rows, "
             f"{len(print_cache)} orders, {len(ctx.print_rate_cache)} rate cache entries")
    ctx._log(f"  Print metadata: Date={meta_filled['date']}, M/c={meta_filled['machine']}, "
             f"Mic={meta_filled['mic']}, Ink GSM={meta_filled['ink_gsm']}")
