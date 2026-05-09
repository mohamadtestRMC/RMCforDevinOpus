"""Fix the template by patching sheet dimensions in the XLSX XML.
OPN_WIP and CLS_WIP claim 1M rows but only have ~250 actual rows.
This patches the dimension tag so openpyxl loads in seconds, not hours."""
import zipfile, os, re, shutil, time, tempfile

def strip_template(src_path: str, dst_path: str = None) -> str:
    """Strip empty rows from the template XLSX by patching XML dimensions.
    Returns path to the stripped file."""
    if dst_path is None:
        dst_path = src_path.replace('.xlsx', '_stripped.xlsx')

    t0 = time.time()
    print(f"Stripping template: {os.path.basename(src_path)}")

    # Read the XLSX as a ZIP
    tmp = tempfile.mktemp(suffix='.xlsx')
    shutil.copy2(src_path, tmp)

    # Known max actual rows per sheet (from scan)
    MAX_ROWS = {
        'OPN_WIP': 300,
        'CLS_WIP': 300,
    }

    with zipfile.ZipFile(tmp, 'r') as zin:
        # Find sheet name -> file mapping from workbook.xml
        wb_xml = zin.read('xl/workbook.xml').decode('utf-8')
        # Find sheet names and their rIds
        sheet_map = {}  # name -> rId
        for m in re.finditer(r'<sheet\s+name="([^"]+)"\s+sheetId="\d+"\s+(?:state="[^"]*"\s+)?r:id="([^"]+)"', wb_xml):
            sheet_map[m.group(1)] = m.group(2)

        # Find rId -> file mapping from relationships
        rels_xml = zin.read('xl/_rels/workbook.xml.rels').decode('utf-8')
        rid_file = {}
        for m in re.finditer(r'<Relationship\s+Id="([^"]+)"\s+[^>]*Target="([^"]+)"', rels_xml):
            rid_file[m.group(1)] = m.group(2)

        # Identify files to patch
        files_to_patch = {}
        for sname, max_row in MAX_ROWS.items():
            if sname in sheet_map:
                rid = sheet_map[sname]
                if rid in rid_file:
                    target = rid_file[rid]
                    filepath = 'xl/' + target if not target.startswith('/') else target.lstrip('/')
                    files_to_patch[filepath] = max_row
                    print(f"  Will patch {sname} -> {filepath} (max_row={max_row})")

        # Create new ZIP with patched files
        with zipfile.ZipFile(dst_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)

                if item.filename in files_to_patch:
                    max_r = files_to_patch[item.filename]
                    xml = data.decode('utf-8')
                    # Patch dimension ref
                    xml = re.sub(
                        r'<dimension\s+ref="([A-Z]+\d+):([A-Z]+)\d+"',
                        lambda m: f'<dimension ref="{m.group(1)}:{m.group(2)}{max_r}"',
                        xml
                    )
                    data = xml.encode('utf-8')
                    print(f"  Patched {item.filename}")

                zout.writestr(item, data)

    os.unlink(tmp)
    elapsed = time.time() - t0
    size_before = os.path.getsize(src_path)
    size_after = os.path.getsize(dst_path)
    print(f"Done in {elapsed:.1f}s: {size_before:,} -> {size_after:,} bytes "
          f"({100*size_after/size_before:.0f}%)")
    return dst_path


if __name__ == '__main__':
    base = r'c:\Users\mohamad.al\OneDrive - AL KHAYYAT INVESTMENTS\Desktop\IPP'
    src = os.path.join(base, 'Files_need_to_study', 'Unfilled', '1 Base RMC _ 2026 February.xlsx')
    dst = os.path.join(base, 'output', 'template_stripped.xlsx')
    os.makedirs(os.path.dirname(dst), exist_ok=True)

    result = strip_template(src, dst)
    print(f"\nStripped template: {result}")

    # Quick test: load with openpyxl
    import openpyxl
    print("\nLoading stripped template with openpyxl...")
    t0 = time.time()
    wb = openpyxl.load_workbook(result, keep_links=False)
    print(f"Loaded in {time.time()-t0:.1f}s!")
    print(f"Sheets: {wb.sheetnames}")
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"  {sn}: max_row={ws.max_row}, max_col={ws.max_column}")
    wb.close()
