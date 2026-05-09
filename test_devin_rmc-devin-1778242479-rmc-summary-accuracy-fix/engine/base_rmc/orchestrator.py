"""
Base RMC Orchestrator — Single entry point that fills the entire Base RMC workbook.

Execution order (strict dependency chain):
  1. Load all input files
  2. Fill Jobtrack with MRR rates (existing engine/fill_jobtrack.py)
  3. Build OPN_WIP (paste + composite keys)
  4. Read Jobtrack data into process-level DataFrames
  5. Fill BFL (Extrusion)
  6. Fill Print + Printing Work
  7. Fill Lam (Lamination)
  8. Build Pivot_Lam Rates
  9. Fill Slit
  10. Fill Bag&Pouch
  11. Fill Spout&Valve
  12. Fill HCI Rew, PTR Rew, Embossing
  13. Build FG
  14. Build CLS_WIP (qty + rate cascade)
  15. Fill RMC Summary
  16. Fill Overall Wastage
  17. Validate
"""
from __future__ import annotations
import io
import logging
import time
import pandas as pd
from pathlib import Path
from typing import Any, Callable, Dict, Optional

from engine.base_rmc.context import RMCContext
from engine.base_rmc.loaders import load_all_into_context, build_ink_cost_by_order
from engine.base_rmc.wip_keys import build_wip_index
from engine.base_rmc.opn_wip_filler import fill_opn_wip

logger = logging.getLogger(__name__)

ProgressCallback = Callable[[int, str], None]


def _safe_float(val) -> float:
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def run_pipeline(
    *,
    base_rmc_template,
    purchase_register_file,
    stores_file,
    jobtrack_file=None,
    filled_jobtrack_file=None,
    granules_file=None,
    prev_granules_file=None,
    ink_consumption_file=None,
    megapack_file=None,
    opn_wip_file=None,
    cls_wip_file=None,
    valve_spout_file=None,
    component_consumption_file=None,
    ink_stock_opening_file=None,
    dispensed_movement_file=None,
    filled_reference=None,
    progress_cb: Optional[ProgressCallback] = None,
) -> Dict[str, Any]:
    """Run the full Base RMC fill pipeline.

    Args:
        base_rmc_template: Unfilled Base RMC workbook (file #1)
        purchase_register_file: Purchase Register (file #2)
        stores_file: Stores Recordings / RM Film Stock (file #3)
        jobtrack_file: Optional enriched Jobtrack (if not provided, uses template's Jobtrack sheet)
        filled_jobtrack_file: Pre-filled Jobtrack from ipp2_ref engine (100% accurate rates)
        granules_file: Granules Recipe current month (file #4)
        prev_granules_file: Granules Recipe previous month (file #4b)
        ink_consumption_file: Ink Consumption (file #5)
        megapack_file: MEGAPACK Rate (file #6)
        opn_wip_file: Opening WIP Stock (file #9)
        cls_wip_file: Closing WIP Stock (file #10)
        valve_spout_file: Valve/Spout/Zipper prices (file #11)
        component_consumption_file: Components Consumption (file #12)
        filled_reference: Optional filled Base RMC for validation
        progress_cb: Optional progress callback(pct, msg)

    Returns:
        Dict with output_bytes, metrics, log, errors
    """
    t0 = time.time()
    ctx = RMCContext()

    def update(pct: int, msg: str):
        ctx._log(f"[{pct}%] {msg}")
        if progress_cb:
            progress_cb(pct, msg)

    # ══════════════════════════════════════════════════════════════
    # PHASE 1: Load all input files (0-10%)
    # ══════════════════════════════════════════════════════════════
    update(2, "Loading all input files...")
    load_all_into_context(
        ctx,
        base_rmc_template=base_rmc_template,
        purchase_register_file=purchase_register_file,
        stores_file=stores_file,
        granules_file=granules_file,
        prev_granules_file=prev_granules_file,
        ink_consumption_file=ink_consumption_file,
        megapack_file=megapack_file,
        opn_wip_file=opn_wip_file,
        cls_wip_file=cls_wip_file,
        valve_spout_file=valve_spout_file,
        component_consumption_file=component_consumption_file,
        ink_stock_opening_file=ink_stock_opening_file,
        dispensed_movement_file=dispensed_movement_file,
    )
    update(10, f"Files loaded. Template: {len(ctx.wb.sheetnames)} sheets")

    # ══════════════════════════════════════════════════════════════
    # PHASE 2: Fill Jobtrack with MRR rates (10-35%)
    # ══════════════════════════════════════════════════════════════
    if filled_jobtrack_file:
        update(12, "Loading PRE-FILLED Jobtrack (from ipp2_ref engine)...")
        _load_filled_jobtrack(ctx, filled_jobtrack_file)
        update(35, f"Pre-filled Jobtrack loaded: {len(ctx.jobtrack_df)} rows with rates")
    else:
        update(12, "Enriching Jobtrack with MRR rates...")
        _fill_jobtrack_phase(ctx, jobtrack_file, stores_file, purchase_register_file,
                             granules_file, prev_granules_file, megapack_file, update)
        update(35, "Jobtrack MRR fill complete")

    # ══════════════════════════════════════════════════════════════
    # PHASE 3: Build OPN_WIP (35-40%)
    # ══════════════════════════════════════════════════════════════
    update(36, "Building OPN_WIP...")
    fill_opn_wip(ctx)
    update(40, f"OPN_WIP: {len(ctx.opn_wip_by_key)} entries")

    # ══════════════════════════════════════════════════════════════
    # PHASE 4: Read enriched Jobtrack into process DataFrames (40-45%)
    # ══════════════════════════════════════════════════════════════
    update(42, "Reading Jobtrack into process-level views...")
    _build_jobtrack_views(ctx)
    update(45, "Jobtrack views built")

    # ══════════════════════════════════════════════════════════════
    # PHASE 5-12: Fill process sheets (45-80%)
    # ══════════════════════════════════════════════════════════════
    update(46, "Filling BFL (Extrusion)...")
    from engine.base_rmc.bfl_filler import fill_bfl
    fill_bfl(ctx)
    update(50, f"BFL done: {len(ctx.bfl_by_order)} orders")

    update(51, "Filling Print...")
    from engine.base_rmc.print_filler import fill_print
    fill_print(ctx)
    update(55, f"Print done: {len(ctx.print_rate_cache)} rates cached")

    update(56, "Filling Lam (Lamination)...")
    from engine.base_rmc.lam_filler import fill_lam
    fill_lam(ctx)
    update(62, f"Lam done: {len(ctx.pivot_lam_rates)} pivot rates")

    update(63, "Building Pivot_Lam Rates sheet...")
    from engine.base_rmc.remaining_fillers import build_pivot_lam_rates
    build_pivot_lam_rates(ctx)
    update(65, "Pivot_Lam Rates built")

    update(66, "Filling Slit...")
    from engine.base_rmc.slit_filler import fill_slit
    fill_slit(ctx)
    update(69, f"Slit done: {len(ctx.slit_rate_cache)} rates cached")

    update(70, "Filling Bag&Pouch...")
    from engine.base_rmc.bag_pouch_filler import fill_bag_pouch
    fill_bag_pouch(ctx)
    update(73, f"Bag&Pouch done: {len(ctx.bp_rate_cache)} rates cached")

    update(74, "Filling Spout&Valve...")
    from engine.base_rmc.remaining_fillers import fill_spout_valve
    fill_spout_valve(ctx)
    update(76, f"Spout&Valve done: {len(ctx.spout_valve_by_order)} orders")

    update(77, "Filling HCI Rew, PTR Rew, Embossing...")
    from engine.base_rmc.remaining_fillers import fill_hci_rew, fill_ptr_rew, fill_embossing
    fill_hci_rew(ctx)
    fill_ptr_rew(ctx)
    fill_embossing(ctx)
    update(80, "Rewinder sheets done")

    # ══════════════════════════════════════════════════════════════
    # PHASE 13-14: FG + CLS_WIP (80-88%)
    # ══════════════════════════════════════════════════════════════
    update(81, "Building FG...")
    from engine.base_rmc.remaining_fillers import fill_fg
    fill_fg(ctx)
    update(84, f"FG done: {len(ctx.fg_by_order)} orders")

    update(85, "Building CLS_WIP (rate cascade)...")
    from engine.base_rmc.remaining_fillers import fill_cls_wip
    fill_cls_wip(ctx)
    update(88, f"CLS_WIP done: {len(ctx.cls_wip_by_key)} keys")

    # ══════════════════════════════════════════════════════════════
    # PHASE 15-16: RMC Summary + Overall Wastage (88-95%)
    # ══════════════════════════════════════════════════════════════
    update(89, "Filling RMC Summary...")
    from engine.base_rmc.rmc_summary_filler import fill_rmc_summary
    fill_rmc_summary(ctx)
    update(93, "RMC Summary done")

    update(94, "Overall Wastage (formulas only)...")
    # Overall Wastage uses cross-sheet formulas — we leave Excel formulas in place
    ctx._log("NOTE: 'prnt wrkg pivot' and 'Printing Work' are Excel PivotTables — "
             "they CANNOT be refreshed by openpyxl. Open the output file in Excel "
             "and right-click → Refresh on each pivot to update them from the filled Print sheet.")
    ctx._log("NOTE: 'Pivot_Lam Rates' sheet is engine-filled by build_pivot_lam_rates().")
    update(95, "Overall Wastage done")

    # ══════════════════════════════════════════════════════════════
    # PHASE 17: Save output + Validate (95-100%)
    # ══════════════════════════════════════════════════════════════
    update(96, "Saving output workbook...")
    output_buf = io.BytesIO()
    ctx.wb.save(output_buf)
    output_bytes = output_buf.getvalue()

    validation = {}
    if filled_reference:
        update(98, "Validating against reference...")
        # from engine.base_rmc.validator import validate_against_reference
        # validation = validate_against_reference(ctx, filled_reference)
        validation = {"status": "placeholder"}

    elapsed = time.time() - t0
    update(100, f"Done! {elapsed:.1f}s")

    return {
        "output_bytes": output_bytes,
        "metrics": {
            "elapsed_seconds": round(elapsed, 1),
            "opn_wip_entries": len(ctx.opn_wip_by_key),
            "sheets_filled": len(ctx.wb.sheetnames),
        },
        "log": ctx.log,
        "errors": ctx.errors,
        "validation": validation,
    }


def _fill_jobtrack_phase(ctx, jobtrack_file, stores_file, pr_file,
                          granules_file, prev_granules_file, megapack_file, update):
    """Use existing engine/fill_jobtrack.py to enrich the Jobtrack sheet."""
    if jobtrack_file is None:
        ctx._log("  No separate Jobtrack file — using template's Jobtrack sheet")
        # Export template's Jobtrack sheet as XLSX and feed through fill_jobtrack()
        if 'Jobtrack' not in ctx.wb.sheetnames:
            ctx._error("  No Jobtrack sheet in template!")
            return

        # Create a temporary workbook with just the Jobtrack sheet
        import openpyxl
        from copy import copy
        tmp_wb = openpyxl.Workbook()
        tmp_ws = tmp_wb.active
        src_ws = ctx.wb['Jobtrack']

        # Copy all data from template Jobtrack to temp workbook
        for r in range(1, src_ws.max_row + 1):
            for c in range(1, src_ws.max_column + 1):
                val = src_ws.cell(row=r, column=c).value
                tmp_ws.cell(row=r, column=c, value=val)

        # Save to bytes
        tmp_buf = io.BytesIO()
        tmp_wb.save(tmp_buf)
        tmp_buf.seek(0)
        jobtrack_file = tmp_buf
        ctx._log(f"  Exported Jobtrack: {src_ws.max_row} rows, {src_ws.max_column} cols")


    # Use the full fill_jobtrack engine
    try:
        from engine.fill_jobtrack import fill_jobtrack

        def jt_progress(pct, msg):
            mapped = 12 + int(pct * 0.23)
            update(mapped, f"Jobtrack: {msg}")

        enriched_bytes, jt_log, jt_stats = fill_jobtrack(
            jt_file=jobtrack_file,
            stores_file=stores_file,
            pr_file=pr_file,
            granules_file=granules_file,
            megapack_file=megapack_file,
            prev_granules_file=prev_granules_file,
            progress_callback=jt_progress,
        )

        ctx._log(f"  Jobtrack fill stats: {jt_stats}")

        # Load enriched Jobtrack into DataFrame
        import pandas as pd
        if hasattr(enriched_bytes, 'read'):
            enriched_bytes.seek(0)
            raw = enriched_bytes.read()
        elif isinstance(enriched_bytes, bytes):
            raw = enriched_bytes
        else:
            raw = enriched_bytes

        ctx._log(f"  Enriched Jobtrack: writing rates back...")

        # Write enriched rate columns back into ctx.wb['Jobtrack']
        import openpyxl as _opx
        enriched_wb = _opx.load_workbook(io.BytesIO(raw))
        enriched_ws = enriched_wb.active
        tgt_ws = ctx.wb['Jobtrack']
        # Rate columns to copy: 54-102 (Film MR, Rate, Value, Fresh1/2, Adh, Hard, Sol)
        RATE_COLS = list(range(54, 103))
        copied = 0
        for r in range(5, enriched_ws.max_row + 1):
            for c in RATE_COLS:
                v = enriched_ws.cell(row=r, column=c).value
                if v is not None:
                    tgt_ws.cell(row=r, column=c, value=v)
                    copied += 1
        enriched_wb.close()
        ctx._log(f"  Wrote {copied} enriched cells back to template Jobtrack")

        # Build ctx.jobtrack_df by reading from ctx.wb['Jobtrack'] with formula eval
        import re as _re

        def _eval_formula(val):
            """Evaluate simple additive Excel formulas like =326+326+22+16."""
            if val is None:
                return None
            if not isinstance(val, str) or not val.startswith('='):
                return val
            expr = val[1:].strip()
            # Handle =SUM(AY5+AZ5) -> strip SUM wrapper
            m = _re.match(r'^SUM\((.+)\)$', expr, _re.IGNORECASE)
            if m:
                expr = m.group(1)
            # Reject if contains cell references (letters)
            if _re.search(r'[A-Za-z]', expr):
                return None  # Can't evaluate cell references
            # Try to evaluate simple arithmetic
            try:
                parts = _re.split(r'(?=[+-])', expr)
                total = sum(float(p.strip()) for p in parts if p.strip())
                return total
            except (ValueError, TypeError):
                return None

        ws_jt = ctx.wb['Jobtrack']
        headers = []
        for c in range(1, ws_jt.max_column + 1):
            h = ws_jt.cell(row=4, column=c).value
            headers.append(str(h).strip() if h else f'Col_{c}')

        data = []
        for r in range(5, ws_jt.max_row + 1):
            row_data = {}
            has_data = False
            for c in range(1, ws_jt.max_column + 1):
                v = ws_jt.cell(row=r, column=c).value
                v = _eval_formula(v)
                if v is not None:
                    has_data = True
                row_data[headers[c - 1]] = v
            if has_data:
                data.append(row_data)

        ctx.jobtrack_df = pd.DataFrame(data)
        ctx._log(f"  Enriched Jobtrack DataFrame: {len(ctx.jobtrack_df)} rows, {len(ctx.jobtrack_df.columns)} cols")

    except Exception as e:
        ctx._error(f"  Jobtrack fill failed: {e}")
        import traceback
        ctx._error(traceback.format_exc())


def _load_filled_jobtrack(ctx, filled_jobtrack_file):
    """Load a pre-filled Jobtrack (from ipp2_ref engine) directly into ctx.jobtrack_df.

    The filled Jobtrack already has correct MRR#, Rate, and Value columns
    computed by the ipp2_ref engine. We just need to read them as a DataFrame.
    """
    import openpyxl
    import re as _re

    ctx._log(f"  Loading filled Jobtrack: {filled_jobtrack_file}")

    # Open with data_only=True to read computed values (not formulas)
    wb_data = openpyxl.load_workbook(filled_jobtrack_file, data_only=True, read_only=True)
    ws = wb_data.active

    # Read headers from row 4
    headers = []
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=4, column=c).value
        headers.append(str(h).strip() if h else f'Col_{c}')

    ctx._log(f"  Filled Jobtrack: {ws.max_row} rows x {ws.max_column} cols")

    # Read all data rows (row 5+) into DataFrame
    data = []
    for r in range(5, ws.max_row + 1):
        row_data = {}
        has_data = False
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                has_data = True
            row_data[headers[c - 1]] = v
        if has_data:
            data.append(row_data)

    wb_data.close()

    ctx.jobtrack_df = pd.DataFrame(data)
    ctx._log(f"  Filled Jobtrack DataFrame: {len(ctx.jobtrack_df)} rows, "
             f"{len(ctx.jobtrack_df.columns)} cols")

    # Quick stats: how many rows have rates filled
    rate_col = None
    film_val_col = None
    for c in ctx.jobtrack_df.columns:
        cl = str(c).lower().strip()
        if cl == 'rate':
            rate_col = c
        if 'film value' in cl:
            film_val_col = c

    if rate_col:
        rate_filled = ctx.jobtrack_df[rate_col].apply(
            lambda x: x is not None and _safe_float(x) > 0
        ).sum()
        ctx._log(f"  Film Rate filled: {rate_filled}/{len(ctx.jobtrack_df)}")
    if film_val_col:
        val_filled = ctx.jobtrack_df[film_val_col].apply(
            lambda x: x is not None and _safe_float(x) > 0
        ).sum()
        ctx._log(f"  Film Value filled: {val_filled}/{len(ctx.jobtrack_df)}")


def _build_jobtrack_views(ctx):
    """Split Jobtrack data into process-level views for each filler."""
    if ctx.jobtrack_df is None or ctx.jobtrack_df.empty:
        ctx._log("  No Jobtrack data to build views from")
        return

    df = ctx.jobtrack_df

    # Find process column
    proc_col = None
    order_col = None
    for c in df.columns:
        cl = str(c).lower().strip()
        if cl == 'process':
            proc_col = c
        elif cl == 'order no' or cl == 'order_no' or 'order' in cl:
            if order_col is None:
                order_col = c

    if not proc_col:
        ctx._log("  Could not find Process column in Jobtrack")
        return

    # Detect report month from date column
    date_col = None
    for c in df.columns:
        if 'date' in str(c).lower():
            date_col = c
            break
    if date_col:
        dates = pd.to_datetime(df[date_col], errors='coerce').dropna()
        if not dates.empty:
            most_common = dates.dt.to_period('M').mode()
            if len(most_common) > 0:
                ctx.report_month_num = most_common[0].month
                ctx.report_year = most_common[0].year
                ctx.report_month = f"{ctx.report_month_num}-{ctx.report_year}"
                ctx._log(f"  Detected report month: {ctx.report_month}")

    # Detect solvent rate
    from engine.rate_lookup import lookup_solvent_rate
    ctx.solvent_rate = lookup_solvent_rate(ctx.purchase_register, report_month=ctx.report_month)
    ctx._log(f"  Solvent rate: {ctx.solvent_rate:.4f}")

    # Build ink cost index
    ink_costs = build_ink_cost_by_order({'summary': ctx.ink_summary} if ctx.ink_summary is not None else {})
    ctx.ink_rate_cache = ink_costs
    ctx._log(f"  Ink costs: {len(ink_costs)} orders")

    # Collect unique orders
    if order_col:
        ctx.order_list = sorted(df[order_col].dropna().astype(str).str.strip().unique().tolist())
        ctx._log(f"  Unique orders: {len(ctx.order_list)}")



    # Process-level split
    process_counts = {}
    for _, row in df.iterrows():
        proc = str(row.get(proc_col, '')).strip().upper()
        process_counts[proc] = process_counts.get(proc, 0) + 1

    ctx._log(f"  Process distribution: {process_counts}")
