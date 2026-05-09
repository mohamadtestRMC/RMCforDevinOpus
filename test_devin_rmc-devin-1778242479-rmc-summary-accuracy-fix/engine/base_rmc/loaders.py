"""
File Loaders — Load and parse all 13 input files into the RMCContext.

Each loader normalizes column names, detects header rows, and returns
clean DataFrames or dicts ready for use by the fillers.
"""
from __future__ import annotations
import logging
import io
import os
import pandas as pd
import openpyxl

from engine.base_rmc.context import RMCContext

logger = logging.getLogger(__name__)


def _safe_float(val) -> float:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _to_bytes(file_input) -> bytes:
    """Convert file input (path, BytesIO, or UploadedFile) to bytes."""
    if isinstance(file_input, bytes):
        return file_input
    if hasattr(file_input, 'read'):
        file_input.seek(0)
        data = file_input.read()
        file_input.seek(0)
        return data
    with open(str(file_input), 'rb') as f:
        return f.read()


# ═══════════════════════════════════════════════════════════════
# 1. Purchase Register
# ═══════════════════════════════════════════════════════════════
def load_purchase_register(file_input) -> pd.DataFrame:
    """Load Purchase Register (file #2). Header at row 3 (0-indexed row 2)."""
    df = pd.read_excel(io.BytesIO(_to_bytes(file_input)), sheet_name=0, header=2)
    col_map = {c: str(c).strip().replace('\n', ' ').replace('\r', '') for c in df.columns}
    df.rename(columns=col_map, inplace=True)
    logger.info(f"Purchase Register: {len(df)} rows, cols: {list(df.columns[:10])}")
    return df


# ═══════════════════════════════════════════════════════════════
# 2. Stores Recordings / RM FILM STOCK
# ═══════════════════════════════════════════════════════════════
def load_stores_recordings(file_input) -> pd.DataFrame:
    """Load Stores Recordings (file #3). Header at row 2 (0-indexed row 1)."""
    df = pd.read_excel(io.BytesIO(_to_bytes(file_input)), sheet_name=0, header=1)
    col_map = {c: str(c).strip().replace('\n', ' ').replace('\r', '') for c in df.columns}
    df.rename(columns=col_map, inplace=True)
    logger.info(f"Stores Recordings: {len(df)} rows, cols: {list(df.columns[:10])}")
    return df


# ═══════════════════════════════════════════════════════════════
# 3. Granules Recipe
# ═══════════════════════════════════════════════════════════════
def load_granules_rates(file_input) -> dict:
    """Load WO# → Rate AED from Granules Recipe (file #4).
    Uses first sheet (most recent month).
    Returns: {WO_upper: rate_float}
    """
    if file_input is None:
        return {}
    data = _to_bytes(file_input)
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Find header row with "WO #"
    header_row, wo_col = None, None
    for r in range(1, 15):
        for c in range(1, 6):
            v = ws.cell(row=r, column=c).value
            if v and isinstance(v, str) and v.strip().upper() in ('WO #', 'WO#', 'WO  #'):
                header_row, wo_col = r, c
                break
        if header_row:
            break
    if not header_row:
        wb.close()
        return {}

    # Find "Rate AED" column
    rate_col = None
    for c in range(1, ws.max_column + 1):
        for sr in [header_row, max(1, header_row - 2), max(1, header_row - 1)]:
            v = ws.cell(row=sr, column=c).value
            if v and isinstance(v, str) and 'rate' in v.lower() and 'aed' in v.lower():
                rate_col = c
                break
        if rate_col:
            break
    if not rate_col:
        wb.close()
        return {}

    rates = {}
    for r in range(header_row + 1, ws.max_row + 1):
        wo = ws.cell(row=r, column=wo_col).value
        rate = ws.cell(row=r, column=rate_col).value
        if wo and rate and isinstance(rate, (int, float)):
            wo_key = str(wo).strip().upper()
            if wo_key not in ('WO #', 'WO#', 'V'):
                rates[wo_key] = float(rate)

    wb.close()
    logger.info(f"Granules Recipe: {len(rates)} WO# rates loaded")
    return rates


# ═══════════════════════════════════════════════════════════════
# 4. Ink Consumption
# ═══════════════════════════════════════════════════════════════
def load_ink_consumption(file_input) -> dict:
    """Load Ink Consumption (file #5).
    Returns dict with:
        'summary': DataFrame (WO# → Ink & Solvent Cost)
        'calculation': DataFrame
        'ink_rates': DataFrame (Feb-2026 rates)
    """
    if file_input is None:
        return {}
    data = _to_bytes(file_input)

    result = {}
    # Summary sheet: WO# -> Ink & Solvent Cost
    try:
        summary = pd.read_excel(io.BytesIO(data), sheet_name='Summary', header=1)
        col_map = {c: str(c).strip().replace('\n', ' ').replace('\r', '') for c in summary.columns}
        summary.rename(columns=col_map, inplace=True)
        result['summary'] = summary
        logger.info(f"Ink Summary: {len(summary)} rows")
    except Exception as e:
        logger.warning(f"Could not load Ink Summary: {e}")

    # Calculation sheet
    try:
        calc = pd.read_excel(io.BytesIO(data), sheet_name='Calculation', header=1)
        result['calculation'] = calc
    except Exception as e:
        logger.warning(f"Could not load Ink Calculation: {e}")

    return result


def build_ink_cost_by_order(ink_data: dict) -> dict:
    """Build {order_no: ink_solvent_cost} from ink consumption data."""
    if not ink_data or 'summary' not in ink_data:
        return {}
    summary = ink_data['summary']

    # Find WO# and Cost columns
    wo_col, cost_col = None, None
    for c in summary.columns:
        cl = str(c).lower().strip()
        if 'wo' in cl:
            wo_col = c
        elif 'ink' in cl and 'solvent' in cl and 'cost' in cl:
            cost_col = c
        elif 'cost' in cl and cost_col is None:
            cost_col = c

    if not wo_col or not cost_col:
        logger.warning(f"Ink summary columns not found. Available: {list(summary.columns)}")
        return {}

    result = {}
    for _, row in summary.iterrows():
        wo = str(row.get(wo_col, '')).strip().upper()
        cost = _safe_float(row.get(cost_col))
        if wo and cost > 0:
            result[wo] = result.get(wo, 0) + cost

    logger.info(f"Ink cost by order: {len(result)} orders")
    return result


# ═══════════════════════════════════════════════════════════════
# 5. MEGAPACK Rate
# ═══════════════════════════════════════════════════════════════
def load_megapack_rates(file_input) -> dict:
    """Load MEGA PACK monthly TPE/WPE rates (file #6).
    Returns: {(year, month): {'TPE': rate, 'WPE': rate}}
    """
    if file_input is None:
        return {}
    data = _to_bytes(file_input)
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.active

    rates = {}
    for r in range(5, ws.max_row + 1):
        date_val = ws.cell(row=r, column=2).value
        tpe_conv = ws.cell(row=r, column=7).value
        wpe_conv = ws.cell(row=r, column=8).value
        if date_val and hasattr(date_val, 'year'):
            rates[(date_val.year, date_val.month)] = {
                'TPE': _safe_float(tpe_conv),
                'WPE': _safe_float(wpe_conv),
            }
    wb.close()
    logger.info(f"MEGAPACK: {len(rates)} monthly rates")
    return rates


# ═══════════════════════════════════════════════════════════════
# 6. Opening WIP Stock
# ═══════════════════════════════════════════════════════════════
def load_opn_wip(file_input) -> pd.DataFrame:
    """Load Opening WIP Stock (file #9).
    Header at row 5 (1-indexed), data starts row 6.
    Cols: W/O, Design Name, Mat Structure, Process, Substrate, Lam Pass, Qty, Rate, Value
    """
    if file_input is None:
        return pd.DataFrame()
    data = _to_bytes(file_input)
    df = pd.read_excel(io.BytesIO(data), sheet_name=0, header=4)
    col_map = {c: str(c).strip().replace('\n', ' ').replace('\r', '') for c in df.columns}
    df.rename(columns=col_map, inplace=True)
    # Drop rows where W/O is NaN (empty rows at bottom)
    wo_col = None
    for c in df.columns:
        if 'w/o' in str(c).lower() or c == 'W/O':
            wo_col = c
            break
    if wo_col:
        df = df.dropna(subset=[wo_col])
    logger.info(f"Opening WIP: {len(df)} rows, cols: {list(df.columns)}")
    return df


# ═══════════════════════════════════════════════════════════════
# 7. Closing WIP Stock
# ═══════════════════════════════════════════════════════════════
def load_cls_wip(file_input) -> pd.DataFrame:
    """Load Closing WIP Stock (file #10). Same structure as OPN_WIP but Rate/Value blank."""
    if file_input is None:
        return pd.DataFrame()
    data = _to_bytes(file_input)
    df = pd.read_excel(io.BytesIO(data), sheet_name=0, header=4)
    col_map = {c: str(c).strip().replace('\n', ' ').replace('\r', '') for c in df.columns}
    df.rename(columns=col_map, inplace=True)
    wo_col = None
    for c in df.columns:
        if 'w/o' in str(c).lower() or c == 'W/O':
            wo_col = c
            break
    if wo_col:
        df = df.dropna(subset=[wo_col])
    logger.info(f"Closing WIP: {len(df)} rows, cols: {list(df.columns)}")
    return df


# ═══════════════════════════════════════════════════════════════
# 8. Valve / Spout / Tin Tie Prices
# ═══════════════════════════════════════════════════════════════
def load_valve_spout_prices(file_input) -> dict:
    """Load component prices from file #11.
    Returns: {
        'valve_spout_tintie': [{name, rate_per_pc, rate_per_kg}, ...],
        'zipper': [{code, ipp_code, rate_per_kg}, ...],
    }
    """
    if file_input is None:
        return {}
    data = _to_bytes(file_input)
    result = {}

    # Sheet "Valve, Spout & Tintie" — latest prices
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)

        # Valve, Spout & Tintie sheet
        for sn in wb.sheetnames:
            if 'valve' in sn.lower() and 'spout' in sn.lower():
                ws = wb[sn]
                items = []
                for r in range(3, ws.max_row + 1):
                    name = ws.cell(row=r, column=3).value
                    rpc = ws.cell(row=r, column=4).value
                    rpk = ws.cell(row=r, column=5).value
                    if name:
                        items.append({
                            'name': str(name).strip(),
                            'rate_per_pc': _safe_float(rpc),
                            'rate_per_kg': _safe_float(rpk),
                        })
                result['valve_spout_tintie'] = items
                break

        # Zipper sheet
        for sn in wb.sheetnames:
            if 'zipper' in sn.lower():
                ws = wb[sn]
                zippers = []
                for r in range(3, ws.max_row + 1):
                    code = ws.cell(row=r, column=2).value
                    ipp_code = ws.cell(row=r, column=3).value
                    rate = ws.cell(row=r, column=4).value
                    if code:
                        zippers.append({
                            'code': str(code).strip(),
                            'ipp_code': str(ipp_code).strip() if ipp_code else '',
                            'rate_per_kg': _safe_float(rate),
                        })
                result['zipper'] = zippers
                break

        wb.close()
    except Exception as e:
        logger.error(f"Error loading valve/spout prices: {e}")

    logger.info(f"Valve/Spout/Zipper: {len(result.get('valve_spout_tintie', []))} items, "
                f"{len(result.get('zipper', []))} zippers")
    return result


# ═══════════════════════════════════════════════════════════════
# 9. Components Consumption Dispensed Details
# ═══════════════════════════════════════════════════════════════
def load_component_consumption(file_input) -> pd.DataFrame:
    """Load Components Consumptions (file #12).
    This is a production report with variable header positions.
    """
    if file_input is None:
        return pd.DataFrame()
    data = _to_bytes(file_input)

    # Try to find header row by scanning for known headers
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.active
    header_row = None
    for r in range(1, 30):
        for c in range(1, 10):
            v = ws.cell(row=r, column=c).value
            if v and isinstance(v, str) and 'work order' in v.lower():
                header_row = r
                break
        if header_row:
            break

    wb.close()

    if header_row:
        df = pd.read_excel(io.BytesIO(data), sheet_name=0, header=header_row - 1)
    else:
        df = pd.read_excel(io.BytesIO(data), sheet_name=0)

    col_map = {c: str(c).strip().replace('\n', ' ').replace('\r', '') for c in df.columns}
    df.rename(columns=col_map, inplace=True)
    logger.info(f"Components Consumption: {len(df)} rows, cols: {list(df.columns[:10])}")
    return df


# ═══════════════════════════════════════════════════════════════
# 10. Ink Stock Opening & Dispensed Movement (files #7, #8)
# ═══════════════════════════════════════════════════════════════
def load_ink_stock_opening(file_input) -> pd.DataFrame:
    """Load Dispense Ink Stock Opening (file #7)."""
    if file_input is None:
        return pd.DataFrame()
    return pd.read_excel(io.BytesIO(_to_bytes(file_input)), sheet_name=0, header=0)


def load_dispensed_movement(file_input) -> pd.DataFrame:
    """Load Dispensed Stock Movement (file #8)."""
    if file_input is None:
        return pd.DataFrame()
    return pd.read_excel(io.BytesIO(_to_bytes(file_input)), sheet_name=0, header=0)


# ═══════════════════════════════════════════════════════════════
# TEMPLATE XML STRIPPER — removes 1M empty rows from OPN_WIP/CLS_WIP
# ═══════════════════════════════════════════════════════════════
def _strip_template_xml(src_path: str, dst_path: str, ctx: RMCContext) -> None:
    """Strip 1M empty rows from OPN_WIP/CLS_WIP sheets inside the XLSX ZIP.

    The Base RMC template has formatting applied to row 1,048,576 in
    OPN_WIP and CLS_WIP sheets, causing openpyxl to process 1M rows.
    This function removes those empty rows via proper XML parsing,
    reducing load time from infinite to ~12 minutes.
    """
    import zipfile, re
    from xml.etree import ElementTree as ET

    NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    # sheets to strip (identified by filename in the ZIP)
    STRIP_SHEETS = {
        'xl/worksheets/sheet26.xml': 300,  # OPN_WIP
        'xl/worksheets/sheet27.xml': 300,  # CLS_WIP
    }
    REMOVE_FILES = {'xl/calcChain.xml'}  # 5MB, openpyxl rebuilds it
    dim_re = re.compile(r'<dimension ref="([A-Z]+\d+):([A-Z]+)\d+"')

    ctx._log("  Stripping template (removing 1M empty rows)...")

    with zipfile.ZipFile(src_path, 'r') as zin, \
         zipfile.ZipFile(dst_path, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            if item.filename in REMOVE_FILES:
                ctx._log(f"    Removed {item.filename} ({item.file_size:,} bytes)")
                continue

            data = zin.read(item.filename)

            if item.filename in STRIP_SHEETS:
                max_row = STRIP_SHEETS[item.filename]
                before = len(data)

                root = ET.fromstring(data)

                # Fix dimension tag
                dim_elem = root.find(f'{{{NS}}}dimension')
                if dim_elem is not None:
                    ref = dim_elem.get('ref', '')
                    parts = ref.split(':')
                    if len(parts) == 2:
                        col_part = ''.join(c for c in parts[1] if c.isalpha())
                        dim_elem.set('ref', f'{parts[0]}:{col_part}{max_row}')

                # Remove rows beyond max_row
                sheet_data = root.find(f'{{{NS}}}sheetData')
                if sheet_data is not None:
                    to_remove = [r for r in sheet_data.findall(f'{{{NS}}}row')
                                 if int(r.get('r', '0')) > max_row]
                    for r in to_remove:
                        sheet_data.remove(r)
                    ctx._log(f"    Stripped {item.filename}: removed {len(to_remove)} rows "
                             f"({before:,} -> est. bytes)")

                # Register namespaces to preserve XML structure
                ET.register_namespace('', NS)
                ET.register_namespace('r', 'http://schemas.openxmlformats.org/officeDocument/2006/relationships')
                ET.register_namespace('mc', 'http://schemas.openxmlformats.org/markup-compatibility/2006')
                ET.register_namespace('x14ac', 'http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac')

                data = ET.tostring(root, xml_declaration=True, encoding='UTF-8')

            zout.writestr(item, data)

    src_size = os.path.getsize(src_path)
    dst_size = os.path.getsize(dst_path)
    ctx._log(f"  Template stripped: {src_size:,} -> {dst_size:,} bytes "
             f"({100*dst_size//src_size}%)")


# ═══════════════════════════════════════════════════════════════
# MASTER LOADER — loads everything into RMCContext
# ═══════════════════════════════════════════════════════════════
def load_all_into_context(
    ctx: RMCContext,
    *,
    base_rmc_template,
    purchase_register_file,
    stores_file,
    granules_file=None,
    prev_granules_file=None,
    ink_consumption_file=None,
    megapack_file=None,
    ink_stock_opening_file=None,
    dispensed_movement_file=None,
    opn_wip_file=None,
    cls_wip_file=None,
    valve_spout_file=None,
    component_consumption_file=None,
) -> None:
    """Load all input files into the RMCContext."""

    ctx._log("Loading Base RMC template (this may take a few minutes)...")
    # Strip 1M empty rows from OPN_WIP/CLS_WIP before openpyxl loads
    import shutil, tempfile
    if isinstance(base_rmc_template, (str, os.PathLike)):
        src_path = str(base_rmc_template)
    else:
        # Write uploaded file to temp
        tmp_src = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        tmp_src.write(_to_bytes(base_rmc_template))
        tmp_src.close()
        src_path = tmp_src.name

    # Strip the template (removes 1M empty rows from OPN_WIP/CLS_WIP)
    stripped_path = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False).name
    _strip_template_xml(src_path, stripped_path, ctx)

    ctx.wb = openpyxl.load_workbook(
        stripped_path, keep_vba=False, keep_links=False,
    )
    ctx._template_tmp = stripped_path  # keep ref for cleanup
    ctx._log(f"  Template sheets: {ctx.wb.sheetnames}")

    ctx._log("Loading Purchase Register...")
    ctx.purchase_register = load_purchase_register(purchase_register_file)

    ctx._log("Loading Stores Recordings...")
    ctx.stores_recordings = load_stores_recordings(stores_file)

    ctx._log("Loading Granules Recipe...")
    ctx.granules_rates = load_granules_rates(granules_file)
    ctx.prev_granules_rates = load_granules_rates(prev_granules_file)

    ctx._log("Loading Ink Consumption...")
    ink_data = load_ink_consumption(ink_consumption_file)
    if 'summary' in ink_data:
        ctx.ink_summary = ink_data['summary']
    if 'calculation' in ink_data:
        ctx.ink_calculation = ink_data['calculation']

    ctx._log("Loading MEGAPACK rates...")
    ctx.megapack_rates = load_megapack_rates(megapack_file)

    ctx._log("Loading Opening WIP...")
    ctx.opn_wip_df = load_opn_wip(opn_wip_file)

    ctx._log("Loading Closing WIP...")
    ctx.cls_wip_df = load_cls_wip(cls_wip_file)

    ctx._log("Loading Valve/Spout/Zipper prices...")
    prices = load_valve_spout_prices(valve_spout_file)
    ctx.valve_spout_prices = prices.get('valve_spout_tintie', [])
    ctx.zipper_prices = prices.get('zipper', [])

    ctx._log("Loading Components Consumption...")
    ctx.component_consumption = load_component_consumption(component_consumption_file)

    # Build MRR → Supplier map from Stores
    from engine.supplier_rates import build_mrr_supplier_map
    ctx.mrr_supplier_map = build_mrr_supplier_map(ctx.stores_recordings)

    ctx._log(f"All files loaded. Template has {len(ctx.wb.sheetnames)} sheets.")
