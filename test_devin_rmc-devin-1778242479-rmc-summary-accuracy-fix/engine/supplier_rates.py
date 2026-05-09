"""
Supplier-specific rate lookups.
- Bandera / CYM → Granules Recipe (Rate AED by WO#)
- Mega Pack → MEGA PACK.xlsx (TPE/WPE converted rate by month)
"""
import logging
import shutil
import tempfile
import os

import openpyxl
import pandas as pd

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────
# Stores → Supplier Map
# ─────────────────────────────────────────────────────────────

SPECIAL_SUPPLIERS = ('BANDERA', 'CYM', 'MEGA PACK')


def build_mrr_supplier_map(stores_df: pd.DataFrame) -> dict:
    """
    Build a dict mapping MRR number → supplier name from Stores Recordings.
    Uses the raw DataFrame loaded with header at row 2.
    """
    # Find the supplier column (Col 6 in the Excel = 'Supplier Name')
    supplier_col = None
    for col in stores_df.columns:
        # The supplier column often has 'Unnamed: 5' due to merged headers
        # We detect it by checking for known supplier names
        vals = stores_df[col].astype(str).str.upper().str.strip()
        if vals.str.contains('BANDERA').any() or vals.str.contains('MEGA PACK').any():
            supplier_col = col
            break

    if supplier_col is None:
        logger.warning("Could not find supplier column in Stores Recordings")
        return {}

    # Find MRR column (Col 16 = 'M.R.R No.')
    mrr_col = None
    for col in stores_df.columns:
        vals = pd.to_numeric(stores_df[col], errors='coerce').dropna()
        if len(vals) > 100:
            # Check if values look like MRR numbers (5-digit)
            sample = vals.head(50)
            if (sample > 10000).sum() > 10 and (sample < 100000).sum() > 10:
                # Confirm by checking the column index position
                col_idx = stores_df.columns.get_loc(col)
                if 14 <= col_idx <= 17:  # MRR column is around position 15-16
                    mrr_col = col
                    break

    if mrr_col is None:
        # Fallback: try column at index 15 (0-based) = Col 16 in Excel
        if len(stores_df.columns) > 15:
            mrr_col = stores_df.columns[15]
            logger.info(f"Using fallback MRR column: {mrr_col}")
        else:
            logger.warning("Could not find MRR column in Stores Recordings")
            return {}

    result = {}
    for _, row in stores_df.iterrows():
        mrr_val = row.get(mrr_col)
        sup_val = str(row.get(supplier_col, '')).strip().upper()
        if pd.notna(mrr_val) and sup_val:
            try:
                mrr_num = int(float(mrr_val))
                result[mrr_num] = sup_val
            except (ValueError, TypeError):
                pass

    logger.info(f"Built MRR->Supplier map: {len(result)} entries, "
                f"special suppliers: "
                f"BANDERA={sum(1 for v in result.values() if 'BANDERA' in v)}, "
                f"CYM={sum(1 for v in result.values() if 'CYM' in v)}, "
                f"MEGA PACK={sum(1 for v in result.values() if 'MEGA' in v)}")
    return result


def get_supplier_for_mrrs(mrr_supplier_map: dict, mrr_numbers: list,
                          mrr_qty: dict = None) -> str:
    """
    Check if the given MRR numbers belong to a special supplier.
    
    When mrr_qty is provided (dict of mrr -> qty), the special supplier
    must account for >30% of total quantity to be considered dominant.
    This prevents a minor-qty BANDERA MRR from triggering INH when the
    dominant MRRs are from standard suppliers.
    
    Returns the supplier name or None.
    """
    if mrr_qty:
        # Qty-aware: aggregate qty per special supplier
        total_qty = sum(mrr_qty.values())
        special_qty = {}  # {supplier_name: total_qty}
        for mrr in mrr_numbers:
            try:
                mrr_num = int(float(mrr))
            except (ValueError, TypeError):
                continue
            sup = mrr_supplier_map.get(mrr_num, '')
            for special in SPECIAL_SUPPLIERS:
                if special in sup:
                    qty = mrr_qty.get(mrr, mrr_qty.get(str(mrr), 0))
                    special_qty[special] = special_qty.get(special, 0) + qty
        
        # Return the dominant special supplier (>30% of total qty)
        if special_qty and total_qty > 0:
            best = max(special_qty, key=special_qty.get)
            ratio = special_qty[best] / total_qty
            if ratio > 0.30:
                logger.info(f"Supplier {best}: {special_qty[best]:.1f}/{total_qty:.1f} "
                           f"= {ratio*100:.1f}% → dominant")
                return best
            else:
                logger.info(f"Supplier {best}: {special_qty[best]:.1f}/{total_qty:.1f} "
                           f"= {ratio*100:.1f}% → NOT dominant, treating as standard")
                return None
        return None
    
    # Fallback: no qty info, return first match (legacy behavior)
    for mrr in mrr_numbers:
        try:
            mrr_num = int(float(mrr))
        except (ValueError, TypeError):
            continue
        sup = mrr_supplier_map.get(mrr_num, '')
        for special in SPECIAL_SUPPLIERS:
            if special in sup:
                return special
    return None


# ─────────────────────────────────────────────────────────────
# Granules Recipe → Rate AED by WO#
# ─────────────────────────────────────────────────────────────

def load_granules_rates(granules_file) -> dict:
    """
    Load WO# → Rate AED mapping from the Granules Recipe file.
    Uses the FIRST sheet (most recent month).

    Returns: dict of {wo_number_upper: rate_aed}
    """
    try:
        # Handle BytesIO or file path — copy to temp if needed for locked files
        if hasattr(granules_file, 'read'):
            # BytesIO object
            tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
            tmp.write(granules_file.read())
            tmp.close()
            granules_file.seek(0)
            path = tmp.name
            cleanup = True
        else:
            path = str(granules_file)
            cleanup = False

        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]  # First sheet = most recent month

        # Find header row: row where a cell in cols 1-5 contains exactly "WO #" or "WO#"
        header_row = None
        wo_col = None
        for r in range(1, 15):
            for c in range(1, 6):
                v = ws.cell(row=r, column=c).value
                if v and isinstance(v, str):
                    v_clean = v.strip().upper()
                    if v_clean in ('WO #', 'WO#', 'WO  #'):
                        header_row = r
                        wo_col = c
                        break
            if header_row:
                break

        if not header_row:
            logger.warning("Granules Recipe: could not find WO# header row")
            wb.close()
            if cleanup:
                os.unlink(path)
            return {}

        # Find "Rate AED" column — search in header_row and nearby rows
        rate_col = None
        for c in range(1, ws.max_column + 1):
            for search_r in [header_row, max(1, header_row - 2), max(1, header_row - 1)]:
                v = ws.cell(row=search_r, column=c).value
                if v and isinstance(v, str) and 'rate' in v.lower() and 'aed' in v.lower():
                    rate_col = c
                    break
            if rate_col:
                break

        if not rate_col:
            logger.warning("Granules Recipe: could not find 'Rate AED' column")
            wb.close()
            if cleanup:
                os.unlink(path)
            return {}

        logger.info(f"Granules Recipe: header_row={header_row}, wo_col={wo_col}, "
                    f"rate_col={rate_col}, sheet='{wb.sheetnames[0]}'")

        # Build WO# → Rate map
        rates = {}
        for r in range(header_row + 1, ws.max_row + 1):
            wo = ws.cell(row=r, column=wo_col).value
            rate = ws.cell(row=r, column=rate_col).value
            if wo and rate and isinstance(rate, (int, float)):
                wo_key = str(wo).strip().upper()
                # Skip the header row itself (e.g. "WO #")
                if wo_key in ('WO #', 'WO#', 'WO  #', 'V'):
                    continue
                rates[wo_key] = float(rate)

        logger.info(f"Granules Recipe: loaded {len(rates)} WO# rates")
        for wo, rate in rates.items():
            logger.debug(f"  Granules: WO={wo} -> Rate={rate:.6f}")

        wb.close()
        if cleanup:
            os.unlink(path)
        return rates

    except Exception as e:
        logger.error(f"Error loading Granules Recipe: {e}")
        return {}


# ─────────────────────────────────────────────────────────────
# MEGA PACK → Rate by material type + month
# ─────────────────────────────────────────────────────────────

def load_megapack_rates(megapack_file) -> dict:
    """
    Load monthly TPE/WPE rates from MEGA PACK.xlsx.

    Structure:
        Row 3: Headers (C5=TPE, C6=WPE, C7=Conversion TPE, C8=Conversion WPE)
        Row 5+: Monthly data (C2=date, C4=label, C5-C8=rates)

    Returns: dict of {(year, month): {'TPE': rate, 'WPE': rate}}
    """
    try:
        if hasattr(megapack_file, 'read'):
            tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
            tmp.write(megapack_file.read())
            tmp.close()
            megapack_file.seek(0)
            path = tmp.name
            cleanup = True
        else:
            path = str(megapack_file)
            cleanup = False

        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active

        rates = {}
        for r in range(5, ws.max_row + 1):
            date_val = ws.cell(row=r, column=2).value
            tpe_conv = ws.cell(row=r, column=7).value  # TPE with conversion
            wpe_conv = ws.cell(row=r, column=8).value  # WPE with conversion

            if date_val and hasattr(date_val, 'year'):
                key = (date_val.year, date_val.month)
                rates[key] = {
                    'TPE': float(tpe_conv) if tpe_conv else 0.0,
                    'WPE': float(wpe_conv) if wpe_conv else 0.0,
                }
                logger.info(f"MEGA PACK: {date_val.year}-{date_val.month:02d} -> "
                            f"TPE={rates[key]['TPE']:.6f}, WPE={rates[key]['WPE']:.6f}")

        wb.close()
        if cleanup:
            os.unlink(path)
        return rates

    except Exception as e:
        logger.error(f"Error loading MEGA PACK: {e}")
        return {}


def lookup_megapack_rate(megapack_rates: dict, material_type: str,
                         report_year: int, report_month: int) -> float:
    """
    Get the Mega Pack rate for a material type (TPE/WPE) and month.
    """
    key = (report_year, report_month)
    if key not in megapack_rates:
        logger.warning(f"MEGA PACK: no rates for {report_year}-{report_month:02d}")
        return 0.0

    mat_upper = str(material_type).strip().upper()
    month_rates = megapack_rates[key]

    if 'WPE' in mat_upper or 'WLDPE' in mat_upper:
        return month_rates.get('WPE', 0.0)
    else:
        # Default to TPE for TPE, TPE EASY TEARABLE, etc.
        return month_rates.get('TPE', 0.0)
