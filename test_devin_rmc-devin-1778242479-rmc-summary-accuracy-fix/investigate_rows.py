"""
Deep investigation of failing rows 42, 43, 45, 46.
For each row: show ground truth, engine output, source data, and exact cause.
Also check row 44 per user request.
"""
import io
import os
import shutil
import json
import openpyxl
import pandas as pd
from engine.fill_jobtrack import fill_jobtrack, COLS, DATA_START_ROW, _safe_str, _safe_float
from engine.mrr_lookup import load_stores_recordings, lookup_mrr_with_qty
from engine.rate_lookup import (
    load_purchase_register, lookup_film_rate_weighted,
    filter_mrr_by_pr, lookup_material_rate_for_month, _find_col, _get_rate_for_mrr
)
from engine.supplier_rates import build_mrr_supplier_map, get_supplier_for_mrrs

BASE = "Template_Files"

# Load all data
stores_df = load_stores_recordings(f"{BASE}/Stores Recordings.xlsx")
pr_df = load_purchase_register(f"{BASE}/Purchase Register - 2021 - 2026 _Feb 26.xlsx")
mrr_sup_map = build_mrr_supplier_map(stores_df)

# Load ground truth (handle locked file)
_tmp_gt = f"{BASE}/Jobtrack_Feb_MRR_tmp.xlsx"
shutil.copy2(f"{BASE}/Jobtrack Feb With MRR.xlsx", _tmp_gt)
wb_gt = openpyxl.load_workbook(_tmp_gt, data_only=True)
ws_gt = wb_gt.active

# Run engine with supplier files
with open(f"{BASE}/Jobtrack Feb Without MRR.xlsx", "rb") as f:
    jt = io.BytesIO(f.read())
with open(f"{BASE}/Stores Recordings.xlsx", "rb") as f:
    stores = io.BytesIO(f.read())
with open(f"{BASE}/Purchase Register - 2021 - 2026 _Feb 26.xlsx", "rb") as f:
    pr = io.BytesIO(f.read())
with open(f"{BASE}/Granules Recipe - February 2026.xlsx", "rb") as f:
    granules = io.BytesIO(f.read())
tmp_mp = f"{BASE}/MEGA_PACK_tmp.xlsx"
shutil.copy2(f"{BASE}/MEGA PACK.xlsx", tmp_mp)
with open(tmp_mp, "rb") as f:
    megapack = io.BytesIO(f.read())
os.remove(tmp_mp)

output_bytes, results_log, fill_stats = fill_jobtrack(
    jt, stores, pr, granules_file=granules, megapack_file=megapack
)
output_bytes.seek(0)
wb_out = openpyxl.load_workbook(output_bytes, data_only=False)
ws_out = wb_out.active

# Columns to inspect
check_cols = {
    'UID': 1, 'Process': 6, 'Order_No': 11,
    'Input_Name': 47, 'Input_Size': 48, 'Input_Mic': 49,
    'Total_Input': 50,
    'Film_MR': 54, 'Film_Rate': 55, 'Film_Value': 56,
    'Fresh1_Name': 71, 'Fresh1_Size': 72, 'Fresh1_Mic': 73,
    'Total_Fresh1': 74,
    'Fresh1_MR': 78, 'Fresh1_Rate': 79, 'Fresh1_Value': 80,
    'Fresh2_Name': 81, 'Fresh2_Size': 82, 'Fresh2_Mic': 83,
    'Fresh2_MR': 88, 'Fresh2_Rate': 89, 'Fresh2_Value': 90,
}

report = []
report.append("=" * 100)
report.append("DETAILED INVESTIGATION: Rows 42, 43, 44, 45, 46")
report.append("=" * 100)

for row in [42, 43, 44, 45, 46]:
    uid = ws_gt.cell(row=row, column=1).value
    if not uid:
        report.append(f"\nRow {row}: EMPTY (no UID)")
        continue
    
    report.append(f"\n{'='*80}")
    report.append(f"ROW {row}")
    report.append(f"{'='*80}")
    
    process = str(ws_gt.cell(row=row, column=COLS['Process']).value or '').strip()
    order = str(ws_gt.cell(row=row, column=COLS['Order_No']).value or '').strip()
    
    report.append(f"  UID:     {uid}")
    report.append(f"  Process: {process}")
    report.append(f"  Order:   {order}")
    
    # Check each material type
    for label, mr_key, rate_key, val_key, name_key, size_key, mic_key in [
        ('Film', 'Film_MR', 'Film_Rate', 'Film_Value', 'Input_Name', 'Input_Size', 'Input_Mic'),
        ('Fresh1', 'Fresh1_MR', 'Fresh1_Rate', 'Fresh1_Value', 'Fresh1_Name', 'Fresh1_Size', 'Fresh1_Mic'),
        ('Fresh2', 'Fresh2_MR', 'Fresh2_Rate', 'Fresh2_Value', 'Fresh2_Name', 'Fresh2_Size', 'Fresh2_Mic'),
    ]:
        gt_mr = ws_gt.cell(row=row, column=check_cols[mr_key]).value
        gt_rate = ws_gt.cell(row=row, column=check_cols[rate_key]).value
        gt_val = ws_gt.cell(row=row, column=check_cols[val_key]).value
        
        eng_mr = ws_out.cell(row=row, column=check_cols[mr_key]).value
        eng_rate = ws_out.cell(row=row, column=check_cols[rate_key]).value
        eng_val = ws_out.cell(row=row, column=check_cols[val_key]).value
        
        mat_name = str(ws_gt.cell(row=row, column=check_cols[name_key]).value or '').strip()
        mat_size = ws_gt.cell(row=row, column=check_cols[size_key]).value
        mat_mic = ws_gt.cell(row=row, column=check_cols[mic_key]).value
        
        if not gt_mr and not gt_rate:
            continue
        
        rate_match = True
        if gt_rate and eng_rate:
            try:
                rate_match = abs(float(gt_rate) - float(eng_rate)) < 0.01
            except:
                rate_match = False
        
        status = "MATCH" if rate_match else "*** MISMATCH ***"
        report.append(f"\n  --- {label} ({mat_name}, Size={mat_size}, Mic={mat_mic}) ---")
        report.append(f"    Ground Truth:  MR={gt_mr}, Rate={gt_rate}, Value={gt_val}")
        report.append(f"    Engine Output: MR={eng_mr}, Rate={eng_rate}, Value={eng_val}")
        report.append(f"    Status: {status}")
        
        if not rate_match:
            report.append(f"\n    *** ROOT CAUSE ANALYSIS ***")
            
            # 1. What MRRs does Stores give for this order+material?
            mrr_qty = lookup_mrr_with_qty(stores_df, mat_name, mat_mic, mat_size,
                                           order, process)
            if not mrr_qty:
                mrr_qty = lookup_mrr_with_qty(stores_df, mat_name, mat_mic, None,
                                               order, process)
            if not mrr_qty:
                mrr_qty = lookup_mrr_with_qty(stores_df, mat_name, mat_mic, None, order)
            
            report.append(f"    Stores MRR discovery: {mrr_qty}")
            
            if mrr_qty:
                # 2. What supplier?
                for mrr_num in mrr_qty:
                    sup = mrr_sup_map.get(int(mrr_num), 'N/A')
                    report.append(f"      MRR {mrr_num} -> Supplier: {sup}")
                
                # 3. What rate does PR give for each MRR?
                tracking_col = _find_col(pr_df, 'tracking')
                rate_col = [c for c in pr_df.columns if str(c).strip().lower() == 'rate']
                material_col = _find_col(pr_df, 'material')
                month_col = _find_col(pr_df, 'month')
                
                if tracking_col and rate_col:
                    report.append(f"\n    PR Rate breakdown per MRR:")
                    for mrr_num, qty in mrr_qty.items():
                        pr_rate = lookup_film_rate_weighted(pr_df, {mrr_num: qty}, mat_name, mat_size, mat_mic)
                        report.append(f"      MRR {mrr_num}: Qty={qty}, PR Rate={pr_rate}")
                        
                        # Show the actual PR rows for this MRR
                        mrr_rows = pr_df[pr_df[tracking_col].astype(str).str.strip() == str(mrr_num)]
                        for _, pr_row in mrr_rows.iterrows():
                            mat = pr_row.get(material_col, 'N/A')
                            rate = pr_row.get(rate_col[0], 'N/A')
                            month = pr_row.get(month_col, 'N/A') if month_col else 'N/A'
                            report.append(f"        PR row: Material={mat}, Rate={rate}, Month={month}")
                
                # 4. What does the engine's weighted average compute?
                filtered = filter_mrr_by_pr(pr_df, mrr_qty.copy(), mat_name, mat_size, mat_mic)
                w_rate = lookup_film_rate_weighted(pr_df, filtered, mat_name, mat_size, mat_mic)
                w_rate_nosize = lookup_film_rate_weighted(pr_df, mrr_qty.copy(), mat_name, None, mat_mic)
                month_rate = lookup_material_rate_for_month(pr_df, mat_name, mat_mic, "2-2026")
                
                report.append(f"\n    Engine calculations:")
                report.append(f"      Filtered MRRs (by PR):     {filtered}")
                report.append(f"      Weighted rate (with size):  {w_rate}")
                report.append(f"      Weighted rate (no size):    {w_rate_nosize}")
                report.append(f"      Month avg rate (2-2026):    {month_rate}")
                
                # 5. How was ground truth likely calculated?
                if gt_rate and gt_mr:
                    gt_mrrs = str(gt_mr).split('/')
                    report.append(f"\n    Ground truth MRR(s): {gt_mrrs}")
                    gt_total_rate = 0
                    gt_total_qty = 0
                    for gm in gt_mrrs:
                        gm = gm.strip()
                        try:
                            gm_int = int(float(gm))
                            gm_qty = mrr_qty.get(gm_int, 0)
                            gm_rate = lookup_film_rate_weighted(pr_df, {gm_int: gm_qty if gm_qty > 0 else 1}, mat_name, mat_size, mat_mic)
                            report.append(f"      GT MRR {gm}: PR Rate={gm_rate}, Qty={gm_qty}")
                            if gm_rate > 0:
                                gt_total_rate += gm_rate * gm_qty
                                gt_total_qty += gm_qty
                        except:
                            pass
                    if gt_total_qty > 0:
                        gt_computed = gt_total_rate / gt_total_qty
                        report.append(f"      GT weighted avg: {gt_computed:.6f}")
                        report.append(f"      GT actual rate:  {gt_rate}")
                        if abs(gt_computed - float(gt_rate)) < 0.01:
                            report.append(f"      --> Ground truth matches weighted avg of its own MRRs")
                        else:
                            report.append(f"      --> Ground truth does NOT match weighted avg of its own MRRs")
                            report.append(f"      --> Possible: manual selection or different calculation method")

wb_gt.close()
wb_out.close()

# Write report
output = "\n".join(report)
print(output)

with open("row_investigation_42_46.txt", "w", encoding="utf-8") as f:
    f.write(output)
print(f"\nReport saved to: row_investigation_42_46.txt")
